# Groww_CP_Bot.py


############### MUST READ #####################
#• To start Bot we just need to assure that config data is proper,
# Need to remove comments from "place_cp_order" method to validate the order status under comment starts from "#STATUS VALIDATION",
# And in last we need to validate that funds are matching with BOT and Groww Account unnder comment "Funds check"

#DOWNLOAD GROWW INSTRUMENTS:- https://growwapi-assets.groww.in/instruments/instrument.csv

import os
import re
import json
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import pyotp
from openpyxl import Workbook, load_workbook
from playsound3 import playsound
from datetime import datetime
from threading import Lock
import requests
import sys
from datetime import datetime
import time
import os
import sys
from datetime import datetime

MOMENTUM_SAMPLES = 5
MOMENTUM_DELAY = 1

def setup_persistent_logger():
    """Creates a local 'logs' folder beside the script and logs all console output there."""
    # Create /logs folder in the same directory as the script
    base_dir = os.path.dirname(os.path.abspath(__file__))
    log_dir = os.path.join(base_dir, "logs")
    os.makedirs(log_dir, exist_ok=True)

    # Create a timestamped log file
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_path = os.path.join(log_dir, f"Groww_Bot_{timestamp}.log")

    # Define a Tee class to write to both console and log file
    class Tee:
        def __init__(self, *streams):
            self.streams = streams

        def write(self, data):
            for s in self.streams:
                try:
                    s.write(data)
                    s.flush()
                except Exception:
                    pass  # Ignore on shutdown

        def flush(self):
            for s in self.streams:
                try:
                    s.flush()
                except Exception:
                    pass

    # Open log file (unbuffered, UTF-8)
    logfile = open(log_path, "a", buffering=1, encoding="utf-8")

    # Redirect both stdout and stderr
    sys.stdout = Tee(sys.stdout, logfile)
    sys.stderr = Tee(sys.stderr, logfile)

    print(f"📝 Logging started. Log file: {log_path}")

    return log_path


# --- Initialize persistent logging ---
LOG_FILE_PATH = setup_persistent_logger()

# Replace with your Groww API key (or leave and use TOTP to fetch access_token)
api_key = "eyJraWQiOiJaTUtjVXciLCJhbGciOiJFUzI1NiJ9.eyJleHAiOjI1NTAwNDY3MzksImlhdCI6MTc2MTY0NjczOSwibmJmIjoxNzYxNjQ2NzM5LCJzdWIiOiJ7XCJ0b2tlblJlZklkXCI6XCI2MmEwMTc4YS0wOTk3LTQ0ZDAtOWRiNC0wZDAzOWM5MzY3YmZcIixcInZlbmRvckludGVncmF0aW9uS2V5XCI6XCJlMzFmZjIzYjA4NmI0MDZjODg3NGIyZjZkODQ5NTMxM1wiLFwidXNlckFjY291bnRJZFwiOlwiMmVlMjYyMjItN2MwNS00Y2IwLWIwM2MtNzAzYWRmNWVmN2RkXCIsXCJkZXZpY2VJZFwiOlwiNWQwYzdjODgtMGI1OS01MDU0LTk5ZTYtYWU5MzY5OTc2ZmRiXCIsXCJzZXNzaW9uSWRcIjpcIjY1NzBiNDUwLWE2YzYtNDMyYi1hYTJmLTA4MjExZjk0YzRiOVwiLFwiYWRkaXRpb25hbERhdGFcIjpcIno1NC9NZzltdjE2WXdmb0gvS0EwYktvMDZXRlpjc241VUNmTWF5aERtNGxSTkczdTlLa2pWZDNoWjU1ZStNZERhWXBOVi9UOUxIRmtQejFFQisybTdRPT1cIixcInJvbGVcIjpcImF1dGgtdG90cFwiLFwic291cmNlSXBBZGRyZXNzXCI6XCIxNzEuNjAuMTY5LjI1MiwxNzIuNjkuOTUuOTMsMzUuMjQxLjIzLjEyM1wiLFwidHdvRmFFeHBpcnlUc1wiOjI1NTAwNDY3Mzk5MTV9IiwiaXNzIjoiYXBleC1hdXRoLXByb2QtYXBwIn0.EKERC7OzG-lblhaOSQPyb44mafdNFpErGbcELiTiLnRk4WEW9p7aBBf6iq-3LGagY4ORdOCnrXbRhyGzbscxSw"
totp_gen = pyotp.TOTP('WI4M7KCAMH5CGN2I6SVB6MN2QDKUXRJF')

# Get project root directory (folder where your script is running)
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
# Build CSV path dynamically
csv_path = os.path.join(PROJECT_ROOT, "instrument.csv")
print(csv_path)

# Instruments CSV/JSON path (script will convert CSV -> JSON if convert_csv_to_json = yes)
# csv_path = r"C:\Users\HITS\Downloads\instrument (1).csv"
convert_csv_to_json = "yes"

# Telegram placeholders (you will replace later)
TELEGRAM_BOT_TOKEN = "PUT_YOUR_TOKEN_HERE"
TELEGRAM_CHAT_ID = "PUT_YOUR_CHAT_ID_HERE"

# Sound files (ensure these exist in script folder or provide full path)
SOUND_PROFIT = "coin.mp3"
SOUND_SL = "SL_HIT.mp3"

# Trade defaults for Groww
DEFAULT_PRODUCT = "MIS"   # intraday; change to "NRML" if you want positional
ORDER_PRODUCT_MAP = {
    "MIS": "MIS",
    "NRML": "NRML"
}
# NOTE: the growwapi wrapper constants are used from the growwapi instance below

# ----------------- import growwapi late (after auth) -----------------
try:
    from growwapi import GrowwAPI
except Exception:
    # If local module not available, user must install or place it in PYTHONPATH
    print("❗ growwapi module not found. Make sure it's installed and importable.")
    # continue; import errors will show when script runs further

# ----------------- Groww auth & wrapper -----------------
def groww_init(api_key):
    """
    Return growwapi client instance (GrowwAPI(access_token))
    This function gets access_token using GrowwAPI.get_access_token if available.
    """
    totp = totp_gen.now()
    try:
        access_token = GrowwAPI.get_access_token(api_key=api_key, totp=totp)
        client = GrowwAPI(access_token)
        print(access_token)
        print("✅ Groww API Initialized Successfully")
        return client, access_token
    except Exception as e:
        print(f"❌ Groww login failed: {e}")
        raise

# Init groww client
groww ,access_token = groww_init(api_key)


# ----------------- Utilities: Telegram, Sound, Excel Logging -----------------

# === TELEGRAM CONFIG ===
BOT_TOKEN = "8226223419:AAGX5fKG21CfceF_0_WjPIrOMx6ON17pZMw"
CHAT_ID = "6012308856"

def send_telegram(message: str):
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        payload = {"chat_id": CHAT_ID, "text": message}
        requests.post(url, data=payload)
    except Exception as e:
        print(f"⚠️ Telegram Error: {e}")

def play_sound_async(filename):
    try:
        if not os.path.exists(filename):
            print(f"🔇 Sound file not found: {filename}")
            return
        threading.Thread(target=playsound, args=(filename,), daemon=True).start()
    except Exception as e:
        print(f"🔇 Sound error: {e}")

def log_trade_to_excel(symbol, buy_price, sell_price, quantity, profit):
    file_name = "Lakshmi.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append(["DateTime", "Symbol", "Buy Price", "Sell Price", "Quantity", "Profit"])
        wb.save(file_name)

    # Load existing workbook
    wb = load_workbook(file_name)
    ws = wb.active

    # Find the next empty row
    next_row = ws.max_row + 1
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.cell(row=next_row, column=1).value = now
    ws.cell(row=next_row, column=2).value = symbol
    ws.cell(row=next_row, column=3).value = buy_price
    ws.cell(row=next_row, column=4).value = sell_price
    ws.cell(row=next_row, column=5).value = quantity
    ws.cell(row=next_row, column=6).value = round(profit, 2)
    wb.save(file_name)


# ----------------- CSV -> JSON loader -----------------
def csv_to_json(csv_file_path, json_file_path=None):
    import csv
    if json_file_path is None:
        json_file_path = os.path.splitext(csv_file_path)[0] + ".json"
    data = []
    with open(csv_file_path, encoding='utf-8') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            data.append(row)
    with open(json_file_path, 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file, indent=4, ensure_ascii=False)
    print(f"✅ Converted '{csv_file_path}' → '{json_file_path}'")
    return data

ltp_lock = threading.Lock()

def get_ltp_for_instrument(instrument, access_token, verbose=True,segment = "FNO"):
    """
    Fetches the latest traded price (LTP) for a given F&O instrument using Groww's authenticated API.
    Thread-safe with a global lock to prevent too-frequent API calls.
    """

    try:
        trading_symbol = instrument.get("trading_symbol")  # e.g. NIFTY25N0425950CE
        if not trading_symbol:
            print("⚠️ Missing trading_symbol in instrument.")
            return None

        exchange_symbol = f"NSE_{trading_symbol}"
        url = f"https://api.groww.in/v1/live-data/ltp?segment={segment}&exchange_symbols={exchange_symbol}"
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0"
        }

        # 🔒 Lock ensures one API call at a time
        with ltp_lock:
            resp = requests.get(url, headers=headers, timeout=10)
            time.sleep(0.5)  # ⏳ short delay to respect Groww API rate limits

        if resp.status_code != 200:
            print(f"⚠️ HTTP {resp.status_code} error fetching LTP: {resp.text}")
            return None

        data = resp.json()
        payload = data.get("payload", {})
        ltp = payload.get(exchange_symbol)

        if ltp is None:
            print(f"⚠️ No LTP found for {exchange_symbol} in payload: {payload}")
            return None
        if verbose:
            print(f"💰 LTP for {exchange_symbol}: ₹{ltp} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            send_telegram(f"💰 LTP for {exchange_symbol}: ₹{ltp} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
        return float(ltp)

    except Exception as e:
        print(f"⚠️ Error fetching LTP for {instrument.get('trading_symbol')}: {e}")
        return None

def get_nifty_spot_price(access_token=None,json_path=None):
    """
    Fetches live NIFTY 50 spot price using Groww instrument data.
    Matches by trading_symbol = 'NIFTY' or groww_symbol = 'NSE-NIFTY'
    """
    global instruments1

    if json_path is None:
        json_path = os.path.splitext(csv_path)[0] + ".json"

    # 🔄 Step 1: Load or convert JSON
    if convert_csv_to_json.lower() == "yes":
        instruments1 = csv_to_json(csv_path, json_path)
    else:
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON not found: {json_path}")
        with open(json_path, "r", encoding="utf-8") as jf:
            instruments1 = json.load(jf)
        print(f"ℹ️ Loaded instruments from existing JSON: {json_path}")

    try:
        nifty_spot_instrument = next(
            (item for item in instruments1
             if item.get("trading_symbol") == "NIFTY"
             or item.get("groww_symbol") == "NSE-NIFTY"
             or item.get("name") == "NIFTY 50"),
            None
        )

        if not nifty_spot_instrument:
            print("⚠️ NIFTY spot instrument not found in instruments1")
            return 0

        spot = get_ltp_for_instrument(nifty_spot_instrument, access_token, verbose=False,segment = "CASH")
        if spot:
            print(f"📊 Live NIFTY Spot: {spot}")
            return float(spot)
        else:
            print("⚠️ Failed to fetch LTP for NIFTY spot")
            return 0
    except Exception as e:
        print(f"⚠️ Error fetching NIFTY spot: {e}")
        return 0


CONFIG = {
    "index": "NIFTY",
    "expiry": "02/12/2025",  #this needs to be same as expiry_date in json file of instruments # format DD/MM/YYYY to match instruments JSON (example)
    "min_premium": 80,
    "max_premium": 130,
    "lots": 3,
    "book_profit": 300,
    "target_pnl": 6000,
    "spot":get_nifty_spot_price(access_token)
}

# Load instruments_data
def load_instruments_from_json(json_path=None):
    """
    Loads instruments from JSON (or CSV → JSON if convert_csv_to_json = 'yes'),
    but only keeps instruments:
      - matching expiry from CONFIG
      - within ±10 strikes of current index spot price
    """
    global instruments
    config = CONFIG
    INDEX = config["index"].upper()
    EXPIRY = config["expiry"].strip()

    if json_path is None:
        json_path = os.path.splitext(csv_path)[0] + ".json"

    # 🔄 Step 1: Load or convert JSON
    if convert_csv_to_json.lower() == "yes":
        instruments = csv_to_json(csv_path, json_path)
    else:
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON not found: {json_path}")
        with open(json_path, "r", encoding="utf-8") as jf:
            instruments = json.load(jf)
        print(f"ℹ️ Loaded instruments from existing JSON: {json_path}")

    # 🧩 Step 2: Get live index spot
    spot = config["spot"]

    # Determine strike step (e.g., NIFTY = 50, BANKNIFTY = 100)
    step = 100 if "BANK" in INDEX else 50

    # Define strike range (±10 strikes)
    nearest_strike = round(spot / step) * step
    lower_bound = nearest_strike - (10 * step)
    upper_bound = nearest_strike + (10 * step)

    print(f"🎯 Filtering {INDEX} {EXPIRY} instruments between {lower_bound}–{upper_bound} (Spot={spot})")

    # 🧹 Step 3: Filter instruments by expiry and strike range
    filtered = []
    for item in instruments:
        try:
            if item.get("underlying_symbol", "").upper() != INDEX:
                continue
            # if item.get("expiry_date", "").strip() != EXPIRY:
            #     continue
            strike = float(item.get("strike_price") or 0)
            if lower_bound <= strike <= upper_bound:
                filtered.append(item)
        except Exception:
            continue

    print(f"✅ Loaded {len(filtered)} filtered instruments (out of {len(instruments)})")
    instruments = filtered
    return instruments


# initialize
instruments_data = load_instruments_from_json()


# Pre-index instruments for quick lookup by internal_trading_symbol or groww_symbol or custom compact symbol
symbol_index = {}
for it in instruments_data:
    # keys: internal_trading_symbol, groww_symbol, compact like NIFTY04NOV2525950CE approximate
    try:
        k1 = it.get("internal_trading_symbol", "") or it.get("trading_symbol", "")
        k2 = it.get("groww_symbol", "")
        if k1:
            symbol_index[k1.upper()] = it
        if k2:
            symbol_index[k2.upper()] = it
    except Exception:
        pass

# ----------------- Helpers: date/expiry normalization -----------------
MONTHS = {
    'JAN': '01','FEB':'02','MAR':'03','APR':'04','MAY':'05','JUN':'06',
    'JUL':'07','AUG':'08','SEP':'09','OCT':'10','NOV':'11','DEC':'12'
}

def cmd_expiry_to_date(expiry_token):
    """
    expiry_token example: 04NOV25 or 04NOV2025 or 28AUG25 or 28AUG2025
    Return string 'DD/MM/YYYY'
    """
    m = re.match(r'(\d{1,2})([A-Z]{3})(\d{2,4})', expiry_token.upper())
    if not m:
        return None
    dd = m.group(1).zfill(2)
    mon_abbr = m.group(2)
    yy = m.group(3)
    if len(yy) == 2:
        yyyy = "20" + yy
    else:
        yyyy = yy
    mm = MONTHS.get(mon_abbr[:3], None)
    if not mm:
        return None
    return f"{dd}/{mm}/{yyyy}"

# ----------------- Command parser -----------------
def parse_cp_command(command):
    """
    Parse strings like:
      Buy 14 NIFTY04NOV2525950CE at CP and Book at 1050
    Returns dict or None
    """
    pattern = r'(?i)^\s*(Buy|Sell)\s+(\d+)\s+([A-Z]+)(\d{1,2}[A-Z]{3}\d{2,4})(\d+)(CE|PE)\s+at\s+CP\s+and\s+Book\s+at\s+(\d+(\.\d+)?)\s*$'
    m = re.match(pattern, command.strip())
    if not m:
        return None
    action = m.group(1).upper()
    lots = int(m.group(2))
    underlying = m.group(3).upper()
    expiry_token = m.group(4).upper()
    strike = m.group(5)
    opt_type = m.group(6).upper()
    target_profit = float(m.group(7))
    expiry_date = cmd_expiry_to_date(expiry_token)
    return {
        "action": action,
        "lots": lots,
        "underlying": underlying,
        "expiry_token": expiry_token,
        "expiry_date": expiry_date,
        "strike": strike,
        "opt_type": opt_type,
        "target_profit": target_profit
    }

# ----------------- Find instrument in instruments_data -----------------
def find_instrument_from_command(command: str, instruments: list):
    import re
    # Example command: Buy 14 NIFTY04NOV2525950CE at CP and Book at 1050
    pattern = r'([A-Z]+)(\d{2})([A-Z]{3})(\d{2})(\d+)(CE|PE)'
    match = re.search(pattern, command.upper())
    if not match:
        print("❌ Could not parse symbol from command.")
        return None

    underlying, day, mon, yr, strike, opt_type = match.groups()
    expiry_date = f"{day}/{mon_to_number(mon)}/20{yr}"

    # Find match in JSON
    for inst in instruments:
        if (
            inst["underlying_symbol"].upper() == underlying
            and inst["expiry_date"] == expiry_date
            and inst["strike_price"] == strike
            and inst["instrument_type"].upper() == opt_type
        ):
            return inst

    print("❌ Instrument not found in instrument master.")
    return None


def mon_to_number(mon: str):
    mapping = {
        "JAN": "01", "FEB": "02", "MAR": "03", "APR": "04",
        "MAY": "05", "JUN": "06", "JUL": "07", "AUG": "08",
        "SEP": "09", "OCT": "10", "NOV": "11", "DEC": "12"
    }
    return mapping.get(mon.upper(), "00")


import requests, time

import requests
import json

import requests

def get_order_status(order_id, access_token):
    """
    Fetch the status of a Groww order (CASH, F&O, etc.)
    Works with official Groww REST API response format.
    """
    url = f"https://api.groww.in/v1/order/status/{order_id}?segment=FNO"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0"
    }

    try:
        resp = requests.get(url, headers=headers, timeout=8)
        resp.raise_for_status()  # raises for non-200 responses

        data = resp.json()  # Proper JSON response from Groww
        print("🔍 Order status response:", data)

        # ✅ Extract status cleanly
        payload = data.get("payload", {})
        status = payload.get("order_status")

        return status

    except requests.exceptions.JSONDecodeError:
        print("⚠️ Error: Non-JSON response received.")
        print("Response text:", resp.text)
        return None

    except Exception as e:
        print(f"⚠️ Error fetching order status: {e}")
        return None


import time
from datetime import datetime, timedelta, timezone

def get_recent_market_direction(symbol, groww):
    """
    Returns 'CE' if recent 5-min direction is upward (bullish),
    'PE' if downward (bearish), or None if uncertain.
    Also prints the equivalent cURL command.
    """
    try:
        # Current time and 5 minutes earlier
        end_time = datetime.now()
        start_time = end_time - timedelta(minutes=3)

        # Convert to string format accepted by Groww API
        end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        start_time_str = start_time.strftime("%Y-%m-%d %H:%M:%S")

        # Construct the Groww candle API URL
        url = (
            f"https://api.groww.in/v1/historical/candles?"
            f"exchange=NSE&segment=FNO&groww_symbol={symbol}"
            f"&start_time={start_time_str}"
            f"&end_time={end_time_str}"
            f"&candle_interval=1minute"
        )

        # Print cURL command for debugging
        print("\n🌀 Generated cURL for Groww Candle API:")
        print(f"curl --location '{url}' \\")
        print("  --header 'Accept: application/json' \\")
        print(f"  --header 'Authorization: Bearer {access_token}' \\")
        print("  --header 'X-API-VERSION: 1.0'\n")

        # Fetch last 5-min candle via Groww SDK
        historical = groww.get_historical_candles(
            groww_symbol=symbol,
            exchange=groww.EXCHANGE_NSE,
            segment=groww.SEGMENT_FNO,
            start_time=start_time_str,
            end_time=end_time_str,
            candle_interval="1minute" # 1-min candles for better precision
        )

        candles = historical.get("candles", [])
        if not candles:
            print("⚠️ No recent candle data found.")
            return None

        first_open = candles[0][1]
        last_close = historical.get("closing_price")

        if "PE" in symbol:
            direction = "PE" if last_close > first_open else "CE"
        else:  # CE symbol
            direction = "CE" if last_close > first_open else "PE"

        print(f"📊 3-min candle trend → {direction} (O1={first_open}, C3={last_close})")
        return direction

    except Exception as e:
        print("⚠️ Error fetching recent market direction:", e)
        return None


# ----------------- Place orders with Groww -----------------
def place_market_order_groww(instrument, quantity, transaction_type="BUY", product="MIS"):
    """
    place market order via growwapi wrapper. Returns order response or raises.
    """
    trading_symbol = instrument.get("internal_trading_symbol") or instrument.get("trading_symbol")
    try:
        order = groww.place_order(
            trading_symbol=trading_symbol,
            quantity=quantity,
            validity=groww.VALIDITY_DAY,
            exchange=groww.EXCHANGE_NSE,
            segment=groww.SEGMENT_FNO,
            product=getattr(groww, f"PRODUCT_{product}") if hasattr(groww, f"PRODUCT_{product}") else getattr(groww, "PRODUCT_MIS", product),
            order_type=groww.ORDER_TYPE_MARKET,
            transaction_type=getattr(groww, f"TRANSACTION_TYPE_{transaction_type}"),
            price=0
        )
        return order
    except Exception as e:
        raise

def place_limit_order_groww(instrument, quantity, price, transaction_type="SELL", product="MIS"):
    trading_symbol = instrument.get("internal_trading_symbol") or instrument.get("trading_symbol")
    try:
        order = groww.place_order(
            trading_symbol=trading_symbol,
            quantity=quantity,
            validity=groww.VALIDITY_DAY,
            exchange=groww.EXCHANGE_NSE,
            segment=groww.SEGMENT_FNO,
            product=getattr(groww, f"PRODUCT_{product}") if hasattr(groww, f"PRODUCT_{product}") else getattr(groww, "PRODUCT_MIS", product),
            order_type=groww.ORDER_TYPE_LIMIT,
            transaction_type=getattr(groww, f"TRANSACTION_TYPE_{transaction_type}"),
            price=price
        )
        return order
    except Exception as e:
        raise

# ----------------- Rounding for limits (5 paise) -----------------
def round_to_nearest_5_paise(price):
    # Round to nearest 0.05
    return round(round(price * 20) / 20, 2)

# ----------------- Momentum sampling -----------------
import numpy as np
import time

def momentum_check_for_symbol(instrument, MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY, threshold=0.25):
    """
    Improved short-term momentum check for Nifty options.
    - Uses multiple intermediate samples
    - Smooths noise
    - Checks direction consistency
    - Returns a cleaner momentum signal
    """
    trading_symbol = instrument.get("trading_symbol")
    prices = []

    print(f"\n🧭 Checking momentum for {trading_symbol} ({MOMENTUM_SAMPLES} samples, every {MOMENTUM_DELAY}s):")

    for i in range(MOMENTUM_SAMPLES):
        p = get_ltp_for_instrument(instrument, access_token, verbose=False)
        if p:
            price = float(p)
            prices.append(price)
            print(f"[{trading_symbol}] ⏱ Sample {i+1}/{MOMENTUM_SAMPLES}: LTP = ₹{price:.2f}")
        else:
            print(f"[{trading_symbol}] ⚠️ Sample {i+1}/{MOMENTUM_SAMPLES}: Failed to fetch LTP")
        time.sleep(MOMENTUM_DELAY)

    if len(prices) < 3:
        print(f"[{trading_symbol}] ❌ Not enough data ({len(prices)} samples)")
        return None, len(prices)

    prices = np.array(prices)

    # 1️⃣ Smooth noise with small moving average
    smooth = np.convolve(prices, np.ones(3)/3, mode="valid")

    # 2️⃣ Compute rate of change (%)
    roc = np.diff(smooth) / smooth[:-1] * 100

    # 3️⃣ Remove outliers (big spikes)
    median = np.median(roc)
    std = np.std(roc)
    filtered = roc[(roc > median - 1.5*std) & (roc < median + 1.5*std)]

    if len(filtered) < 2:
        print(f"[{trading_symbol}] ⚠️ Too noisy for reliable momentum reading")
        return None, len(prices)

    # 4️⃣ Average change and direction consistency
    avg_change = np.mean(filtered)
    direction_signs = np.sign(filtered)
    consistency = np.mean(direction_signs == np.sign(avg_change)) * 100

    # 5️⃣ Decision
    if avg_change > threshold and consistency > 70:
        direction = "UP"
    elif avg_change < -threshold and consistency > 70:
        direction = "DOWN"
    else:
        direction = "FLAT"

    print(f"[{trading_symbol}] 📊 Avg Δ = {avg_change:.3f}%, Consistency = {consistency:.1f}% → {direction}")
    print(f"[{trading_symbol}] 📈 Range ₹{prices[0]:.2f} → ₹{prices[-1]:.2f}\n")

    return {"symbol": trading_symbol,
            "avg_change": round(avg_change, 3),
            "consistency": round(consistency, 1),
            "direction": direction}, len(prices)



# ----------------- Find option by premium (parallel) -----------------

def find_option_by_premium_parallel(option_type, min_premium, max_premium,
                                    lots=1, funds_buffer=0.9, momentum_threshold_pct=0.25,
                                    MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY):
    """
    Filters instruments using INDEX and EXPIRY from config,
    matches by option_type, filters by premium range,
    and runs momentum checks in parallel.
    Returns: (instrument, ltp, lot_size) or (None, None, None)
    """
    config = CONFIG
    INDEX = config["index"].upper()
    EXPIRY = config["expiry"].strip()
    candidates = []

    # 🔍 Filter based on index + expiry + type
    for item in instruments_data:
        try:
            if item.get("underlying_symbol", "").upper() != INDEX:
                continue
            if item.get("instrument_type", "").upper() != option_type.upper():
                continue
            if item.get("expiry_date", "").strip() != EXPIRY:
                continue

            lot_size = int(item.get("lot_size") or item.get("lotsize") or 1)
            ltp = get_ltp_for_instrument(item, access_token, verbose=False)
            if ltp is None:
                continue
            if not (min_premium <= ltp <= max_premium):
                continue

            candidates.append({
                "instrument": item,
                "ltp": float(ltp),
                "lot_size": lot_size
            })

        except Exception as e:
            print(f"⚠️ Error while scanning: {e}")
            continue

    if not candidates:
        print(f"⚠️ No instruments for {INDEX} {EXPIRY} {option_type}")
        return None, None, None

    # ✅ Funds check (fallback = 1.2L if not available)
    try:
        margins = getattr(groww, "get_margins", lambda: {"availablecash": 130000})()
        available_cash = float(margins.get("availablecash", 130000))
    except Exception:
        available_cash = 130000

    affordable = []
    for c in candidates:
        qty = lots * c["lot_size"]
        est_cost = c["ltp"] * qty
        if available_cash <= 0 or est_cost <= available_cash * funds_buffer:
            affordable.append(c)

    if not affordable:
        print(f"⚠️ No affordable instruments for {INDEX} {EXPIRY} {option_type}")
        return None, None, None

    if option_type.upper() == "PE":
        momentum_threshold_pct = 0.30  # PEs move sharper
    else:
        momentum_threshold_pct = 0.25

    # ✅ Sort candidates closest to mid-premium
    mid = (min_premium + max_premium) / 2.0
    affordable.sort(key=lambda x: abs(x["ltp"] - mid))
    probe_list = affordable[:12]

    momentum_results = []

    def check_momentum(cand):
        mom_result, ticks = momentum_check_for_symbol(
            cand["instrument"],
            MOMENTUM_SAMPLES=MOMENTUM_SAMPLES,
            MOMENTUM_DELAY=MOMENTUM_DELAY
        )
        if mom_result and isinstance(mom_result, dict):
            slope_pct = mom_result.get("avg_change", 0)
            direction = mom_result.get("direction", "FLAT")
            consistency = mom_result.get("consistency", 0)

            # ✅ Apply momentum filter right here
            if abs(slope_pct) >= momentum_threshold_pct and consistency >= 70 and direction != "FLAT":
                return {
                    "instrument": cand["instrument"],
                    "ltp": cand["ltp"],
                    "lot_size": cand["lot_size"],
                    "slope_pct": slope_pct,
                    "direction": direction,
                    "consistency": consistency,
                    "ticks": ticks
                }
        return None

    print(f"⚙️ Checking momentum for top {len(probe_list)} {option_type} candidates...")

    with ThreadPoolExecutor(max_workers=min(len(probe_list), 8)) as executor:
        futures = {executor.submit(check_momentum, c): c for c in probe_list}
        for future in as_completed(futures):
            res = future.result()
            if res:
                momentum_results.append(res)

    if not momentum_results:
        print(f"⚠️ No strong momentum found for {option_type} (>{momentum_threshold_pct}%, consistency >70%)")
        # fallback: pick the one closest to mid-premium
        pick = probe_list[0]
        return pick["instrument"], pick["ltp"], pick["lot_size"]

    # ✅ Rank: strongest slope first, then consistency
    momentum_results.sort(key=lambda x: (x["slope_pct"], x["consistency"]), reverse=True)
    pick = momentum_results[0]

    print(f"🏆 Selected {option_type}: {pick['instrument']['trading_symbol']} "
          f"({pick['direction']} | {pick['slope_pct']:.2f}% | Consistency {pick['consistency']}%)")

    return pick["instrument"], pick["ltp"], pick["lot_size"]


# ----------------- Detect CE/PE (parallel) -----------------
def detect_option_type_parallel(index, expiry, min_p, max_p, lots, funds_buffer=0.9):
    print(f"🔍 Detecting best option between CE and PE for {index} {expiry}…")

    def worker(opt_type):
        print(f"➡️  Searching {opt_type} between {min_p}-{max_p}")
        inst, ltp, lot_size = find_option_by_premium_parallel(opt_type, min_p, max_p, lots, funds_buffer)
        mom = None
        if inst:
            print(f"📊 Running momentum check for {opt_type} ({inst.get('trading_symbol')})")
            mom, _ = momentum_check_for_symbol(inst, MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY)
            print(f"✅ Momentum for {opt_type}: {mom}")
        else:
            print(f"⚠️ No instrument found for {opt_type}")
        return opt_type, inst, ltp, lot_size, mom

    results = {}
    with ThreadPoolExecutor(max_workers=2) as ex:
        futures = {ex.submit(worker, t): t for t in ["CE", "PE"]}
        for future in as_completed(futures):
            opt_type, inst, ltp, lot_size, mom = future.result()
            results[opt_type] = {"instrument": inst, "ltp": ltp, "lot_size": lot_size, "momentum": mom}
            print(f"🧩 Finished {opt_type}: {inst.get('trading_symbol') if inst else 'None'}, momentum={mom}")

    print("🧮 Comparing CE vs PE momentum...")
    ce_mom = results.get("CE", {}).get("momentum")
    pe_mom = results.get("PE", {}).get("momentum")

    # Handle missing momentum
    if not ce_mom and not pe_mom:
        print("❌ No momentum data found for CE or PE.")
        return None
    if not ce_mom:
        r = results["PE"]
        return "PE", r["instrument"], r["ltp"], r["lot_size"]
    if not pe_mom:
        r = results["CE"]
        return "CE", r["instrument"], r["ltp"], r["lot_size"]

    ce_val = ce_mom["avg_change"]
    pe_val = pe_mom["avg_change"]

    print(f"📈 CE momentum: {ce_val:.3f}% ({ce_mom['direction']}, {ce_mom['consistency']}%)")
    print(f"📉 PE momentum: {pe_val:.3f}% ({pe_mom['direction']}, {pe_mom['consistency']}%)")

    # selection logic
    if abs(ce_val - pe_val) >= 0.25 and ce_val > pe_val and ce_val >= 0.10:
        print("✅ Selected CE (stronger momentum)")
        r = results["CE"]
        return "CE", r["instrument"], r["ltp"], r["lot_size"]
    if abs(pe_val - ce_val) >= 0.25 and pe_val > ce_val and pe_val >= 0.10:
        print("✅ Selected PE (stronger momentum)")
        r = results["PE"]
        return "PE", r["instrument"], r["ltp"], r["lot_size"]

    # fallback
    if ce_val >= pe_val:
        print("⚖️  Momentum similar — choosing CE fallback")
        r = results["CE"]
        return "CE", r["instrument"], r["ltp"], r["lot_size"]
    else:
        print("⚖️  Momentum similar — choosing PE fallback")
        r = results["PE"]
        return "PE", r["instrument"], r["ltp"], r["lot_size"]



def wait_for_order_status(order_id, access_token, order_type="BUY"):
    """
    Wait indefinitely until a Groww order reaches EXECUTED / COMPLETED / DELIVERY_AWAITED.
    Returns final status (string).
    """
    print(f"🔎 Waiting for {order_type} order ({order_id}) to finish...")

    while True:
        status = get_order_status(order_id, access_token)
        print(f"🕒 {order_type} status: {status}")

        if status in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
            print(f"✅ {order_type} order executed successfully.")
            send_telegram(f"✅ {order_type} order executed successfully.")
            return status

        elif status in ["FAILED", "REJECTED", "CANCELLED"]:
            print(f"❌ {order_type} order failed with status {status}.")
            send_telegram(f"❌ {order_type} order failed ({status}).")
            return status

        # wait before next check (adjust if needed)
        time.sleep(2)


import requests

def get_order_executed_price(order_id, access_token, segment="FNO"):
    """
    Fetch executed trades for a given Groww order_id and return average price & total quantity.
    """
    try:
        url = f"https://api.groww.in/v1/order/trades/{order_id}?segment={segment}&page=0&page_size=50"
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0"
        }

        print(f"\n📦 Fetching trade details for order: {order_id}")
        response = requests.get(url, headers=headers)
        data = response.json()

        if data.get("status") != "SUCCESS":
            print("⚠️ Failed to fetch trade info:", data)
            return None, None

        trades = data.get("payload", {}).get("trade_list", [])
        if not trades:
            print("⚠️ No trades found for order ID.")
            return None, None

        # Compute average price & total quantity
        total_qty = sum(t["quantity"] for t in trades)
        total_value = sum(t["price"] * t["quantity"] for t in trades)
        avg_price = round(total_value / total_qty, 2)

        symbol = trades[0]["trading_symbol"]
        side = trades[0]["transaction_type"]

        print(f"✅ {side} {symbol} | Total Qty={total_qty} | Avg Price=₹{avg_price}")
        return avg_price, total_qty

    except Exception as e:
        print("❌ Error fetching order trades:", e)
        return None, None



# ----------------- Place CP order workflow (mirrors AngelOne logic) -----------------
def place_cp_order(command, is_auto=False):
    global buy_status
    if is_auto:
        order = command  # dict form
        symbol = order["symbol"]
        qty = order["lots"] * order["lot_size"]
        book_profit = order["book_profit"]

        # get instrument info directly from master
        instrument = next((inst for inst in instruments_data if inst["internal_trading_symbol"] == symbol), None)
        if not instrument:
            print(f"❌ Instrument {symbol} not found in master.")
            return

        print(f"🔹 Auto order => {symbol}, qty={qty}, book@{book_profit} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")

        # --- Fetch LTP ---
        ltp_before = get_ltp_for_instrument(instrument, access_token)
        if ltp_before is None:
            print("❌ Could not fetch LTP before placing order.")
            return

        entry_price = round(float(ltp_before), 2)


        # === BUY @ MARKET ===
        try:
            order_resp = place_market_order_groww(instrument, qty, transaction_type="BUY", product="MIS")
            order_id = order_resp.get("payload", {}).get("groww_order_id") or order_resp.get("groww_order_id")
            print(f"✅ Auto Buy placed: :{order_resp} ======= [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            send_telegram(f"✅ Auto Buy placed: :{order_resp} ======= [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
        except Exception as e:
            print(f"❌ Auto BUY failed: {e}")
            send_telegram(f"❌ Auto BUY failed: {e}")
            return

        #STATUS VALIDATION
        # --- Wait until BUY order is EXECUTED or COMPLETED ---
        if order_id:
            buy_status = wait_for_order_status(order_id, access_token, "BUY")
            if buy_status not in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                print(f"⚠️ Skipping SELL due to BUY status: {buy_status}")
                return

        avg_price, qty = get_order_executed_price(order_id, access_token)
        print(f"\n🎯 Executed avg price: ₹{avg_price}, Total Qty: {qty}")

        target_price = round_to_nearest_5_paise(avg_price + book_profit / qty)

        # === SELL @ LIMIT ===
        try:
            sell_resp = place_limit_order_groww(instrument, qty, price=target_price, transaction_type="SELL",
                                                product="MIS")
            sell_order_id = (
                    sell_resp.get("payload", {}).get("groww_order_id")
                    or sell_resp.get("groww_order_id")
            )
            print(f"✅ Auto Target SELL @ {target_price} placed: {sell_resp} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            send_telegram(f"✅ Auto Target SELL @ {target_price} placed: {sell_resp} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
        except Exception as e:
            print(f"❌ Auto SELL failed: {e}")
            send_telegram(f"❌ Auto SELL failed: {e}")
            return

        # STATUS VALIDATION
        # --- Wait until SELL order EXECUTED or COMPLETED ---
        if sell_order_id:
            sell_status = wait_for_order_status(sell_order_id, access_token, "SELL")
            if sell_status in ["EXECUTED", "COMPLETED"]:
                print(f"💰 Target HIT @ {target_price}")
                send_telegram(f"💰 Target HIT @ {target_price}")
                print(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
                send_telegram(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
                play_sound_async(SOUND_PROFIT)
                total_profit = (target_price - entry_price) * qty
                log_trade_to_excel(
                    instrument.get("internal_trading_symbol"),
                    entry_price, target_price, qty, total_profit
                )
                print("Waiting for 1 min to get another data.")
                time.sleep(60)
                return

            elif sell_status in ["FAILED", "REJECTED", "CANCELLED"]:
                print(f"⚠️ SELL order failed with status {sell_status}")
                send_telegram(f"⚠️ SELL order failed ({sell_status})")
            else:
                print("⏱ Target not hit yet — switching to LTP monitoring.")

        # while True:
        #     ltp_now = get_ltp_for_instrument(instrument, access_token, verbose=False)
        #
        #     if ltp_now is None:
        #         continue
        #
        #     if ltp_now >= target_price:
        #         print(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
        #         send_telegram(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
        #         play_sound_async(SOUND_PROFIT)
        #         total_profit = (target_price - entry_price) * qty
        #         log_trade_to_excel(
        #             instrument.get("internal_trading_symbol"),
        #             entry_price, target_price, qty, total_profit
        #         )
        #         break
        #
        #     time.sleep(.5)
        return  # ✅ end of auto mode execution

    else:
        parsed = parse_cp_command(command)
        if not parsed:
            print("❌ Could not parse command.")
            return

        instrument = find_instrument_from_command(command, instruments_data)
        if not instrument:
            print("❌ Instrument not found in instrument master.")
            return

        lot_size = int(instrument.get("lot_size") or instrument.get("lotsize") or 1)
        quantity = parsed["lots"] * lot_size

        ltp_before = get_ltp_for_instrument(instrument, access_token)
        if ltp_before is None:
            print("❌ Could not fetch LTP before placing order.")
            return

        entry_price = round(float(ltp_before), 2)
        send_telegram(f"entry price: {entry_price} | {instrument.get('internal_trading_symbol')} | qty={quantity}")
        print(f"entry price: {entry_price}")

        # compute target per unit and target price
        profit_total = parsed["target_profit"]
        profit_per_unit = profit_total / quantity
        target_price = round_to_nearest_5_paise(entry_price + profit_per_unit)

        # Place BUY @ MARKET
        try:
            order_resp = place_market_order_groww(instrument, quantity, transaction_type="BUY", product="MIS")
            order_id = order_resp.get("payload", {}).get("groww_order_id") or order_resp.get("groww_order_id")
            print("✅ Buy Order placed:", order_resp)
        except Exception as e:
            print(f"❌ Buy order failed: {e}")
            send_telegram(f"❌ Buy order failed: {e}")
            return

        # STATUS VALIDATION
        # --- Wait until BUY order is EXECUTED or COMPLETED ---
        if order_id:
            buy_status = wait_for_order_status(order_id, access_token, "BUY")
            if buy_status not in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                print(f"⚠️ Skipping SELL due to BUY status: {buy_status}")
                return

        # === SELL @ LIMIT (target) ===
        try:
            sell_resp = place_limit_order_groww(
                instrument, quantity, price=target_price,
                transaction_type="SELL", product="MIS"
            )
            sell_order_id = (
                    sell_resp.get("payload", {}).get("groww_order_id")
                    or sell_resp.get("groww_order_id")
            )
            print(f"✅ Target SELL placed @ {target_price}. Response:", sell_resp)
            send_telegram(f"✅ Target SELL placed @ {target_price}")
        except Exception as e:
            print(f"❌ Failed to place SELL order: {e}")
            send_telegram(f"❌ Failed to place SELL order: {e}")
            return

        # STATUS VALIDATION
        # --- Wait until SELL order EXECUTED or COMPLETED ---
        if sell_order_id:
            sell_status = wait_for_order_status(sell_order_id, access_token, "SELL")
            if sell_status in ["EXECUTED", "COMPLETED"]:
                print(f"💰 Target HIT @ {target_price}")
                send_telegram(f"💰 Target HIT @ {target_price}")
                print(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
                send_telegram(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
                play_sound_async(SOUND_PROFIT)
                total_profit = (target_price - entry_price) * quantity
                log_trade_to_excel(
                    instrument.get("internal_trading_symbol"),
                    entry_price, target_price, quantity, total_profit
                )
                print("Waiting for 1 min to get another data.")
                time.sleep(60)
                return

            elif sell_status in ["FAILED", "REJECTED", "CANCELLED"]:
                print(f"⚠️ SELL order failed with status {sell_status}")
                send_telegram(f"⚠️ SELL order failed ({sell_status})")
            else:
                print("⏱ Target not hit yet — switching to LTP monitoring.")

        # while True:
        #     ltp_now = get_ltp_for_instrument(instrument, access_token)
        #     time.sleep(5)
        #     if ltp_now is None:
        #         continue
        #
        #     if ltp_now >= target_price:
        #         print(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
        #         send_telegram(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
        #         play_sound_async(SOUND_PROFIT)
        #         total_profit = (target_price - entry_price) * quantity
        #         log_trade_to_excel(
        #             instrument.get("internal_trading_symbol"),
        #             entry_price, target_price, quantity, total_profit
        #         )
        #         break




# ----------------- Auto mode runner (momentum + premium) -----------------


def auto_mode_runner():
    cfg = CONFIG
    print("\n--- AUTO MODE (momentum + premium) ---")
    send_telegram("\n--- AUTO MODE (momentum + premium) ---")
    index = cfg["index"]
    expiry = cfg["expiry"]
    min_p = cfg["min_premium"]
    max_p = cfg["max_premium"]
    lots = cfg["lots"]
    book_profit = cfg["book_profit"]

    target_pnl = cfg["target_pnl"]

    while True:
        opt_result = detect_option_type_parallel(index, expiry, min_p, max_p, lots)
        if not opt_result:
            print("❌ Could not determine CE/PE momentum side. Retrying in 3 minutes...")
            send_telegram("❌ Could not determine CE/PE momentum side. Retrying in 3 minutes...")
            time.sleep(180)
            continue

        opt_type, instrument, ltp, lot_size = opt_result
        if not instrument:
            print("❌ No matching/affordable option found. Retrying...")
            send_telegram("❌ No matching/affordable option found. Retrying...")
            time.sleep(60)
            continue

        symbol = instrument.get("tradingsymbol") or instrument.get("symbol") or instrument.get("internal_trading_symbol")
        instrument_type = instrument.get("instrument_type", "NA")
        groww_symbol = instrument.get("groww_symbol")
        print(f"✅ Selected: {symbol} ({instrument_type}) | LTP={ltp} | lot_size={lot_size} | groww_symbol={groww_symbol}")
        send_telegram(f"✅ Selected: {symbol} ({instrument_type}) | LTP={ltp} | lot_size={lot_size} | groww_symbol={groww_symbol}")

        # 🚀 Directly place the order (no string parsing)
        order_details = {
            "symbol": symbol,
            "ltp": ltp,
            "lots": lots,
            "book_profit": float(book_profit),
            "lot_size": lot_size,
            "side": "BUY"
        }

        market_direction = get_recent_market_direction(groww_symbol, groww)
        print(f"Market Direction: {market_direction}")
        send_telegram(f"Market Direction: {market_direction}")

        if market_direction == instrument_type:
            print("✅ Market direction CONFIRMS momentum → proceeding with order.")
            print(f"➡️ Placing auto order: {order_details} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            send_telegram(f"➡️ Placing auto order: {order_details} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            place_cp_order(order_details, is_auto=True)
        else:
            print("❌ Skipping trade — market direction conflicts with momentum.")
            print("Waiting for 3 mins to get another data.")
            time.sleep(180)

        time.sleep(2)


# ----------------- Main menu -----------------
if __name__ == "__main__":
    print("\n✨ Groww NIFTY CP Bot Ready (Groww backend)")
    print("You can run in MANUAL or AUTO mode.")
    print("Manual example: Buy 14 NIFTY04NOV2525950CE at CP and Book at 1050\n")
    while True:
        mode = input("Choose mode: (m)anual / (a)uto / (q)uit: ").strip().lower()
        if mode in ["q", "quit", "exit"]:
            print("Exiting.")
            break
        if mode in ["a", "auto"]:
            auto_mode_runner()
            continue
        if mode in ["m", "manual"]:
            user_input = input("\nEnter command (or press Enter for status, type 'back' to menu): ").strip()
            if user_input.lower() in ["back"]:
                continue
            if user_input == "":
                print("Status check not implemented for Groww PnL in this script.")
                continue
            place_cp_order(user_input)
            continue
        print("Invalid input. Choose 'm' or 'a' or 'q'.")
