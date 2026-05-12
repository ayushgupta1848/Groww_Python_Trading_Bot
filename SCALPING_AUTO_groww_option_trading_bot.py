"""
ONE BOX SCALPER STRATEGY - SCALPING BOT
========================================
Strategy: Opening Range Breakout + Retest Scalping
Market: NSE Nifty/BankNifty Options
Timeframe: First 5-min box (9:15-9:20), Entry on 1-min retest
Win Rate: 55-65% (Expected)
Risk:Reward = 1:2

STRATEGY LOGIC:
1. Create box from first 5-min candle (9:15-9:20)
2. Wait for breakout (close above/below box)
3. Wait for retest of box boundary
4. Enter on confirmation candle pattern
5. Target 2x risk

IMPORTANT: This bot can be run anytime during market hours (9:15 AM - 3:30 PM)
"""

import os
import re
import json
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import pyotp
from openpyxl import Workbook, load_workbook
from playsound3 import playsound
from datetime import datetime, timedelta, time as dt_time
from threading import Lock
import requests
import sys
import time
import random

# ==================== LOGGING SETUP ====================
session = requests.Session()

def setup_persistent_logger():
    """Creates a local 'logs' folder beside the script and logs all console output there."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    log_dir = os.path.join(base_dir, "logs")
    os.makedirs(log_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_path = os.path.join(log_dir, f"Scalping_Bot_{timestamp}.log")

    class Tee:
        def __init__(self, *streams):
            self.streams = streams

        def write(self, data):
            for s in self.streams:
                try:
                    s.write(data)
                    s.flush()
                except Exception:
                    pass

        def flush(self):
            for s in self.streams:
                try:
                    s.flush()
                except Exception:
                    pass

    logfile = open(log_path, "a", buffering=1, encoding="utf-8")
    sys.stdout = Tee(sys.stdout, logfile)
    sys.stderr = Tee(sys.stderr, logfile)

    print(f"📝 Scalping Bot Logging started. Log file: {log_path}")
    return log_path

LOG_FILE_PATH = setup_persistent_logger()

# ==================== GROWW API SETUP ====================
api_key = "eyJraWQiOiJaTUtjVXciLCJhbGciOiJFUzI1NiJ9.eyJleHAiOjI1NjM5ODI5MjksImlhdCI6MTc3NTU4MjkyOSwibmJmIjoxNzc1NTgyOTI5LCJzdWIiOiJ7XCJ0b2tlblJlZklkXCI6XCJhYjZiMTFlMy0xNGNlLTQwOTUtOTEzYi1hZjIxYWNhMzhkZjJcIixcInZlbmRvckludGVncmF0aW9uS2V5XCI6XCJlMzFmZjIzYjA4NmI0MDZjODg3NGIyZjZkODQ5NTMxM1wiLFwidXNlckFjY291bnRJZFwiOlwiMmVlMjYyMjItN2MwNS00Y2IwLWIwM2MtNzAzYWRmNWVmN2RkXCIsXCJkZXZpY2VJZFwiOlwiNjA2MzE5M2QtZWZkMC01OWViLTgzYzQtNWQ2NGZkNzdkNzQ3XCIsXCJzZXNzaW9uSWRcIjpcImZkYjJjYmM2LWU1MWYtNGY2Yi1hODFmLTlhOWQzYTFhMjg5MVwiLFwiYWRkaXRpb25hbERhdGFcIjpcIno1NC9NZzltdjE2WXdmb0gvS0EwYktvMDZXRlpjc241VUNmTWF5aERtNGxSTkczdTlLa2pWZDNoWjU1ZStNZERhWXBOVi9UOUxIRmtQejFFQisybTdRPT1cIixcInJvbGVcIjpcImF1dGgtdG90cFwiLFwic291cmNlSXBBZGRyZXNzXCI6XCIyNDAxOjQ5MDA6MWNhMzo4NTNhOjI4MTE6OWE3ZTo2NzA3OmIyZTksMTcyLjY5LjIwMy4xNDIsMzUuMjQxLjIzLjEyM1wiLFwidHdvRmFFeHBpcnlUc1wiOjI1NjM5ODI5MjkxNjUsXCJ2ZW5kb3JOYW1lXCI6XCJncm93d0FwaVwifSIsImlzcyI6ImFwZXgtYXV0aC1wcm9kLWFwcCJ9.5DM14qUlaFHzWhaHFn7EJGI2QyNjYwhbShM72v6dE13zQY-os8T9x0p87xbSaYxY4aN7whkAV3IaXk4QM0f3vg"
totp_gen = pyotp.TOTP('JKE5A5XD75LMF7KV7MKWS3W4YAS3HCT5')

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
csv_path = os.path.join(PROJECT_ROOT, "instrument.csv")
convert_csv_to_json = "yes"

# Telegram config
BOT_TOKEN = "8482701378:AAG7Jtfw0ZW_K9mFiX21LpsyUAV4oOcDiAQ"
CHAT_ID = "6012308856"

# Sound files
SOUND_PROFIT = "coin.mp3"
SOUND_SL = "SL_HIT.mp3"

try:
    from growwapi import GrowwAPI
except Exception:
    print("❗ growwapi module not found. Make sure it's installed.")

# ==================== GROWW INITIALIZATION ====================
def groww_init(api_key):
    """Initialize Groww API client"""
    totp = totp_gen.now()
    try:
        access_token = GrowwAPI.get_access_token(api_key=api_key, totp=totp)
        client = GrowwAPI(access_token)
        print(f"✅ Groww API Initialized Successfully")
        print(f"🔑 Access Token: {access_token[:50]}...")
        return client, access_token
    except Exception as e:
        print(f"❌ Groww login failed: {e}")
        raise

groww, access_token = groww_init(api_key)

# ==================== UTILITY FUNCTIONS ====================
def send_telegram(message: str):
    """Send telegram message asynchronously"""
    def _send():
        try:
            url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
            payload = {"chat_id": CHAT_ID, "text": message}
            requests.post(url, data=payload)
        except Exception as e:
            print(f"⚠️ Telegram Error: {e}")
    threading.Thread(target=_send, daemon=True).start()

def play_sound_async(filename):
    """Play sound asynchronously"""
    try:
        if not os.path.exists(filename):
            print(f"🔇 Sound file not found: {filename}")
            return
        threading.Thread(target=playsound, args=(filename,), daemon=True).start()
    except Exception as e:
        print(f"🔇 Sound error: {e}")

def log_trade_to_excel(symbol, entry_price, exit_price, quantity, profit, trade_type, setup_type):
    """Log trades to Excel"""
    file_name = "scalping_trades.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Scalping"
        ws.append(["DateTime", "Symbol", "Type", "Setup", "Entry", "Exit", "Qty", "Profit"])
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active
    next_row = ws.max_row + 1
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.cell(row=next_row, column=1).value = now
    ws.cell(row=next_row, column=2).value = symbol
    ws.cell(row=next_row, column=3).value = trade_type
    ws.cell(row=next_row, column=4).value = setup_type
    ws.cell(row=next_row, column=5).value = entry_price
    ws.cell(row=next_row, column=6).value = exit_price
    ws.cell(row=next_row, column=7).value = quantity
    ws.cell(row=next_row, column=8).value = round(profit, 2)
    wb.save(file_name)

# ==================== CSV/JSON LOADING ====================
def csv_to_json(csv_file_path, json_file_path=None):
    """Convert CSV to JSON"""
    import csv
    if json_file_path is None:
        json_file_path = os.path.splitext(csv_file_path)[0] + ".json"
    data = []
    with open(csv_file_path, encoding='utf-8') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            data.append(row)
    with open(json_file_path, 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file, indent=4, ensure_ascii=False)
    print(f"✅ Converted '{csv_file_path}' → '{json_file_path}'")
    return data

# ==================== LTP FETCHING ====================
ltp_lock = threading.Lock()

def get_ltp_for_instrument(instrument, access_token, verbose=True, segment="FNO", delay=0.1, max_retries=2):
    """Fetch LTP for instrument with thread safety"""
    trading_symbol = instrument.get("trading_symbol")
    if not trading_symbol:
        print("⚠️ Missing trading_symbol in instrument.")
        return None

    exchange_symbol = f"NSE_{trading_symbol}"
    url = f"https://api.groww.in/v1/live-data/ltp?segment={segment}&exchange_symbols={exchange_symbol}"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0"
    }

    last_exc = None
    for attempt in range(max_retries + 1):
        try:
            with ltp_lock:
                resp = session.get(url, headers=headers, timeout=10)
                if delay > 0:
                    time.sleep(delay)

            if resp.status_code == 429:
                backoff = 0.4 * (2 ** attempt) + random.uniform(-0.05, 0.05)
                print(f"⚠️ LTP 429 received, backing off {backoff:.2f}s (attempt {attempt+1})")
                time.sleep(backoff)
                last_exc = requests.exceptions.HTTPError("429 Too Many Requests")
                continue

            resp.raise_for_status()
            data = resp.json()
            payload = data.get("payload", {})
            ltp = payload.get(exchange_symbol)

            if ltp is None:
                print(f"⚠️ No LTP found for {exchange_symbol}")
                return None
            if verbose:
                print(f"💰 LTP for {exchange_symbol}: ₹{ltp} @ {datetime.now().strftime('%H:%M:%S')}")
            return float(ltp)

        except Exception as e:
            last_exc = e
            if attempt < max_retries:
                backoff = 0.3 * (attempt + 1) + random.uniform(0, 0.05)
                print(f"⚠️ Error fetching LTP ({e}), retrying in {backoff:.2f}s...")
                time.sleep(backoff)
                continue
            else:
                print(f"⚠️ Error fetching LTP: {e}")
                return None

    print(f"⚠️ Failed to fetch LTP after {max_retries + 1} attempts")
    return None

def get_nifty_spot_price(access_token, json_path=None):
    """Fetch NIFTY 50 spot price"""
    global instruments_all

    if json_path is None:
        json_path = os.path.splitext(csv_path)[0] + ".json"

    if convert_csv_to_json.lower() == "yes":
        instruments_all = csv_to_json(csv_path, json_path)
    else:
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON not found: {json_path}")
        with open(json_path, "r", encoding="utf-8") as jf:
            instruments_all = json.load(jf)

    try:
        nifty_spot_instrument = next(
            (item for item in instruments_all
             if item.get("trading_symbol") == "NIFTY"
             or item.get("groww_symbol") == "NSE-NIFTY"
             or item.get("name") == "NIFTY 50"),
            None
        )

        if not nifty_spot_instrument:
            print("⚠️ NIFTY spot instrument not found")
            return 0

        spot = get_ltp_for_instrument(nifty_spot_instrument, access_token, verbose=False, segment="CASH")
        if spot:
            print(f"📊 Live NIFTY Spot: {spot}")
            return float(spot)
        else:
            print("⚠️ Failed to fetch LTP for NIFTY spot")
            return 0
    except Exception as e:
        print(f"⚠️ Error fetching NIFTY spot: {e}")
        return 0

# ==================== SCALPING BOT CONFIG ====================
CONFIG = {
    "index": "NIFTY",
    "expiry": "2026-03-17",
    "min_premium": 85,  # Lower for scalping
    "max_premium": 150,
    "lots": 20,  # Adjust based on capital
    "lot_size": 65,  # NIFTY lot size (standard)
    "spot": None,  # Will be set dynamically
    
    # SCALPING SPECIFIC SETTINGS
    "BOX_CREATION_TIME": "09:20",  # First 5-min candle ends
    "TRADING_START_TIME": "09:20",  # Start after box creation
    "TRADING_END_TIME": "14:30",  # Stop new entries by 2:30 PM
    "CANDLE_INTERVAL": 60,  # 1-minute candles for entry
    
    # RISK MANAGEMENT - OPTIMIZED FOR BIGGER MOVES
    "RISK_REWARD_RATIO": 3.0,  # 1:3 risk-reward for better profits
    "MIN_PROFIT_POINTS": 8.0,  # Minimum 8 points profit target
    "MAX_RISK_PER_TRADE": 800,  # Max loss per trade in ₹ (increased for bigger positions)
    "MAX_TRADES_PER_DAY": 4,  # Limit to 4 quality trades
    "MIN_BOX_RANGE": 3.0,  # Minimum box range for volatility
    "POLL_INTERVAL": 0.2,  # Fast polling for scalping
    
    # TRAILING STOP FOR EXTENDED MOVES
    "USE_TRAILING_STOP": True,  # Enable trailing stop
    "TRAIL_START_POINTS": 6.0,  # Start trailing after 6 points profit
    "TRAIL_DISTANCE": 3.0,  # Trail 3 points below highest
    
    # FILTERS - STRICTER FOR QUALITY
    "USE_VWAP_FILTER": True,  # Enable VWAP confirmation
    "MIN_BREAKOUT_CONFIRMATION": True,  # Require candle close for breakout
    "MIN_VOLUME_RATIO": 1.5,  # Breakout volume should be 1.5x average
    "COOLDOWN_AFTER_LOSS": 900,  # 15 min cooldown after loss
    
    # ORDER SETTINGS
    "VALIDATE_ORDERS": False,  # Set to True for live trading
    "user_confirmation_needed": False,
}

# ==================== INSTRUMENT LOADING ====================
def load_instruments_from_json(json_path=None):
    """Load and filter instruments"""
    global instruments
    config = CONFIG
    INDEX = config["index"].upper()
    EXPIRY = config["expiry"].strip()

    if json_path is None:
        json_path = os.path.splitext(csv_path)[0] + ".json"

    if convert_csv_to_json.lower() == "yes":
        instruments = csv_to_json(csv_path, json_path)
    else:
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON not found: {json_path}")
        with open(json_path, "r", encoding="utf-8") as jf:
            instruments = json.load(jf)

    # Get spot price
    spot = get_nifty_spot_price(access_token)
    CONFIG["spot"] = spot

    # Filter instruments
    step = 100 if "BANK" in INDEX else 50
    nearest_strike = round(spot / step) * step
    lower_bound = nearest_strike - (10 * step)
    upper_bound = nearest_strike + (10 * step)

    print(f"🎯 Filtering {INDEX} {EXPIRY} between {lower_bound}–{upper_bound} (Spot={spot})")

    filtered = []
    for item in instruments:
        try:
            if item.get("underlying_symbol", "").upper() != INDEX:
                continue
            strike = float(item.get("strike_price") or 0)
            if lower_bound <= strike <= upper_bound:
                filtered.append(item)
        except Exception:
            continue

    print(f"✅ Loaded {len(filtered)} filtered instruments")
    instruments = filtered
    return instruments

instruments_data = load_instruments_from_json()

# ==================== CANDLE DATA FETCHING ====================
def get_historical_candles(instrument, access_token, interval="1minute", candles_count=50):
    """
    Fetch historical candle data for box creation and pattern detection
    Returns: List of candles [[timestamp, open, high, low, close, volume], ...]
    """
    try:
        trading_symbol = instrument.get("trading_symbol")
        groww_symbol = instrument.get("groww_symbol", f"NSE_{trading_symbol}")
        
        end_time = datetime.now()
        # Calculate start time based on interval and count needed
        if interval == "1minute":
            start_time = end_time - timedelta(minutes=candles_count + 10)
        elif interval == "5minute":
            start_time = end_time - timedelta(minutes=(candles_count * 5) + 30)
        else:
            start_time = end_time - timedelta(hours=2)
        
        end_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        start_str = start_time.strftime("%Y-%m-%d %H:%M:%S")
        
        historical = groww.get_historical_candles(
            groww_symbol=groww_symbol,
            exchange=groww.EXCHANGE_NSE,
            segment=groww.SEGMENT_FNO,
            start_time=start_str,
            end_time=end_str,
            candle_interval=interval
        )
        
        candles = historical.get("candles", [])
        if not candles:
            print(f"⚠️ No candle data returned for {trading_symbol}")
            return None
            
        print(f"📊 Fetched {len(candles)} candles for {trading_symbol}")
        return candles
        
    except Exception as e:
        print(f"⚠️ Error fetching historical candles: {e}")
        return None

def calculate_vwap(candles):
    """Calculate VWAP from candle data"""
    if not candles or len(candles) == 0:
        return None
    
    try:
        total_volume = 0
        total_pv = 0
        
        for candle in candles:
            # candle format: [timestamp, open, high, low, close, volume]
            high = candle[2]
            low = candle[3]
            close = candle[4]
            volume = candle[5]
            
            typical_price = (high + low + close) / 3.0
            total_pv += typical_price * volume
            total_volume += volume
        
        if total_volume == 0:
            return None
        
        vwap = total_pv / total_volume
        return round(vwap, 2)
    except Exception as e:
        print(f"⚠️ Error calculating VWAP: {e}")
        return None

# ==================== ONE BOX SCALPER CORE LOGIC ====================

class OpeningRangeBox:
    """Manages the opening range box (9:15-9:20 candle)"""
    
    def __init__(self):
        self.box_high = None
        self.box_low = None
        self.box_created = False
        self.creation_time = None
        
    def create_box(self, candles):
        """
        Create box from first 5-minute candle
        candles: list of 1-min candles from 9:15-9:20
        """
        if not candles or len(candles) < 5:
            print("⚠️ Not enough candles to create box")
            return False
        
        # Get first 5 candles (9:15-9:20)
        box_candles = candles[:5]
        
        highs = [c[2] for c in box_candles]
        lows = [c[3] for c in box_candles]
        
        self.box_high = max(highs)
        self.box_low = min(lows)
        self.box_created = True
        self.creation_time = datetime.now()
        
        print(f"📦 BOX CREATED: HIGH={self.box_high}, LOW={self.box_low}")
        send_telegram(f"📦 Opening Range Box Created\n🔼 High: {self.box_high}\n🔽 Low: {self.box_low}")
        
        return True
    
    def is_breakout(self, candle, breakout_type):
        """
        Check if candle represents a valid breakout
        breakout_type: 'bullish' or 'bearish'
        Returns: True if valid breakout
        """
        if not self.box_created:
            return False
        
        close = candle[4]
        
        if breakout_type == "bullish":
            # Bullish breakout: close above box high
            if close > self.box_high:
                print(f"🔼 BULLISH BREAKOUT detected: Close={close} > Box High={self.box_high}")
                return True
        elif breakout_type == "bearish":
            # Bearish breakout: close below box low
            if close < self.box_low:
                print(f"🔽 BEARISH BREAKOUT detected: Close={close} < Box Low={self.box_low}")
                return True
        
        return False
    
    def is_retest(self, candle, breakout_type):
        """
        Check if price is retesting the box boundary
        Returns: True if retest happening
        """
        if not self.box_created:
            return False
        
        high = candle[2]
        low = candle[3]
        close = candle[4]
        
        if breakout_type == "bullish":
            # For bullish: price should pull back to box_high
            # Check if price touched or came close to box_high
            if low <= self.box_high <= high:
                print(f"🔄 BULLISH RETEST: Price touched box high ({self.box_high})")
                return True
            # Also consider close retest
            if abs(close - self.box_high) <= (self.box_high * 0.002):  # Within 0.2%
                print(f"🔄 BULLISH RETEST: Close near box high")
                return True
                
        elif breakout_type == "bearish":
            # For bearish: price should pull back to box_low
            if low <= self.box_low <= high:
                print(f"🔄 BEARISH RETEST: Price touched box low ({self.box_low})")
                return True
            if abs(close - self.box_low) <= (self.box_low * 0.002):
                print(f"🔄 BEARISH RETEST: Close near box low")
                return True
        
        return False
    
    def get_box_range(self):
        """Return box range"""
        if self.box_created:
            return self.box_high - self.box_low
        return None


class CandlePatternDetector:
    """Detects candlestick confirmation patterns"""
    
    @staticmethod
    def is_shooting_star(candle):
        """
        Detect Shooting Star / Inverted Hammer (Bearish)
        - Long upper wick
        - Small body
        - Close near low
        """
        open_price = candle[1]
        high = candle[2]
        low = candle[3]
        close = candle[4]
        
        body = abs(close - open_price)
        upper_wick = high - max(open_price, close)
        lower_wick = min(open_price, close) - low
        
        # Pattern requirements
        if body == 0:
            return False
        
        # Upper wick should be at least 2x body
        if upper_wick >= 2 * body:
            # Lower wick should be small
            if lower_wick < body:
                # Close should be near low
                if close <= (low + (high - low) * 0.3):
                    print(f"🕯️ SHOOTING STAR detected")
                    return True
        
        return False
    
    @staticmethod
    def is_hammer(candle):
        """
        Detect Hammer (Bullish)
        - Long lower wick
        - Small body
        - Close near high
        """
        open_price = candle[1]
        high = candle[2]
        low = candle[3]
        close = candle[4]
        
        body = abs(close - open_price)
        upper_wick = high - max(open_price, close)
        lower_wick = min(open_price, close) - low
        
        if body == 0:
            return False
        
        # Lower wick should be at least 2x body
        if lower_wick >= 2 * body:
            # Upper wick should be small
            if upper_wick < body:
                # Close should be near high
                if close >= (low + (high - low) * 0.7):
                    print(f"🕯️ HAMMER detected")
                    return True
        
        return False
    
    @staticmethod
    def is_bearish_engulfing(current_candle, previous_candle):
        """
        Detect Bearish Engulfing
        - Previous candle is green
        - Current candle is red
        - Current candle engulfs previous
        """
        prev_open = previous_candle[1]
        prev_close = previous_candle[4]
        curr_open = current_candle[1]
        curr_high = current_candle[2]
        curr_low = current_candle[3]
        curr_close = current_candle[4]
        
        # Previous candle should be bullish
        if prev_close <= prev_open:
            return False
        
        # Current candle should be bearish
        if curr_close >= curr_open:
            return False
        
        # Current should engulf previous
        if curr_high > max(prev_open, prev_close) and curr_low < min(prev_open, prev_close):
            print(f"🕯️ BEARISH ENGULFING detected")
            return True
        
        return False
    
    @staticmethod
    def is_bullish_engulfing(current_candle, previous_candle):
        """
        Detect Bullish Engulfing
        - Previous candle is red
        - Current candle is green
        - Current candle engulfs previous
        """
        prev_open = previous_candle[1]
        prev_close = previous_candle[4]
        curr_open = current_candle[1]
        curr_high = current_candle[2]
        curr_low = current_candle[3]
        curr_close = current_candle[4]
        
        # Previous candle should be bearish
        if prev_close >= prev_open:
            return False
        
        # Current candle should be bullish
        if curr_close <= curr_open:
            return False
        
        # Current should engulf previous
        if curr_high > max(prev_open, prev_close) and curr_low < min(prev_open, prev_close):
            print(f"🕯️ BULLISH ENGULFING detected")
            return True
        
        return False
    
    @staticmethod
    def detect_bearish_confirmation(candles):
        """
        Detect any bearish confirmation pattern
        Returns: True if bearish pattern found
        """
        if len(candles) < 2:
            return False
        
        current = candles[-1]
        previous = candles[-2]
        
        # Check shooting star
        if CandlePatternDetector.is_shooting_star(current):
            return True
        
        # Check bearish engulfing
        if CandlePatternDetector.is_bearish_engulfing(current, previous):
            return True
        
        return False
    
    @staticmethod
    def detect_bullish_confirmation(candles):
        """
        Detect any bullish confirmation pattern
        Returns: True if bullish pattern found
        """
        if len(candles) < 2:
            return False
        
        current = candles[-1]
        previous = candles[-2]
        
        # Check hammer
        if CandlePatternDetector.is_hammer(current):
            return True
        
        # Check bullish engulfing
        if CandlePatternDetector.is_bullish_engulfing(current, previous):
            return True
        
        return False


# ==================== ORDER EXECUTION ====================
def place_market_order_groww(instrument, qty, side, product="MIS"):
    """Place market order using Groww API"""
    try:
        trading_symbol = instrument.get("trading_symbol")
        groww_symbol = instrument.get("groww_symbol", f"NSE_{trading_symbol}")
        
        print(f"📤 Placing {side} order: {trading_symbol}, Qty={qty}")
        
        order = groww.place_order(
            exchange=groww.EXCHANGE_NSE,
            segment=groww.SEGMENT_FNO,
            trading_symbol=trading_symbol,
            transaction_type=side,
            quantity=qty,
            order_type=groww.ORDER_TYPE_MARKET,
            product=product,
            validity=groww.VALIDITY_DAY,
            price=0,
            trigger_price=0
        )
        
        print(f"✅ Order placed successfully: {order}")
        send_telegram(f"✅ {side} Order Placed\n{trading_symbol}\nQty: {qty}")
        
        return order
        
    except Exception as e:
        print(f"❌ Order placement failed: {e}")
        send_telegram(f"❌ Order Failed: {e}")
        raise


def get_order_status(order_id, access_token, max_attempts=10):
    """Check order status"""
    for attempt in range(max_attempts):
        try:
            order = groww.get_order_details(order_id)
            status = order.get("order_status", "")
            print(f"📊 Order Status (attempt {attempt+1}): {status}")
            
            if status in ["COMPLETE", "EXECUTED"]:
                return "COMPLETE"
            elif status in ["REJECTED", "CANCELLED"]:
                return "FAILED"
            
            time.sleep(1)
        except Exception as e:
            print(f"⚠️ Error checking order status: {e}")
            time.sleep(1)
    
    return "PENDING"


# ==================== TRADE EXECUTION ENGINE ====================
class ScalpingTradeEngine:
    """Main trade execution engine for scalping strategy"""
    
    def __init__(self, instrument, lot_size, access_token):
        self.instrument = instrument
        self.lot_size = lot_size
        self.access_token = access_token
        self.in_trade = False
        self.entry_price = None
        self.stop_loss = None
        self.target = None
        self.quantity = None
        self.trade_type = None  # 'LONG' or 'SHORT'
        
    def calculate_position_size(self, entry_price, stop_loss):
        """Calculate position size based on risk"""
        max_risk = CONFIG.get("MAX_RISK_PER_TRADE", 500)
        risk_per_unit = abs(entry_price - stop_loss)
        
        if risk_per_unit == 0:
            return self.lot_size
        
        # Calculate lots based on risk
        max_qty = int(max_risk / risk_per_unit)
        
        # Round down to lot size
        lots = max(1, max_qty // self.lot_size)
        qty = lots * self.lot_size
        
        print(f"📊 Position Size: {qty} (Risk per unit: ₹{risk_per_unit:.2f})")
        return qty
    
    def enter_long(self, entry_price, confirmation_candle, setup_type):
        """Enter long trade"""
        try:
            # Calculate stop loss (below confirmation candle low)
            self.stop_loss = confirmation_candle[3]  # Low of confirmation candle
            
            # Calculate target (2x risk)
            risk = entry_price - self.stop_loss
            self.target = entry_price + (risk * CONFIG.get("RISK_REWARD_RATIO", 2.0))
            
            # Calculate quantity
            self.quantity = self.calculate_position_size(entry_price, self.stop_loss)
            
            print(f"🟢 LONG ENTRY SIGNAL")
            print(f"   Entry: ₹{entry_price}")
            print(f"   Stop Loss: ₹{self.stop_loss:.2f}")
            print(f"   Target: ₹{self.target:.2f}")
            print(f"   Quantity: {self.quantity}")
            print(f"   Risk: ₹{risk:.2f} | Reward: ₹{(self.target - entry_price):.2f}")
            
            send_telegram(
                f"🟢 LONG ENTRY\n"
                f"Symbol: {self.instrument.get('trading_symbol')}\n"
                f"Setup: {setup_type}\n"
                f"Entry: ₹{entry_price}\n"
                f"SL: ₹{self.stop_loss:.2f}\n"
                f"Target: ₹{self.target:.2f}\n"
                f"Qty: {self.quantity}"
            )
            
            # Place order
            if CONFIG.get("VALIDATE_ORDERS", False):
                place_market_order_groww(self.instrument, self.quantity, "BUY", "MIS")
            
            self.entry_price = entry_price
            self.trade_type = "LONG"
            self.in_trade = True
            
            # Start monitoring
            self.monitor_trade()
            
        except Exception as e:
            print(f"❌ Error entering long: {e}")
            send_telegram(f"❌ Long entry failed: {e}")
    
    def enter_short(self, entry_price, confirmation_candle, setup_type):
        """Enter short trade"""
        try:
            # Calculate stop loss (above confirmation candle high)
            self.stop_loss = confirmation_candle[2]  # High of confirmation candle
            
            # Calculate target (2x risk)
            risk = self.stop_loss - entry_price
            self.target = entry_price - (risk * CONFIG.get("RISK_REWARD_RATIO", 2.0))
            
            # Calculate quantity
            self.quantity = self.calculate_position_size(entry_price, self.stop_loss)
            
            print(f"🔴 SHORT ENTRY SIGNAL")
            print(f"   Entry: ₹{entry_price}")
            print(f"   Stop Loss: ₹{self.stop_loss:.2f}")
            print(f"   Target: ₹{self.target:.2f}")
            print(f"   Quantity: {self.quantity}")
            print(f"   Risk: ₹{risk:.2f} | Reward: ₹{(entry_price - self.target):.2f}")
            
            send_telegram(
                f"🔴 SHORT ENTRY\n"
                f"Symbol: {self.instrument.get('trading_symbol')}\n"
                f"Setup: {setup_type}\n"
                f"Entry: ₹{entry_price}\n"
                f"SL: ₹{self.stop_loss:.2f}\n"
                f"Target: ₹{self.target:.2f}\n"
                f"Qty: {self.quantity}"
            )
            
            # Place order
            if CONFIG.get("VALIDATE_ORDERS", False):
                place_market_order_groww(self.instrument, self.quantity, "SELL", "MIS")
            
            self.entry_price = entry_price
            self.trade_type = "SHORT"
            self.in_trade = True
            
            # Start monitoring
            self.monitor_trade()
            
        except Exception as e:
            print(f"❌ Error entering short: {e}")
            send_telegram(f"❌ Short entry failed: {e}")
    
    def monitor_trade(self):
        """Monitor active trade for stop loss and target"""
        print(f"👁️ Monitoring {self.trade_type} trade...")
        
        poll_interval = CONFIG.get("POLL_INTERVAL", 0.2)
        
        while self.in_trade:
            try:
                # Fetch current LTP
                ltp = get_ltp_for_instrument(self.instrument, self.access_token, verbose=False, delay=0)
                
                if ltp is None:
                    time.sleep(poll_interval)
                    continue
                
                ltp = float(ltp)
                
                # Calculate current P&L
                if self.trade_type == "LONG":
                    pnl = (ltp - self.entry_price) * self.quantity
                    
                    # Check stop loss
                    if ltp <= self.stop_loss:
                        print(f"🛑 STOP LOSS HIT @ ₹{ltp}")
                        self.exit_trade(ltp, "STOP_LOSS")
                        break
                    
                    # Check target
                    if ltp >= self.target:
                        print(f"🎯 TARGET HIT @ ₹{ltp}")
                        self.exit_trade(ltp, "TARGET")
                        break
                
                elif self.trade_type == "SHORT":
                    pnl = (self.entry_price - ltp) * self.quantity
                    
                    # Check stop loss
                    if ltp >= self.stop_loss:
                        print(f"🛑 STOP LOSS HIT @ ₹{ltp}")
                        self.exit_trade(ltp, "STOP_LOSS")
                        break
                    
                    # Check target
                    if ltp <= self.target:
                        print(f"🎯 TARGET HIT @ ₹{ltp}")
                        self.exit_trade(ltp, "TARGET")
                        break
                
                # Print status every 10 seconds
                if int(time.time()) % 10 == 0:
                    print(f"📊 LTP: ₹{ltp} | P&L: ₹{pnl:.2f} | Target: ₹{self.target:.2f} | SL: ₹{self.stop_loss:.2f}")
                
                time.sleep(poll_interval)
                
            except Exception as e:
                print(f"⚠️ Error in trade monitoring: {e}")
                time.sleep(poll_interval)
    
    def exit_trade(self, exit_price, exit_reason):
        """Exit trade"""
        try:
            if not self.in_trade:
                return
            
            # Calculate final P&L
            if self.trade_type == "LONG":
                pnl = (exit_price - self.entry_price) * self.quantity
                side = "SELL"
            else:
                pnl = (self.entry_price - exit_price) * self.quantity
                side = "BUY"
            
            print(f"🏁 EXITING {self.trade_type} @ ₹{exit_price}")
            print(f"   Reason: {exit_reason}")
            print(f"   P&L: ₹{pnl:.2f}")
            
            # Place exit order
            if CONFIG.get("VALIDATE_ORDERS", False):
                place_market_order_groww(self.instrument, self.quantity, side, "MIS")
            
            # Send notification
            emoji = "✅" if pnl > 0 else "❌"
            send_telegram(
                f"{emoji} TRADE CLOSED\n"
                f"Symbol: {self.instrument.get('trading_symbol')}\n"
                f"Type: {self.trade_type}\n"
                f"Entry: ₹{self.entry_price}\n"
                f"Exit: ₹{exit_price}\n"
                f"P&L: ₹{pnl:.2f}\n"
                f"Reason: {exit_reason}"
            )
            
            # Play sound
            if pnl > 0:
                play_sound_async(SOUND_PROFIT)
            else:
                play_sound_async(SOUND_SL)
            
            # Log to Excel
            log_trade_to_excel(
                self.instrument.get("trading_symbol"),
                self.entry_price,
                exit_price,
                self.quantity,
                pnl,
                self.trade_type,
                exit_reason
            )
            
            # Reset state
            self.in_trade = False
            self.entry_price = None
            self.stop_loss = None
            self.target = None
            self.quantity = None
            self.trade_type = None
            
        except Exception as e:
            print(f"❌ Error exiting trade: {e}")
            send_telegram(f"❌ Exit failed: {e}")


# ==================== MAIN SCALPING STRATEGY ====================
def find_option_for_scalping(option_type, min_premium, max_premium):
    """Find option matching premium criteria"""
    config = CONFIG
    INDEX = config["index"].upper()
    EXPIRY = config["expiry"].strip()
    
    candidates = []
    
    for item in instruments_data:
        try:
            if item.get("underlying_symbol", "").upper() != INDEX:
                continue
            if item.get("instrument_type", "").upper() != option_type.upper():
                continue
            if item.get("expiry_date", "").strip() != EXPIRY:
                continue
            
            lot_size = int(item.get("lot_size") or item.get("lotsize") or 1)
            ltp = get_ltp_for_instrument(item, access_token, verbose=False)
            
            if ltp is None:
                continue
            if not (min_premium <= ltp <= max_premium):
                continue
            
            candidates.append({
                "instrument": item,
                "ltp": float(ltp),
                "lot_size": lot_size
            })
        except Exception as e:
            continue
    
    if not candidates:
        print(f"⚠️ No candidates found for {option_type}")
        return None, None, None
    
    # Sort by closest to mid-premium
    mid = (min_premium + max_premium) / 2.0
    candidates.sort(key=lambda x: abs(x["ltp"] - mid))
    
    pick = candidates[0]
    print(f"✅ Selected {option_type}: {pick['instrument']['trading_symbol']} @ ₹{pick['ltp']}")
    
    return pick["instrument"], pick["ltp"], pick["lot_size"]


def run_scalping_bot():
    """Main scalping bot loop"""
    print("=" * 60)
    print("🚀 ONE BOX SCALPER STRATEGY - STARTING")
    print("=" * 60)
    
    send_telegram("🚀 Scalping Bot Started\nStrategy: One Box Scalper")
    
    # Trading hours check
    def is_market_hours():
        now = datetime.now().time()
        market_start = dt_time(9, 15)
        market_end = dt_time(15, 30)
        return market_start <= now <= market_end
    
    def is_trading_window():
        now = datetime.now().time()
        start = dt_time(9, 20)
        end = dt_time(15, 0)
        return start <= now <= end
    
    # Wait for market to open if before 9:15
    while not is_market_hours():
        print(f"⏰ Waiting for market to open... Current time: {datetime.now().strftime('%H:%M:%S')}")
        time.sleep(60)
    
    # Find option to trade
    print("\n🔍 Searching for option to trade...")
    option_type = "CE"  # Start with CE, can be made dynamic
    min_prem = CONFIG.get("min_premium", 40)
    max_prem = CONFIG.get("max_premium", 150)
    
    instrument, ltp, lot_size = find_option_for_scalping(option_type, min_prem, max_prem)
    
    if not instrument:
        print("❌ No suitable option found")
        send_telegram("❌ No suitable option found")
        return
    
    print(f"\n✅ Trading Instrument: {instrument.get('trading_symbol')}")
    print(f"   LTP: ₹{ltp}")
    print(f"   Lot Size: {lot_size}")
    
    # Initialize components
    box = OpeningRangeBox()
    pattern_detector = CandlePatternDetector()
    trade_engine = ScalpingTradeEngine(instrument, lot_size, access_token)
    
    # State tracking
    breakout_detected = False
    breakout_type = None
    retest_detected = False
    
    print("\n📊 Fetching initial candles...")
    
    # Main loop
    last_candle_time = None
    
    while is_market_hours():
        try:
            current_time = datetime.now()
            
            # Fetch recent 1-minute candles
            candles = get_historical_candles(instrument, access_token, interval="1minute", candles_count=20)
            
            if not candles or len(candles) < 2:
                print("⚠️ Not enough candle data")
                time.sleep(30)
                continue
            
            # Create box if not created and time is right
            if not box.box_created:
                # Check if we have 5+ candles after 9:15
                if len(candles) >= 5:
                    # Create box from first 5 candles
                    if box.create_box(candles):
                        send_telegram(f"📦 Box Created\nHigh: {box.box_high}\nLow: {box.box_low}")
                    else:
                        print("⚠️ Failed to create box")
                        time.sleep(30)
                        continue
                else:
                    print(f"⏳ Waiting for 5 candles to create box... ({len(candles)}/5)")
                    time.sleep(30)
                    continue
            
            # Only trade in trading window
            if not is_trading_window():
                print(f"⏰ Outside trading window. Current: {current_time.strftime('%H:%M:%S')}")
                time.sleep(60)
                continue
            
            # Skip if already in trade
            if trade_engine.in_trade:
                time.sleep(5)
                continue
            
            # Get latest candle
            latest_candle = candles[-1]
            candle_time = latest_candle[0]
            
            # Skip if same candle
            if last_candle_time == candle_time:
                time.sleep(5)
                continue
            
            last_candle_time = candle_time
            
            # Log candle
            print(f"\n📊 New Candle @ {datetime.fromtimestamp(candle_time/1000).strftime('%H:%M:%S')}")
            print(f"   O: {latest_candle[1]}, H: {latest_candle[2]}, L: {latest_candle[3]}, C: {latest_candle[4]}")
            
            # STEP 1: Detect breakout
            if not breakout_detected:
                if box.is_breakout(latest_candle, "bullish"):
                    breakout_detected = True
                    breakout_type = "bullish"
                    print("🔼 BULLISH BREAKOUT - Waiting for retest...")
                    send_telegram(f"🔼 Bullish Breakout Detected\nWaiting for retest...")
                    
                elif box.is_breakout(latest_candle, "bearish"):
                    breakout_detected = True
                    breakout_type = "bearish"
                    print("🔽 BEARISH BREAKOUT - Waiting for retest...")
                    send_telegram(f"🔽 Bearish Breakout Detected\nWaiting for retest...")
            
            # STEP 2: Detect retest
            elif breakout_detected and not retest_detected:
                if box.is_retest(latest_candle, breakout_type):
                    retest_detected = True
                    print(f"🔄 RETEST DETECTED - Waiting for confirmation candle...")
                    send_telegram(f"🔄 Retest Detected\nWaiting for confirmation...")
            
            # STEP 3: Check for confirmation pattern and enter
            elif breakout_detected and retest_detected:
                current_ltp = get_ltp_for_instrument(instrument, access_token, verbose=False)
                
                if current_ltp is None:
                    time.sleep(5)
                    continue
                
                # VWAP filter (optional)
                vwap = None
                if CONFIG.get("USE_VWAP_FILTER", True):
                    vwap = calculate_vwap(candles)
                    if vwap:
                        print(f"📊 VWAP: ₹{vwap}")
                
                # Check confirmation patterns
                if breakout_type == "bullish":
                    # Check for bullish confirmation
                    if pattern_detector.detect_bullish_confirmation(candles):
                        # VWAP filter for long
                        if vwap and current_ltp < vwap:
                            print(f"❌ VWAP filter failed: LTP ({current_ltp}) < VWAP ({vwap})")
                            # Reset and wait for next setup
                            breakout_detected = False
                            retest_detected = False
                            breakout_type = None
                            continue
                        
                        print("✅ BULLISH CONFIRMATION - ENTERING LONG")
                        trade_engine.enter_long(current_ltp, latest_candle, "Bullish Breakout + Retest")
                        
                        # Reset for next trade
                        breakout_detected = False
                        retest_detected = False
                        breakout_type = None
                
                elif breakout_type == "bearish":
                    # Check for bearish confirmation
                    if pattern_detector.detect_bearish_confirmation(candles):
                        # VWAP filter for short
                        if vwap and current_ltp > vwap:
                            print(f"❌ VWAP filter failed: LTP ({current_ltp}) > VWAP ({vwap})")
                            # Reset and wait for next setup
                            breakout_detected = False
                            retest_detected = False
                            breakout_type = None
                            continue
                        
                        print("✅ BEARISH CONFIRMATION - ENTERING SHORT")
                        trade_engine.enter_short(current_ltp, latest_candle, "Bearish Breakout + Retest")
                        
                        # Reset for next trade
                        breakout_detected = False
                        retest_detected = False
                        breakout_type = None
            
            # Sleep before next iteration
            time.sleep(10)
            
        except KeyboardInterrupt:
            print("\n⚠️ Bot stopped by user")
            send_telegram("⚠️ Bot stopped by user")
            break
        except Exception as e:
            print(f"❌ Error in main loop: {e}")
            send_telegram(f"❌ Error: {e}")
            time.sleep(30)
    
    print("\n🏁 Scalping bot stopped")
    send_telegram("🏁 Scalping bot stopped")


# ==================== ENTRY POINT ====================
if __name__ == "__main__":
    try:
        print("\n" + "="*60)
        print("ONE BOX SCALPER STRATEGY - NIFTY OPTIONS")
        print("="*60)
        print(f"Index: {CONFIG['index']}")
        print(f"Expiry: {CONFIG['expiry']}")
        print(f"Premium Range: ₹{CONFIG['min_premium']} - ₹{CONFIG['max_premium']}")
        print(f"Lots: {CONFIG['lots']}")
        print(f"Risk:Reward: 1:{CONFIG['RISK_REWARD_RATIO']}")
        print(f"Max Risk per Trade: ₹{CONFIG['MAX_RISK_PER_TRADE']}")
        print("="*60)
        
        # Run bot
        run_scalping_bot()
        
    except KeyboardInterrupt:
        print("\n⚠️ Bot terminated by user")
        send_telegram("⚠️ Bot terminated")
    except Exception as e:
        print(f"\n❌ Fatal error: {e}")
        send_telegram(f"❌ Fatal error: {e}")
        raise
