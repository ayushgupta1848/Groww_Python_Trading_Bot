#!/usr/bin/env python3
from __future__ import annotations
"""
MOMENTUM_AUTO_BOT.py
====================
Short-term premium velocity auto-trader for NIFTY/SENSEX options.

Strategy
--------
Phase 1 — Discover: fetch all ATM±range CE and PE LTPs; keep only
  those whose premium falls within [min_premium, max_premium].
Phase 2 — Observe: poll all discovered strikes every scan_poll_sec
  for scan_seconds, printing a live table each second.
Phase 3 — Decide: score each side (CE / PE) by velocity × consistency
  over the observation window; enter BUY on the winning strike.
Phase 4 — Manage: trail SL with hard floor; log and notify on exit.

OI Integration (optional)
--------------------------
Reads oi_snapshot.json written by calculate_oi_pcr.py.
Uses OI writer bias as a soft directional filter:
  BULLISH → prefer CE entries  |  BEARISH → prefer PE entries
  NEUTRAL → scan both sides and let momentum decide.

Config Keys (edit CONFIG dict below)
-------------------------------------
index, lots, strike_step, atm_range
scan_seconds        — observe all premiums for this many seconds (20)
scan_poll_sec       — poll interval during observation window (1s)
poll_seconds        — poll interval during trail SL loop (3s)
velocity_pct        — min total % premium move over scan window (0.5%)
consistency_pct     — min % of ticks moving in signal direction (70%)
HARD_SL_POINTS      — premium points hard SL below entry (8 pts)
TRAIL_START_PROFIT  — start trailing after this profit in points (1)
TRAIL_STEP          — trail gap behind peak, fixed mode (0.75 pts)
TRAIL_SL_ATR_BASED  — if True, trail step = ATR × TRAIL_SL_ATR_MULTIPLIER
max_hold_min        — force exit after N minutes (30)
cooldown_sec        — wait after trade before next scan (90s)
use_oi_filter       — use oi_snapshot.json as bias (True)
validate_orders     — wait for EXECUTED status (True for live)
"""

import os
import re
import sys
import csv
import json
import time
import threading
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pyotp
import requests

# ============================================================
# 1. LOGGING
# ============================================================
def _ts() -> str:
    """Current time as HH:MM:SS.mmm (millisecond precision)."""
    n = datetime.now()
    return n.strftime(f"%H:%M:%S.{n.microsecond // 1000:03d}")


def _setup_logger():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    log_dir = os.path.join(base_dir, "logs", "momentum_bot")
    os.makedirs(log_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_path = os.path.join(log_dir, f"Momentum_Bot_{ts}.log")

    class Tee:
        def __init__(self, *streams):
            self.streams = streams
        def write(self, d):
            for s in self.streams:
                try: s.write(d)
                except Exception: pass
        def flush(self):
            for s in self.streams:
                try: s.flush()
                except Exception: pass

    lf = open(log_path, "a", buffering=1, encoding="utf-8")
    sys.stdout = Tee(sys.stdout, lf)
    sys.stderr = Tee(sys.stderr, lf)
    print(f"[MOM-BOT] Logging → {log_path}")
    return log_path

LOG_PATH = _setup_logger()

# ============================================================
# 2. CREDENTIALS  — fill these in
# ============================================================
API_KEY     = "eyJraWQiOiJaTUtjVXciLCJhbGciOiJFUzI1NiJ9.eyJleHAiOjI1NjQ2NTczODEsImlhdCI6MTc3NjI1NzM4MSwibmJmIjoxNzc2MjU3MzgxLCJzdWIiOiJ7XCJ0b2tlblJlZklkXCI6XCJjMjAzMmM5MS04ZGYzLTRkZDUtYjc5NS0yMGVlOWRhZDhhZjlcIixcInZlbmRvckludGVncmF0aW9uS2V5XCI6XCJlMzFmZjIzYjA4NmI0MDZjODg3NGIyZjZkODQ5NTMxM1wiLFwidXNlckFjY291bnRJZFwiOlwiMmVlMjYyMjItN2MwNS00Y2IwLWIwM2MtNzAzYWRmNWVmN2RkXCIsXCJkZXZpY2VJZFwiOlwiNjA2MzE5M2QtZWZkMC01OWViLTgzYzQtNWQ2NGZkNzdkNzQ3XCIsXCJzZXNzaW9uSWRcIjpcIjI0OWQ2OGRlLTNjZTgtNGQ4OS05ODJkLWM0N2NmYmI1YzdlNFwiLFwiYWRkaXRpb25hbERhdGFcIjpcIno1NC9NZzltdjE2WXdmb0gvS0EwYktvMDZXRlpjc241VUNmTWF5aERtNGxSTkczdTlLa2pWZDNoWjU1ZStNZERhWXBOVi9UOUxIRmtQejFFQisybTdRPT1cIixcInJvbGVcIjpcImF1dGgtdG90cFwiLFwic291cmNlSXBBZGRyZXNzXCI6XCIyNDA5OjQwYzQ6MTBhMzozN2UzOjE4NGI6N2IyOTpiMzBlOjIwZTUsMTcyLjcwLjIxOC4xMzUsMzUuMjQxLjIzLjEyM1wiLFwidHdvRmFFeHBpcnlUc1wiOjI1NjQ2NTczODE2ODYsXCJ2ZW5kb3JOYW1lXCI6XCJncm93d0FwaVwifSIsImlzcyI6ImFwZXgtYXV0aC1wcm9kLWFwcCJ9.3kotfZI_EC0lzszHKlXiRdqEQv-O8ubYFh0pgoAT0KsSfdQ1sHmts5UtlaAq4PB6DEwY4X2jZUCD8uBgc2nwXQ"
TOTP_SECRET = "SC3YMFLEGLHBWUPHRBOYLPEEOVAT2PZ4"
from whatsapp_gateway import send_whatsapp as send_telegram, start_webhook_server

# ============================================================
# 3. CONFIG
# ============================================================
PROJECT_ROOT     = os.path.dirname(os.path.abspath(__file__))
CSV_PATH         = os.path.join(PROJECT_ROOT, "instrument.csv")
OI_SNAPSHOT_PATH = os.path.join(PROJECT_ROOT, "oi_snapshot.json")
EXCEL_FILE       = "Lakshmi.xlsx"

def _next_expiry(weekday=1):
    """Next occurrence of weekday (0=Mon…6=Sun). NIFTY=Tuesday(1)."""
    today = datetime.now().date()
    days = (weekday - today.weekday()) % 7 or 7
    return (today + timedelta(days=days)).strftime("%Y-%m-%d")

CONFIG = {
    "index":         "NIFTY",
    "expiry":        _next_expiry(weekday=1),  # Tuesday for NIFTY
    "strike_step":   50,    # 50 for NIFTY; 100 for SENSEX/BANKNIFTY
    "atm_range":     3,     # scan ATM ± this many strikes (CE and PE)
    "lots":          1,
    "min_premium":   50,    # ignore strikes cheaper than this (₹)
    "max_premium":   200,   # ignore strikes more expensive than this (₹)

    # --- momentum detection ---
    "scan_seconds":    10,   # observe all premiums for this many seconds before deciding
    "scan_poll_sec":    1,   # poll interval during observation window (seconds)
    "poll_seconds":     1,   # poll interval during trail SL loop (seconds)
    "velocity_pct":   0.5,   # min total % premium move over scan window to trigger
    "consistency_pct": 55,   # min % of ticks that moved in the signal direction

    # --- trade management ---
    "HARD_SL_POINTS":         8.0,   # hard SL below entry (premium pts)
    "TRAIL_START_PROFIT":     1.0,   # start trailing after this profit (pts)
    "TRAIL_STEP":             0.75,  # trail gap behind peak when not ATR-based (pts)
    "TRAIL_SL_ATR_BASED":     False, # True → trail step = ATR × multiplier
    "TRAIL_SL_ATR_MULTIPLIER":1.0,   # ATR multiplier (only when ATR-based)
    "QUICK_TRAIL_BUFFER":     1.0,   # pts above target before switching to tight trail
    "QUICK_TRAIL_GAP":        1.5,   # pts below peak for tight trail stop
    "max_hold_min":           30,    # force-exit after N minutes
    "cooldown_sec":          120,    # wait after a trade before scanning again (2 min)
    "no_signal_wait_sec":     60,    # wait after no-signal scan before restarting (1 min)

    # --- safety ---
    "max_trades_day":     5,
    "validate_orders":    True,  # True = live, False = test/simulate
    "use_oi_filter":      True,  # use oi_snapshot.json writer bias
    "oi_max_age_sec":     120,   # ignore OI snapshot older than this

    # --- trade mode ---
    # "paper": no real orders, no Telegram; trail loop uses real LTP
    # "mock":  no real orders, Telegram IS sent (full simulation with notifications)
    # "live":  real orders via Groww API (default for production)
    "trade_mode": "paper",

    # --- exit mode ---
    # "manual": full trailing SL (default)
    # "quick":  exit immediately when profit >= TRAIL_START_PROFIT
    "exit_mode": "manual",

    # --- consecutive Hard SL circuit breaker (always active unless disabled) ---
    "consec_sl_brake":      True,  # pause entries after N consecutive Hard SLs
    "consec_sl_pause_min":  30,    # minutes to pause when circuit trips

    # --- ATR-based Hard SL ---
    "HARD_SL_ATR_BASED":      False,  # True = hard SL = ATR × multiplier instead of fixed pts
    "HARD_SL_ATR_MULTIPLIER":   1.5,  # multiplier applied to ATR
    "atr_source":          "candle",  # "candle" = real 14-period EMA ATR from 1-min historical candles (PROD10 style)
                                      # "scan"   = observation-window high-low range × multiplier (original, floor 3 pts)

    # --- score / velocity filters ---
    "min_score_filter":          True,  # True = winning side needs score ≥ vel×cons threshold (0.275); False = pick highest positive score regardless
    "velocity_filter":           True,  # True = best strike must clear vel_thresh (%); False = pick highest-score rising strike regardless

    # --- choppiness detection ---
    "choppiness_enabled":        True,  # master toggle — False = skip all choppiness logic
    "choppiness_window":         6,     # number of recent scans to evaluate
    "choppiness_flip_threshold": 0.55,  # direction flip rate above this = HIGH
    "choppiness_spread_min":     1.5,   # avg |CE-PE| score below this = indecisive
    "choppiness_pause_min":      15,    # minutes to pause entries when HIGH choppiness
    "max_consecutive_hard_sl":   2,     # consecutive Hard SLs triggers HIGH + pause
}

# ── config override written by LIVE_DASHBOARD (on launch and live toggle changes) ──
_override_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "momentum_config_override.json")
_OVERRIDE_CAST = {
    "trade_mode":      str,   "index":          str,
    "expiry":          str,   "lots":           int,
    "exit_mode":       str,   "min_premium":    float,
    "max_premium":     float, "atm_range":      int,
    "velocity_pct":    float, "consistency_pct":float,
    "validate_orders": bool,
    "scan_seconds":    int,   "poll_seconds":   int,
    "consec_sl_brake":        bool,
    "consec_sl_pause_min":    int,
    "HARD_SL_ATR_BASED":      bool,
    "HARD_SL_ATR_MULTIPLIER": float,
    "atr_source":             str,
    "min_score_filter":           bool,
    "velocity_filter":            bool,
    "choppiness_enabled":         bool,
    "choppiness_window":          int,
    "choppiness_flip_threshold":  float,
    "choppiness_spread_min":      float,
    "choppiness_pause_min":       int,
    "max_consecutive_hard_sl":    int,
}


_last_vix_note = ""   # track last logged VIX note to avoid duplicate spam

def _reload_override(verbose=True):
    """Re-read momentum_config_override.json and update CONFIG in-place.
    Called once at startup (verbose=True) and silently at each scan cycle."""
    global _last_vix_note
    if not os.path.exists(_override_path):
        return
    try:
        with open(_override_path) as _f:
            _ov = json.load(_f)
        _applied = {}
        for _k, _cast in _OVERRIDE_CAST.items():
            if _k in _ov:
                try:
                    CONFIG[_k] = _cast(_ov[_k])
                    _applied[_k] = CONFIG[_k]
                except Exception:
                    pass
        if _applied and verbose:
            print(f"[CONFIG] Override applied from UI: {_applied}")
        # Log VIX auto config note whenever it changes (not just on startup)
        note = _ov.get("_vix_config_note", "")
        if note and note != _last_vix_note:
            print(f"[VIX AUTO CONFIG] {note}")
            _last_vix_note = note
    except Exception as _oe:
        if verbose:
            print(f"[CONFIG] Override read error: {_oe}")


_reload_override(verbose=True)

# ── Choppiness Tracker ────────────────────────────────────────────────────────
# Detects sideways / choppy market using rolling scan history.
# Three signals: direction flip rate, CE/PE score spread, consecutive Hard SL count.
_last_scores: dict = {"ce_net": 0.0, "pe_net": 0.0}


class _ChopTracker:
    """Rolling choppiness evaluator. Holds last N scan results and Hard SL streak."""

    def __init__(self):
        self._scans          = []   # list of {dir: 'CE'/'PE'/None, spread: float}
        self._hard_sl_streak = 0
        self._pause_until    = None  # datetime or None

    def record(self, direction, ce_net, pe_net):
        """Call once per scan cycle after analyze_momentum() returns."""
        self._scans.append({"dir": direction, "spread": abs(ce_net - pe_net)})
        window = CONFIG.get("choppiness_window", 6)
        if len(self._scans) > window:
            self._scans.pop(0)

    def record_hard_sl(self):
        self._hard_sl_streak += 1

    def reset_hard_sl(self):
        self._hard_sl_streak = 0

    def is_expiry_day(self):
        return datetime.now().strftime("%Y-%m-%d") == CONFIG.get("expiry", "")

    def trigger_pause(self):
        mins = CONFIG.get("choppiness_pause_min", 15)
        self._pause_until = datetime.now() + timedelta(minutes=mins)
        return self._pause_until.strftime("%H:%M")

    def evaluate(self):
        result = {
            "level": "LOW", "flip_rate": 0.0, "avg_spread": 0.0,
            "hard_sl_streak": self._hard_sl_streak,
            "paused": False, "pause_until": None, "reason": "",
        }

        # Check active pause
        if self._pause_until:
            if datetime.now() < self._pause_until:
                result["paused"]      = True
                result["pause_until"] = self._pause_until.strftime("%H:%M")
                result["level"]       = "HIGH"
                result["reason"]      = f"choppiness pause until {result['pause_until']}"
                return result
            self._pause_until = None  # expired

        n = len(self._scans)
        if n < 3:
            result["reason"] = f"accumulating scan history ({n}/3)"
            return result

        # Direction flip rate
        dirs      = [s["dir"] for s in self._scans if s["dir"] is not None]
        flips     = sum(1 for i in range(1, len(dirs)) if dirs[i] != dirs[i - 1])
        flip_rate = flips / max(len(dirs) - 1, 1) if len(dirs) >= 2 else 0.0

        # Average CE/PE score spread
        avg_spread = sum(s["spread"] for s in self._scans) / n

        result["flip_rate"]  = round(flip_rate, 2)
        result["avg_spread"] = round(avg_spread, 3)

        # Thresholds — tighter on expiry day
        flip_thresh = CONFIG.get("choppiness_flip_threshold", 0.55)
        spread_min  = CONFIG.get("choppiness_spread_min", 1.5)
        max_hard_sl = CONFIG.get("max_consecutive_hard_sl", 2)
        on_expiry   = self.is_expiry_day()
        if on_expiry:
            flip_thresh = round(flip_thresh * 0.85, 3)  # 0.55 → ~0.47
            spread_min  = round(spread_min  * 1.25, 3)  # 1.50 → ~1.88

        chop_score = 0
        reasons    = []

        if flip_rate > flip_thresh:
            chop_score += 2
            reasons.append(f"direction flipping {flip_rate:.0%} of scans (>{flip_thresh:.0%})")
        elif flip_rate > flip_thresh * 0.75:
            chop_score += 1
            reasons.append(f"direction unstable ({flip_rate:.0%} flip rate)")

        if avg_spread < spread_min:
            chop_score += 2
            reasons.append(f"CE/PE scores too close — avg spread {avg_spread:.2f} < {spread_min:.2f}")
        elif avg_spread < spread_min * 1.3:
            chop_score += 1
            reasons.append(f"CE/PE indecisive — avg spread {avg_spread:.2f}")

        if self._hard_sl_streak >= max_hard_sl:
            chop_score += 3
            reasons.append(f"{self._hard_sl_streak} consecutive Hard SLs — circuit breaker")

        if on_expiry and datetime.now().hour >= 13 and datetime.now().minute >= 30:
            chop_score += 1
            reasons.append("expiry day after 1:30 PM (premium decay risk)")

        if chop_score >= 3:
            result["level"] = "HIGH"
        elif chop_score >= 1:
            result["level"] = "MEDIUM"

        result["reason"] = ", ".join(reasons) if reasons else "clean trend"
        return result


_chop = _ChopTracker()

# Consecutive Hard SL circuit breaker — independent of choppiness tracker
_sl_circuit: dict = {"count": 0, "pause_until": None}


def _print_choppiness(state: dict) -> None:
    level  = state["level"]
    icon   = {"LOW": "✅", "MEDIUM": "⚡", "HIGH": "🔴"}.get(level, "⏸")
    flip   = state.get("flip_rate", 0)
    spread = state.get("avg_spread", 0)
    streak = state.get("hard_sl_streak", 0)
    status = f"PAUSED until {state['pause_until']}" if state.get("paused") else level
    print(f"\n  {icon} Choppiness: {status:<30}  "
          f"flip={flip:.0%}  spread={spread:.2f}  hard_sl_streak={streak}")
    if level in ("MEDIUM", "HIGH") or state.get("paused"):
        print(f"     💬 {state['reason']}")


# ============================================================
# 4. GROWW INIT
# ============================================================
try:
    from growwapi import GrowwAPI
except ImportError:
    print("❌ growwapi not found. pip install growwapi or place it in PYTHONPATH.")
    sys.exit(1)

_totp_gen = pyotp.TOTP(TOTP_SECRET)
_session  = requests.Session()
_ltp_lock = threading.Lock()

groww       = None
access_token = None

from groww_token import get_access_token as get_cached_access_token


def groww_init():
    global groww, access_token
    access_token = get_cached_access_token(API_KEY, TOTP_SECRET)
    groww = GrowwAPI(access_token)
    print(f"✅ Groww API initialised  [{datetime.now().strftime('%H:%M:%S')}]")

# ============================================================
# 5. WHATSAPP (see whatsapp_gateway.py)
# ============================================================

# ============================================================
# 6. INSTRUMENTS
# ============================================================
instruments_data: list = []

def _load_instruments():
    global instruments_data
    rows = []
    with open(CSV_PATH, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    INDEX  = CONFIG["index"].upper()
    EXPIRY = CONFIG["expiry"]
    # Auto-resolve strike step from index — overrides CONFIG default so SENSEX (100) works correctly
    step   = _INDEX_STRIKE_STEP.get(INDEX, CONFIG["strike_step"])
    CONFIG["strike_step"] = step

    spot = _get_spot()
    atm  = round(spot / step) * step
    lo   = atm - 20 * step
    hi   = atm + 20 * step

    filtered = []
    for r in rows:
        if r.get("underlying_symbol", "").upper() != INDEX:
            continue
        if r.get("expiry_date", "").strip() != EXPIRY:
            continue
        try:
            strike = float(r.get("strike_price") or 0)
        except ValueError:
            continue
        if lo <= strike <= hi:
            filtered.append(r)

    instruments_data = filtered
    print(f"✅ Loaded {len(instruments_data)} instruments  (spot={spot}, ATM={atm})")
    return spot, atm

def _find_instrument(strike: float, opt_type: str):
    for inst in instruments_data:
        if (float(inst.get("strike_price") or 0) == strike
                and inst.get("instrument_type", "").upper() == opt_type.upper()):
            return inst
    return None

# ============================================================
# 7. LTP HELPERS
# ============================================================
def _get_ltp(instrument) -> float | None:
    trading_symbol = (instrument.get("internal_trading_symbol")
                      or instrument.get("trading_symbol"))
    if not trading_symbol:
        return None
    exchange_symbol = f"NSE_{trading_symbol}"
    url = (f"https://api.groww.in/v1/live-data/ltp"
           f"?segment=FNO&exchange_symbols={exchange_symbol}")
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0",
    }
    with _ltp_lock:
        resp = _session.get(url, headers=headers, timeout=8)
        time.sleep(0.05)
    if resp.status_code != 200:
        return None
    payload = resp.json().get("payload", {})
    val = payload.get(exchange_symbol)
    return float(val) if val is not None else None

_INDEX_SPOT_SYMBOL = {
    "NIFTY":     "NSE_NIFTY",
    "BANKNIFTY": "NSE_BANKNIFTY",
    "FINNIFTY":  "NSE_FINNIFTY",
    "SENSEX":    "BSE_SENSEX",
    "BANKEX":    "BSE_BANKEX",
}
_INDEX_STRIKE_STEP = {
    "NIFTY":     50,
    "BANKNIFTY": 100,
    "FINNIFTY":  50,
    "SENSEX":    100,
    "BANKEX":    100,
}

def _get_spot() -> float:
    """Fetch index spot LTP based on CONFIG['index']."""
    idx = CONFIG.get("index", "NIFTY").upper()
    sym = _INDEX_SPOT_SYMBOL.get(idx, "NSE_NIFTY")
    try:
        url = (f"https://api.groww.in/v1/live-data/ltp"
               f"?segment=CASH&exchange_symbols={sym}")
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0",
        }
        resp = _session.get(url, headers=headers, timeout=8)
        val = resp.json().get("payload", {}).get(sym)
        return float(val) if val else 0.0
    except Exception:
        return 0.0

# ============================================================
# 8. ORDER HELPERS  (mirrors master bot pattern)
# ============================================================
def _place_market_order(instrument, qty, side="BUY"):
    trading_symbol = (instrument.get("internal_trading_symbol")
                      or instrument.get("trading_symbol"))
    return groww.place_order(
        trading_symbol=trading_symbol,
        quantity=qty,
        validity=groww.VALIDITY_DAY,
        exchange=groww.EXCHANGE_NSE,
        segment=groww.SEGMENT_FNO,
        product=groww.PRODUCT_MIS,
        order_type=groww.ORDER_TYPE_MARKET,
        transaction_type=getattr(groww, f"TRANSACTION_TYPE_{side}"),
        price=0,
    )

def _get_order_status(order_id) -> str | None:
    url = f"https://api.groww.in/v1/order/status/{order_id}?segment=FNO"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0",
    }
    try:
        data = _session.get(url, headers=headers, timeout=8).json()
        return data.get("payload", {}).get("order_status")
    except Exception:
        return None

def _wait_executed(order_id, side="BUY") -> bool:
    print(f"  ⏳ Waiting for {side} order {order_id} …")
    while True:
        status = _get_order_status(order_id)
        print(f"    status: {status}")
        if status in ("EXECUTED", "COMPLETED", "DELIVERY_AWAITED"):
            return True
        if status in ("FAILED", "REJECTED", "CANCELLED"):
            print(f"  ❌ {side} order {status}")
            return False
        time.sleep(2)

def _get_executed_price(order_id) -> tuple[float | None, int | None]:
    """Fetch actual executed price from Groww trades API — same logic as PROD10FEB."""
    import requests as _req
    url = (f"https://api.groww.in/v1/order/trades/{order_id}"
           f"?segment=FNO&page=0&page_size=50")
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0",
    }
    try:
        print(f"\n  📦 Fetching trade details for order: {order_id}")
        trades = []
        for _attempt in range(4):          # retry up to 4×, 500ms apart
            if _attempt > 0:
                time.sleep(0.5)
            response = _req.get(url, headers=headers)
            data = response.json()
            if data.get("status") != "SUCCESS":
                print(f"  ⚠️ Failed to fetch trade info: {data}")
                return None, None
            trades = data.get("payload", {}).get("trade_list", [])
            if trades:
                break
            print(f"  ⏳ Trade list empty — retrying ({_attempt + 1}/4)…")
        if not trades:
            print(f"  ⚠️ No trades found for order ID after retries.")
            return None, None
        total_qty   = sum(t["quantity"] for t in trades)
        total_value = sum(t["price"] * t["quantity"] for t in trades)
        avg_price   = round(total_value / total_qty, 2)
        symbol = trades[0]["trading_symbol"]
        side   = trades[0]["transaction_type"]
        print(f"  ✅ {side} {symbol} | Total Qty={total_qty} | Avg Price=₹{avg_price}")
        return avg_price, total_qty
    except Exception as e:
        print(f"  ❌ Error fetching order trades: {e}")
        return None, None

# ============================================================
# 9. EXCEL LOGGING
# ============================================================
def _log_excel(symbol, buy_px, sell_px, qty, profit):
    from openpyxl import Workbook, load_workbook
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Momentum_Trades"
        ws.append(["DateTime", "Symbol", "Buy Price", "Sell Price",
                   "Qty", "Profit", "Source"])
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
               symbol, buy_px, sell_px, qty, round(profit, 2), "MOMENTUM_BOT"])
    wb.save(EXCEL_FILE)

# ============================================================
# 10. OI SNAPSHOT READER
# ============================================================
def load_oi_snapshot() -> dict | None:
    """Load oi_snapshot.json written by calculate_oi_pcr.py.
    Returns None if file missing or too stale."""
    if not CONFIG.get("use_oi_filter"):
        return None
    try:
        with open(OI_SNAPSHOT_PATH) as f:
            snap = json.load(f)
        age = time.time() - snap.get("timestamp", 0)
        if age > CONFIG["oi_max_age_sec"]:
            print(f"  ⚠️  OI snapshot is {age:.0f}s old — ignoring bias")
            return None
        return snap
    except Exception:
        return None

def oi_bias(snap) -> str:
    """Returns 'BULLISH', 'BEARISH', or 'NEUTRAL' from OI snapshot."""
    if snap is None:
        return "NEUTRAL"
    # Prefer writer_bias (from tick-over-tick analysis) if available
    writer = snap.get("writer_bias", "NEUTRAL")
    sentiment = snap.get("sentiment", "NEUTRAL")
    # If they agree, strong signal; if they disagree, use writer_bias
    if writer == sentiment:
        return writer
    # PCR as tiebreak
    pcr = snap.get("pcr_atm", 1.0)
    if pcr > 1.1:
        return "BULLISH"
    if pcr < 0.9:
        return "BEARISH"
    return writer  # writer_bias wins by default

# ============================================================
# 11. SCAN HELPERS
# ============================================================

def _ordinal(n: int) -> str:
    if 11 <= n <= 13:
        return f"{n}th"
    return f"{n}{('th','st','nd','rd','th','th','th','th','th','th')[n % 10]}"


def discover_candidates(spot: float) -> list:
    """
    Fetch LTP for every ATM±range CE and PE strike.
    Return list of (strike, opt_type, instrument_dict, ltp) for those
    whose premium falls inside [min_premium, max_premium].
    """
    cfg  = CONFIG
    step = cfg["strike_step"]
    n    = cfg["atm_range"]
    atm  = round(spot / step) * step

    # Scan ALL ±n strikes from ATM for BOTH CE and PE (OI bias is logged in
    # the main loop for info only — filtering by bias here blocks one entire
    # side from scanning, missing ITM options that may be in the premium range).
    seen = set()
    pairs = []
    for i in range(0, n + 1):
        for strike in (atm + i * step, atm - i * step):
            for opt_type in ("CE", "PE"):
                key = (strike, opt_type)
                if key not in seen:
                    seen.add(key)
                    pairs.append(key)

    def _fetch(strike, opt_type):
        inst = _find_instrument(strike, opt_type)
        if inst is None:
            return None
        ltp = _get_ltp(inst)
        if ltp and cfg["min_premium"] <= float(ltp) <= cfg["max_premium"]:
            return (strike, opt_type, inst, float(ltp))
        return None

    with ThreadPoolExecutor(max_workers=min(len(pairs), 20)) as ex:
        results = list(ex.map(lambda p: _fetch(*p), pairs))

    return [r for r in results if r is not None]


def run_observation(candidates: list) -> dict:
    """
    Poll all candidates every scan_poll_sec for scan_seconds.
    Prints a table each second like the user's log format.
    Returns history dict: {(strike, opt_type): [ltp1, ltp2, ...]}
    """
    cfg        = CONFIG
    scan_secs  = cfg["scan_seconds"]
    poll_sec   = cfg["scan_poll_sec"]
    inst_map   = {(s, t): inst for s, t, inst, _ in candidates}
    _is_mock   = cfg.get("trade_mode", "live").lower() == "mock"

    # For mock simulation: use discovered prices as base, then apply per-tick delta
    # CE side rises (+0.6 pts/tick) → strong upward momentum → wins signal
    # PE side drifts down (-0.2 pts/tick) → no upward momentum
    _sim_bases = {(s, t): ltp0 for s, t, _, ltp0 in candidates} if _is_mock else {}

    ce_keys = sorted([(s, t) for s, t, _, _ in candidates if t == "CE"])
    pe_keys = sorted([(s, t) for s, t, _, _ in candidates if t == "PE"])

    history: dict = {(s, t): [] for s, t, _, _ in candidates}

    if _is_mock:
        print(f"\nScanning---- 🎭 [MOCK SIMULATION — prices are simulated]")
    else:
        print(f"\nScanning----")

    for tick in range(1, scan_secs + 1):
        t0 = time.time()

        def _poll_one(key, _tick=tick):
            try:
                if _is_mock:
                    base = _sim_bases.get(key, 100.0)
                    _, t = key
                    delta = _tick * 0.6 if t == "CE" else -_tick * 0.2
                    return (key, round(base + delta, 2))
                ltp = _get_ltp(inst_map[key])
                return (key, float(ltp) if ltp else None)
            except Exception:
                return (key, None)

        with ThreadPoolExecutor(max_workers=min(len(candidates), 20)) as ex:
            tick_results = list(ex.map(_poll_one, list(inst_map.keys())))

        for key, ltp in tick_results:
            if ltp is not None:
                history[key].append(ltp)

        ce_vals = [f"{history[k][-1]:.2f}" for k in ce_keys if history[k]]
        pe_vals = [f"{history[k][-1]:.2f}" for k in pe_keys if history[k]]

        sim_tag = " 🎭" if _is_mock else ""
        print(f"\n{_ordinal(tick)} second{sim_tag}  [{datetime.now().strftime('%H:%M:%S')}]")
        if ce_vals:
            print(f"  CE = {',  '.join(ce_vals)}")
        if pe_vals:
            print(f"  PE = {',  '.join(pe_vals)}")

        elapsed = time.time() - t0
        wait = poll_sec - elapsed
        if wait > 0:
            time.sleep(wait)

    return history


def analyze_momentum(candidates: list, history: dict):
    """
    After observation, score each strike by velocity + consistency.
    Pick the winning side (CE or PE) and the best strike on that side.
    Returns a signal dict (with 'inst' key) or None.
    """
    cfg         = CONFIG
    vel_thresh  = cfg["velocity_pct"]
    cons_thresh = cfg["consistency_pct"]

    strike_scores = {}

    for s, t, inst, _ in candidates:
        ticks = history.get((s, t), [])
        if len(ticks) < 3:
            continue

        first, last = ticks[0], ticks[-1]
        if first <= 0:
            continue

        velocity  = (last - first) / first * 100
        deltas    = [ticks[i + 1] - ticks[i] for i in range(len(ticks) - 1)]
        direction = 1 if velocity > 0 else -1
        same      = sum(1 for d in deltas if (d > 0) == (direction > 0))
        consistency = same / len(deltas) * 100 if deltas else 0.0
        score     = abs(velocity) * (consistency / 100)

        strike_scores[(s, t)] = {
            "strike":          s,
            "opt_type":        t,
            "inst":            inst,
            "velocity_pct":    round(velocity, 3),
            "consistency_pct": round(consistency, 1),
            "score":           round(score, 4),
            "direction":       "UP" if velocity > 0 else "DOWN",
            "entry_ltp":       last,
        }

    # Side net scores (positive = net upward momentum on that side)
    def _side_net(side):
        vals = [v["velocity_pct"] * (v["consistency_pct"] / 100)
                for v in strike_scores.values() if v["opt_type"] == side]
        return sum(vals) / len(vals) if vals else 0.0

    ce_net = _side_net("CE")
    pe_net = _side_net("PE")
    _last_scores["ce_net"] = ce_net   # expose to choppiness tracker
    _last_scores["pe_net"] = pe_net

    print(f"\n  📊 Momentum analysis after {cfg['scan_seconds']}s:")
    print(f"     CE net score = {ce_net:+.3f}  |  PE net score = {pe_net:+.3f}")

    if ce_net <= 0 and pe_net <= 0:
        print("  ❌ Both sides showing no upward momentum")
        return None

    winning_side = "CE" if ce_net >= pe_net else "PE"
    winning_net  = ce_net if winning_side == "CE" else pe_net

    # Net score quality gate — skipped when min_score_filter is OFF
    if cfg.get("min_score_filter", True):
        min_net = vel_thresh * (cons_thresh / 100)
        if winning_net < min_net:
            print(f"  ❌ {winning_side} net score {winning_net:+.3f} below min {min_net:.3f} — too weak")
            return None
    else:
        print(f"  ℹ️  Min score filter OFF — picking highest positive side regardless of score")

    # Best strike on winning side — velocity filter toggleable
    if cfg.get("velocity_filter", True):
        candidates_on_side = [
            v for v in strike_scores.values()
            if v["opt_type"] == winning_side
            and v["direction"] == "UP"
            and abs(v["velocity_pct"]) >= vel_thresh
        ]
        if not candidates_on_side:
            print(f"  ❌ {winning_side} side: no strike cleared vel≥{vel_thresh}%")
            return None
    else:
        candidates_on_side = [
            v for v in strike_scores.values()
            if v["opt_type"] == winning_side and v["direction"] == "UP"
        ]
        if not candidates_on_side:
            print(f"  ❌ {winning_side} side: no rising strike found")
            return None
        print(f"  ℹ️  Velocity filter OFF — picking highest-score rising strike on {winning_side}")

    candidates_on_side.sort(key=lambda x: x["score"], reverse=True)
    best = candidates_on_side[0]

    # Attach scan-window ATR (high-low range of observed ticks) to signal
    best_ticks = history.get((best["strike"], best["opt_type"]), [])
    best["scan_atr"] = round(max(best_ticks) - min(best_ticks), 2) if len(best_ticks) >= 2 else None

    print(f"  ✅ Signal → {winning_side} {int(best['strike'])}  "
          f"vel={best['velocity_pct']:+.2f}%  "
          f"consistency={best['consistency_pct']:.0f}%  "
          f"LTP=₹{best['entry_ltp']:.2f}  [{_ts()}]")
    return best

# ============================================================
# 12. TRAIL HELPERS
# ============================================================
def _resolve_trail_step(atr_value):
    """Return effective trail step: ATR×multiplier if ATR-based mode, else fixed TRAIL_STEP."""
    cfg = CONFIG
    if cfg.get("TRAIL_SL_ATR_BASED", False) and atr_value and atr_value > 0:
        step = round(float(atr_value) * cfg.get("TRAIL_SL_ATR_MULTIPLIER", 1.0), 2)
        print(f"  📐 ATR trail step: ₹{step:.2f}  (ATR={atr_value:.2f} × {cfg.get('TRAIL_SL_ATR_MULTIPLIER', 1.0)})")
        return step
    return cfg["TRAIL_STEP"]

def _ema_of(values: list, period: int) -> float | None:
    """Compute EMA of a list of floats. Returns last EMA value or None."""
    if len(values) < period:
        return None
    k   = 2.0 / (period + 1)
    ema = sum(values[:period]) / period
    for v in values[period:]:
        ema = v * k + ema * (1 - k)
    return ema


def _real_atr_from_candles(highs: list, lows: list, closes: list, period: int = 14) -> float | None:
    """14-period EMA ATR from 1-min OHLC candle data (same method as PROD10)."""
    if len(highs) < period + 1:
        return None
    trs = []
    for i in range(1, len(highs)):
        tr = max(
            highs[i]  - lows[i],
            abs(highs[i]  - closes[i - 1]),
            abs(lows[i]   - closes[i - 1]),
        )
        trs.append(tr)
    return _ema_of(trs, period)


def _fetch_real_atr(instrument, timeout: int = 6) -> float | None:
    """Fetch last 60 min of 1-min candles and return 14-period EMA ATR.
    Non-blocking — designed to run in a background thread.
    Returns None on failure/timeout."""
    import queue as _queue
    q = _queue.Queue()

    def _worker():
        try:
            trading_symbol = (instrument.get("internal_trading_symbol")
                              or instrument.get("trading_symbol"))
            exchange  = instrument.get("exchange", "NSE").upper()
            exch_const = (groww.EXCHANGE_BSE if exchange == "BSE"
                          else groww.EXCHANGE_NSE)
            end_dt   = datetime.now()
            start_dt = end_dt - timedelta(minutes=60)
            hist = groww.get_historical_candles(
                groww_symbol=trading_symbol,
                exchange=exch_const,
                segment=groww.SEGMENT_FNO,
                start_time=start_dt.strftime("%Y-%m-%d %H:%M:%S"),
                end_time=end_dt.strftime("%Y-%m-%d %H:%M:%S"),
                candle_interval="1minute",
            )
            candles = (hist or {}).get("candles", [])
            if len(candles) < 20:
                q.put(None)
                return
            highs  = [c[2] for c in candles]
            lows   = [c[3] for c in candles]
            closes = [c[4] for c in candles]
            atr = _real_atr_from_candles(highs, lows, closes, period=14)
            q.put(round(atr, 2) if atr else None)
        except Exception:
            q.put(None)

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    t.join(timeout=timeout)
    return q.get() if not q.empty() else None

# ============================================================
# 13. TRADE MANAGER — entry + trail SL loop
# ============================================================

def _oi_verdict(oi_bias: str, opt_type: str, profit: float):
    """
    Return (display_line, verdict_tag) showing whether the OI filter
    would have helped or hurt on this trade.

    verdict_tag values (used in JSONL log + UI):
      ALIGNED_WIN    — OI agreed with trade direction AND trade won
      ALIGNED_LOSS   — OI agreed with trade direction BUT trade lost (OI was wrong)
      OPPOSED_WIN    — OI disagreed (filter would have blocked) but trade WON
      OPPOSED_LOSS   — OI disagreed and trade LOST (filter would have SAVED this)
      NEUTRAL        — OI snapshot was stale/unavailable
    """
    pnl_str = f"₹{profit:+,.0f}"
    if not oi_bias or oi_bias == "NEUTRAL":
        return (
            f"OI NEUTRAL/Stale — no directional data  ({pnl_str})",
            "NEUTRAL"
        )
    # aligned = trade direction agrees with OI bias
    aligned = (
        (oi_bias == "BEARISH" and opt_type == "PE") or
        (oi_bias == "BULLISH" and opt_type == "CE")
    )
    win = profit > 0
    if aligned and win:
        line = (f"✅ OI ALIGNED & WON   │ OI={oi_bias}  Trade={opt_type}  {pnl_str}"
                f"  →  Filter would ALLOW  ✓")
        tag  = "ALIGNED_WIN"
    elif aligned and not win:
        line = (f"⚠️  OI ALIGNED, STILL LOST │ OI={oi_bias}  Trade={opt_type}  {pnl_str}"
                f"  →  OI was wrong about the move")
        tag  = "ALIGNED_LOSS"
    elif not aligned and win:
        line = (f"🚫 OI OPPOSED, BUT WON │ OI={oi_bias}  Trade={opt_type}  {pnl_str}"
                f"  →  Filter WOULD HAVE BLOCKED this winner")
        tag  = "OPPOSED_WIN"
    else:
        line = (f"🛡️  OI OPPOSED & LOST   │ OI={oi_bias}  Trade={opt_type}  {pnl_str}"
                f"  →  Filter WOULD HAVE SAVED ₹{abs(profit):,.0f}")
        tag  = "OPPOSED_LOSS"
    return line, tag


def _log_trade_history(symbol: str, buy_price: float, sell_price: float,
                        qty: int, lots: int, pnl: float,
                        exit_reason: str, mode: str,
                        time_entry: str, time_exit: str,
                        oi_bias: str = "NEUTRAL",
                        oi_verdict_tag: str = "NEUTRAL") -> None:
    """Append a completed trade record to the daily trade-history JSONL log."""
    os.makedirs("logs/trade_history", exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    record = {
        "date":           date_str,
        "time_entry":     time_entry,
        "time_exit":      time_exit,
        "bot":            "Auto",
        "mode":           mode,
        "index":          CONFIG["index"],
        "symbol":         symbol,
        "expiry":         CONFIG["expiry"],
        "buy_price":      buy_price,
        "sell_price":     sell_price,
        "qty":            qty,
        "lots":           lots,
        "pnl":            round(pnl, 2),
        "exit_reason":    exit_reason or "",
        "oi_bias":        oi_bias,
        "oi_verdict_tag": oi_verdict_tag,
    }
    try:
        path = os.path.join("logs", "trade_history", f"{date_str}.jsonl")
        with open(path, "a", encoding="utf-8") as f:
            f.write(json.dumps(record) + "\n")
    except Exception as e:
        print(f"  ⚠️  Could not write trade history: {e}")


def execute_trade(instrument, signal: dict) -> bool:
    """Buy, monitor with trail SL, sell.  Returns True if trade attempted."""
    cfg        = CONFIG
    validate   = cfg["validate_orders"]
    mode       = cfg.get("trade_mode", "live").lower()   # "paper" | "mock" | "live"
    is_sim     = mode in ("paper", "mock")               # no real orders
    notify     = mode in ("mock", "live")                # send Telegram
    symbol     = (instrument.get("internal_trading_symbol")
                  or instrument.get("trading_symbol"))
    lot_size   = int(instrument.get("lot_size") or instrument.get("lotsize") or 1)
    qty        = cfg["lots"] * lot_size
    entry_ltp  = signal["entry_ltp"]
    opt_type   = signal["opt_type"]

    mode_tag = f"[{mode.upper()}] " if is_sim else ""
    print(f"\n{'='*65}")
    print(f"  📈  {mode_tag}MOMENTUM ENTRY — {symbol}  ({opt_type})")
    print(f"  Velocity={signal['velocity_pct']:+.2f}%  "
          f"Consistency={signal['consistency_pct']:.0f}%  "
          f"LTP=₹{entry_ltp:.2f}")
    print(f"{'='*65}")
    if notify:
        send_telegram(f"📈 {mode_tag}MOMENTUM ENTRY\n{symbol} ({opt_type})\n"
                      f"Velocity={signal['velocity_pct']:+.2f}%  LTP=₹{entry_ltp:.2f}")

    # ---- BUY ----
    avg_price  = entry_ltp
    _atr_queue = None   # will hold real ATR result queue when HARD_SL_ATR_BASED=True

    if is_sim:
        print(f"  📋 {mode_tag}BUY simulated @ ₹{avg_price:.2f}  Qty={qty}  [{_ts()}]")
    else:
        try:
            print(f"  📤 BUY request sent to Groww  [{_ts()}]")
            resp     = _place_market_order(instrument, qty, "BUY")
            order_id = (resp.get("payload", {}).get("groww_order_id")
                        or resp.get("groww_order_id"))
            print(f"  ✅ BUY placed  order_id={order_id}  [{_ts()}]")
        except Exception as e:
            print(f"  ❌ BUY failed: {e}  [{_ts()}]")
            send_telegram(f"❌ BUY failed: {e}")
            return False

        # Start real ATR fetch in background — overlaps with BUY validation wait
        # Only needed when atr_source == "candle"; scan ATR comes from signal dict
        if cfg.get("HARD_SL_ATR_BASED") and cfg.get("atr_source", "candle") == "candle":
            import queue as _q
            _atr_queue = _q.Queue()
            threading.Thread(
                target=lambda: _atr_queue.put(_fetch_real_atr(instrument, timeout=6)),
                daemon=True,
            ).start()
            print(f"  🔄 Fetching real ATR (1-min candles) in background…")

        if validate and order_id:
            ok = _wait_executed(order_id, "BUY")
            if not ok:
                send_telegram(f"❌ BUY not executed: {symbol}")
                return False
            p, q = _get_executed_price(order_id)
            if p:
                avg_price = p
                qty       = q or qty
            print(f"  ✅ BUY confirmed (executed) @ ₹{avg_price}  Qty={qty}  [{_ts()}]")
            send_telegram(f"✅ BUY EXECUTED @ ₹{avg_price}  Qty={qty}")

    entry_ts = _ts()

    # Scan-window ATR still used for trail step (short-term volatility measure)
    scan_atr    = signal.get("scan_atr")
    trail_step  = _resolve_trail_step(scan_atr)
    trail_start = cfg["TRAIL_START_PROFIT"]
    max_time_sec = cfg["max_hold_min"] * 60

    # Hard SL: two sources selectable via atr_source config key
    if cfg.get("HARD_SL_ATR_BASED") and not is_sim:
        mult = cfg.get("HARD_SL_ATR_MULTIPLIER", 1.5)
        atr_src = cfg.get("atr_source", "candle")
        if atr_src == "candle":
            # 14-period EMA ATR from 1-min historical candles (PROD10-style, no floor)
            _real_atr = None
            if _atr_queue is not None:
                try:
                    _real_atr = _atr_queue.get(timeout=4)
                except Exception:
                    pass
            if _real_atr:
                hard_sl_pts = round(_real_atr * mult, 2)
                print(f"  📐 Hist ATR Hard SL: {hard_sl_pts:.2f} pts  "
                      f"(1-min EMA ATR=₹{_real_atr:.2f} × {mult:.1f})")
            else:
                hard_sl_pts = cfg["HARD_SL_POINTS"]
                print(f"  ⚠️ Hist ATR unavailable — fixed {hard_sl_pts:.1f} pts Hard SL")
        else:
            # Tick Range ATR: high-low range from 15–25 sec live tick scan window
            if scan_atr:
                raw = round(scan_atr * mult, 2)
                hard_sl_pts = max(3.0, raw)
                print(f"  📐 Tick Range Hard SL: {hard_sl_pts:.2f} pts  "
                      f"(tick range={scan_atr:.2f} × {mult:.1f}, floor=3)")
            else:
                hard_sl_pts = cfg["HARD_SL_POINTS"]
                print(f"  ⚠️ Tick Range ATR unavailable — fixed {hard_sl_pts:.1f} pts Hard SL")
    else:
        hard_sl_pts = cfg["HARD_SL_POINTS"]

    hard_sl = round(avg_price - hard_sl_pts, 2)

    print(f"  Entry=₹{avg_price}  Qty={qty}  Hard SL=₹{hard_sl:.2f} ({hard_sl_pts:.1f} pts)  "
          f"Trail start=+{trail_start} pts  Trail step={trail_step} pts")
    if notify:
        send_telegram(f"📈 Trailing started | Entry=₹{avg_price:.2f} | Hard SL=₹{hard_sl:.2f}")

    # ---- TRAIL LOOP ----
    highest_price    = avg_price
    last_trail_exit  = None
    start_time       = time.time()
    last_heartbeat   = time.time()
    sell_reason      = None
    ltp              = avg_price
    _ltp_fail_streak = 0

    # Mock trail: scripted tick sequence (same pattern as PROD10 MOCK_LTP_RUN)
    # stable → rise → new high → pullback → trail SL hit
    _mock_trail_tick = 0
    def _next_mock_trail_ltp():
        nonlocal _mock_trail_tick
        t = _mock_trail_tick; _mock_trail_tick += 1
        if   t < 3: offset = 0.0    # ticks 0-2: stable at entry
        elif t < 7: offset = 5.0    # ticks 3-6: +5 pts, trail activates
        elif t < 9: offset = 8.0    # ticks 7-8: +8 pts, new high
        else:        offset = 3.0   # tick 9+:   +3 pts, below trail_exit → hit
        v = round(avg_price + offset, 2)
        print(f"  🎭 MOCK TRAIL tick={t}: ₹{v:.2f}")
        time.sleep(cfg["poll_seconds"])
        return v

    while True:
        # Heartbeat every 30s
        if time.time() - last_heartbeat >= 30:
            print(f"\n  💓 Monitoring... LTP last seen: ₹{ltp:.2f}")
            last_heartbeat = time.time()

        if mode == "mock":
            fetched = _next_mock_trail_ltp()
        else:
            try:
                fetched = _get_ltp(instrument)
            except Exception as e:
                _ltp_fail_streak += 1
                backoff = min(30, cfg["poll_seconds"] * (2 ** min(_ltp_fail_streak, 6)))
                if _ltp_fail_streak <= 3 or _ltp_fail_streak % 10 == 0:
                    print(f"\n  ⚠️ LTP error (streak={_ltp_fail_streak}): {e}")
                time.sleep(backoff)
                continue

            if fetched is None:
                _ltp_fail_streak += 1
                backoff = min(30, cfg["poll_seconds"] * (2 ** min(_ltp_fail_streak, 6)))
                time.sleep(backoff)
                continue

        _ltp_fail_streak = 0
        ltp = float(fetched)
        sell_reason = None

        # Check exit conditions
        if ltp <= hard_sl:
            sell_reason = f"🛑 HARD SL hit @ ₹{ltp:.2f}  [detected {_ts()}]"
        elif time.time() - start_time >= max_time_sec:
            sell_reason = f"⏰ Max hold time ({cfg['max_hold_min']}m) reached  [detected {_ts()}]"
        elif cfg.get("exit_mode") == "quick" and ltp >= avg_price + cfg["TRAIL_START_PROFIT"]:
            sell_reason = (f"🎯 Quick target hit @ ₹{ltp:.2f}  "
                           f"(+{cfg['TRAIL_START_PROFIT']} pts, P&L=₹{(ltp-avg_price)*qty:+.2f})"
                           f"  [detected {_ts()}]")
        else:
            if ltp > highest_price:
                highest_price = ltp
                print(f"\n  🔼 New High: ₹{highest_price:.2f}")

            if highest_price >= avg_price + trail_start:
                trail_exit = round(highest_price - trail_step, 2)
                if trail_exit != last_trail_exit:
                    print(f"\n  📉 Trail active | LTP=₹{ltp:.2f}  "
                          f"peak=₹{highest_price:.2f}  exit=₹{trail_exit:.2f}")
                    if notify:
                        send_telegram(f"📉 Trail active | LTP=₹{ltp:.2f} | "
                                      f"peak=₹{highest_price:.2f} | exit=₹{trail_exit:.2f}")
                    last_trail_exit = trail_exit
                if ltp <= trail_exit:
                    sell_reason = (f"🔻 Trail SL hit @ ₹{ltp:.2f}  "
                                   f"(peak=₹{highest_price:.2f}  exit=₹{trail_exit:.2f})"
                                   f"  [detected {_ts()}]")
                else:
                    print(f"  Trail | LTP=₹{ltp:.2f}  peak=₹{highest_price:.2f}  "
                          f"exit=₹{trail_exit:.2f}  P&L=₹{(ltp-avg_price)*qty:+.2f}", end="\r")
            else:
                print(f"  Monitoring | LTP=₹{ltp:.2f}  "
                      f"P&L=₹{(ltp - avg_price) * qty:+.2f}  "
                      f"(trail at ₹{avg_price + trail_start:.2f})", end="\r")

        if sell_reason:
            print(f"\n  {sell_reason}")
            if notify:
                send_telegram(sell_reason)
            break

        if mode != "mock":   # mock sleeps inside _next_mock_trail_ltp
            time.sleep(cfg["poll_seconds"])

    exit_ts = _ts()

    # Update Hard SL streak — both choppiness tracker and circuit breaker
    if sell_reason and "HARD SL" in sell_reason:
        _chop.record_hard_sl()
        _sl_circuit["count"] += 1
    else:
        _chop.reset_hard_sl()
        _sl_circuit["count"] = 0

    # ---- SELL ----
    sell_price = ltp
    if is_sim:
        print(f"  📋 {mode_tag}SELL simulated @ ₹{sell_price:.2f}  [{_ts()}]")
    else:
        try:
            print(f"  📤 SELL request sent to Groww  [{_ts()}]")
            sell_resp = _place_market_order(instrument, qty, "SELL")
            sell_id   = (sell_resp.get("payload", {}).get("groww_order_id")
                         or sell_resp.get("groww_order_id"))
            print(f"  ✅ SELL placed  order_id={sell_id}  [{_ts()}]")
        except Exception as e:
            print(f"  ❌ SELL failed: {e}  [{_ts()}]")
            send_telegram(f"❌ SELL FAILED: {e}\nManually close {symbol} qty={qty}")
            _log_trade_history(symbol, avg_price, sell_price, qty,
                               cfg["lots"], (sell_price - avg_price) * qty,
                               f"SELL_FAILED: {e}", mode, entry_ts, exit_ts,
                               oi_bias=signal.get("oi_bias","NEUTRAL"),
                               oi_verdict_tag="SELL_FAILED")
            return True

        if validate and sell_id:
            # Full blocking wait — get confirmed fill price
            ok = _wait_executed(sell_id, "SELL")
            if ok:
                sp, _ = _get_executed_price(sell_id)
                if sp:
                    sell_price = sp
                print(f"  ✅ SELL confirmed (executed) @ ₹{sell_price:.2f}  [{_ts()}]")
        elif sell_id:
            # validate=OFF: non-blocking fill check in background thread
            # Waits 2s for Groww to fill, then logs actual price + slippage vs detection LTP
            _det_ltp = sell_price
            def _bg_fill_check(_sid=sell_id, _det=_det_ltp):
                try:
                    time.sleep(2.0)
                    status = _get_order_status(_sid)
                    fill_ts = _ts()
                    if status in ("EXECUTED", "COMPLETED", "DELIVERY_AWAITED"):
                        sp, _ = _get_executed_price(_sid)
                        if sp:
                            slip = round(sp - _det, 2)
                            print(f"\n  📌 Groww actual fill: ₹{sp:.2f}  "
                                  f"slippage={slip:+.2f} vs detection ₹{_det:.2f}  [{fill_ts}]")
                        else:
                            print(f"\n  📌 SELL status: {status}  [{fill_ts}]")
                    else:
                        print(f"\n  ⚠️ SELL status after 2s: {status}  [{fill_ts}]")
                except Exception as _ex:
                    print(f"\n  ⚠️ Fill check error: {_ex}")
            threading.Thread(target=_bg_fill_check, daemon=True).start()

    profit = (sell_price - avg_price) * qty
    emoji  = "💰" if profit > 0 else "🔴"
    print(f"\n  {emoji} {mode_tag}CLOSED @ ₹{sell_price:.2f}  "
          f"Profit=₹{profit:+.2f}  (entry=₹{avg_price:.2f})  [{_ts()}]")
    if notify:
        send_telegram(f"{emoji} {mode_tag}CLOSED {symbol}\n"
                      f"Entry=₹{avg_price:.2f}  Exit=₹{sell_price:.2f}\n"
                      f"Profit=₹{profit:+.2f}")

    # ── OI filter effectiveness verdict ──────────────────────────
    _oi_b   = signal.get("oi_bias", "NEUTRAL") or "NEUTRAL"
    _vline, _vtag = _oi_verdict(_oi_b, opt_type, profit)
    print(f"  🔎 OI VERDICT │ {_vline}")

    _log_excel(symbol, avg_price, sell_price, qty, profit)
    _log_trade_history(symbol, avg_price, sell_price, qty,
                       cfg["lots"], profit, sell_reason or "exit", mode, entry_ts, exit_ts,
                       oi_bias=_oi_b, oi_verdict_tag=_vtag)
    print(f"{'='*65}\n")
    return True

# ============================================================
# 13. MAIN LOOP
# ============================================================
def main():
    print("\n" + "=" * 65)
    print("  MOMENTUM AUTO BOT — starting up")
    print("=" * 65)

    groww_init()

    spot, atm = _load_instruments()
    if not instruments_data:
        print("❌ No instruments loaded. Check expiry date and CSV.")
        return

    trades_today = 0
    last_atm     = atm
    next_scan_at = 0.0

    trail_mode = (f"ATR×{CONFIG['TRAIL_SL_ATR_MULTIPLIER']}" if CONFIG['TRAIL_SL_ATR_BASED']
                  else f"{CONFIG['TRAIL_STEP']} pts fixed")
    print(f"\n  Index={CONFIG['index']}  Expiry={CONFIG['expiry']}")
    print(f"  Lots={CONFIG['lots']}  ATM range=±{CONFIG['atm_range']} strikes")
    print(f"  Premium range=₹{CONFIG['min_premium']}–₹{CONFIG['max_premium']}")
    print(f"  Scan window={CONFIG['scan_seconds']}s @ {CONFIG['scan_poll_sec']}s/tick  "
          f"Trail poll={CONFIG['poll_seconds']}s")
    print(f"  Velocity>={CONFIG['velocity_pct']}%  "
          f"Consistency>={CONFIG['consistency_pct']}%")
    print(f"  Hard SL={CONFIG['HARD_SL_POINTS']} pts  "
          f"Trail start=+{CONFIG['TRAIL_START_PROFIT']} pts  "
          f"Trail step={trail_mode}")
    print(f"  Max hold={CONFIG['max_hold_min']} min  "
          f"Post-trade wait={CONFIG['cooldown_sec']}s  "
          f"No-signal wait={CONFIG['no_signal_wait_sec']}s  "
          f"Max trades/day=unlimited")
    print(f"  OI filter={'ON' if CONFIG['use_oi_filter'] else 'OFF'}  "
          f"Validate orders={CONFIG['validate_orders']}\n")
    start_webhook_server()
    send_telegram("🤖 MOMENTUM AUTO BOT started")

    while True:
        now = time.time()

        # Respect cooldown after a trade
        if now < next_scan_at:
            remain = int(next_scan_at - now)
            print(f"  💤 Cooldown — next scan in {remain}s", end="\r")
            time.sleep(2)
            continue

        # ── Consecutive Hard SL circuit breaker (always active) ──────────────────
        if CONFIG.get("consec_sl_brake", True):
            max_sl    = CONFIG.get("max_consecutive_hard_sl", 2)
            pause_min = CONFIG.get("consec_sl_pause_min", 30)
            # Clear expired pause
            if _sl_circuit["pause_until"] and datetime.now() >= _sl_circuit["pause_until"]:
                _sl_circuit.update({"count": 0, "pause_until": None})
                print(f"\n  ✅ Circuit breaker cleared — resuming scans")
            # Still in pause
            if _sl_circuit["pause_until"] and datetime.now() < _sl_circuit["pause_until"]:
                remain_m = int((_sl_circuit["pause_until"] - datetime.now()).total_seconds() // 60) + 1
                print(f"  🛑 Circuit break — {remain_m}m left  ({_sl_circuit['count']} consec Hard SLs)", end="\r")
                next_scan_at = time.time() + 60
                continue
            # Newly tripped
            if _sl_circuit["count"] >= max_sl and _sl_circuit["pause_until"] is None:
                _sl_circuit["pause_until"] = datetime.now() + timedelta(minutes=pause_min)
                resume_at = _sl_circuit["pause_until"].strftime("%H:%M")
                print(f"\n  🛑 CIRCUIT BREAKER — {_sl_circuit['count']} consecutive Hard SLs "
                      f"→ pausing new entries until {resume_at}")
                send_telegram(f"🛑 CIRCUIT BREAKER\n"
                              f"{_sl_circuit['count']} consecutive Hard SLs\n"
                              f"No new entries until {resume_at}")
                next_scan_at = time.time() + 60
                continue

        # Re-fetch spot; reload instruments if ATM drifted by ≥ 2 strikes
        spot    = _get_spot()
        step    = CONFIG["strike_step"]
        new_atm = round(spot / step) * step
        if abs(new_atm - last_atm) >= 2 * step:
            print(f"\n  🔄 ATM shifted {last_atm}→{new_atm}, reloading instruments …")
            _load_instruments()
            last_atm = new_atm

        snap     = load_oi_snapshot()
        bias     = oi_bias(snap)
        bias_tag = f"[OI bias: {bias}]" if snap else "[OI: n/a]"

        # Re-read config override so live dashboard toggle changes take effect
        _reload_override(verbose=False)

        # ── Phase 1: discover premiums in range ──────────────────
        print(f"\n{'='*65}")
        print(f"  🔍 spot={spot:.2f}  ATM={new_atm}  {bias_tag}  "
              f"[{datetime.now().strftime('%H:%M:%S')}]")

        candidates = discover_candidates(spot)

        if not candidates:
            print(f"  No premiums found in ₹{CONFIG['min_premium']}–"
                  f"₹{CONFIG['max_premium']} range. Retrying in 5s …")
            time.sleep(5)
            continue

        ce_cnt = sum(1 for _, t, _, _ in candidates if t == "CE")
        pe_cnt = sum(1 for _, t, _, _ in candidates if t == "PE")
        print(f"  Scanning for Premiums under range "
              f"{CONFIG['min_premium']}-{CONFIG['max_premium']} "
              f"found {len(candidates)} premiums  (CE: {ce_cnt}, PE: {pe_cnt})")

        # ── Phase 2: observe for scan_seconds ────────────────────
        history = run_observation(candidates)

        # ── Phase 3: pick momentum winner ────────────────────────
        signal = analyze_momentum(candidates, history)

        # ── Choppiness evaluation ─────────────────────────────────
        if CONFIG.get("choppiness_enabled", True):
            direction = signal["opt_type"] if signal else None
            _chop.record(direction, _last_scores["ce_net"], _last_scores["pe_net"])
            chop_state = _chop.evaluate()
            _print_choppiness(chop_state)

            if chop_state["level"] == "HIGH":
                if not chop_state["paused"]:
                    resume_at = _chop.trigger_pause()
                    msg = (f"⏸  HIGH CHOPPINESS — pausing new entries until {resume_at}\n"
                           f"     {chop_state['reason']}\n"
                           f"     Market needs time to settle before next entry.")
                    print(f"\n  {msg}")
                    send_telegram(
                        f"⚠️ CHOPPINESS ALERT\n{chop_state['reason']}\n"
                        f"Pausing entries until {resume_at}"
                    )
                else:
                    print(f"\n  ⏸  Choppiness pause active — resuming at {chop_state['pause_until']}")
                # Skip any signal and check again in 60s
                next_scan_at = time.time() + 60
                continue

            if chop_state["level"] == "MEDIUM" and signal:
                print(f"  ⚡ Caution: MEDIUM choppiness — {chop_state['reason']}")

        if signal:
            signal["oi_bias"] = bias   # carry current OI bias into trade context for verdict logging
            print(f"\n  🚨 SIGNAL: {signal['opt_type']} {int(signal['strike'])}  "
                  f"vel={signal['velocity_pct']:+.2f}%  "
                  f"consistency={signal['consistency_pct']:.0f}%  "
                  f"LTP=₹{signal['entry_ltp']:.2f}")
            ok = execute_trade(signal["inst"], signal)
            if ok:
                trades_today += 1
                next_scan_at = time.time() + CONFIG["cooldown_sec"]
                print(f"  Trades today: {trades_today}")
        else:
            wait = CONFIG["no_signal_wait_sec"]
            print(f"  No momentum signal — waiting {wait}s before next scan …")
            for remaining in range(wait, 0, -1):
                print(f"  ⏳ Next scan in {remaining}s …", end="\r")
                time.sleep(1)
            print()

    print("\n[MOM-BOT] Session complete.")
    send_telegram("🤖 MOMENTUM BOT session complete.")

# ============================================================
if __name__ == "__main__":
    main()
