#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║            PERSONAL TRADING INTELLIGENCE AI  —  v1.0                       ║
║  Analyzes YOUR 3-year trade history + live market conditions to determine   ║
║  whether you should trade today and how.                                    ║
║                                                                             ║
║  Usage:  python3 PERSONAL_TRADING_AI.py                                     ║
║  Setup:  Add ANTHROPIC_API_KEY to ai_config.json for AI narrative           ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import os, json, sys, time, warnings, re
from datetime import datetime, timedelta, date, time as dtime
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from collections import defaultdict

import requests
import numpy as np
import pandas as pd
import openpyxl

warnings.filterwarnings("ignore")

try:
    import yfinance as yf
    YF_OK = True
except ImportError:
    YF_OK = False

# ═══════════════════════════════════════════════════════════════════════════════
#  PATHS & CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
ROOT       = Path(__file__).parent
DATA_DIR   = ROOT / "ayush_previous_data"
CACHE_FILE = ROOT / ".trading_ai_cache.json"
AI_CONFIG  = ROOT / "ai_config.json"
LAKSHMI    = ROOT / "Lakshmi.xlsx"

def _load_ai_cfg() -> dict:
    try:
        return json.loads(AI_CONFIG.read_text())
    except Exception:
        return {}

# ═══════════════════════════════════════════════════════════════════════════════
#  TERMINAL COLORS
# ═══════════════════════════════════════════════════════════════════════════════
class C:
    RST = "\033[0m";  BD = "\033[1m";  DIM = "\033[2m"
    GRN = "\033[92m"; RED = "\033[91m"; YLW = "\033[93m"
    CYN = "\033[96m"; WHT = "\033[97m"; MAG = "\033[95m"
    BLU = "\033[94m"; ORG = "\033[38;5;208m"

    @staticmethod
    def g(t): return f"{C.GRN}{t}{C.RST}"
    @staticmethod
    def r(t): return f"{C.RED}{t}{C.RST}"
    @staticmethod
    def y(t): return f"{C.YLW}{t}{C.RST}"
    @staticmethod
    def c(t): return f"{C.CYN}{t}{C.RST}"
    @staticmethod
    def b(t): return f"{C.BD}{t}{C.RST}"
    @staticmethod
    def m(t): return f"{C.MAG}{t}{C.RST}"
    @staticmethod
    def dim(t): return f"{C.DIM}{t}{C.RST}"
    @staticmethod
    def o(t): return f"{C.ORG}{t}{C.RST}"

W = 82

def hdr(title, color=C.CYN):
    bar = "═" * W
    pad = (W - len(title) - 2) // 2
    return f"\n{color}{C.BD}{bar}\n{' '*pad} {title} \n{bar}{C.RST}"

def section(title, color=C.YLW):
    bar = "─" * W
    return f"\n{color}{C.BD}  ▸ {title}{C.RST}\n{C.DIM}  {bar}{C.RST}"

def row(label, val, color=C.WHT, label_w=30):
    lbl = f"{C.DIM}{label:<{label_w}}{C.RST}"
    return f"  {lbl}  {color}{val}{C.RST}"

def progress_bar(score, max_score=100, width=28):
    filled = int(round(score / max_score * width))
    bar = "█" * filled + "░" * (width - filled)
    if score >= 61:
        color = C.GRN
    elif score >= 41:
        color = C.YLW
    else:
        color = C.RED
    return f"{color}{bar}{C.RST}"

def pnl_color(val):
    if val > 0:   return C.g(f"₹{val:>+,.0f}")
    elif val < 0: return C.r(f"₹{val:>+,.0f}")
    else:         return C.y("₹0")

# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 1 — PARSE PERSONAL TRADE HISTORY
# ═══════════════════════════════════════════════════════════════════════════════
_DATE_FMT = "%d %b %Y"   # "17 Apr 2023"

def _parse_date(s) -> Optional[date]:
    if not s or not isinstance(s, str):
        return None
    try:
        return datetime.strptime(s.strip(), _DATE_FMT).date()
    except Exception:
        return None

def _is_expiry_day(scrip: str, trade_date: date) -> bool:
    """Heuristic: trade day matches contract expiry if scrip date == trade date."""
    # Scrip looks like "NIFTY 05 MAY 26 24000 Call"
    months = {"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,
              "JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}
    m = re.search(r'(\d{2})\s+([A-Z]{3})\s+(\d{2})', scrip.upper())
    if not m:
        return False
    try:
        day = int(m.group(1))
        mon = months.get(m.group(2), 0)
        yr  = 2000 + int(m.group(3))
        return trade_date == date(yr, mon, day)
    except Exception:
        return False

def parse_excel_history() -> Tuple[Dict, Dict, Dict]:
    """
    Returns:
        daily_pnl    {date: float}  — net realized P&L per day
        daily_trades {date: int}    — number of trade legs per day
        expiry_days  {date: bool}   — True if traded on expiry
    """
    daily_pnl    = defaultdict(float)
    daily_trades = defaultdict(int)
    expiry_days  = defaultdict(bool)

    if not DATA_DIR.exists():
        print(C.r("  ✗ ayush_previous_data/ folder not found"))
        return {}, {}, {}

    for fpath in sorted(DATA_DIR.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(str(fpath), read_only=True)
            ws = wb["Trade Level"]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            print(C.y(f"  ⚠ Could not read {fpath.name}: {e}"))
            continue

        for row_data in rows:
            scrip  = row_data[0]
            qty    = row_data[1]
            s_date = row_data[5]   # Sell Date
            pnl_v  = row_data[8]   # Realized P&L

            if (not scrip or not isinstance(scrip, str) or
                    scrip in ("Scrip Name", "Futures", "Options", "Total", "Name",
                              "Unique Client Code", "Summary", "Realised P&L", "Charges")
                    or pnl_v is None or s_date is None):
                continue

            d = _parse_date(s_date)
            if d is None:
                continue

            daily_pnl[d]    += float(pnl_v)
            daily_trades[d] += 1
            if _is_expiry_day(scrip, d):
                expiry_days[d] = True

    return dict(daily_pnl), dict(daily_trades), dict(expiry_days)


def parse_lakshmi_intraday() -> Dict:
    """
    Parses Lakshmi.xlsx (intraday with timestamps) for behavioral patterns.
    Returns {date: {early_trades, pre10_pnl, total_trades, trade_times, hourly_pnl}}
    """
    result = defaultdict(lambda: {"early_trades":0, "pre_10_pnl":0.0,
                                   "total":0, "trades":[], "hourly_pnl":defaultdict(float)})
    if not LAKSHMI.exists():
        return {}

    try:
        wb = openpyxl.load_workbook(str(LAKSHMI), read_only=True)
        ws = wb["Lakshmi"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception:
        return {}

    for r in rows:
        dt_raw, sym, buy_p, sell_p, qty, pnl_v = r[0], r[1], r[2], r[3], r[4], r[5]
        if not dt_raw or not sym or pnl_v is None:
            continue
        if isinstance(pnl_v, str):   # formula string
            continue
        try:
            if isinstance(dt_raw, datetime):
                dt = dt_raw
            else:
                dt = datetime.strptime(str(dt_raw), "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue

        d   = dt.date()
        hr  = dt.hour
        pnl = float(pnl_v)

        result[d]["total"] += 1
        result[d]["trades"].append({"time": dt.strftime("%H:%M"), "pnl": pnl})
        result[d]["hourly_pnl"][hr] += pnl

        if hr < 10 or (hr == 9 and dt.minute < 30):
            result[d]["early_trades"] += 1
            result[d]["pre_10_pnl"]   += pnl

    return dict(result)


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 2 — HISTORICAL MARKET DATA (NIFTY + VIX via yfinance)
# ═══════════════════════════════════════════════════════════════════════════════
def build_market_db(force_refresh=False) -> pd.DataFrame:
    """
    Builds / loads a DataFrame with one row per trading day:
    date | vix_close | nifty_open | nifty_close | nifty_prev_close | gap_pct | trend_5d | dow
    """
    if not force_refresh and CACHE_FILE.exists():
        try:
            cache = json.loads(CACHE_FILE.read_text())
            ts    = cache.get("built_at", 0)
            # Refresh if older than 12 hours
            if time.time() - ts < 43200 and cache.get("records"):
                df = pd.DataFrame(cache["records"])
                df["date"] = pd.to_datetime(df["date"]).dt.date
                return df
        except Exception:
            pass

    if not YF_OK:
        print(C.y("  yfinance unavailable — historical market context skipped."))
        return pd.DataFrame()

    print(C.dim("  Downloading NIFTY50 + India VIX history (2023-present)…"))
    try:
        nifty = yf.download("^NSEI",      start="2023-04-01", progress=False)
        vix   = yf.download("^INDIAVIX",  start="2023-04-01", progress=False)
    except Exception as e:
        print(C.y(f"  yfinance error: {e}"))
        return pd.DataFrame()

    # Flatten multi-index columns
    nifty.columns = ["_".join(c).strip("_") if isinstance(c, tuple) else c
                     for c in nifty.columns]
    vix.columns   = ["_".join(c).strip("_") if isinstance(c, tuple) else c
                     for c in vix.columns]

    close_col = [c for c in nifty.columns if "Close" in c][0]
    open_col  = [c for c in nifty.columns if "Open"  in c][0]
    vix_col   = [c for c in vix.columns   if "Close" in c][0]

    nifty["prev_close"] = nifty[close_col].shift(1)
    nifty["gap_pct"]    = (nifty[open_col] - nifty["prev_close"]) / nifty["prev_close"] * 100
    nifty["trend_5d"]   = nifty[close_col].pct_change(5) * 100   # 5-day return %

    df = pd.DataFrame({
        "date":        nifty.index.date,
        "nifty_open":  nifty[open_col].values,
        "nifty_close": nifty[close_col].values,
        "nifty_prev":  nifty["prev_close"].values,
        "gap_pct":     nifty["gap_pct"].values,
        "trend_5d":    nifty["trend_5d"].values,
        "dow":         [d.weekday() for d in nifty.index.date],  # 0=Mon…4=Fri
    })
    df = df.merge(
        pd.DataFrame({"date": vix.index.date, "vix": vix[vix_col].values}),
        on="date", how="left"
    ).dropna(subset=["nifty_close"])

    # Cache it
    records = df.copy()
    records["date"] = records["date"].astype(str)
    CACHE_FILE.write_text(json.dumps({
        "built_at": time.time(),
        "records":  records.to_dict(orient="records")
    }, default=str))

    return df


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 3 — LIVE MARKET DATA (NSE)
# ═══════════════════════════════════════════════════════════════════════════════
_NSE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36",
    "Accept":     "application/json, text/plain, */*",
    "Referer":    "https://www.nseindia.com",
}

def _nse_session():
    s = requests.Session()
    s.headers.update(_NSE_HEADERS)
    try:
        s.get("https://www.nseindia.com", timeout=5)
    except Exception:
        pass
    return s

def fetch_live_market() -> dict:
    """Returns live VIX, NIFTY level, PCR, gap info."""
    result = {
        "vix": None, "vix_prev": None, "vix_chg_pct": None,
        "nifty": None, "nifty_prev": None, "nifty_open": None, "gap_pct": None,
        "pcr": None, "as_of": None, "market_open": False,
    }
    try:
        s = _nse_session()
        r = s.get("https://www.nseindia.com/api/allIndices", timeout=7)
        data = r.json().get("data", [])
        now  = datetime.now()
        result["as_of"] = now.strftime("%H:%M")
        result["market_open"] = (
            now.weekday() < 5 and
            dtime(9, 15) <= now.time() <= dtime(15, 30)
        )

        for item in data:
            idx = item.get("index", "")
            if idx == "INDIA VIX":
                result["vix"]         = item.get("last")
                result["vix_prev"]    = item.get("previousClose")
                result["vix_chg_pct"] = item.get("percentChange")
            elif idx == "NIFTY 50":
                result["nifty"]      = item.get("last")
                result["nifty_prev"] = item.get("previousClose")
                result["nifty_open"] = item.get("open")
                if result["nifty_open"] and result["nifty_prev"]:
                    result["gap_pct"] = round(
                        (result["nifty_open"] - result["nifty_prev"]) /
                        result["nifty_prev"] * 100, 2
                    )
    except Exception as e:
        pass

    # PCR via option chain
    try:
        s2 = _nse_session()
        r2 = s2.get(
            "https://www.nseindia.com/api/option-chain-indices?symbol=NIFTY",
            timeout=8
        )
        oc   = r2.json()
        recs = oc.get("records", {}).get("data", [])
        total_ce_oi = sum(
            r.get("CE", {}).get("openInterest", 0) for r in recs if r.get("CE")
        )
        total_pe_oi = sum(
            r.get("PE", {}).get("openInterest", 0) for r in recs if r.get("PE")
        )
        if total_ce_oi > 0:
            result["pcr"] = round(total_pe_oi / total_ce_oi, 3)
    except Exception:
        pass

    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 4 — MARKET CONDITION SCORE (0-100)
# ═══════════════════════════════════════════════════════════════════════════════
DOW_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

def market_condition_score(live: dict) -> Tuple[int, dict]:
    """
    Returns (score_0_to_100, breakdown_dict)
    Higher = more favorable / predictable market environment.
    """
    breakdown = {}
    score     = 0

    # ── VIX Level (0–25) ──────────────────────────────────────────────────────
    vix = live.get("vix")
    if vix is not None:
        if   vix < 11:
            pts     = 25
            meaning = "Ultra-low fear. Premiums are cheap — great for buying options, but moves may be small."
        elif vix < 13:
            pts     = 23
            meaning = "Very calm market. Options are cheap, clean trends likely."
        elif vix < 15:
            pts     = 20
            meaning = "Normal volatility. Balanced premium — good conditions for most strategies."
        elif vix < 17:
            pts     = 17
            meaning = f"Slightly elevated fear (VIX {vix:.1f}). Premiums higher, moves can be sharp — stay disciplined."
        elif vix < 19:
            pts     = 14
            meaning = f"Moderate fear (VIX {vix:.1f}). Market is nervous, fakeouts common. Smaller size recommended."
        elif vix < 22:
            pts     = 10
            meaning = f"High volatility (VIX {vix:.1f}). Whipsaw risk is real. Only experienced setups, wide stops."
        elif vix < 26:
            pts     = 6
            meaning = f"Very high fear (VIX {vix:.1f}). Market in panic mode — most retail traders lose here."
        else:
            pts     = 3
            meaning = f"Extreme fear (VIX {vix:.1f}). Circuit-breaker territory — do NOT trade unless you are a pro."
        breakdown["VIX Level"] = (pts, 25, f"{vix:.2f}", meaning)
        score += pts
    else:
        breakdown["VIX Level"] = (12, 25, "N/A", "VIX data unavailable — treating as moderate volatility.")
        score += 12

    # ── VIX Daily Change (0–20) ───────────────────────────────────────────────
    vix_chg = live.get("vix_chg_pct")
    if vix_chg is not None:
        av = abs(vix_chg)
        if   av < 3:
            pts     = 20
            meaning = f"VIX stable ({vix_chg:+.1f}%). Fear level unchanged — market environment predictable."
        elif av < 5:
            pts     = 16
            meaning = f"VIX moved {vix_chg:+.1f}%. Minor shift in fear — watch for direction confirmation after open."
        elif av < 8:
            pts     = 11
            meaning = f"VIX jumped {vix_chg:+.1f}% today. Noticeable fear injection — gaps and whipsaws more likely."
        elif av < 12:
            pts     = 6
            meaning = f"VIX spike of {vix_chg:+.1f}%! Sudden panic in market. First 30 min will be chaotic — wait."
        else:
            pts     = 2
            meaning = f"VIX exploded {vix_chg:+.1f}%! Extreme panic event. Avoid all directional trades."
        breakdown["VIX Stability"] = (pts, 20, f"{vix_chg:+.1f}%", meaning)
        score += pts
    else:
        breakdown["VIX Stability"] = (10, 20, "N/A", "VIX change unavailable — stability unknown.")
        score += 10

    # ── NIFTY Gap (0–20) ──────────────────────────────────────────────────────
    gap = live.get("gap_pct")
    if gap is not None:
        ag = abs(gap)
        if   ag < 0.2:
            pts     = 20
            meaning = f"Flat open ({gap:+.2f}%). No gap to chase. Price action will be technical and clean."
        elif ag < 0.5:
            pts     = 17
            meaning = f"Small gap ({gap:+.2f}%). Minor overnight positioning — usually fills within 30 min."
        elif ag < 0.8:
            pts     = 13
            meaning = f"Moderate gap ({gap:+.2f}%). First 15–20 min may see gap-fill attempt before real trend starts."
        elif ag < 1.2:
            pts     = 8
            meaning = f"Big gap ({gap:+.2f}%). Market opened with strong conviction — wait for 9:30 stabilization."
        else:
            pts     = 4
            meaning = f"Huge gap ({gap:+.2f}%)! Panic/euphoria open. Very high reversal risk in first 30 min — do NOT enter at open."
        breakdown["NIFTY Gap"] = (pts, 20, f"{gap:+.2f}%", meaning)
        score += pts
    else:
        breakdown["NIFTY Gap"] = (10, 20, "N/A", "Gap data unavailable.")
        score += 10

    # ── PCR (0–20) ────────────────────────────────────────────────────────────
    pcr = live.get("pcr")
    if pcr is not None:
        if   0.9 <= pcr <= 1.3:
            pts     = 20
            meaning = f"PCR {pcr:.2f} — healthy balance of puts and calls. No extreme positioning, cleaner moves."
        elif 0.8 <= pcr < 0.9:
            pts     = 16
            meaning = f"PCR {pcr:.2f} — slightly call-heavy. Mild bullish bias in OI. Market may face resistance."
        elif 1.3 < pcr <= 1.5:
            pts     = 15
            meaning = f"PCR {pcr:.2f} — slightly put-heavy. Mild bearish hedge in OI. Support likely stronger."
        elif 0.7 <= pcr < 0.8:
            pts     = 10
            meaning = f"PCR {pcr:.2f} — too many calls written. Extreme complacency — market vulnerable to sharp drop."
        elif pcr > 1.5:
            pts     = 8
            meaning = f"PCR {pcr:.2f} — extreme put buying. Panic hedging. Contrarian signal — possible bounce but risky."
        else:
            pts     = 6
            meaning = f"PCR {pcr:.2f} — extreme call dominance. Overconfidence — sharp reversal risk."
        breakdown["PCR"] = (pts, 20, f"{pcr:.3f}", meaning)
        score += pts
    else:
        breakdown["PCR"] = (12, 20, "N/A", "PCR unavailable (market closed or API issue). Assuming neutral balance.")
        score += 12

    # ── Day of Week (0–15) ────────────────────────────────────────────────────
    dow = datetime.now().weekday()
    if   dow == 2:
        pts     = 15
        meaning = "Wednesday — historically most stable mid-week session. Best day for technical setups."
    elif dow == 1:
        pts     = 14
        meaning = "Tuesday — market has found its footing after Monday. Good for trend continuation trades."
    elif dow == 3:
        pts     = 13
        meaning = "Thursday — often expiry day (weekly). Premiums decay fast. Watch for pinning near ATM strikes."
    elif dow == 0:
        pts     = 11
        meaning = "Monday — direction set by global weekend news. Wait for 9:30–10:00 to confirm true trend."
    elif dow == 4:
        pts     = 8
        meaning = "Friday — weekly expiry + weekend premium bleed. Your personal Friday win rate is only 46%."
    else:
        pts     = 5
        meaning = "Weekend — market is closed. Use today for analysis and planning, not trading."
    breakdown["Day of Week"] = (pts, 15, DOW_NAMES[dow], meaning)
    score += pts

    return min(score, 100), breakdown


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 5 — SIMILAR DAY FINDER
# ═══════════════════════════════════════════════════════════════════════════════
def find_similar_days(
    live: dict,
    market_db: pd.DataFrame,
    daily_pnl: dict,
    top_n: int = 30,
) -> Tuple[pd.DataFrame, dict]:
    """
    Match today's conditions against all historical days.
    Returns (similar_days_df, similarity_report).
    """
    if market_db.empty:
        return pd.DataFrame(), {}

    today_vix  = live.get("vix") or 16.0
    today_gap  = live.get("gap_pct") or 0.0
    today_dow  = datetime.now().weekday()

    rows = []
    for _, mrow in market_db.iterrows():
        d   = mrow["date"]
        if not isinstance(d, date):
            try:
                d = datetime.strptime(str(d), "%Y-%m-%d").date()
            except Exception:
                continue
        # We only care about days we actually traded
        if d not in daily_pnl:
            continue

        hist_vix = mrow.get("vix", np.nan)
        hist_gap = mrow.get("gap_pct", np.nan)
        hist_dow = mrow.get("dow", -1)

        if pd.isna(hist_vix):
            continue

        # Similarity score (0–100)
        vix_sim = max(0, 100 - abs(today_vix - hist_vix) * 10)  # ±10 VIX = 0 sim
        gap_sim = max(0, 100 - abs(today_gap - hist_gap) * 25)   # ±4% = 0 sim
        dow_sim = 100 if hist_dow == today_dow else (60 if abs(hist_dow - today_dow) == 1 else 30)
        sim     = round(vix_sim * 0.45 + gap_sim * 0.35 + dow_sim * 0.20, 1)

        rows.append({
            "date":      d,
            "vix":       round(hist_vix, 2),
            "gap_pct":   round(hist_gap,  2),
            "dow":       DOW_NAMES[int(hist_dow)],
            "trend_5d":  round(mrow.get("trend_5d", 0), 2),
            "nifty_cls": round(mrow.get("nifty_close", 0), 2),
            "pnl":       daily_pnl[d],
            "sim_score": sim,
        })

    if not rows:
        return pd.DataFrame(), {}

    df = pd.DataFrame(rows).sort_values("sim_score", ascending=False).head(top_n)

    wins    = df[df["pnl"] > 0]
    losses  = df[df["pnl"] <= 0]
    win_rt  = len(wins) / len(df) * 100 if len(df) else 0
    avg_win = wins["pnl"].mean()    if len(wins) > 0  else 0
    avg_los = losses["pnl"].mean()  if len(losses) > 0 else 0
    best_d  = df.loc[df["pnl"].idxmax()]  if len(df) else None
    worst_d = df.loc[df["pnl"].idxmin()]  if len(df) else None

    report = {
        "count":    len(df),
        "win_rate": round(win_rt, 1),
        "avg_win":  round(avg_win,  2),
        "avg_loss": round(avg_los,  2),
        "total_pnl": round(df["pnl"].sum(), 2),
        "best_day":  best_d,
        "worst_day": worst_d,
        "top5":      df.head(5).to_dict("records"),
    }
    return df, report


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 6 — BEHAVIORAL RISK ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
def behavioral_analysis(daily_pnl: dict, daily_trades: dict,
                         expiry_days: dict, intraday: dict) -> dict:
    risks    = []
    insights = []

    if not daily_pnl:
        return {"risks": [], "insights": [], "risk_score": 0}

    # ── Friday performance ────────────────────────────────────────────────────
    friday_pnls = [v for k, v in daily_pnl.items() if k.weekday() == 4]
    if friday_pnls:
        fri_wr  = sum(1 for v in friday_pnls if v > 0) / len(friday_pnls) * 100
        fri_avg = np.mean(friday_pnls)
        if fri_wr < 45 or fri_avg < 0:
            risks.append({
                "type":   "FRIDAY_PATTERN",
                "detail": f"You win only {fri_wr:.0f}% of Fridays (avg {fri_avg:+,.0f} ₹). "
                          f"Total Fridays traded: {len(friday_pnls)}",
                "weight": 3,
            })

    # ── Overtrading ───────────────────────────────────────────────────────────
    if daily_trades:
        avg_trades = np.mean(list(daily_trades.values()))
        over_days  = {d: t for d, t in daily_trades.items() if t > avg_trades * 2}
        if over_days:
            over_pnls = [daily_pnl.get(d, 0) for d in over_days]
            over_wr   = sum(1 for v in over_pnls if v > 0) / len(over_pnls) * 100
            risks.append({
                "type":   "OVERTRADING",
                "detail": f"On {len(over_days)} days you traded 2× the average ({avg_trades:.0f}). "
                          f"Win rate on those days: {over_wr:.0f}%",
                "weight": 2,
            })
        insights.append(f"Avg legs/day: {avg_trades:.1f}  |  Max: {max(daily_trades.values())}")

    # ── Expiry day trading ────────────────────────────────────────────────────
    exp_pnls = [daily_pnl.get(d, 0) for d in expiry_days if d in daily_pnl]
    if exp_pnls:
        exp_wr  = sum(1 for v in exp_pnls if v > 0) / len(exp_pnls) * 100
        exp_avg = np.mean(exp_pnls)
        if exp_wr < 50 or exp_avg < -5000:
            risks.append({
                "type":   "EXPIRY_DAY_RISK",
                "detail": f"You traded on {len(exp_pnls)} expiry days. "
                          f"Win rate: {exp_wr:.0f}%, avg P&L: ₹{exp_avg:+,.0f}",
                "weight": 2,
            })

    # ── Revenge trading (intraday) ────────────────────────────────────────────
    if intraday:
        revenge_days = 0
        for d, info in intraday.items():
            trades = info.get("trades", [])
            if len(trades) < 4:
                continue
            # Detect: after 2+ consecutive losses, did trades accelerate?
            consec_loss = 0
            for t in trades:
                if t["pnl"] < 0:
                    consec_loss += 1
                else:
                    consec_loss = 0
                if consec_loss >= 2:
                    revenge_days += 1
                    break
        if revenge_days:
            pct = revenge_days / len(intraday) * 100
            risks.append({
                "type":   "REVENGE_TRADING",
                "detail": f"Detected consecutive-loss then continue-trading pattern on "
                          f"{revenge_days} days ({pct:.0f}% of intraday log)",
                "weight": 3,
            })

    # ── Early trading ─────────────────────────────────────────────────────────
    if intraday:
        early_days = [d for d, i in intraday.items() if i.get("early_trades", 0) > 0]
        if early_days:
            early_pnls = [intraday[d]["pre_10_pnl"] for d in early_days]
            early_avg  = np.mean(early_pnls)
            if early_avg < 0:
                risks.append({
                    "type":   "EARLY_TRADING",
                    "detail": f"Traded before 10:00 AM on {len(early_days)} days. "
                              f"Avg P&L from those early trades: ₹{early_avg:+,.0f}",
                    "weight": 2,
                })

    # ── Best trading hours (insight) ─────────────────────────────────────────
    if intraday:
        hourly_totals = defaultdict(float)
        hourly_counts = defaultdict(int)
        for d, info in intraday.items():
            for hr, p in info.get("hourly_pnl", {}).items():
                hourly_totals[hr] += p
                hourly_counts[hr] += 1
        if hourly_totals:
            best_hr  = max(hourly_totals, key=hourly_totals.get)
            worst_hr = min(hourly_totals, key=hourly_totals.get)
            insights.append(
                f"Best trading hour:  {best_hr:02d}:00–{best_hr+1:02d}:00  "
                f"(avg ₹{hourly_totals[best_hr]/max(hourly_counts[best_hr],1):+,.0f})"
            )
            insights.append(
                f"Worst trading hour: {worst_hr:02d}:00–{worst_hr+1:02d}:00  "
                f"(avg ₹{hourly_totals[worst_hr]/max(hourly_counts[worst_hr],1):+,.0f})"
            )

    # ── Loss streak pattern ───────────────────────────────────────────────────
    sorted_days = sorted(daily_pnl.items())
    max_streak  = 0
    cur_streak  = 0
    for _, p in sorted_days:
        if p < 0:
            cur_streak += 1
            max_streak  = max(max_streak, cur_streak)
        else:
            cur_streak  = 0
    insights.append(f"Max consecutive loss days: {max_streak}")

    # ── Recent momentum (last 10 trading days) ────────────────────────────────
    recent = sorted_days[-10:]
    if recent:
        r_wr  = sum(1 for _, p in recent if p > 0) / len(recent) * 100
        r_avg = np.mean([p for _, p in recent])
        insights.append(
            f"Last 10 trading days: {sum(1 for _,p in recent if p>0)}W / "
            f"{sum(1 for _,p in recent if p<=0)}L  |  avg ₹{r_avg:+,.0f}"
        )

    risk_score = min(sum(r["weight"] for r in risks) * 5, 40)

    return {
        "risks":      risks,
        "insights":   insights,
        "risk_score": risk_score,
        "recent_wr":  round(r_wr if recent else 50.0, 1),
        "recent_avg": round(r_avg if recent else 0.0,  2),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 7 — TRADING PERMISSION SCORE (0-100)
# ═══════════════════════════════════════════════════════════════════════════════
def trading_permission_score(
    mkt_score:   int,
    sim_report:  dict,
    behav:       dict,
    daily_pnl:   dict,
) -> Tuple[int, str, dict]:
    """
    Returns (score, verdict, score_breakdown)
    verdict: "NO_TRADE" | "CAUTION" | "NORMAL" | "HIGH_CONFIDENCE"
    """
    breakdown = {}

    # Component 1: Market condition (35%)
    mkt_pts = round(mkt_score * 0.35)
    breakdown["Market Conditions (35%)"] = mkt_pts

    # Component 2: Similar days personal win rate (40%)
    sim_wr  = sim_report.get("win_rate", 50)
    sim_pts = round(sim_wr * 0.40)
    breakdown["Similar Days Win Rate (40%)"] = sim_pts

    # Component 3: Recent momentum (20%)
    rec_wr  = behav.get("recent_wr", 50)
    rec_pts = round(rec_wr * 0.20)
    breakdown["Recent Momentum (20%)"] = rec_pts

    # Component 4: Behavioral risk deduction (up to −20)
    risk_deduct = -min(behav.get("risk_score", 0), 20)
    breakdown["Behavioral Risk Deduction"] = risk_deduct

    # Component 5: Today is expiry (−10 bonus deduction)
    today    = date.today()
    # Check if today is a Thursday (weekly expiry day for NIFTY)
    exp_deduct = -10 if today.weekday() == 3 else 0
    if exp_deduct:
        breakdown["Expiry Day Penalty"] = exp_deduct

    raw_score = mkt_pts + sim_pts + rec_pts + risk_deduct + exp_deduct
    score     = max(0, min(100, raw_score))

    if   score >= 81: verdict = "HIGH_CONFIDENCE"
    elif score >= 61: verdict = "NORMAL"
    elif score >= 41: verdict = "CAUTION"
    else:             verdict = "NO_TRADE"

    return score, verdict, breakdown


# ═══════════════════════════════════════════════════════════════════════════════
#  STEP 8 — AI NARRATIVE (ANTHROPIC CLAUDE)
# ═══════════════════════════════════════════════════════════════════════════════
def generate_ai_narrative(
    live:        dict,
    mkt_score:   int,
    mkt_bkdwn:   dict,
    sim_report:  dict,
    behav:       dict,
    perm_score:  int,
    verdict:     str,
    daily_pnl:   dict,
) -> str:
    import subprocess, shutil

    today_dow  = DOW_NAMES[datetime.now().weekday()]
    total_days = len(daily_pnl)
    total_pnl  = sum(daily_pnl.values())
    win_days   = sum(1 for v in daily_pnl.values() if v > 0)
    overall_wr = round(win_days / total_days * 100, 1) if total_days else 0

    prompt = f"""You are a Personal Trading Intelligence AI advisor for Ayush Gupta,
an experienced F&O options trader on GROWW (NSE India).

Today: {date.today()} ({today_dow})

LIVE MARKET DATA:
- NIFTY 50: {live.get('nifty', 'N/A')}  (Prev close: {live.get('nifty_prev', 'N/A')})
- India VIX: {live.get('vix', 'N/A')} ({live.get('vix_chg_pct', 0):+.1f}% change today)
- NIFTY Open Gap: {live.get('gap_pct', 0):+.2f}%
- PCR (Put/Call Ratio): {live.get('pcr', 'N/A')}

MARKET CONDITION SCORE: {mkt_score}/100
Score breakdown: {json.dumps(mkt_bkdwn, default=str)}

AYUSH'S 3-YEAR PERSONAL HISTORY:
- Total trading days: {total_days}
- Overall win rate: {overall_wr}%
- Total P&L (3 years): ₹{total_pnl:+,.0f}

SIMILAR DAYS ANALYSIS (top {sim_report.get('count', 0)} historically similar days):
- Win rate on similar days: {sim_report.get('win_rate', 0)}%
- Avg win on similar days: ₹{sim_report.get('avg_win', 0):+,.0f}
- Avg loss on similar days: ₹{sim_report.get('avg_loss', 0):+,.0f}

BEHAVIORAL RISKS DETECTED:
{json.dumps([r['type'] + ': ' + r['detail'] for r in behav.get('risks', [])], indent=2)}

BEHAVIORAL INSIGHTS:
{json.dumps(behav.get('insights', []), indent=2)}

TRADING PERMISSION SCORE: {perm_score}/100
VERDICT: {verdict}

Based on all this data, provide a concise (under 200 words) trading advisory:
1. RECOMMENDATION (Trade / No Trade / Caution)
2. Direction bias (Bullish/Bearish/Neutral) with brief reasoning
3. If trading: recommended setup + best time window + position sizing
4. One key behavioral warning relevant to today
5. Key levels to watch

Be direct, specific, and data-driven. Reference actual historical dates when relevant.
Speak to Ayush directly — he is an experienced F&O trader."""

    # Use Claude Code CLI (already authenticated, no API key needed)
    claude_bin = shutil.which("claude")
    if claude_bin:
        try:
            result = subprocess.run(
                [claude_bin, "-p", prompt],
                capture_output=True, text=True, timeout=60
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
            if result.stderr:
                return f"(Claude CLI error: {result.stderr.strip()[:200]})"
        except subprocess.TimeoutExpired:
            return "(Claude CLI timed out after 60s)"
        except Exception as e:
            return f"(Claude CLI failed: {e})"

    return "(claude CLI not found — install Claude Code to enable AI narrative)"


# ═══════════════════════════════════════════════════════════════════════════════
#  OVERALL STATS HELPER
# ═══════════════════════════════════════════════════════════════════════════════
def overall_stats(daily_pnl: dict) -> dict:
    if not daily_pnl:
        return {}
    pnls = list(daily_pnl.values())
    wins = [v for v in pnls if v > 0]
    loss = [v for v in pnls if v <= 0]
    sorted_items = sorted(daily_pnl.items())

    # Year-by-year breakdown
    yearly = defaultdict(float)
    for d, p in sorted_items:
        yearly[d.year] += p

    return {
        "total_days":  len(pnls),
        "win_days":    len(wins),
        "loss_days":   len(loss),
        "win_rate":    round(len(wins) / len(pnls) * 100, 1),
        "total_pnl":   round(sum(pnls), 2),
        "avg_win":     round(np.mean(wins),  2) if wins else 0,
        "avg_loss":    round(np.mean(loss),  2) if loss else 0,
        "best_day":    sorted_items[np.argmax(pnls)],
        "worst_day":   sorted_items[np.argmin(pnls)],
        "yearly":      dict(yearly),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  DISPLAY ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
def display_header(live: dict):
    today = date.today()
    dow   = DOW_NAMES[today.weekday()]
    print(hdr(f"PERSONAL TRADING INTELLIGENCE AI  •  {today}  ({dow})"))

def display_live_market(live: dict, mkt_score: int, mkt_bkdwn: dict):
    print(section("LIVE MARKET CONDITIONS"))
    nifty   = live.get("nifty")
    vix     = live.get("vix")
    gap     = live.get("gap_pct")
    pcr     = live.get("pcr")
    vchg    = live.get("vix_chg_pct")

    nifty_s = f"{nifty:,.2f}" if nifty else "N/A"
    vix_s   = f"{vix:.2f} ({vchg:+.1f}%)" if vix and vchg is not None else (str(vix) if vix else "N/A")
    gap_s   = (f"{gap:+.2f}%" if gap is not None else "N/A")
    gap_lbl = "FLAT" if gap is not None and abs(gap) < 0.3 else (
               "GAP UP" if gap is not None and gap > 0 else "GAP DOWN")
    pcr_s   = f"{pcr:.3f}" if pcr else "N/A"

    nifty_c = C.g(nifty_s) if (live.get("nifty") or 0) >= (live.get("nifty_prev") or 0) else C.r(nifty_s)
    gap_c   = C.g(f"{gap_s} ({gap_lbl})") if gap is not None and gap >= 0 else C.r(f"{gap_s} ({gap_lbl})")
    vix_c   = C.g(vix_s) if vix and vix < 16 else (C.y(vix_s) if vix and vix < 20 else C.r(vix_s))

    print(f"\n  {'NIFTY 50':<16}  {nifty_c}")
    print(f"  {'India VIX':<16}  {vix_c}")
    print(f"  {'NIFTY Gap':<16}  {gap_c}")
    print(f"  {'PCR':<16}  {C.c(pcr_s)}")
    print(f"  {'As of':<16}  {C.dim(live.get('as_of', 'N/A'))}")

    print(f"\n  {C.b('MARKET CONDITION SCORE')}  "
          f"{progress_bar(mkt_score)}  "
          f"{C.b(str(mkt_score))}/100")

    cat = ("VOLATILE / HIGH RISK"  if mkt_score < 41 else
           "MIXED CONDITIONS"      if mkt_score < 61 else
           "CALM / STABLE"         if mkt_score < 81 else "IDEAL CONDITIONS")
    cat_c = (C.r(cat) if mkt_score < 41 else C.y(cat) if mkt_score < 61 else C.g(cat))
    print(f"  {'Market Environment':<16}  {cat_c}")
    print(f"  {C.dim('(Market calm/volatile — does NOT mean you should trade)')}")

    print(f"\n  {C.dim('Score breakdown:')}")
    for k, entry in mkt_bkdwn.items():
        pts, mx, detail = entry[0], entry[1], entry[2]
        meaning = entry[3] if len(entry) > 3 else ""
        bar   = "▪" * pts + "·" * (mx - pts)
        pct_c = C.g(f"{pts}/{mx}") if pts >= mx * 0.75 else (
                C.y(f"{pts}/{mx}") if pts >= mx * 0.5  else C.r(f"{pts}/{mx}"))
        print(f"    {k:<22}  [{bar}] {pct_c}  {C.dim(detail)}")
        if meaning:
            print(f"    {' '*22}  {C.dim('→ ' + meaning)}")


def display_personal_stats(stats: dict):
    print(section("YOUR 3-YEAR TRADING PROFILE"))
    if not stats:
        print(C.y("  No trade history loaded."))
        return

    wr   = stats["win_rate"]
    wr_c = C.g(f"{wr}%") if wr >= 55 else (C.y(f"{wr}%") if wr >= 45 else C.r(f"{wr}%"))

    print(f"\n  {'Total trading days':<28}  {C.b(str(stats['total_days']))}")
    print(f"  {'Win days / Loss days':<28}  {C.g(str(stats['win_days']))} / {C.r(str(stats['loss_days']))}")
    print(f"  {'Overall win rate':<28}  {wr_c}")
    print(f"  {'Total P&L (3 years)':<28}  {pnl_color(stats['total_pnl'])}")
    print(f"  {'Avg profit (win days)':<28}  {pnl_color(stats['avg_win'])}")
    print(f"  {'Avg loss  (loss days)':<28}  {pnl_color(stats['avg_loss'])}")

    best_d, best_p  = stats["best_day"]
    worst_d, worst_p = stats["worst_day"]
    print(f"  {'Best single day':<28}  {pnl_color(best_p)}  {C.dim(str(best_d))}")
    print(f"  {'Worst single day':<28}  {pnl_color(worst_p)}  {C.dim(str(worst_d))}")

    print(f"\n  {C.dim('Year-by-year P&L:')}")
    for yr, pnl in sorted(stats["yearly"].items()):
        bar = "█" * min(int(abs(pnl) / 50000), 20)
        c   = C.g if pnl >= 0 else C.r
        print(f"    {yr}  {c(bar)}  {pnl_color(pnl)}")


def display_similar_days(sim_df: pd.DataFrame, sim_report: dict):
    print(section(f"SIMILAR DAYS FROM YOUR HISTORY  (found: {sim_report.get('count', 0)})"))
    if sim_df.empty or not sim_report:
        print(C.y("  Not enough historical data for similarity matching."))
        return

    wr    = sim_report["win_rate"]
    wr_c  = C.g(f"{wr}%") if wr >= 55 else (C.y(f"{wr}%") if wr >= 45 else C.r(f"{wr}%"))

    print(f"\n  {'Win rate on similar days':<32}  {wr_c}")
    print(f"  {'Avg profit (similar win days)':<32}  {pnl_color(sim_report['avg_win'])}")
    print(f"  {'Avg loss  (similar loss days)':<32}  {pnl_color(sim_report['avg_loss'])}")
    print(f"  {'Combined P&L on similar days':<32}  {pnl_color(sim_report['total_pnl'])}")

    best  = sim_report.get("best_day")
    worst = sim_report.get("worst_day")
    if best is not None:
        print(f"  {'Best similar day':<32}  {pnl_color(best['pnl'])}  {C.dim(str(best['date']))}")
    if worst is not None:
        print(f"  {'Worst similar day':<32}  {pnl_color(worst['pnl'])}  {C.dim(str(worst['date']))}")

    print(f"\n  {C.dim('Top 5 most similar historical days:')}")
    print(f"  {C.DIM}{'Date':<14} {'VIX':>6} {'Gap%':>7} {'Day':<11} {'P&L':>12} {'Sim':>6}{C.RST}")
    print(f"  {C.DIM}{'─'*60}{C.RST}")
    for rec in sim_report.get("top5", []):
        d_s = str(rec["date"])
        p_c = C.g(f"₹{rec['pnl']:>+,.0f}") if rec["pnl"] >= 0 else C.r(f"₹{rec['pnl']:>+,.0f}")
        vix_c = C.y(f"{rec['vix']:>5.1f}") if rec['vix'] > 15 else C.g(f"{rec['vix']:>5.1f}")
        print(f"  {d_s:<14} {vix_c}  {rec['gap_pct']:>+6.2f}%  {rec['dow']:<11}  {p_c:>12}  {C.dim(str(rec['sim_score'])+'/100')}")


def display_behavioral(behav: dict):
    print(section("BEHAVIORAL RISK ANALYSIS"))
    risks    = behav.get("risks", [])
    insights = behav.get("insights", [])

    if risks:
        print(f"\n  {C.b(C.y('⚠  WARNINGS'))}")
        for risk in risks:
            weight_s = "●" * risk["weight"] + "○" * (4 - risk["weight"])
            print(f"\n  {C.y(risk['type']):<30}  Risk: {C.r(weight_s)}")
            # Word-wrap detail at 70 chars
            detail = risk["detail"]
            while len(detail) > 70:
                cut = detail[:70].rfind(" ")
                print(f"    {C.dim(detail[:cut])}")
                detail = detail[cut+1:]
            print(f"    {C.dim(detail)}")
    else:
        print(f"\n  {C.g('✓  No major behavioral risks detected.')}")

    if insights:
        print(f"\n  {C.b(C.c('ℹ  INSIGHTS'))}")
        for ins in insights:
            print(f"    {C.dim('•')} {ins}")


def display_permission_score(perm_score: int, verdict: str, perm_bkdwn: dict,
                              live: dict, sim_report: dict):
    print(section("TRADING PERMISSION SCORE  (should YOU trade — based on YOUR history)"))
    print(f"  {C.dim('Weights: Market 35% + Your similar-day win rate 40% + Recent form 20% − Behavioral risks')}")

    verdict_display = {
        "NO_TRADE":        (C.r,  "🔴  NO TRADE TODAY"),
        "CAUTION":         (C.y,  "🟡  TRADE WITH CAUTION"),
        "NORMAL":          (C.g,  "🟢  NORMAL TRADING"),
        "HIGH_CONFIDENCE": (C.g,  "✅  HIGH CONFIDENCE DAY"),
    }
    color_fn, label = verdict_display.get(verdict, (C.y, "⚪  UNKNOWN"))

    print(f"\n  {C.b('PERMISSION SCORE')}   "
          f"{progress_bar(perm_score)}  "
          f"{C.b(str(perm_score))}/100")
    print(f"\n  {color_fn(C.b(label))}")

    print(f"\n  {C.dim('Score breakdown:')}")
    for k, v in perm_bkdwn.items():
        v_c = C.g(str(v)) if v > 0 else C.r(str(v))
        print(f"    {k:<38}  {v_c}")

    # Direction bias
    pcr = live.get("pcr")
    vix = live.get("vix")
    gap = live.get("gap_pct")

    signals = []
    if gap is not None:
        if gap > 0.4:  signals.append(("BULLISH", 1))
        elif gap < -0.4: signals.append(("BEARISH", 1))

    if pcr is not None:
        if pcr > 1.2:  signals.append(("BULLISH", 1))
        elif pcr < 0.8: signals.append(("BEARISH", 1))

    sim_wr = sim_report.get("win_rate", 50)
    if sim_wr > 60:   signals.append(("BULLISH", 1))
    elif sim_wr < 40: signals.append(("BEARISH", 1))

    bull_score = sum(w for s, w in signals if s == "BULLISH")
    bear_score = sum(w for s, w in signals if s == "BEARISH")

    if   bull_score > bear_score: bias = C.g("BULLISH BIAS")
    elif bear_score > bull_score: bias = C.r("BEARISH BIAS")
    else:                         bias = C.y("NEUTRAL / WAIT FOR CONFIRMATION")

    print(f"\n  {'Direction Bias':<22}  {bias}")

    # Risk and position sizing suggestions
    if verdict == "NO_TRADE":
        print(f"  {'Position Sizing':<22}  {C.r('0%  —  Do not trade')}")
    elif verdict == "CAUTION":
        print(f"  {'Position Sizing':<22}  {C.y('25–40%  of normal size')}")
    elif verdict == "NORMAL":
        print(f"  {'Position Sizing':<22}  {C.g('50–75%  of normal size')}")
    else:
        print(f"  {'Position Sizing':<22}  {C.g('75–100% of normal size')}")


def display_ai_narrative(narrative: str):
    print(section("AI ANALYSIS & RECOMMENDATION"))
    print()
    lines = narrative.split("\n")
    for line in lines:
        if line.strip():
            # Word-wrap long lines
            while len(line) > 78:
                cut = line[:78].rfind(" ")
                if cut < 0:
                    cut = 78
                print(f"  {line[:cut]}")
                line = line[cut+1:]
            print(f"  {line}")
        else:
            print()


def display_footer():
    print(f"\n{C.DIM}{'═' * W}{C.RST}")
    print(f"  {C.dim('Data: ayush_previous_data/ + Lakshmi.xlsx + NSE live API + yfinance')}")
    print(f"  {C.dim('Refresh cache:  python3 PERSONAL_TRADING_AI.py --refresh')}")
    print(f"  {C.dim('Run this every morning before 9:15 AM for best results.')}")
    print(f"{C.DIM}{'═' * W}{C.RST}\n")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    force_refresh = "--refresh" in sys.argv

    # ── Splash ─────────────────────────────────────────────────────────────────
    print(f"\n{C.CYN}{C.BD}{'═'*W}{C.RST}")
    print(f"{C.CYN}{C.BD}  Loading Personal Trading Intelligence AI…{C.RST}")
    print(f"{C.DIM}  Parsing 3+ years of your F&O trade history…{C.RST}")

    # ── Parse historical trade data ────────────────────────────────────────────
    daily_pnl, daily_trades, expiry_days = parse_excel_history()
    intraday = parse_lakshmi_intraday()

    print(f"  {C.g('✓')} Loaded {len(daily_pnl)} trading days  |  "
          f"{len(intraday)} intraday-detailed days")

    # ── Fetch historical market data ───────────────────────────────────────────
    print(f"{C.DIM}  Fetching historical NIFTY + VIX data…{C.RST}")
    market_db = build_market_db(force_refresh=force_refresh)
    if not market_db.empty:
        print(f"  {C.g('✓')} Market DB: {len(market_db)} trading days (NIFTY + VIX)")
    else:
        print(f"  {C.y('⚠')} Market DB unavailable — similarity matching limited")

    # ── Fetch live market data ──────────────────────────────────────────────────
    print(f"{C.DIM}  Fetching live market data from NSE…{C.RST}")
    live = fetch_live_market()
    if live.get("vix"):
        print(f"  {C.g('✓')} Live: VIX={live['vix']:.2f}  NIFTY={live.get('nifty', 'N/A')}"
              f"  PCR={live.get('pcr', 'N/A')}")
    else:
        print(f"  {C.y('⚠')} NSE live data unavailable (market may be closed)")

    # ── Run analysis ───────────────────────────────────────────────────────────
    mkt_score, mkt_bkdwn   = market_condition_score(live)
    sim_df,    sim_report   = find_similar_days(live, market_db, daily_pnl)
    behav                   = behavioral_analysis(daily_pnl, daily_trades, expiry_days, intraday)
    perm_score, verdict, perm_bkdwn = trading_permission_score(
        mkt_score, sim_report, behav, daily_pnl
    )
    stats = overall_stats(daily_pnl)

    print(f"  {C.g('✓')} Analysis complete.  Generating report…\n")

    # ── Display ────────────────────────────────────────────────────────────────
    display_header(live)
    display_live_market(live, mkt_score, mkt_bkdwn)
    display_personal_stats(stats)
    display_similar_days(sim_df, sim_report)
    display_behavioral(behav)
    display_permission_score(perm_score, verdict, perm_bkdwn, live, sim_report)

    # ── AI narrative ───────────────────────────────────────────────────────────
    print(f"{C.DIM}  Generating AI narrative…{C.RST}")
    narrative = generate_ai_narrative(
        live, mkt_score, mkt_bkdwn, sim_report, behav, perm_score, verdict, daily_pnl
    )
    display_ai_narrative(narrative)
    display_footer()


if __name__ == "__main__":
    main()
