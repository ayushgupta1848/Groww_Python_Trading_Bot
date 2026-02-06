# Groww_CP_Bot.py


############### MUST READ #####################
#• To start Bot we just need to assure that config data is proper,
# Need to remove comments from "place_cp_order" method to validate the order status under comment starts from "#STATUS VALIDATION",
# And in last we need to validate that funds are matching with BOT and Groww Account unnder comment "Funds check"

#DOWNLOAD GROWW INSTRUMENTS:- https://growwapi-assets.groww.in/instruments/instrument.csv

import os
import re
import json
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import pyotp
from openpyxl import Workbook, load_workbook
from playsound3 import playsound
from datetime import datetime
from threading import Lock
import requests
import sys
from datetime import datetime
import time
import os
import sys
from datetime import datetime
import random

# ENHANCEMENT: Use a session for persistent HTTP connections (faster polling)
session = requests.Session()

MOMENTUM_SAMPLES = 5
MOMENTUM_DELAY = 1

def setup_persistent_logger():
    """Creates a local 'logs' folder beside the script and logs all console output there."""
    # Create /logs folder in the same directory as the script
    base_dir = os.path.dirname(os.path.abspath(__file__))
    log_dir = os.path.join(base_dir, "logs")
    os.makedirs(log_dir, exist_ok=True)

    # Create a timestamped log file
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_path = os.path.join(log_dir, f"Groww_Bot_{timestamp}.log")

    # Define a Tee class to write to both console and log file
    class Tee:
        def __init__(self, *streams):
            self.streams = streams

        def write(self, data):
            for s in self.streams:
                try:
                    s.write(data)
                    s.flush()
                except Exception:
                    pass  # Ignore on shutdown

        def flush(self):
            for s in self.streams:
                try:
                    s.flush()
                except Exception:
                    pass

    # Open log file (unbuffered, UTF-8)
    logfile = open(log_path, "a", buffering=1, encoding="utf-8")

    # Redirect both stdout and stderr
    sys.stdout = Tee(sys.stdout, logfile)
    sys.stderr = Tee(sys.stderr, logfile)

    print(f"📝 Logging started. Log file: {log_path}")

    return log_path


# --- Initialize persistent logging ---
LOG_FILE_PATH = setup_persistent_logger()

EXCEL_FILE = "technical_logs.xlsx"

# Replace with your Groww API key (or leave and use TOTP to fetch access_token)
api_key = "eyJraWQiOiJaTUtjVXciLCJhbGciOiJFUzI1NiJ9.eyJleHAiOjI1NTAwNDY3MzksImlhdCI6MTc2MTY0NjczOSwibmJmIjoxNzYxNjQ2NzM5LCJzdWIiOiJ7XCJ0b2tlblJlZklkXCI6XCI2MmEwMTc4YS0wOTk3LTQ0ZDAtOWRiNC0wZDAzOWM5MzY3YmZcIixcInZlbmRvckludGVncmF0aW9uS2V5XCI6XCJlMzFmZjIzYjA4NmI0MDZjODg3NGIyZjZkODQ5NTMxM1wiLFwidXNlckFjY291bnRJZFwiOlwiMmVlMjYyMjItN2MwNS00Y2IwLWIwM2MtNzAzYWRmNWVmN2RkXCIsXCJkZXZpY2VJZFwiOlwiNWQwYzdjODgtMGI1OS01MDU0LTk5ZTYtYWU5MzY5OTc2ZmRiXCIsXCJzZXNzaW9uSWRcIjpcIjY1NzBiNDUwLWE2YzYtNDMyYi1hYTJmLTA4MjExZjk0YzRiOVwiLFwiYWRkaXRpb25hbERhdGFcIjpcIno1NC9NZzltdjE2WXdmb0gvS0EwYktvMDZXRlpjc241VUNmTWF5aERtNGxSTkczdTlLa2pWZDNoWjU1ZStNZERhWXBOVi9UOUxIRmtQejFFQisybTdRPT1cIixcInJvbGVcIjpcImF1dGgtdG90cFwiLFwic291cmNlSXBBZGRyZXNzXCI6XCIxNzEuNjAuMTY5LjI1MiwxNzIuNjkuOTUuOTMsMzUuMjQxLjIzLjEyM1wiLFwidHdvRmFFeHBpcnlUc1wiOjI1NTAwNDY3Mzk5MTV9IiwiaXNzIjoiYXBleC1hdXRoLXByb2QtYXBwIn0.EKERC7OzG-lblhaOSQPyb44mafdNFpErGbcELiTiLnRk4WEW9p7aBBf6iq-3LGagY4ORdOCnrXbRhyGzbscxSw"
totp_gen = pyotp.TOTP('WI4M7KCAMH5CGN2I6SVB6MN2QDKUXRJF')

# Get project root directory (folder where your script is running)
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
# Build CSV path dynamically
csv_path = os.path.join(PROJECT_ROOT, "instrument.csv")
print(csv_path)

# Instruments CSV/JSON path (script will convert CSV -> JSON if convert_csv_to_json = yes)
# csv_path = r"C:\Users\HITS\Downloads\instrument (6).csv"
convert_csv_to_json = "yes"

# Telegram placeholders (you will replace later)
TELEGRAM_BOT_TOKEN = "PUT_YOUR_TOKEN_HERE"
TELEGRAM_CHAT_ID = "PUT_YOUR_CHAT_ID_HERE"

# Sound files (ensure these exist in script folder or provide full path)
SOUND_PROFIT = "coin.mp3"
SOUND_SL = "SL_HIT.mp3"
SOUND_user_input = "User_input.WAV"

# Trade defaults for Groww
DEFAULT_PRODUCT = "MIS"   # intraday; change to "NRML" if you want positional
ORDER_PRODUCT_MAP = {
    "MIS": "MIS",
    "NRML": "NRML"
}
# NOTE: the growwapi wrapper constants are used from the growwapi instance below

# ----------------- import growwapi late (after auth) -----------------
try:
    from growwapi import GrowwAPI
except Exception:
    # If local module not available, user must install or place it in PYTHONPATH
    print("❗ growwapi module not found. Make sure it's installed and importable.")
    # continue; import errors will show when script runs further

# ----------------- Groww auth & wrapper -----------------
def groww_init(api_key):
    """
    Return growwapi client instance (GrowwAPI(access_token))
    This function gets access_token using GrowwAPI.get_access_token if available.
    """
    totp = totp_gen.now()
    try:
        access_token = GrowwAPI.get_access_token(api_key=api_key, totp=totp)
        client = GrowwAPI(access_token)
        print(access_token)
        print("✅ Groww API Initialized Successfully")
        return client, access_token
    except Exception as e:
        print(f"❌ Groww login failed: {e}")
        raise

# Init groww client
groww ,access_token = groww_init(api_key)


# ----------------- Utilities: Telegram, Sound, Excel Logging -----------------

# === TELEGRAM CONFIG ===
BOT_TOKEN = "8482701378:AAG7Jtfw0ZW_K9mFiX21LpsyUAV4oOcDiAQ"
CHAT_ID = "6012308856"

def send_telegram(message: str):
    """Sends a telegram message asynchronously to avoid blocking the main thread."""
    def _send():
        try:
            url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
            payload = {"chat_id": CHAT_ID, "text": message}
            requests.post(url, data=payload)
        except Exception as e:
            print(f"⚠️ Telegram Error: {e}")

    # Fire and forget thread
    threading.Thread(target=_send, daemon=True).start()

def play_sound_async(filename):
    try:
        if not os.path.exists(filename):
            print(f"🔇 Sound file not found: {filename}")
            return
        threading.Thread(target=playsound, args=(filename,), daemon=True).start()
    except Exception as e:
        print(f"🔇 Sound error: {e}")

def log_technical_to_excel(symbol, ltp, ema_9, rsi, adx, vwap, sma_20):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Create file if not exists
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Technicals"
        ws.append([
            "Time", "Symbol", "LTP",
            "EMA_9", "SMA_20", "RSI", "ADX", "VWAP"
        ])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Technicals"]

    ws.append([
        now, symbol, ltp,
        round(ema_9, 2),
        round(sma_20, 2),
        round(rsi, 2),
        round(adx, 2),
        round(vwap, 2)
    ])

    wb.save(EXCEL_FILE)

def log_trade_to_excel(symbol, buy_price, sell_price, quantity, profit , volume , oi):
    file_name = "Lakshmi1.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append(["DateTime", "Symbol", "Buy Price", "Sell Price", "Quantity", "Profit", "Volume" , "oi"])
        wb.save(file_name)

    # Load existing workbook
    wb = load_workbook(file_name)
    ws = wb.active

    # Find the next empty row
    next_row = ws.max_row + 1
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.cell(row=next_row, column=1).value = now
    ws.cell(row=next_row, column=2).value = symbol
    ws.cell(row=next_row, column=3).value = buy_price
    ws.cell(row=next_row, column=4).value = sell_price
    ws.cell(row=next_row, column=5).value = quantity
    ws.cell(row=next_row, column=6).value = round(profit, 2)
    ws.cell(row=next_row, column=7).value = volume
    ws.cell(row=next_row, column=8).value = oi
    wb.save(file_name)


# ----------------- CSV -> JSON loader -----------------
def csv_to_json(csv_file_path, json_file_path=None):
    import csv
    if json_file_path is None:
        json_file_path = os.path.splitext(csv_file_path)[0] + ".json"
    data = []
    with open(csv_file_path, encoding='utf-8') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            data.append(row)
    with open(json_file_path, 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file, indent=4, ensure_ascii=False)
    print(f"✅ Converted '{csv_file_path}' → '{json_file_path}'")
    return data

ltp_lock = threading.Lock()

def get_ltp_for_instrument(instrument, access_token, verbose=True, segment="FNO", delay=0.2, max_retries=2):
    """
    Fetches the latest traded price (LTP) for a given F&O instrument using Groww's authenticated API.
    Thread-safe with a global lock to prevent too-frequent API calls.
    Added 'delay' parameter to control sleep time (default 0.1s). Set to 0 for instant return.
    """
    trading_symbol = instrument.get("trading_symbol")  # e.g. NIFTY25N0425950CE
    if not trading_symbol:
        print("⚠️ Missing trading_symbol in instrument.")
        return None

    exchange_symbol = f"NSE_{trading_symbol}"
    url = f"https://api.groww.in/v1/live-data/ltp?segment={segment}&exchange_symbols={exchange_symbol}"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0"
    }

    # 🔒 Lock ensures one API call at a time
    last_exc = None
    for attempt in range(max_retries + 1):
        try:
            with ltp_lock:
                # Use session for faster connection reuse
                resp = session.get(url, headers=headers, timeout=10)
                if delay > 0:
                    time.sleep(delay)  # ⏳ delay to respect Groww API rate limits

            if resp.status_code == 429:
                # small exponential backoff + jitter
                backoff = 0.4 * (2 ** attempt) + random.uniform(-0.05, 0.05)
                print(f"⚠️ LTP 429 received, backing off {backoff:.2f}s (attempt {attempt+1})")
                time.sleep(backoff)
                last_exc = requests.exceptions.HTTPError("429 Too Many Requests")
                continue

            resp.raise_for_status()
            data = resp.json()
            payload = data.get("payload", {})
            ltp = payload.get(exchange_symbol)

            if ltp is None:
                print(f"⚠️ No LTP found for {exchange_symbol} in payload: {payload}")
                return None
            if verbose:
                print(f"💰 LTP for {exchange_symbol}: ₹{ltp} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
                send_telegram(f"💰 LTP for {exchange_symbol}: ₹{ltp} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            return float(ltp)

        except Exception as e:
            last_exc = e
            # retry small number of times on transient errors
            if attempt < max_retries:
                backoff = 0.3 * (attempt + 1) + random.uniform(0, 0.05)
                print(f"⚠️ Error fetching LTP ({e}), retrying in {backoff:.2f}s...")
                time.sleep(backoff)
                continue
            else:
                print(f"⚠️ Error fetching LTP for {instrument.get('trading_symbol')}: {e}")
                return None

    # If we exhaust retries
    print(f"⚠️ Failed to fetch LTP after {max_retries + 1} attempts")
    return None

def get_historical_candles(instrument, access_token, interval="5minute", days=2):
    """
    Fetch historical candle data for ATR calculation.
    Returns list of candles: [[timestamp, open, high, low, close, volume], ...]
    
    Note: This is a placeholder. Implement using your broker's historical data API.
    For Groww, you may need to use their chart/candle endpoint.
    """
    try:
        # Placeholder implementation - replace with actual Groww API call
        # Example structure for when you implement:
        # url = f"https://groww.in/v1/api/charting_service/v2/chart/exchange/segment/{instrument['groww_symbol']}"
        # params = {"intervalInMinutes": 5, "days": days}
        # resp = requests.get(url, headers={"Authorization": f"Bearer {access_token}"}, params=params)
        # data = resp.json()
        # return data.get("candles", [])
        
        # For now, return None to skip ATR calculation if not implemented
        print("⚠️ Historical candles API not implemented yet. Skipping ATR calculation.")
        return None
        
    except Exception as e:
        print(f"⚠️ Error fetching historical candles: {e}")
        return None

def get_nifty_spot_price(access_token=None,json_path=None):
    """
    Fetches live NIFTY 50 spot price using Groww instrument data.
    Matches by trading_symbol = 'NIFTY' or groww_symbol = 'NSE-NIFTY'
    """
    global instruments1

    if json_path is None:
        json_path = os.path.splitext(csv_path)[0] + ".json"

    # 🔄 Step 1: Load or convert JSON
    if convert_csv_to_json.lower() == "yes":
        instruments1 = csv_to_json(csv_path, json_path)
    else:
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON not found: {json_path}")
        with open(json_path, "r", encoding="utf-8") as jf:
            instruments1 = json.load(jf)
        print(f"ℹ️ Loaded instruments from existing JSON: {json_path}")

    try:
        nifty_spot_instrument = next(
            (item for item in instruments1
             if item.get("trading_symbol") == "NIFTY"
             or item.get("groww_symbol") == "NSE-NIFTY"
             or item.get("name") == "NIFTY 50"),
            None
        )

        if not nifty_spot_instrument:
            print("⚠️ NIFTY spot instrument not found in instruments1")
            return 0

        spot = get_ltp_for_instrument(nifty_spot_instrument, access_token, verbose=False,segment = "CASH")
        if spot:
            print(f"📊 Live NIFTY Spot: {spot}")
            return float(spot)
        else:
            print("⚠️ Failed to fetch LTP for NIFTY spot")
            return 0
    except Exception as e:
        print(f"⚠️ Error fetching NIFTY spot: {e}")
        return 0


CONFIG = {
    "index": "NIFTY",
    "expiry": "2026-02-10",  # Updated to DD/MM/YYYY to match instruments JSON
    "min_premium": 90,
    "max_premium": 155,
    "lots": 22,
    "book_profit": 1050,
    "target_pnl": 6000,
    "spot":get_nifty_spot_price(access_token),
    "TRAIL_START_PROFIT": 1.25,  # Start trailing after this profit per unit (in points) #NEWCHANGE
    "TRAIL_STEP": .75,  # Trailing step (in points) #NEWCHANGE
    "POLL_INTERVAL": 0.2,  # Poll interval in seconds (Reduced for faster SL hit)
    "MAX_TRAIL_TIME": 3600,  # Max trailing time in seconds (1 hour)
    "HARD_SL_POINTS": 6.0,  # Hard stop loss points below entry
    "user_confirmation_needed": False,   # or False
    "enable_technical_filters": False,  # Set to False to skip technical filters (EMA, RSI, ADX, VWAP)
    "enable_option_filters": False,  # Set to False to skip option Greeks filters (IV, Delta, Gamma, OI, Volume)
    # Volume + OI Risk Filter Settings
    "enable_volume_oi_filter": False,  # Enable Volume + OI based risk management
    "VOLUME_SPIKE_FACTOR": 2.0,  # Volume spike detection: curr_volume > prev_volume * factor (relaxed from 1.5)
    "HIGH_OI_THRESHOLD": 200000,  # High OI threshold for trap zone detection (relaxed from 150k)
    "SMALL_PROFIT_TARGET": 1500,  # Quick exit profit target (₹)
    "MIN_MOMENTUM_SCORE": 0.15,  # Momentum score threshold (relaxed from 0.2 for more entries)
    # Risk & Prefetch settings
    "MAX_EXPOSURE_PCT": 98,        # percent of equity per trade
    "MAX_DAILY_LOSS_PCT": 2.0,     # percent of equity to stop trading for the day
    "MAX_CONSECUTIVE_LOSSES": 3,
    "COOLDOWN_SECONDS": 1800,      # cooldown after consecutive losses (seconds)
    "PREFETCH_INTERVAL": 8,        # seconds between option-chain prefetches
    "PREFETCH_JITTER": 1.5,        # jitter for prefetch interval
    # ATR & Dynamic Stops
    "ATR_MULTIPLIER_SL": 2.0,
    "ATR_MULTIPLIER_TP": 3.0,
    "ATR_TIGHTEN_FACTOR": 0.5,
    "PROFIT_TIGHTEN_THRESHOLD": 0.10,
    "USE_VOLATILITY_SIZING": True,
    # Greeks Thresholds (relaxed for more trades)
    "MIN_DELTA_ENTRY": 0.35,  # Relaxed from 0.45
    "MAX_IV_PERCENTILE": 0.75  # Relaxed from 0.70
}

# ----------------- Volume + OI Risk Filter Functions -----------------
def is_volume_spike(curr_volume, prev_volume, spike_factor=None):
    """
    Detect if current volume is a spike compared to previous volume.
    Returns True if curr_volume > prev_volume * spike_factor
    """
    if spike_factor is None:
        spike_factor = CONFIG.get("VOLUME_SPIKE_FACTOR", 1.5)
    
    if prev_volume is None or prev_volume <= 0:
        return False
    
    return curr_volume > (prev_volume * spike_factor)

def is_high_oi(oi, threshold=None):
    """
    Check if Open Interest is above high threshold.
    Returns True if OI exceeds threshold (indicates crowded trade).
    """
    if threshold is None:
        threshold = CONFIG.get("HIGH_OI_THRESHOLD", 150000)
    
    return oi > threshold

def should_exit_trade(profit, curr_volume, prev_volume, oi, curr_ltp, prev_ltp):
    """
    Determine if trade should be exited based on Volume + OI + PnL conditions.
    
    EXIT CONDITIONS (return True if ANY):
    1. Small profit target reached → book profit
    2. TRAP ZONE: Volume spike + High OI + (negative PnL OR stalled momentum)
    
    Returns: (should_exit: bool, reason: str)
    """
    cfg = CONFIG
    small_profit_target = cfg.get("SMALL_PROFIT_TARGET", 500)
    
    # ✅ CONDITION 1: Small profit target reached
    if profit >= small_profit_target:
        return True, f"✅ Small profit target reached: ₹{profit:.2f} >= ₹{small_profit_target}"
    
    # ❌ CONDITION 2: TRAP ZONE Detection
    volume_spike = is_volume_spike(curr_volume, prev_volume)
    high_oi = is_high_oi(oi)
    
    # Check momentum stall: price not moving up
    momentum_stalled = False
    if prev_ltp is not None and prev_ltp > 0:
        price_change_pct = ((curr_ltp - prev_ltp) / prev_ltp) * 100
        momentum_stalled = abs(price_change_pct) < 0.1  # Less than 0.1% movement
    
    negative_pnl = profit < 0
    
    if volume_spike and high_oi and (negative_pnl or momentum_stalled):
        reason = (
            f"⚠️ TRAP ZONE DETECTED:\n"
            f"  • Volume Spike: {curr_volume:,} vs {prev_volume:,} (factor {curr_volume/prev_volume:.2f}x)\n"
            f"  • High OI: {oi:,}\n"
            f"  • {'Negative PnL' if negative_pnl else 'Momentum Stalled'}\n"
            f"  → Exiting to avoid further loss"
        )
        return True, reason
    
    return False, ""

def calculate_momentum_score(current_price, ma_50, rsi, adx=None):
    """
    Calculate composite momentum score based on research.
    Formula: Score = (Price/MA) + (RSI/50) - 1
    Research suggests entry when score > 0.2
    
    Returns: momentum_score (float)
    """
    if ma_50 is None or ma_50 <= 0:
        return 0.0
    
    price_ma_ratio = (current_price / ma_50)
    rsi_normalized = (rsi / 50.0) if rsi else 1.0
    
    momentum_score = price_ma_ratio + rsi_normalized - 1.0
    
    # Bonus for strong trend (ADX > 25)
    if adx and adx > 25:
        momentum_score += 0.1
    
    return round(momentum_score, 3)

def should_enter_trade(curr_volume, prev_volume, oi, price_momentum_positive, momentum_score=None):
    """
    Determine if trade entry is safe based on Volume + OI conditions.
    Enhanced with momentum score validation.
    
    SAFE ENTRY CONDITIONS (return True if ALL):
    1. Volume is NOT spiking (gradual increase)
    2. Open Interest is not extremely high
    3. Price momentum is in trade direction
    4. Momentum score > threshold (research: 0.2)
    
    Returns: (should_enter: bool, reason: str)
    """
    volume_spike = is_volume_spike(curr_volume, prev_volume)
    high_oi = is_high_oi(oi)
    min_score = CONFIG.get("MIN_MOMENTUM_SCORE", 0.2)
    
    # ❌ Reject if volume spike detected
    if volume_spike:
        return False, f"❌ Volume spike detected: {curr_volume:,} vs {prev_volume:,} - Likely trap zone"
    
    # ❌ Reject if OI too high (crowded trade)
    if high_oi:
        return False, f"❌ High OI detected: {oi:,} - Overcrowded trade"
    
    # ❌ Reject if momentum not positive
    if not price_momentum_positive:
        return False, "❌ Price momentum not favorable"
    
    # ❌ Reject if momentum score too weak
    if momentum_score is not None and momentum_score < min_score:
        return False, f"❌ Weak momentum score: {momentum_score:.3f} < {min_score} (research threshold)"
    
    # ✅ Safe to enter
    score_msg = f", Score={momentum_score:.3f}" if momentum_score else ""
    return True, f"✅ Safe entry: Volume gradual, OI acceptable ({oi:,}), momentum positive{score_msg}"

def calculate_atr_from_candles(candles, period=14):
    """
    Calculate Average True Range from candle data.
    ATR measures volatility for dynamic stop placement.
    """
    if len(candles) < period + 1:
        return None
    
    true_ranges = []
    for i in range(1, len(candles)):
        high = candles[i][2]
        low = candles[i][3]
        prev_close = candles[i-1][4]
        
        tr = max(
            high - low,
            abs(high - prev_close),
            abs(low - prev_close)
        )
        true_ranges.append(tr)
    
    # Calculate ATR as average of true ranges
    atr = sum(true_ranges[-period:]) / period
    return round(atr, 2)

def calculate_dynamic_stops(entry_price, atr, profit_current=0, profit_threshold=0.10):
    """
    Calculate dynamic SL/TP based on ATR (research-based).
    - Initial: SL = Entry - (ATR × 2), TP = Entry + (ATR × 3)
    - After 10% profit: Tighten to SL = Entry - (ATR × 1)
    
    Returns: (stop_loss, take_profit)
    """
    cfg = CONFIG
    initial_sl_mult = cfg.get("ATR_MULTIPLIER_SL", 2.0)
    tp_mult = cfg.get("ATR_MULTIPLIER_TP", 3.0)
    tighten_factor = cfg.get("ATR_TIGHTEN_FACTOR", 0.5)
    
    # Check if profit threshold reached
    if profit_current >= profit_threshold:
        # Tighten stop loss
        sl_multiplier = initial_sl_mult * tighten_factor
    else:
        sl_multiplier = initial_sl_mult
    
    stop_loss = entry_price - (atr * sl_multiplier)
    take_profit = entry_price + (atr * tp_mult)
    
    return round(stop_loss, 2), round(take_profit, 2)

def calculate_volatility_adjusted_position_size(entry_price, atr, account_balance, risk_pct=1.5):
    """
    Calculate position size based on volatility (ATR).
    Formula: Position = (Account × Risk%) / (ATR × Multiplier)
    Research suggests 1-2% risk per trade.
    
    Returns: quantity
    """
    risk_amount = account_balance * (risk_pct / 100.0)
    risk_per_unit = atr * 2.0  # Use 2x ATR as risk per unit
    
    if risk_per_unit <= 0:
        return 0
    
    quantity = int(risk_amount / (risk_per_unit * entry_price))
    return max(quantity, 1)  # At least 1 unit

# ----------------- End Volume + OI Risk Filter -----------------

# Load instruments_data
def load_instruments_from_json(json_path=None):
    """
    Loads instruments from JSON (or CSV → JSON if convert_csv_to_json = 'yes'),
    but only keeps instruments:
      - matching expiry from CONFIG
      - within ±10 strikes of current index spot price
    """
    global instruments
    config = CONFIG
    INDEX = config["index"].upper()
    EXPIRY = config["expiry"].strip()

    if json_path is None:
        json_path = os.path.splitext(csv_path)[0] + ".json"

    # 🔄 Step 1: Load or convert JSON
    if convert_csv_to_json.lower() == "yes":
        instruments = csv_to_json(csv_path, json_path)
    else:
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON not found: {json_path}")
        with open(json_path, "r", encoding="utf-8") as jf:
            instruments = json.load(jf)
        print(f"ℹ️ Loaded instruments from existing JSON: {json_path}")

    # 🧩 Step 2: Get live index spot
    spot = config["spot"]

    # Determine strike step (e.g., NIFTY = 50, BANKNIFTY = 100)
    step = 100 if "BANK" in INDEX else 50

    # Define strike range (±10 strikes)
    nearest_strike = round(spot / step) * step
    lower_bound = nearest_strike - (10 * step)
    upper_bound = nearest_strike + (10 * step)

    print(f"🎯 Filtering {INDEX} {EXPIRY} instruments between {lower_bound}–{upper_bound} (Spot={spot})")

    # 🧹 Step 3: Filter instruments by expiry and strike range
    filtered = []
    for item in instruments:
        try:
            if item.get("underlying_symbol", "").upper() != INDEX:
                continue
            # if item.get("expiry_date", "").strip() != EXPIRY:
            #     continue
            strike = float(item.get("strike_price") or 0)
            if lower_bound <= strike <= upper_bound:
                filtered.append(item)
        except Exception:
            continue

    print(f"✅ Loaded {len(filtered)} filtered instruments (out of {len(instruments)})")
    instruments = filtered
    return instruments


# initialize
instruments_data = load_instruments_from_json()


# Pre-index instruments for quick lookup by internal_trading_symbol or groww_symbol or custom compact symbol
symbol_index = {}
for it in instruments_data:
    # keys: internal_trading_symbol, groww_symbol, compact like NIFTY04NOV2525950CE approximate
    try:
        k1 = it.get("internal_trading_symbol", "") or it.get("trading_symbol", "")
        k2 = it.get("groww_symbol", "")
        if k1:
            symbol_index[k1.upper()] = it
        if k2:
            symbol_index[k2.upper()] = it
    except Exception:
        pass

# ----------------- Helpers: date/expiry normalization -----------------
MONTHS = {
    'JAN': '01','FEB':'02','MAR':'03','APR':'04','MAY':'05','JUN':'06',
    'JUL':'07','AUG':'08','SEP':'09','OCT':'10','NOV':'11','DEC':'12'
}

def cmd_expiry_to_date(expiry_token):
    """
    expiry_token example: 04NOV25 or 04NOV2025 or 28AUG25 or 28AUG2025
    Return string 'DD/MM/YYYY'
    """
    m = re.match(r'(\d{1,2})([A-Z]{3})(\d{2,4})', expiry_token.upper())
    if not m:
        return None
    dd = m.group(1).zfill(2)
    mon_abbr = m.group(2)
    yy = m.group(3)
    if len(yy) == 2:
        yyyy = "20" + yy
    else:
        yyyy = yy
    mm = MONTHS.get(mon_abbr[:3], None)
    if not mm:
        return None
    return f"{dd}/{mm}/{yyyy}"

# ----------------- Command parser -----------------
def parse_cp_command(command):
    """
    Parse strings like:
      Buy 14 NIFTY04NOV2525950CE at CP and Book at 1050
    Returns dict or None
    """
    pattern = r'(?i)^\s*(Buy|Sell)\s+(\d+)\s+([A-Z]+)(\d{1,2}[A-Z]{3}\d{2,4})(\d+)(CE|PE)\s+at\s+CP\s+and\s+Book\s+at\s+(\d+(\.\d+)?)\s*$'
    m = re.match(pattern, command.strip())
    if not m:
        return None
    action = m.group(1).upper()
    lots = int(m.group(2))
    underlying = m.group(3).upper()
    expiry_token = m.group(4).upper()
    strike = m.group(5)
    opt_type = m.group(6).upper()
    target_profit = float(m.group(7))
    expiry_date = cmd_expiry_to_date(expiry_token)
    return {
        "action": action,
        "lots": lots,
        "underlying": underlying,
        "expiry_token": expiry_token,
        "expiry_date": expiry_date,
        "strike": strike,
        "opt_type": opt_type,
        "target_profit": target_profit
    }

# ----------------- Find instrument in instruments_data -----------------
def find_instrument_from_command(command: str, instruments: list):
    import re
    # Example command: Buy 14 NIFTY04NOV2525950CE at CP and Book at 1050
    pattern = r'([A-Z]+)(\d{1,2})([A-Z]{3})(\d{2,4})(\d+)(CE|PE)'
    match = re.search(pattern, command.upper())
    if not match:
        print("❌ Could not parse symbol from command.")
        return None

    underlying, day, mon, yr, strike, opt_type = match.groups()
    expiry_date = f"20{yr}-{mon_to_number(mon)}-{day}"

    # Find match in JSON
    for inst in instruments:
        if (
            inst["underlying_symbol"].upper() == underlying
            and inst["expiry_date"] == expiry_date
            and str(inst["strike_price"]) == strike
            and inst["instrument_type"].upper() == opt_type
        ):
            return inst

    print("❌ Instrument not found in instrument master.")
    return None


def mon_to_number(mon: str):
    mapping = {
        "JAN": "01", "FEB": "02", "MAR": "03", "APR": "04",
        "MAY": "05", "JUN": "06", "JUL": "07", "AUG": "08",
        "SEP": "09", "OCT": "10", "NOV": "11", "DEC": "12"
    }
    return mapping.get(mon.upper(), "00")


import requests, time

import requests
import json

import requests

def get_order_status(order_id, access_token):
    """
    Fetch the status of a Groww order (CASH, F&O, etc.)
    Works with official Groww REST API response format.
    """
    url = f"https://api.groww.in/v1/order/status/{order_id}?segment=FNO"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0"
    }

    try:
        # Use session for faster connection reuse
        resp = session.get(url, headers=headers, timeout=8)
        resp.raise_for_status()  # raises for non-200 responses

        data = resp.json()  # Proper JSON response from Groww
        print("🔍 Order status response:", data)

        # ✅ Extract status cleanly
        payload = data.get("payload", {})
        status = payload.get("order_status")

        return status

    except requests.exceptions.JSONDecodeError:
        print("⚠️ Error: Non-JSON response received.")
        print("Response text:", resp.text)
        return None

    except Exception as e:
        print(f"⚠️ Error fetching order status: {e}")
        return None


import time
from datetime import datetime, timedelta, timezone

def get_recent_market_direction(symbol, groww):
    """
    Returns 'CE' if recent 5-min direction is upward (bullish),
    'PE' if downward (bearish), or None if uncertain.
    Also prints the equivalent cURL command.
    """
    try:
        # Current time and 5 minutes earlier
        end_time = datetime.now()
        start_time = end_time - timedelta(minutes=3)

        # Convert to string format accepted by Groww API
        end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        start_time_str = start_time.strftime("%Y-%m-%d %H:%M:%S")

        # Construct the Groww candle API URL
        url = (
            f"https://api.groww.in/v1/historical/candles?"
            f"exchange=NSE&segment=FNO&groww_symbol={symbol}"
            f"&start_time={start_time_str}"
            f"&end_time={end_time_str}"
            f"&candle_interval=1minute"
        )

        # Print cURL command for debugging
        print("\n🌀 Generated cURL for Groww Candle API:")
        print(f"curl --location '{url}' \\")
        print("  --header 'Accept: application/json' \\")
        print(f"  --header 'Authorization: Bearer {access_token}' \\")
        print("  --header 'X-API-VERSION: 1.0'\n")

        # Fetch last 5-min candle via Groww SDK
        historical = groww.get_historical_candles(
            groww_symbol=symbol,
            exchange=groww.EXCHANGE_NSE,
            segment=groww.SEGMENT_FNO,
            start_time=start_time_str,
            end_time=end_time_str,
            candle_interval="1minute" # 1-min candles for better precision
        )

        candles = historical.get("candles", [])
        if not candles:
            print("⚠️ No recent candle data found.")
            return None

        first_open = candles[0][1]
        last_close = historical.get("closing_price")

        if "PE" in symbol:
            direction = "PE" if last_close > first_open else "CE"
        else:  # CE symbol
            direction = "CE" if last_close > first_open else "PE"

        print(f"📊 3-min candle trend → {direction} (O1={first_open}, C3={last_close})")
        return direction

    except Exception as e:
        print("⚠️ Error fetching recent market direction:", e)
        return None


import numpy as np

def calculate_sma(prices, period):
    if len(prices) < period:
        return None
    return np.mean(prices[-period:])

def calculate_ema(prices, period):
    if len(prices) < period:
        return None
    # Initial SMA
    ema = np.mean(prices[:period])
    multiplier = 2 / (period + 1)
    for price in prices[period:]:
        ema = (price - ema) * multiplier + ema
    return ema

# ----------------- Advanced Technicals (RSI, ADX, VWAP) -----------------
def calculate_rsi(prices, period=14):
    prices = np.array(prices)
    if len(prices) < period + 1:
        return 50 # Default neutral
    
    deltas = np.diff(prices)
    gains = np.maximum(deltas, 0)
    losses = -np.minimum(deltas, 0)
    
    avg_gain = np.mean(gains[:period])
    avg_loss = np.mean(losses[:period])
    
    if avg_loss == 0:
        return 100
        
    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    
    # Wilder's Smoothing
    for i in range(period, len(deltas)):
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period
        
        if avg_loss == 0:
            rsi = 100
        else:
            rs = avg_gain / avg_loss
            rsi = 100 - (100 / (1 + rs))
            
    return rsi

def calculate_adx(highs, lows, closes, period=14):
    if len(highs) < period * 2:
        return 25 # Default
        
    highs = np.array(highs)
    lows = np.array(lows)
    closes = np.array(closes)
    
    tr = np.zeros(len(highs))
    plus_dm = np.zeros(len(highs))
    minus_dm = np.zeros(len(highs))
    
    for i in range(1, len(highs)):
        tr[i] = max(highs[i] - lows[i], abs(highs[i] - closes[i-1]), abs(lows[i] - closes[i-1]))
        
        up_move = highs[i] - highs[i-1]
        down_move = lows[i-1] - lows[i]
        
        if up_move > down_move and up_move > 0:
            plus_dm[i] = up_move
        else:
            plus_dm[i] = 0
            
        if down_move > up_move and down_move > 0:
            minus_dm[i] = down_move
        else:
            minus_dm[i] = 0
            
    # Smoothing
    def smooth(data, period):
        smoothed = np.zeros_like(data)
        if len(data) > period:
            smoothed[period] = np.mean(data[1:period+1]) # Initial SMA
            for i in range(period+1, len(data)):
                smoothed[i] = (smoothed[i-1] * (period - 1) + data[i]) / period
        return smoothed

    atr = smooth(tr, period)
    plus_di = 100 * smooth(plus_dm, period) / (atr + 1e-9) # Avoid div by zero
    minus_di = 100 * smooth(minus_dm, period) / (atr + 1e-9)
    
    dx = 100 * np.abs(plus_di - minus_di) / (plus_di + minus_di + 1e-9)
    adx = smooth(dx, period)
    
    return adx[-1]

def calculate_vwap(prices, volumes):
    prices = np.array(prices)
    volumes = np.array(volumes)
    if len(prices) == 0 or len(volumes) == 0:
        return prices[-1] if len(prices) > 0 else 0
    
    vwap = np.cumsum(prices * volumes) / np.cumsum(volumes)
    return vwap[-1]

def get_technicals(symbol, groww_client, interval="1minute"):
    try:
        # Fetch enough data for EMA 20/SMA 20/RSI 14/ADX 14. 
        # Increased to 120 mins for better ADX smoothing
        end_time = datetime.now()
        start_time = end_time - timedelta(minutes=120) 
        
        end_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        start_str = start_time.strftime("%Y-%m-%d %H:%M:%S")
        
        historical = groww_client.get_historical_candles(
            groww_symbol=symbol,
            exchange=groww_client.EXCHANGE_NSE,
            segment=groww_client.SEGMENT_FNO,
            start_time=start_str,
            end_time=end_str,
            candle_interval=interval
        )
        
        candles = historical.get("candles", [])
        if not candles or len(candles) < 30:
            return None
            
        # Groww candles: [timestamp, open, high, low, close, volume]
        opens = [c[1] for c in candles]
        highs = [c[2] for c in candles]
        lows = [c[3] for c in candles]
        close_prices = [c[4] for c in candles]
        volumes = [c[5] for c in candles]
        
        sma_20 = calculate_sma(close_prices, 20)
        ema_9 = calculate_ema(close_prices, 9)
        ema_50 = calculate_ema(close_prices, 50)  # For momentum score calculation
        rsi_14 = calculate_rsi(close_prices, 14)
        adx_14 = calculate_adx(highs, lows, close_prices, 14)
        vwap = calculate_vwap(close_prices, volumes)
        
        current_price = close_prices[-1]
        
        return {
            "sma_20": sma_20,
            "ema_9": ema_9,
            "ema_50": ema_50,  # Added for momentum score
            "rsi": rsi_14,
            "adx": adx_14,
            "vwap": vwap,
            "ltp": current_price
        }
    except Exception as e:
        print(f"⚠️ Error fetching technicals: {e}")
        return None

# --- START: Caching layer to prevent API rate limiting ---
_option_chain_cache = {}
_option_chain_cache_lock = threading.Lock()
CACHE_EXPIRY_SECONDS = 15  # Cache the full option chain for 15 seconds (was 5s)
_last_api_call_time = 0  # Track last API call time globally
_api_call_lock = threading.Lock()  # Serialize API calls to avoid 429 errors
def _get_full_option_chain_cached(underlying, expiry_date, access_token, max_retries=3):
    """
    Fetches the full option chain from API or a time-based cache.
    Uses a short serialized window and exponential backoff on 429 to avoid rate limiting.
    """
    global _last_api_call_time
    cache_key = (underlying, expiry_date)
    now = time.time()

    # Fast-path: return if cached and fresh
    cached_payload, timestamp = _option_chain_cache.get(cache_key, (None, 0))
    if cached_payload and (now - timestamp) < CACHE_EXPIRY_SECONDS:
        return cached_payload

    # Acquire cache lock to coordinate fetchers (double-checked locking)
    with _option_chain_cache_lock:
        cached_payload, timestamp = _option_chain_cache.get(cache_key, (None, 0))
        if cached_payload and (time.time() - timestamp) < CACHE_EXPIRY_SECONDS:
            return cached_payload

        # Build request
        url = (
            f"https://api.groww.in/v1/option-chain/exchange/NSE"
            f"/underlying/{underlying}?expiry_date={expiry_date}"
        )
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0"
        }

        # Serialize external API calls so concurrent CE/PE workers don't flood the endpoint
        with _api_call_lock:
            # Respect a minimum inter-call interval to avoid short-burst 429s
            MIN_API_CALL_INTERVAL = 0.35  # seconds
            elapsed = time.time() - _last_api_call_time
            if elapsed < MIN_API_CALL_INTERVAL:
                time.sleep(MIN_API_CALL_INTERVAL - elapsed)

            last_exception = None
            for attempt in range(max_retries):
                try:
                    resp = session.get(url, headers=headers, timeout=12)
                    if resp.status_code == 429:
                        # Exponential backoff on 429
                        backoff = 2 ** attempt
                        print(f"⚠️ 429 received from option-chain API, backing off {backoff}s (attempt {attempt+1}/{max_retries})")
                        time.sleep(backoff)
                        last_exception = requests.exceptions.HTTPError("429 Too Many Requests")
                        continue

                    resp.raise_for_status()
                    data = resp.json()

                    if data.get("status") != "SUCCESS":
                        raise Exception(f"API returned non-SUCCESS status: {data.get('status')}")

                    payload = data["payload"]

                    # Cache result and update timestamp
                    _option_chain_cache[cache_key] = (payload, time.time())
                    _last_api_call_time = time.time()
                    return payload

                except requests.exceptions.RequestException as e:
                    last_exception = e
                    # For transient network errors, small sleep then retry
                    backoff = 1 + attempt * 1.5
                    print(f"⚠️ Option-chain fetch error ({e}), retrying in {backoff}s...")
                    time.sleep(backoff)

            # If we exhausted retries, raise the last exception
            print(f"❌ Exhausted option-chain retries: {last_exception}")
            raise last_exception

# --- END: Caching layer ---

# ----------------- Prefetcher (background) -----------------
def _option_chain_prefetcher_loop():
    """Daemon loop to keep option-chain cache warm and avoid bursts on demand."""
    cfg = CONFIG
    underlying = cfg.get("index", "NIFTY")
    expiry = cfg.get("expiry")
    interval = float(cfg.get("PREFETCH_INTERVAL", 8))
    jitter = float(cfg.get("PREFETCH_JITTER", 1.5))

    while True:
        try:
            # call cached helper (it will handle serialization/backoff)
            _get_full_option_chain_cached(underlying, expiry, access_token)
        except Exception as e:
            print(f"⚠️ Prefetcher: error refreshing option-chain: {e}")
        # sleep with jitter
        sleep_for = max(1.0, interval + random.uniform(-jitter, jitter))
        time.sleep(sleep_for)

def start_option_chain_prefetcher():
    t = threading.Thread(target=_option_chain_prefetcher_loop, daemon=True, name="OptionChainPrefetcher")
    t.start()

# Start background prefetcher to warm option-chain cache
try:
    start_option_chain_prefetcher()
    print("🔁 Option-chain prefetcher started.")
except Exception as e:
    print(f"⚠️ Failed to start prefetcher: {e}")

# ----------------- End prefetcher -----------------

#NEWCHANGE
def get_option_data_from_trading_symbol(
    trading_symbol: str,
    exchange: str = "NSE",
    underlying: str = "NIFTY"
):
    """
    Fetch delta, theta, OI, LTP, IV, volume etc. for a given trading_symbol
    using a cached Groww Option Chain API call to prevent rate limiting.
    """
    expiry_date = CONFIG["expiry"].strip()

    try:
        # Use the new cached helper function to get the entire option chain payload
        payload = _get_full_option_chain_cached(underlying, expiry_date, access_token)
        if not payload:
            raise ValueError("Failed to get option chain payload from cache or API.")
    except Exception as e:
        # If fetching fails, we can't proceed. Re-raise the exception.
        # The concurrent.futures framework will catch this and report it.
        raise e

    strikes = payload.get("strikes", {})
    underlying_ltp = payload.get("underlying_ltp")

    # 🔍 Find this trading_symbol in the fetched option chain
    for strike_str, opt_data in strikes.items():
        for opt_type in ("CE", "PE"):
            opt = opt_data.get(opt_type)
            if not opt:
                continue

            if opt.get("trading_symbol") == trading_symbol:
                greeks = opt.get("greeks", {})

                return {
                    "trading_symbol": trading_symbol,
                    "option_type": opt_type,
                    "strike": int(strike_str),
                    "expiry": expiry_date,
                    "ltp": opt.get("ltp"),
                    "open_interest": opt.get("open_interest"),
                    "volume": opt.get("volume"),
                    "delta": greeks.get("delta"),
                    "theta": greeks.get("theta"),
                    "iv": greeks.get("iv"),
                    "gamma": greeks.get("gamma"),
                    "vega": greeks.get("vega"),
                    "rho": greeks.get("rho"),
                    "underlying_ltp": underlying_ltp
                }

    raise ValueError(f"{trading_symbol} not found in option chain")




# ----------------- Place orders with Groww -----------------
class RiskManager:
    """Simple in-memory risk manager for per-trade and daily limits."""
    def __init__(self, cfg):
        self.cfg = cfg
        self.max_exposure_pct = float(cfg.get("MAX_EXPOSURE_PCT", 0.5)) / 100.0
        self.max_daily_loss_pct = float(cfg.get("MAX_DAILY_LOSS_PCT", 2.0)) / 100.0
        self.max_consec_losses = int(cfg.get("MAX_CONSECUTIVE_LOSSES", 3))
        self.cooldown_seconds = int(cfg.get("COOLDOWN_SECONDS", 1800))

        self.consecutive_losses = 0
        self.daily_loss = 0.0
        self.last_loss_time = 0

    def _get_equity(self):
        try:
            margins = getattr(groww, "get_margins", lambda: {"availablecash": 250000})()
            return float(margins.get("availablecash", 250000))
        except Exception:
            return 250000.0

    def can_open_trade(self, estimated_cost):
        equity = self._get_equity()
        if estimated_cost > equity * self.max_exposure_pct:
            print(f"❌ RiskManager: estimated cost {estimated_cost} exceeds max exposure {equity * self.max_exposure_pct}")
            return False

        # daily loss check
        if self.daily_loss <= -abs(equity * self.max_daily_loss_pct):
            print(f"❌ RiskManager: daily loss {self.daily_loss} exceeded max allowed {equity * self.max_daily_loss_pct}")
            return False

        # consecutive loss cooldown
        if self.consecutive_losses >= self.max_consec_losses:
            if time.time() - self.last_loss_time < self.cooldown_seconds:
                print("❌ RiskManager: in cooldown due to consecutive losses")
                return False
            else:
                # cooldown expired
                self.consecutive_losses = 0
        return True

    def record_trade(self, pnl):
        # pnl positive means profit
        self.daily_loss -= pnl if pnl < 0 else 0
        if pnl < 0:
            self.consecutive_losses += 1
            self.last_loss_time = time.time()
        else:
            self.consecutive_losses = 0


risk_manager = RiskManager(CONFIG)

def place_market_order_groww(instrument, quantity, transaction_type="BUY", product="MIS"):
    """
    place market order via growwapi wrapper. Returns order response or raises.
    """
    trading_symbol = instrument.get("internal_trading_symbol") or instrument.get("trading_symbol")
    try:
        order = groww.place_order(
            trading_symbol=trading_symbol,
            quantity=quantity,
            validity=groww.VALIDITY_DAY,
            exchange=groww.EXCHANGE_NSE,
            segment=groww.SEGMENT_FNO,
            product=getattr(groww, f"PRODUCT_{product}") if hasattr(groww, f"PRODUCT_{product}") else getattr(groww, "PRODUCT_MIS", product),
            order_type=groww.ORDER_TYPE_MARKET,
            transaction_type=getattr(groww, f"TRANSACTION_TYPE_{transaction_type}"),
            price=0
        )
        return order
    except Exception as e:
        raise

def place_limit_order_groww(instrument, quantity, price, transaction_type="SELL", product="MIS"):
    trading_symbol = instrument.get("internal_trading_symbol") or instrument.get("trading_symbol")
    try:
        order = groww.place_order(
            trading_symbol=trading_symbol,
            quantity=quantity,
            validity=groww.VALIDITY_DAY,
            exchange=groww.EXCHANGE_NSE,
            segment=groww.SEGMENT_FNO,
            product=getattr(groww, f"PRODUCT_{product}") if hasattr(groww, f"PRODUCT_{product}") else getattr(groww, "PRODUCT_MIS", product),
            order_type=groww.ORDER_TYPE_LIMIT,
            transaction_type=getattr(groww, f"TRANSACTION_TYPE_{transaction_type}"),
            price=price
        )
        return order
    except Exception as e:
        raise

# ----------------- Rounding for limits (5 paise) -----------------
def round_to_nearest_5_paise(price):
    # Round to nearest 0.05
    return round(round(price * 20) / 20, 2)

# ----------------- Momentum sampling -----------------
import numpy as np
import time

def momentum_check_for_symbol(instrument, MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY, threshold=0.25):
    """
    Improved short-term momentum check for Nifty options.
    - Uses multiple intermediate samples
    - Smooths noise
    - Checks direction consistency
    - Returns a cleaner momentum signal
    """
    trading_symbol = instrument.get("trading_symbol")
    prices = []

    print(f"\n🧭 Checking momentum for {trading_symbol} ({MOMENTUM_SAMPLES} samples, every {MOMENTUM_DELAY}s):")
    #NEWCHANGE
    opt = get_option_data_from_trading_symbol(trading_symbol)
    print(
        f"delta = {opt['delta']}, "
        f"theta = {opt['theta']}, "
        f"iv = {opt['iv']}, "
        f"gamma = {opt['gamma']}, "
        f"vega = {opt['vega']}, "
        f"rho = {opt['rho']}, "
        f"open_interest = {opt['open_interest']}, "
        f"ltp = {opt['ltp']}, "
        f"volume = {opt['volume']}"
    )

    for i in range(MOMENTUM_SAMPLES):
        p = get_ltp_for_instrument(instrument, access_token, verbose=False)
        if p:
            price = float(p)
            prices.append(price)
            print(f"[{trading_symbol}] ⏱ Sample {i+1}/{MOMENTUM_SAMPLES}: LTP = ₹{price:.2f}")
        else:
            print(f"[{trading_symbol}] ⚠️ Sample {i+1}/{MOMENTUM_SAMPLES}: Failed to fetch LTP")
        time.sleep(MOMENTUM_DELAY)

    if len(prices) < 3:
        print(f"[{trading_symbol}] ❌ Not enough data ({len(prices)} samples)")
        return None, len(prices)

    prices = np.array(prices)
    print(f"[{trading_symbol}] 🔍 Raw prices: {prices}")

    # 1️⃣ Smooth noise with small moving average
    smooth = np.convolve(prices, np.ones(3)/3, mode="valid")
    print(f"[{trading_symbol}] 🔍 Smoothed: {smooth}")

    # 2️⃣ Compute rate of change (%)
    roc = np.diff(smooth) / smooth[:-1] * 100
    print(f"[{trading_symbol}] 🔍 Rate of Change (%): {roc}")

    # 3️⃣ Remove outliers (big spikes)
    median = np.median(roc)
    std = np.std(roc)
    filtered = roc[(roc > median - 1.5*std) & (roc < median + 1.5*std)]
    print(f"[{trading_symbol}] 🔍 Filtered ROC (after outlier removal): {filtered} (length={len(filtered)})")

    if len(filtered) < 2:
        print(f"[{trading_symbol}] ⚠️ Too noisy for reliable momentum reading")
        print(f"[{trading_symbol}] 💡 TIP: All prices were identical ({prices[0]:.2f}) - no movement detected")
        return None, len(prices)

    # 4️⃣ Average change and direction consistency
    avg_change = np.mean(filtered)
    direction_signs = np.sign(filtered)
    consistency = np.mean(direction_signs == np.sign(avg_change)) * 100

    # 5️⃣ Decision
    if avg_change > threshold and consistency > 70:
        direction = "UP"
    elif avg_change < -threshold and consistency > 70:
        direction = "DOWN"
    else:
        direction = "FLAT"

    print(f"[{trading_symbol}] 📊 Avg Δ = {avg_change:.3f}%, Consistency = {consistency:.1f}% → {direction}")
    print(f"[{trading_symbol}] 📈 Range ₹{prices[0]:.2f} → ₹{prices[-1]:.2f}\n")

    return {"symbol": trading_symbol,
            "avg_change": round(avg_change, 3),
            "consistency": round(consistency, 1),
            "direction": direction}, len(prices)


# ----------------- Find option by premium (parallel) -----------------

def find_option_by_premium_parallel(option_type, min_premium, max_premium,
                                    lots=1, funds_buffer=0.9, momentum_threshold_pct=0.25,
                                    MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY):
    """
    Filters instruments using INDEX and EXPIRY from config,
    matches by option_type, filters by premium range,
    and runs momentum checks in parallel.
    Returns: (instrument, ltp, lot_size) or (None, None, None)
    """
    config = CONFIG
    INDEX = config["index"].upper()
    EXPIRY = config["expiry"].strip()
    candidates = []

    # 🔍 Filter based on index + expiry + type
    for item in instruments_data:
        try:
            if item.get("underlying_symbol", "").upper() != INDEX:
                continue
            if item.get("instrument_type", "").upper() != option_type.upper():
                continue
            if item.get("expiry_date", "").strip() != EXPIRY:
                continue

            lot_size = int(item.get("lot_size") or item.get("lotsize") or 1)
            ltp = get_ltp_for_instrument(item, access_token, verbose=False)
            if ltp is None:
                continue
            if not (min_premium <= ltp <= max_premium):
                continue

            candidates.append({
                "instrument": item,
                "ltp": float(ltp),
                "lot_size": lot_size
            })

        except Exception as e:
            print(f"⚠️ Error while scanning: {e}")
            continue

    if not candidates:
        print(f"⚠️ No instruments for {INDEX} {EXPIRY} {option_type}")
        return None, None, None

    # ✅ Funds check (fallback = 1.2L if not available)
    try:
        margins = getattr(groww, "get_margins", lambda: {"availablecash": 250000})()
        available_cash = float(margins.get("availablecash", 250000))
    except Exception:
        available_cash = 250000

    affordable = []
    for c in candidates:
        qty = lots * c["lot_size"]
        est_cost = c["ltp"] * qty
        if available_cash <= 0 or est_cost <= available_cash * funds_buffer:
            affordable.append(c)

    if not affordable:
        print(f"⚠️ No affordable instruments for {INDEX} {EXPIRY} {option_type}")
        return None, None, None

    if option_type.upper() == "PE":
        momentum_threshold_pct = 0.30  # PEs move sharper
    else:
        momentum_threshold_pct = 0.25

    # ✅ Sort candidates closest to mid-premium
    mid = (min_premium + max_premium) / 2.0
    affordable.sort(key=lambda x: abs(x["ltp"] - mid))
    # Limit number of simultaneous candidates to reduce API pressure
    probe_list = affordable[:10]

    momentum_results = []

    def check_momentum(cand):
        mom_result, ticks = momentum_check_for_symbol(
            cand["instrument"],
            MOMENTUM_SAMPLES=MOMENTUM_SAMPLES,
            MOMENTUM_DELAY=MOMENTUM_DELAY
        )
        if mom_result and isinstance(mom_result, dict):
            slope_pct = mom_result.get("avg_change", 0)
            direction = mom_result.get("direction", "FLAT")
            consistency = mom_result.get("consistency", 0)

            # ✅ Apply momentum filter right here
            if abs(slope_pct) >= momentum_threshold_pct and consistency >= 70 and direction != "FLAT":
                return {
                    "instrument": cand["instrument"],
                    "ltp": cand["ltp"],
                    "lot_size": cand["lot_size"],
                    "slope_pct": slope_pct,
                    "direction": direction,
                    "consistency": consistency,
                    "ticks": ticks
                }
        return None

    print(f"⚙️ Checking momentum for top {len(probe_list)} {option_type} candidates...")

    # Reduce per-search parallelism (still allow CE vs PE to run in parallel)
    with ThreadPoolExecutor(max_workers=min(len(probe_list), 3)) as executor:
        futures = {executor.submit(check_momentum, c): c for c in probe_list}
        for future in as_completed(futures):
            res = future.result()
            if res:
                momentum_results.append(res)

    if not momentum_results:
        print(f"⚠️ No strong momentum found for {option_type} (>{momentum_threshold_pct}%, consistency >70%)")
        # fallback: pick the one closest to mid-premium
        pick = probe_list[0]
        return pick["instrument"], pick["ltp"], pick["lot_size"]

    # ✅ Rank: strongest slope first, then consistency
    momentum_results.sort(key=lambda x: (x["slope_pct"], x["consistency"]), reverse=True)
    pick = momentum_results[0]

    print(f"🏆 Selected {option_type}: {pick['instrument']['trading_symbol']} "
          f"({pick['direction']} | {pick['slope_pct']:.2f}% | Consistency {pick['consistency']}%)")

    return pick["instrument"], pick["ltp"], pick["lot_size"]


#NEWCHANGE
# ----------------- Detect CE/PE (parallel) -----------------
def detect_option_type_parallel(index, expiry, min_p, max_p, lots, funds_buffer=0.9):
    print(f"🔍 Detecting best option between CE and PE for {index} {expiry}…")

    def worker(opt_type):
        print(f"➡️  Searching {opt_type} between {min_p}-{max_p}")
        inst, ltp, lot_size = find_option_by_premium_parallel(
            opt_type, min_p, max_p, lots, funds_buffer
        )
        mom = None
        if inst:
            print(f"📊 Running momentum check for {opt_type} ({inst.get('trading_symbol')})")
            mom, _ = momentum_check_for_symbol(
                inst,
                MOMENTUM_SAMPLES=MOMENTUM_SAMPLES,
                MOMENTUM_DELAY=MOMENTUM_DELAY
            )
            print(f"✅ Momentum for {opt_type}: {mom}")
        else:
            print(f"⚠️ No instrument found for {opt_type}")
        return opt_type, inst, ltp, lot_size, mom

    # ---------------- PARALLEL EXECUTION ----------------
    results = {}
    with ThreadPoolExecutor(max_workers=2) as ex:
        futures = {ex.submit(worker, t): t for t in ["CE", "PE"]}
        for future in as_completed(futures):
            opt_type, inst, ltp, lot_size, mom = future.result()
            results[opt_type] = {
                "instrument": inst,
                "ltp": ltp,
                "lot_size": lot_size,
                "momentum": mom
            }
            print(f"🧩 Finished {opt_type}: {inst.get('trading_symbol') if inst else 'None'}, momentum={mom}")

    print("🧮 Comparing CE vs PE momentum...")

    ce = results.get("CE")
    pe = results.get("PE")
    ce_mom = ce.get("momentum") if ce else None
    pe_mom = pe.get("momentum") if pe else None

    # ---------------- SAFETY CHECKS ----------------
    if not ce_mom and not pe_mom:
        print("❌ No momentum data found for CE or PE.")
        return None

    if ce_mom and not pe_mom:
        print("⚠️ Only CE has momentum — selecting CE")
        return {
            "selected": {**ce, "type": "CE"},
            "rejected": {**pe, "type": "PE"}
        }

    if pe_mom and not ce_mom:
        print("⚠️ Only PE has momentum — selecting PE")
        return {
            "selected": {**pe, "type": "PE"},
            "rejected": {**ce, "type": "CE"}
        }

    # ---------------- MOMENTUM COMPARISON ----------------
    ce_val = ce_mom["avg_change"]
    pe_val = pe_mom["avg_change"]

    print(f"📈 CE momentum: {ce_val:.3f}% ({ce_mom['direction']}, {ce_mom['consistency']}%)")
    print(f"📉 PE momentum: {pe_val:.3f}% ({pe_mom['direction']}, {pe_mom['consistency']}%)")

    # ---------------- DECISION LOGIC ----------------
    if abs(ce_val - pe_val) >= 0.25 and ce_val > pe_val and ce_val >= 0.10:
        print("✅ Selected CE (stronger momentum)")
        selected_type, rejected_type = "CE", "PE"

    elif abs(pe_val - ce_val) >= 0.25 and pe_val > ce_val and pe_val >= 0.10:
        print("✅ Selected PE (stronger momentum)")
        selected_type, rejected_type = "PE", "CE"

    else:
        if ce_val >= pe_val:
            print("⚖️ Momentum similar — choosing CE fallback")
            selected_type, rejected_type = "CE", "PE"
        else:
            print("⚖️ Momentum similar — choosing PE fallback")
            selected_type, rejected_type = "PE", "CE"

    # ---------------- FINAL RETURN ----------------
    return {
        "selected": {
            "type": selected_type,
            "instrument": results[selected_type]["instrument"],
            "ltp": results[selected_type]["ltp"],
            "lot_size": results[selected_type]["lot_size"],
            "momentum": results[selected_type]["momentum"],
        },
        "rejected": {
            "type": rejected_type,
            "instrument": results[rejected_type]["instrument"],
            "ltp": results[rejected_type]["ltp"],
            "lot_size": results[rejected_type]["lot_size"],
            "momentum": results[rejected_type]["momentum"],
        }
    }


def wait_for_order_status(order_id, access_token, order_type="BUY"):
    """
    Wait indefinitely until a Groww order reaches EXECUTED / COMPLETED / DELIVERY_AWAITED.
    Returns final status (string).
    """
    print(f"🔎 Waiting for {order_type} order ({order_id}) to finish...")

    while True:
        status = get_order_status(order_id, access_token)
        print(f"🕒 {order_type} status: {status}")

        if status in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
            print(f"✅ {order_type} order executed successfully.")
            send_telegram(f"✅ {order_type} order executed successfully.")
            return status

        elif status in ["FAILED", "REJECTED", "CANCELLED"]:
            print(f"❌ {order_type} order failed with status {status}.")
            send_telegram(f"❌ {order_type} order failed ({status}).")
            return status

        # wait before next check (adjust if needed)
        time.sleep(2)


import requests

def get_order_executed_price(order_id, access_token, segment="FNO"):
    """
    Fetch executed trades for a given Groww order_id and return average price & total quantity.
    """
    try:
        url = f"https://api.groww.in/v1/order/trades/{order_id}?segment={segment}&page=0&page_size=50"
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0"
        }

        print(f"\n📦 Fetching trade details for order: {order_id}")
        # Use session for faster connection reuse
        response = session.get(url, headers=headers)
        data = response.json()

        if data.get("status") != "SUCCESS":
            print("⚠️ Failed to fetch trade info:", data)
            return None, None

        trades = data.get("payload", {}).get("trade_list", [])
        if not trades:
            print("⚠️ No trades found for order ID.")
            return None, None

        # Compute average price & total quantity
        total_qty = sum(t["quantity"] for t in trades)
        total_value = sum(t["price"] * t["quantity"] for t in trades)
        avg_price = round(total_value / total_qty, 2)

        symbol = trades[0]["trading_symbol"]
        side = trades[0]["transaction_type"]

        print(f"✅ {side} {symbol} | Total Qty={total_qty} | Avg Price=₹{avg_price}")
        return avg_price, total_qty

    except Exception as e:
        print("❌ Error fetching order trades:", e)
        return None, None



# ----------------- Place CP order workflow (mirrors AngelOne logic) -----------------
def place_cp_order(command, is_auto=False):
    global buy_status
    if is_auto:
        order = command  # dict form
        symbol = order["symbol"]
        qty = order["lots"] * order["lot_size"]
        book_profit = order["book_profit"]
        volume = order.get("volume")
        oi = order.get("oi")

        # get instrument info directly from master
        instrument = next((inst for inst in instruments_data if inst["internal_trading_symbol"] == symbol), None)
        if not instrument:
            print(f"❌ Instrument {symbol} not found in master.")
            return

        print(f"🔹 Auto order => {symbol}, qty={qty}, book@{book_profit} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")

        # --- Fetch LTP ---
        ltp_before = get_ltp_for_instrument(instrument, access_token)
        if ltp_before is None:
            print("❌ Could not fetch LTP before placing order.")
            return

        entry_price = round(float(ltp_before), 2)


        # === BUY @ MARKET ===
        # Risk check: estimated exposure
        est_cost = entry_price * qty
        if not risk_manager.can_open_trade(est_cost):
            send_telegram(f"❌ Trade blocked by RiskManager. est_cost={est_cost}")
            print("❌ Trade blocked by RiskManager. Skipping order.")
            return

        try:
            order_resp = place_market_order_groww(instrument, qty, transaction_type="BUY", product="MIS")
            order_id = order_resp.get("payload", {}).get("groww_order_id") or order_resp.get("groww_order_id")
            print(f"✅ Auto Buy placed: :{order_resp} ======= [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            send_telegram(f"✅ Auto Buy placed: :{order_resp} ======= [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
        except Exception as e:
            print(f"❌ Auto BUY failed: {e}")
            send_telegram(f"❌ Auto BUY failed: {e}")
            return

        target_price = round_to_nearest_5_paise(entry_price + book_profit / qty)

        # STATUS VALIDATION
        # --- Wait until BUY order is EXECUTED or COMPLETED ---
        # if order_id:
        #     buy_status = wait_for_order_status(order_id, access_token, "BUY")
        #     if buy_status not in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
        #         print(f"⚠️ Skipping SELL due to BUY status: {buy_status}")
        #         return
        #
        # avg_price, qty = get_order_executed_price(order_id, access_token)
        # print(f"🎯 Executed avg price: ₹{avg_price}, Qty: {qty}")
        # send_telegram(f"🎯 BUY EXECUTED @ ₹{avg_price} | Qty={qty}")
        # ================= TRAILING LOGIC =================
        # highest_price = avg_price


        highest_price = entry_price

        start_time = time.time()

        trail_start = CONFIG["TRAIL_START_PROFIT"]
        trail_step = CONFIG["TRAIL_STEP"]
        poll = CONFIG["POLL_INTERVAL"]
        max_time = CONFIG["MAX_TRAIL_TIME"]
        hard_sl = entry_price - CONFIG.get("HARD_SL_POINTS")

        # 📊 ATR-based dynamic stops (research-based enhancement)
        atr = None
        dynamic_sl = hard_sl
        dynamic_tp = target_price
        if CONFIG.get("USE_VOLATILITY_SIZING", True):
            try:
                # Get candles for ATR calculation
                candles = get_historical_candles(instrument, access_token, interval="5minute", days=2)
                if candles:
                    atr = calculate_atr_from_candles(candles)
                    if atr:
                        dynamic_sl, dynamic_tp = calculate_dynamic_stops(entry_price, atr)
                        print(f"📊 ATR-based stops: ATR={atr}, SL={dynamic_sl:.2f}, TP={dynamic_tp:.2f}")
                        send_telegram(f"📊 ATR={atr} | SL={dynamic_sl:.2f} | TP={dynamic_tp:.2f}")
            except Exception as e:
                print(f"⚠️ Could not calculate ATR: {e}")
                atr = None

        print("📈 Trailing started...")
        send_telegram("📈 Trailing started")

        # Volume + OI tracking for risk filter
        prev_ltp = entry_price
        prev_volume = volume if volume else 0

        while True:
            # ENHANCEMENT: Zero delay in get_ltp to process immediately
            start_poll = time.time()
            ltp = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
            
            if ltp is None:
                time.sleep(0.2)
                continue

            ltp = float(ltp)
            
            # 🔍 Fetch current volume + OI for risk filter
            curr_volume = volume
            curr_oi = oi
            if CONFIG.get("enable_volume_oi_filter", True):
                try:
                    opt_data = get_option_data_from_trading_symbol(symbol)
                    curr_volume = opt_data.get("volume", volume)
                    curr_oi = opt_data.get("open_interest", oi)
                except Exception as e:
                    print(f"⚠️ Could not fetch live volume/OI: {e}")
            
            # 💰 Calculate current profit
            profit = (ltp - entry_price) * qty
            
            # 🛡️ Volume + OI Risk Filter: Check exit conditions
            if CONFIG.get("enable_volume_oi_filter", True):
                should_exit, exit_reason = should_exit_trade(
                    profit, curr_volume, prev_volume, curr_oi, ltp, prev_ltp
                )
                if should_exit:
                    print(exit_reason)
                    send_telegram(exit_reason)
                    
                    # Place exit order
                    for attempt in range(3):
                        try:
                            place_market_order_groww(instrument, qty, "SELL", "MIS")
                            print(f"✅ Risk Filter Exit Order placed (Attempt {attempt+1})")
                            break
                        except Exception as e:
                            print(f"❌ Exit Order failed (Attempt {attempt+1}): {e}")
                            time.sleep(0.1)
                    
                    if profit >= CONFIG.get("SMALL_PROFIT_TARGET", 500):
                        play_sound_async(SOUND_PROFIT)
                    
                    log_trade_to_excel(
                        instrument.get("internal_trading_symbol"),
                        entry_price, ltp, qty, profit, curr_volume, curr_oi
                    )
                    try:
                        risk_manager.record_trade(profit)
                    except Exception:
                        pass
                    break
            
            # Update tracking variables
            prev_ltp = ltp
            prev_volume = curr_volume
            
            # 📊 Update dynamic stops based on profit (research: tighten after 10% gain)
            if atr and CONFIG.get("USE_VOLATILITY_SIZING", True):
                profit_pct = ((ltp - entry_price) / entry_price)
                if profit_pct >= CONFIG.get("PROFIT_TIGHTEN_THRESHOLD", 0.10):
                    # Tighten stops
                    new_sl, new_tp = calculate_dynamic_stops(entry_price, atr, profit_pct)
                    if new_sl > dynamic_sl:  # Only move SL up, never down
                        dynamic_sl = new_sl
                        print(f"🔒 Stop tightened to {dynamic_sl:.2f} (profit {profit_pct*100:.1f}%)")

            # 🔴 HARD STOP LOSS (use dynamic SL if available)
            active_sl = dynamic_sl if atr else hard_sl
            if ltp <= active_sl:
                print(f"🛑 {'DYNAMIC' if atr else 'HARD'} SL HIT @ {ltp} (SL={active_sl:.2f})")
                send_telegram(f"🛑 SL HIT @ {ltp}")
                
                # Retry logic for SL order
                order_placed = False
                for attempt in range(3):
                    try:
                        place_market_order_groww(instrument, qty, "SELL", "MIS")
                        order_placed = True
                        print(f"✅ SL Order placed (Attempt {attempt+1})")
                        break
                    except Exception as e:
                        print(f"❌ SL Order failed (Attempt {attempt+1}): {e}")
                        time.sleep(0.1)
                
                if not order_placed:
                     send_telegram("🚨 CRITICAL: SL Order FAILED 3 times! Check manually!")
                     play_sound_async(SOUND_SL)
                else:
                     play_sound_async(SOUND_SL)
                
                profit = (ltp - entry_price) * qty
                log_trade_to_excel(
                    instrument.get("internal_trading_symbol"),
                    entry_price, ltp, qty, profit, volume , oi
                )
                try:
                    risk_manager.record_trade(profit)
                except Exception:
                    pass
                break

            # 🔼 Update highest price
            if ltp > highest_price:
                highest_price = ltp
                print(f"🔼 New High: ₹{highest_price}")
                send_telegram(f"🔼 New High: ₹{highest_price}")
            # 🟢 Start trailing after ₹1 profit
            if highest_price >= entry_price + trail_start:
                trail_exit = round_to_nearest_5_paise(highest_price - trail_step)
                print(f"📉 Trail Active | LTP={ltp} | Exit={trail_exit}")
                send_telegram(f"📉 Trail Active | LTP={ltp} | Exit={trail_exit}")

                #NEWCHANGE
                # print("Waiting for 8 sec to have momentum")
                # send_telegram("Waiting for 8 sec to have momentum")
                # time.sleep(8)
                #NEWCHANGE
                # Check again immediately
                ltp = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
                if ltp and float(ltp) <= trail_exit:
                    print(f"🔻 Trailing HIT @ {ltp}")
                    send_telegram(f"🔻 Trailing HIT @ {ltp}")
                    
                    # Retry logic for Target order
                    for attempt in range(3):
                        try:
                            place_market_order_groww(instrument, qty, "SELL", "MIS")
                            print(f"✅ Target Order placed (Attempt {attempt+1})")
                            break
                        except Exception as e:
                            print(f"❌ Target Order failed (Attempt {attempt+1}): {e}")
                            time.sleep(0.1)
                    
                    play_sound_async(SOUND_PROFIT)

                    profit = (float(ltp) - entry_price) * qty
                    log_trade_to_excel(
                        instrument.get("internal_trading_symbol"),
                        entry_price, float(ltp), qty, profit, volume , oi
                    )
                    try:
                        risk_manager.record_trade(profit)
                    except Exception:
                        pass
                    break

            # ⏰ SAFETY TIME EXIT
            if time.time() - start_time >= max_time:
                print("⏰ Max trail time reached — exiting")
                send_telegram("⏰ Max trail time reached — exiting")
                
                # Retry logic for Time Exit
                for attempt in range(3):
                    try:
                        place_market_order_groww(instrument, qty, "SELL", "MIS")
                        print(f"✅ Time Exit Order placed (Attempt {attempt+1})")
                        break
                    except Exception as e:
                        print(f"❌ Time Exit Order failed (Attempt {attempt+1}): {e}")
                        time.sleep(0.1)
                        
                play_sound_async(SOUND_PROFIT)

                ltp_now = get_ltp_for_instrument(instrument, access_token, verbose=False) or entry_price
                profit = (ltp_now - entry_price) * qty
                log_trade_to_excel(
                    instrument.get("internal_trading_symbol"),
                    entry_price, ltp_now, qty, profit , volume , oi
                )
                try:
                    risk_manager.record_trade(profit)
                except Exception:
                    pass
                break

            # Manual sleep to respect poll interval
            elapsed = time.time() - start_poll
            if elapsed < poll:
                time.sleep(poll - elapsed)

        print("Waiting for 1 min to get another data.")
        time.sleep(60)
        return  # ✅ end of auto mode execution

    else:
        parsed = parse_cp_command(command)
        order = command  # dict form
        symbol = order["symbol"]
        qty = order["lots"] * order["lot_size"]
        book_profit = order["book_profit"]
        volume = order.get("volume")
        oi = order.get("oi")
        if not parsed:
            print("❌ Could not parse command.")
            return

        instrument = find_instrument_from_command(command, instruments_data)
        if not instrument:
            print("❌ Instrument not found in instrument master.")
            return

        lot_size = int(instrument.get("lot_size") or instrument.get("lotsize") or 1)
        quantity = parsed["lots"] * lot_size

        ltp_before = get_ltp_for_instrument(instrument, access_token)
        if ltp_before is None:
            print("❌ Could not fetch LTP before placing order.")
            return

        entry_price = round(float(ltp_before), 2)
        send_telegram(f"entry price: {entry_price} | {instrument.get('internal_trading_symbol')} | qty={quantity}")
        print(f"entry price: {entry_price}")

        # Place BUY @ MARKET
        try:
            order_resp = place_market_order_groww(instrument, quantity, transaction_type="BUY", product="MIS")
            order_id = order_resp.get("payload", {}).get("groww_order_id") or order_resp.get("groww_order_id")
            print("✅ Buy Order placed:", order_resp)
        except Exception as e:
            print(f"❌ Buy order failed: {e}")
            send_telegram(f"❌ Buy order failed: {e}")
            return

        # STATUS VALIDATION
        # --- Wait until BUY order is EXECUTED or COMPLETED ---
        # if order_id:
        #     buy_status = wait_for_order_status(order_id, access_token, "BUY")
        #     if buy_status not in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
        #         print(f"⚠️ Skipping SELL due to BUY status: {buy_status}")
        #         return
        #
        # avg_price, quantity = get_order_executed_price(order_id, access_token)
        # if avg_price is None:
        #     print("❌ Could not get executed price. Aborting.")
        #     return
        # print(f"🎯 Executed avg price: ₹{avg_price}, Qty: {quantity}")
        # send_telegram(f"🎯 BUY EXECUTED @ ₹{avg_price} | Qty={quantity}")

        # ================= TRAILING LOGIC =================
        highest_price = entry_price+1
        start_time = time.time()

        trail_start = CONFIG["TRAIL_START_PROFIT"]
        trail_step = CONFIG["TRAIL_STEP"]
        poll = CONFIG["POLL_INTERVAL"]
        max_time = CONFIG["MAX_TRAIL_TIME"]
        hard_sl = entry_price - CONFIG.get("HARD_SL_POINTS")

        print("📈 Trailing started...")
        send_telegram("📈 Trailing started")

        while True:
            # ENHANCEMENT: Zero delay in get_ltp to process immediately
            start_poll = time.time()
            ltp = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
            
            if ltp is None:
                time.sleep(0.2)
                continue

            ltp = float(ltp)

            # 🔴 HARD STOP LOSS
            if ltp <= hard_sl:
                print(f"🛑 HARD SL HIT @ {ltp}")
                send_telegram(f"🛑 HARD SL HIT @ {ltp}")
                
                # Retry logic for SL order
                order_placed = False
                for attempt in range(3):
                    try:
                        place_market_order_groww(instrument, quantity, "SELL", "MIS")
                        order_placed = True
                        print(f"✅ SL Order placed (Attempt {attempt+1})")
                        break
                    except Exception as e:
                        print(f"❌ SL Order failed (Attempt {attempt+1}): {e}")
                        time.sleep(0.1)
                
                if not order_placed:
                     send_telegram("🚨 CRITICAL: SL Order FAILED 3 times! Check manually!")
                     play_sound_async(SOUND_SL)
                else:
                     play_sound_async(SOUND_SL)
                
                profit = (ltp - entry_price) * quantity
                log_trade_to_excel(
                    instrument.get("internal_trading_symbol"),
                    entry_price, ltp, quantity, profit, volume , oi
                )
                break

            # 🔼 Update highest price
            if ltp > highest_price:
                highest_price = ltp
                print(f"🔼 New High: ₹{highest_price}")

            # 🟢 Start trailing after ₹1 profit
            if highest_price >= entry_price + trail_start:
                trail_exit = round_to_nearest_5_paise(highest_price - trail_step)
                print(f"📉 Trail Active | LTP={ltp} | Exit={trail_exit}")

                if ltp <= trail_exit:
                    print(f"🔻 Trailing HIT @ {ltp}")
                    send_telegram(f"🔻 Trailing HIT @ {ltp}")
                    
                    # Retry logic for Target order
                    for attempt in range(3):
                        try:
                            place_market_order_groww(instrument, quantity, "SELL", "MIS")
                            print(f"✅ Target Order placed (Attempt {attempt+1})")
                            break
                        except Exception as e:
                            print(f"❌ Target Order failed (Attempt {attempt+1}): {e}")
                            time.sleep(0.1)
                    
                    print(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
                    send_telegram(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
                    play_sound_async(SOUND_PROFIT)

                    profit = (ltp - entry_price) * quantity
                    log_trade_to_excel(
                        instrument.get("internal_trading_symbol"),
                        entry_price, ltp, quantity, profit , volume, oi
                    )
                    break

            # ⏰ SAFETY TIME EXIT
            if time.time() - start_time >= max_time:
                print("⏰ Max trail time reached — exiting")
                send_telegram("⏰ Max trail time reached — exiting")
                
                # Retry logic for Time Exit
                for attempt in range(3):
                    try:
                        place_market_order_groww(instrument, quantity, "SELL", "MIS")
                        print(f"✅ Time Exit Order placed (Attempt {attempt+1})")
                        break
                    except Exception as e:
                        print(f"❌ Time Exit Order failed (Attempt {attempt+1}): {e}")
                        time.sleep(0.1)
                        
                play_sound_async(SOUND_PROFIT)

                ltp_now = get_ltp_for_instrument(instrument, access_token, verbose=False) or entry_price
                profit = (ltp_now - entry_price) * quantity
                log_trade_to_excel(
                    instrument.get("internal_trading_symbol"),
                    entry_price, ltp_now, quantity, profit , volume , oi
                )
                break

            # Manual sleep to respect poll interval
            elapsed = time.time() - start_poll
            if elapsed < poll:
                time.sleep(poll - elapsed)

        print("Waiting for 1 min to get another data.")
        time.sleep(60)
        return

# ----------------- Auto mode runner (momentum + premium) -----------------

# Global flag for graceful shutdown
shutdown_requested = False

def signal_handler(signum, frame):
    """Handle Ctrl+C gracefully"""
    global shutdown_requested
    shutdown_requested = True
    print("\n🛑 Shutdown requested. Completing current iteration...")
    send_telegram("🛑 Bot shutdown requested by user")

#NEWCHANGE
def auto_mode_runner():
    cfg = CONFIG
    print("\n--- AUTO MODE (momentum + premium) ---")
    send_telegram("\n--- AUTO MODE (momentum + premium) ---")

    index = cfg["index"]
    expiry = cfg["expiry"]
    min_p = cfg["min_premium"]
    max_p = cfg["max_premium"]
    lots = cfg["lots"]
    book_profit = cfg["book_profit"]
    
    # Error tracking
    consecutive_failures = 0
    MAX_CONSECUTIVE_FAILURES = 10
    last_heartbeat = time.time()
    HEARTBEAT_INTERVAL = 300  # 5 minutes

    while not shutdown_requested:
        # Heartbeat log every 5 minutes
        if time.time() - last_heartbeat > HEARTBEAT_INTERVAL:
            print(f"💚 Bot alive - {datetime.now().strftime('%H:%M:%S')}")
            send_telegram(f"💚 Heartbeat: Bot running | Failures: {consecutive_failures}/{MAX_CONSECUTIVE_FAILURES}")
            last_heartbeat = time.time()
        
        opt = detect_option_type_parallel(index, expiry, min_p, max_p, lots)
        print(f"opt {opt}")

        if not opt:
            consecutive_failures += 1
            print(f"❌ Could not determine CE/PE. Retrying in 60s... (Failure {consecutive_failures}/{MAX_CONSECUTIVE_FAILURES})")
            send_telegram(f"❌ Could not determine CE/PE. Retrying in 60s... (Failure {consecutive_failures}/{MAX_CONSECUTIVE_FAILURES})")
            
            if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                print(f"🛑 CRITICAL: {MAX_CONSECUTIVE_FAILURES} consecutive failures. Stopping bot for safety.")
                send_telegram(f"🛑 CRITICAL: Bot stopped after {MAX_CONSECUTIVE_FAILURES} consecutive failures. Manual intervention required!")
                break
            
            time.sleep(60)
            continue
        
        # Reset failure counter on success
        consecutive_failures = 0

        selected = opt["selected"]
        rejected = opt["rejected"]

        print(f"\n<UNK> Selected options: {selected['instrument']} ")
        print(f"\n<UNK> Rejected options: {rejected['instrument']} ")

        sel_inst = selected["instrument"]
        rej_inst = rejected["instrument"]

        if not sel_inst or not rej_inst:
            consecutive_failures += 1
            print(f"❌ Invalid instrument data. Retrying... (Failure {consecutive_failures}/{MAX_CONSECUTIVE_FAILURES})")
            time.sleep(30)
            continue

        instrument_type = selected["type"]
        symbol = sel_inst.get("internal_trading_symbol")
        groww_symbol = sel_inst.get("groww_symbol")
        lot_size = selected["lot_size"]
        ltp = selected["ltp"]

        print(
            f"✅ Selected: {symbol} ({instrument_type}) | "
            f"LTP={ltp} | lot_size={lot_size}"
        )
        print(
            f"🚫 Rejected: {rej_inst.get('internal_trading_symbol')} "
            f"({rejected['type']})"
        )

        send_telegram(
            f"✅ Selected: {symbol} ({instrument_type}) | LTP={ltp}\n"
            f"🚫 Rejected: {rej_inst.get('internal_trading_symbol')} ({rejected['type']})"
        )

        # ---------- Order payloads ----------
        selected_order = {
            "symbol": symbol,
            "ltp": selected["ltp"],
            "lots": lots,
            "book_profit": float(book_profit),
            "lot_size": selected["lot_size"],
            "side": "BUY",
        }

        rejected_order = {
            "symbol": rej_inst.get("internal_trading_symbol"),
            "ltp": rejected["ltp"],
            "lots": lots,
            "book_profit": float(book_profit),
            "lot_size": rejected["lot_size"],
            "side": "BUY",
        }

        # ---------- Market direction check ----------
        market_direction = get_recent_market_direction(groww_symbol, groww)
        print(f"📊 Market Direction: {market_direction}")
        send_telegram(f"📊 Market Direction: {market_direction}")

        if market_direction != instrument_type:
            print("❌ Skipping — market direction conflicts with momentum.")
            send_telegram("❌ Skipping — market direction conflicts with momentum.")
            time.sleep(10)
            continue

        print(f"✅ Market confirms {instrument_type}")

        # Check Technicals (EMA/SMA/RSI/ADX/VWAP)
        if cfg.get("enable_technical_filters", True):
            print(f"📊 Checking Technicals for {symbol}...")
            techs = get_technicals(groww_symbol, groww)
            if techs:
                ema_9 = techs["ema_9"]
                sma_20 = techs["sma_20"]
                curr_ltp = techs["ltp"]
                rsi = techs["rsi"]
                adx = techs["adx"]
                vwap = techs["vwap"]
                
                ema_50 = techs.get("ema_50")  # Get EMA_50 for momentum score
                
                print(f"   LTP: {curr_ltp}, EMA(9): {ema_9:.2f}, RSI: {rsi:.2f}, ADX: {adx:.2f}, VWAP: {vwap:.2f} , SMA(20): {sma_20:.2f}")

                log_technical_to_excel(
                    symbol=symbol,
                    ltp=curr_ltp,
                    ema_9=ema_9,
                    sma_20=sma_20,
                    rsi=rsi,
                    adx=adx,
                    vwap=vwap
                )
                
                # 📊 Store technical indicators in selected_order for momentum score calculation
                selected_order["ema_50"] = ema_50
                selected_order["rsi"] = rsi
                selected_order["adx"] = adx
                
                # 1. EMA Check (Warning only, not blocking)
                if ema_9 and curr_ltp < ema_9:
                    print(f"⚠️ Technical Warning: Price {curr_ltp} is below EMA 9 {ema_9:.2f}. Proceeding with caution.")
                    # Don't skip - let momentum score decide
                    
                # 2. ADX Check (Relaxed threshold: 15 instead of 20)
                if adx < 15:
                    print(f"❌ Technical Filter: ADX {adx:.2f} is too low (Very Choppy Market). Skipping.")
                    send_telegram(f"❌ Technical Filter: ADX too low ({adx:.2f}). Skipping.")
                    time.sleep(5)
                    continue
                    
                # 3. RSI Check (Relaxed range: 40-80 instead of 45-75)
                if rsi > 80:
                    print(f"❌ Technical Filter: RSI {rsi:.2f} is Extremely Overbought (>80). Skipping.")
                    send_telegram(f"❌ Technical Filter: RSI Extremely Overbought ({rsi:.2f}). Skipping.")
                    time.sleep(5)
                    continue
                
                if rsi < 40:
                    print(f"⚠️ Technical Warning: RSI {rsi:.2f} is weak (<40). Proceeding with caution.")
                    # Don't skip - allow entry if momentum score is strong
                    
                # 4. VWAP Check (Warning only, not blocking)
                if vwap and curr_ltp < vwap:
                    print(f"⚠️ Technical Warning: Price {curr_ltp} is below VWAP {vwap:.2f}. Proceeding with caution.")
                    # Don't skip - let momentum score decide

            else:
                print("⚠️ Could not calculate technicals. Proceeding with caution.")
                # Set defaults if technicals fail
                selected_order["ema_50"] = None
                selected_order["rsi"] = None
                selected_order["adx"] = None
        else:
            print("⏩ Technical filters disabled. Skipping technical analysis.")
            # Set defaults if filters disabled
            selected_order["ema_50"] = None
            selected_order["rsi"] = None
            selected_order["adx"] = None

        opt = get_option_data_from_trading_symbol(selected_order["symbol"])
        print(f"Checking delta/theta/OI for selected option === {selected_order['symbol']}")

        iv = opt.get("iv", 0)
        delta = opt.get("delta", 0)
        oi = opt.get("open_interest", 0)
        volume = opt.get("volume", 0)
        gamma = opt.get("gamma", 0)

        # add volume here 👇
        selected_order["volume"] = volume
        selected_order["delta"] = delta
        selected_order["iv"] = iv
        selected_order["oi"] = oi
        selected_order["gamma"] = gamma
        
        # 🛡️ Volume + OI Risk Filter: Pre-entry validation with momentum scoring
        if cfg.get("enable_volume_oi_filter", True):
            print("\n" + "="*70)
            print("🔍 PRE-ENTRY VALIDATION: Volume + OI + Momentum Score")
            print("="*70)
            
            # Get previous volume data (use a conservative estimate if not available)
            prev_volume_estimate = volume * 0.7  # Assume previous was ~70% of current
            
            # Check momentum - price moving up means positive momentum for calls, down for puts
            price_momentum_positive = (instrument_type == "CE")  # CE = bullish, PE = bearish
            
            # Calculate momentum score if technical data available
            momentum_score = None
            print(f"\n📊 MOMENTUM SCORE CALCULATION:")
            print(f"   Current Price (LTP): {selected_order.get('ltp', 0)}")
            print(f"   EMA_50: {selected_order.get('ema_50')}")
            print(f"   RSI: {selected_order.get('rsi')}")
            print(f"   ADX: {selected_order.get('adx')}")
            
            if hasattr(selected_order, 'get'):
                current_price = selected_order.get("ltp", 0)
                ma_50 = selected_order.get("ema_50", None)
                rsi = selected_order.get("rsi", None)
                adx = selected_order.get("adx", None)
                
                if current_price > 0 and ma_50 is not None and rsi is not None:
                    momentum_score = calculate_momentum_score(current_price, ma_50, rsi, adx)
                    print(f"\n   ✅ FORMULA: ({current_price}/{ma_50:.2f}) + ({rsi}/50) - 1")
                    print(f"   ✅ RESULT: {momentum_score:.3f}")
                    print(f"   Threshold: {cfg.get('MIN_MOMENTUM_SCORE', 0.15)}")
                    if momentum_score >= cfg.get('MIN_MOMENTUM_SCORE', 0.15):
                        print(f"   ✅ PASS: Score {momentum_score:.3f} >= {cfg.get('MIN_MOMENTUM_SCORE', 0.15)}")
                    else:
                        print(f"   ❌ FAIL: Score {momentum_score:.3f} < {cfg.get('MIN_MOMENTUM_SCORE', 0.15)}")
                else:
                    print(f"   ⚠️ SKIPPED: Missing data (price={current_price}, ma_50={ma_50}, rsi={rsi})")
            
            print(f"\n📊 VOLUME/OI VALIDATION:")
            print(f"   Current Volume: {volume:,}")
            print(f"   Previous Volume (est): {prev_volume_estimate:,}")
            print(f"   Open Interest: {oi:,}")
            print(f"   Price Momentum: {'Bullish (CE)' if price_momentum_positive else 'Bearish (PE)'}")
            
            should_enter, entry_reason = should_enter_trade(
                volume, prev_volume_estimate, oi, price_momentum_positive, momentum_score
            )
            print(f"\n{'✅ VALIDATION RESULT: PASS' if should_enter else '❌ VALIDATION RESULT: BLOCKED'}")
            print("="*70 + "\n")
            
            if not should_enter:
                print(f"🚫 Entry Blocked by Volume+OI Filter:")
                print(entry_reason)
                send_telegram(f"🚫 Entry Blocked:\n{entry_reason}")
                time.sleep(10)
                continue
            else:
                print(f"✅ Entry Validation Passed:")
                print(entry_reason)

        # ✅ Check option conditions if enabled
        if cfg.get("enable_option_filters", True):
            # Enhanced delta check for momentum options (relaxed threshold)
            min_delta = cfg.get("MIN_DELTA_ENTRY", 0.35)  # Relaxed from 0.45
            
            # Check if IV is too high (research: avoid extreme premium environments)
            max_iv_pct = cfg.get("MAX_IV_PERCENTILE", 0.75)  # Relaxed from 0.70
            iv_too_high = (iv > 15) or (iv > max_iv_pct * 20)  # Relaxed from 12
            
            # Relaxed thresholds for real momentum trading
            if iv_too_high or abs(delta) < min_delta or oi < 20000 or volume < 25000 or volume < 0.3 * oi or gamma < 0.0015:
                print(
                    f"delta = {opt.get('delta')}, volume = {opt.get('volume')}, iv = {opt.get('iv')}, "
                    f"gamma = {opt.get('gamma')}, vega = {opt.get('vega')}, rho = {opt.get('rho')}, "
                    f"open_interest = {opt.get('open_interest')}, ltp = {opt.get('ltp')}"
                )
                print("❌ Option conditions not satisfied, skipping...")
                failed_reasons = []

                if iv > 15:  # Relaxed from 12
                    failed_reasons.append(f"IV too high ({iv:.2f})")

                if abs(delta) < min_delta:  # Now uses min_delta (0.35)
                    failed_reasons.append(f"Delta too low ({delta:.3f})")

                if oi < 20000:  # Relaxed from 25000
                    failed_reasons.append(f"OI too low ({oi})")

                if volume < 25000:  # Relaxed from 35000
                    failed_reasons.append(f"Volume too low ({volume})")

                if volume < 0.3 * oi:  # Relaxed from 0.4
                    failed_reasons.append(f"Volume/OI weak ({volume}/{oi})")

                if gamma < 0.0015:  # Relaxed from 0.0018
                    failed_reasons.append(f"Gamma too low ({gamma:.5f})")

                if failed_reasons:
                    print(
                        f"❌ Skipping {opt['trading_symbol']} | "
                        f"LTP={opt['ltp']} | "
                        f"IV={iv:.2f}, Δ={delta:.3f}, Γ={gamma:.5f}, OI={oi}, Vol={volume}"
                    )
                    print("   Reasons:")
                    for r in failed_reasons:
                        print(f"   • {r}")
                    continue
        else:
            print("⏩ Option filters disabled. Skipping Greeks validation.")

        print(
            f"delta = {opt.get('delta')}, volume = {opt.get('volume')}, iv = {opt.get('iv')}, "
            f"gamma = {opt.get('gamma')}, vega = {opt.get('vega')}, rho = {opt.get('rho')}, "
            f"open_interest = {opt.get('open_interest')}, ltp = {opt.get('ltp')}"
        )

        user_confirmation_needed = cfg.get("user_confirmation_needed", False)
        print(f"user_confirmation_needed : {user_confirmation_needed}")

        if user_confirmation_needed:
            play_sound_async(SOUND_user_input)

            user_input = input(
                f"Confirm trade? "
                f"[Y = {selected['type']}, "
                f"O = {rejected['type']}, "
                f"anything else = Skip]: "
            ).strip().lower()

            if user_input.lower() in ("y", "yes"):

                print(f"➡️ Placing SELECTED order ({selected['type']})")
                send_telegram(f"➡️ Placing SELECTED order ({selected['type']})")
                place_cp_order(selected_order, is_auto=True)
                break  # exit loop AFTER placing order


            elif user_input == "o":
                opt = get_option_data_from_trading_symbol(rejected_order["symbol"])
                print(f"Checking delta/theta/OI for rejected option === {rejected_order['symbol']}")

                iv = opt.get("iv", 0)
                delta = opt.get("delta", 0)
                oi = opt.get("open_interest", 0)
                volume = opt.get("volume", 0)
                gamma = opt.get("gamma", 0)

                selected_order["volume"] = volume
                selected_order["delta"] = delta
                selected_order["iv"] = iv
                selected_order["oi"] = oi
                selected_order["gamma"] = gamma

                # ✅ Correct logical conditions
                if iv > 8 or abs(delta) < 0.45 or oi < 25000 or volume < 35000 or volume < 0.4 * oi or gamma < 0.0018:
                    print(
                        f"delta = {opt.get('delta')}, volume = {opt.get('volume')}, iv = {opt.get('iv')}, "
                        f"gamma = {opt.get('gamma')}, vega = {opt.get('vega')}, rho = {opt.get('rho')}, "
                        f"open_interest = {opt.get('open_interest')}, ltp = {opt.get('ltp')}"
                    )
                    print("❌ Option conditions not satisfied, skipping...")
                    failed_reasons = []

                    if iv > 12:
                        failed_reasons.append(f"IV too high ({iv:.2f})")

                    if abs(delta) < 0.45:
                        failed_reasons.append(f"Delta too low ({delta:.3f})")

                    if oi < 25000:
                        failed_reasons.append(f"OI too low ({oi})")

                    if volume < 35000:
                        failed_reasons.append(f"Volume too low ({volume})")

                    if volume < 0.4 * oi:
                        failed_reasons.append(f"Volume/OI weak ({volume}/{oi})")

                    if gamma < 0.0018:
                        failed_reasons.append(f"Gamma too low ({gamma:.5f})")

                    if failed_reasons:
                        print(
                            f"❌ Skipping {opt['trading_symbol']} | "
                            f"LTP={opt['ltp']} | "
                            f"IV={iv:.2f}, Δ={delta:.3f}, Γ={gamma:.5f}, OI={oi}, Vol={volume}"
                        )
                        print("   Reasons:")
                        for r in failed_reasons:
                            print(f"   • {r}")
                        continue

                else:
                    print(
                        f"delta = {opt.get('delta')}, volume = {opt.get('volume')}, iv = {opt.get('iv')}, "
                        f"gamma = {opt.get('gamma')}, vega = {opt.get('vega')}, rho = {opt.get('rho')}, "
                        f"open_interest = {opt.get('open_interest')}, ltp = {opt.get('ltp')}"
                    )
                    print(f"🔁 Placing OPPOSITE order ({rejected['type']})")
                    send_telegram(f"🔁 Placing OPPOSITE order ({rejected['type']})")
                    place_cp_order(rejected_order, is_auto=True)
                    break  # exit loop AFTER placing order

            else:
                print("❌ Trade skipped by user")
                send_telegram("❌ Trade skipped by user")
                time.sleep(30)
                continue

        else:
            opt = get_option_data_from_trading_symbol(selected_order["symbol"])
            print(f"Checking delta/theta/OI for selected option === {selected_order['symbol']}")

            iv = opt.get("iv", 0)
            delta = opt.get("delta", 0)
            oi = opt.get("open_interest", 0)
            volume = opt.get("volume", 0)
            gamma = opt.get("gamma", 0)

            selected_order["volume"] = volume
            selected_order["delta"] = delta
            selected_order["iv"] = iv
            selected_order["oi"] = oi
            selected_order["gamma"] = gamma

            print(f"➡️ Placing SELECTED order ({selected['type']})")
            send_telegram(f"➡️ Placing SELECTED order ({selected['type']})")
            place_cp_order(selected_order, is_auto=True)
            # exit loop AFTER placing order

        # Small cooldown before next cycle
        time.sleep(2)

        # ---------- User confirmation ----------






# ----------------- Main menu -----------------
if __name__ == "__main__":
    print("\n✨ Groww NIFTY CP Bot Ready (Groww backend)")
    print("You can run in MANUAL or AUTO mode.")
    print("Manual example: Buy 14 NIFTY04NOV2525950CE at CP and Book at 1050\n")
    while True:
        mode = input("Choose mode: (m)anual / (a)uto / (q)uit: ").strip().lower()
        if mode in ["q", "quit", "exit"]:
            print("Exiting.")
            break
        if mode in ["a", "auto"]:
            auto_mode_runner()
            continue
        if mode in ["m", "manual"]:
            user_input = input("\nEnter command (or press Enter for status, type 'back' to menu): ").strip()
            if user_input.lower() in ["back"]:
                continue
            if user_input == "":
                print("Status check not implemented for Groww PnL in this script.")
                continue
            place_cp_order(user_input)
            continue
        print("Invalid input. Choose 'm' or 'a' or 'q'.")