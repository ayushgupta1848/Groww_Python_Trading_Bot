# Groww_CP_Bot.py


############### MUST READ #####################
#• To start Bot we just need to assure that config data is proper,
# Need to remove comments from "place_cp_order" method to validate the order status under comment starts from "#STATUS VALIDATION",
# And in last we need to validate that funds are matching with BOT and Groww Account unnder comment "Funds check"

#DOWNLOAD GROWW INSTRUMENTS:- https://growwapi-assets.groww.in/instruments/instrument.csv

import os
import re
import json
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import pyotp
from openpyxl import Workbook, load_workbook
from playsound3 import playsound
from datetime import datetime
from datetime import time as dtime
from datetime import timedelta
from zoneinfo import ZoneInfo
from threading import Lock
import requests
import sys
from datetime import datetime
import time
import os
import sys
from datetime import datetime
session = requests.Session()
MOMENTUM_SAMPLES = 5
MOMENTUM_DELAY = 1

def setup_persistent_logger():
    """Creates a local 'logs/qa_pass_bot' folder beside the script and logs all console output there."""
    # Create /logs/qa_pass_bot folder in the same directory as the script
    base_dir = os.path.dirname(os.path.abspath(__file__))
    log_dir = os.path.join(base_dir, "logs", "qa_pass_bot")
    os.makedirs(log_dir, exist_ok=True)

    # Create a timestamped log file
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_path = os.path.join(log_dir, f"Groww_Bot_{timestamp}.log")

    # Define a Tee class to write to both console and log file.
    # It stamps every line with a [HH:MM:SS.mmm] prefix automatically, so all
    # existing print()s get a timestamp without touching each call site.
    class Tee:
        def __init__(self, *streams):
            self.streams = streams
            self._at_line_start = True

        def _stamp(self, data):
            if not data:
                return data
            out = []
            for ch in data:
                if self._at_line_start and ch not in ("\n", "\r"):
                    out.append(datetime.now().strftime("[%H:%M:%S.%f")[:-3] + "] ")
                    self._at_line_start = False
                out.append(ch)
                if ch == "\n":
                    self._at_line_start = True
            return "".join(out)

        def write(self, data):
            data = self._stamp(data)
            for s in self.streams:
                try:
                    s.write(data)
                    s.flush()
                except Exception:
                    pass  # Ignore on shutdown

        def flush(self):
            for s in self.streams:
                try:
                    s.flush()
                except Exception:
                    pass

    # Open log file (unbuffered, UTF-8)
    logfile = open(log_path, "a", buffering=1, encoding="utf-8")

    # Redirect both stdout and stderr
    sys.stdout = Tee(sys.stdout, logfile)
    sys.stderr = Tee(sys.stderr, logfile)

    print(f"📝 Logging started. Log file: {log_path}")

    return log_path


# --- Initialize persistent logging ---
LOG_FILE_PATH = setup_persistent_logger()

# Replace with your Groww API key (or leave and use TOTP to fetch access_token)
api_key = "eyJraWQiOiJaTUtjVXciLCJhbGciOiJFUzI1NiJ9.eyJleHAiOjI1NjQ2NTczODEsImlhdCI6MTc3NjI1NzM4MSwibmJmIjoxNzc2MjU3MzgxLCJzdWIiOiJ7XCJ0b2tlblJlZklkXCI6XCJjMjAzMmM5MS04ZGYzLTRkZDUtYjc5NS0yMGVlOWRhZDhhZjlcIixcInZlbmRvckludGVncmF0aW9uS2V5XCI6XCJlMzFmZjIzYjA4NmI0MDZjODg3NGIyZjZkODQ5NTMxM1wiLFwidXNlckFjY291bnRJZFwiOlwiMmVlMjYyMjItN2MwNS00Y2IwLWIwM2MtNzAzYWRmNWVmN2RkXCIsXCJkZXZpY2VJZFwiOlwiNjA2MzE5M2QtZWZkMC01OWViLTgzYzQtNWQ2NGZkNzdkNzQ3XCIsXCJzZXNzaW9uSWRcIjpcIjI0OWQ2OGRlLTNjZTgtNGQ4OS05ODJkLWM0N2NmYmI1YzdlNFwiLFwiYWRkaXRpb25hbERhdGFcIjpcIno1NC9NZzltdjE2WXdmb0gvS0EwYktvMDZXRlpjc241VUNmTWF5aERtNGxSTkczdTlLa2pWZDNoWjU1ZStNZERhWXBOVi9UOUxIRmtQejFFQisybTdRPT1cIixcInJvbGVcIjpcImF1dGgtdG90cFwiLFwic291cmNlSXBBZGRyZXNzXCI6XCIyNDA5OjQwYzQ6MTBhMzozN2UzOjE4NGI6N2IyOTpiMzBlOjIwZTUsMTcyLjcwLjIxOC4xMzUsMzUuMjQxLjIzLjEyM1wiLFwidHdvRmFFeHBpcnlUc1wiOjI1NjQ2NTczODE2ODYsXCJ2ZW5kb3JOYW1lXCI6XCJncm93d0FwaVwifSIsImlzcyI6ImFwZXgtYXV0aC1wcm9kLWFwcCJ9.3kotfZI_EC0lzszHKlXiRdqEQv-O8ubYFh0pgoAT0KsSfdQ1sHmts5UtlaAq4PB6DEwY4X2jZUCD8uBgc2nwXQ"
totp_gen = pyotp.TOTP('SC3YMFLEGLHBWUPHRBOYLPEEOVAT2PZ4')

# Get project root directory (folder where your script is running)
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
# Build CSV path dynamically
csv_path = os.path.join(PROJECT_ROOT, "instrument.csv")
print(csv_path)

# Instruments CSV/JSON path (script will convert CSV -> JSON if convert_csv_to_json = yes)
# csv_path = r"C:\Users\HITS\Downloads\instrument (6).csv"
convert_csv_to_json = "yes"

# Telegram placeholders (you will replace later)
TELEGRAM_BOT_TOKEN = "PUT_YOUR_TOKEN_HERE"
TELEGRAM_CHAT_ID = "PUT_YOUR_CHAT_ID_HERE"

# Sound files (ensure these exist in script folder or provide full path)
SOUND_PROFIT = "coin.mp3"
SOUND_SL = "SL_HIT.mp3"
SOUND_user_input = "User_input.WAV"

# Trade defaults for Groww
DEFAULT_PRODUCT = "MIS"   # intraday; change to "NRML" if you want positional
ORDER_PRODUCT_MAP = {
    "MIS": "MIS",
    "NRML": "NRML"
}
# NOTE: the growwapi wrapper constants are used from the growwapi instance below

# ----------------- import growwapi late (after auth) -----------------
try:
    from growwapi import GrowwAPI
except Exception:
    # If local module not available, user must install or place it in PYTHONPATH
    print("❗ growwapi module not found. Make sure it's installed and importable.")
    # continue; import errors will show when script runs further

from groww_token import get_access_token as get_cached_access_token

# ----------------- Groww auth & wrapper -----------------
def groww_init(api_key):
    """
    Return growwapi client instance (GrowwAPI(access_token))
    This function gets access_token using GrowwAPI.get_access_token if available.
    """
    try:
        access_token = get_cached_access_token(api_key, 'SC3YMFLEGLHBWUPHRBOYLPEEOVAT2PZ4')
        client = GrowwAPI(access_token)
        print(access_token)
        print("✅ Groww API Initialized Successfully")
        return client, access_token
    except Exception as e:
        print(f"❌ Groww login failed: {e}")
        raise

# Init groww client
groww ,access_token = groww_init(api_key)


# ----------------- Utilities: Telegram, Sound, Excel Logging -----------------


#test

# === TELEGRAM CONFIG ===
BOT_TOKEN = "8666941668:AAEObDodwWqDwdVJVXy8WvFx_lyreq8p7fI"
CHAT_ID = "6012308856"

def _send_telegram_sync(message: str):
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        payload = {"chat_id": CHAT_ID, "text": message}
        requests.post(url, data=payload, timeout=5)
    except Exception as e:
        print(f"⚠️ Telegram Error: {e}")

def send_telegram(message: str):
    # Fire-and-forget in a daemon thread so a slow/blocking Telegram POST never
    # adds latency to the order hot-path (previously a no-timeout blocking call).
    threading.Thread(target=_send_telegram_sync, args=(message,), daemon=True).start()

def _fmt_timedelta(td):
    total_seconds = max(0, int(td.total_seconds()))
    hours, rem = divmod(total_seconds, 3600)
    minutes = rem // 60
    return f"{hours}h {minutes}m"

def _next_market_open_ist(now_ist, market_open):
    d = now_ist
    for _ in range(8):  # search up to a week ahead, skipping weekends
        open_dt = datetime.combine(d.date(), market_open, tzinfo=ZoneInfo("Asia/Kolkata"))
        if d.weekday() < 5 and open_dt > now_ist:
            return open_dt
        d = d + timedelta(days=1)
    return None

MARKET_OPEN_TIME = dtime(9, 15)
MARKET_CLOSE_TIME = dtime(15, 30)

def get_market_status_line():
    """Returns a one-line IST market status string, e.g. '🔔 Market OPEN | closes in 5h 30m'."""
    now_ist = datetime.now(ZoneInfo("Asia/Kolkata"))
    is_market_hours = now_ist.weekday() < 5 and MARKET_OPEN_TIME <= now_ist.time() <= MARKET_CLOSE_TIME
    if is_market_hours:
        close_dt = datetime.combine(now_ist.date(), MARKET_CLOSE_TIME, tzinfo=ZoneInfo("Asia/Kolkata"))
        remaining = _fmt_timedelta(close_dt - now_ist)
        return f"🔔 Market OPEN | closes in {remaining}"
    else:
        next_open = _next_market_open_ist(now_ist, MARKET_OPEN_TIME)
        remaining = _fmt_timedelta(next_open - now_ist) if next_open else "unknown"
        return f"🔕 Market CLOSED | opens in {remaining}"

def alert_market_status():
    """Prints and Telegram-alerts the current market status line."""
    line = get_market_status_line()
    print(line)
    send_telegram(line)

def market_hours_watcher():
    """Continuously checks IST time and alerts when market open/closed status changes."""
    last_status = None
    while True:
        line = get_market_status_line()
        status = "OPEN" if line.startswith("🔔") else "CLOSED"
        if status != last_status:
            alert_market_status()
            last_status = status
        time.sleep(30)

def play_sound_async(filename):
    try:
        if not os.path.exists(filename):
            print(f"🔇 Sound file not found: {filename}")
            return
        threading.Thread(target=playsound, args=(filename,), daemon=True).start()
    except Exception as e:
        print(f"🔇 Sound error: {e}")

def log_trade_to_excel(symbol, buy_price, sell_price, quantity, profit):
    file_name = "Lakshmi.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append(["DateTime", "Symbol", "Buy Price", "Sell Price", "Quantity", "Profit"])
        wb.save(file_name)

    # Load existing workbook
    wb = load_workbook(file_name)
    ws = wb.active

    # Find the next empty row
    next_row = ws.max_row + 1
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.cell(row=next_row, column=1).value = now
    ws.cell(row=next_row, column=2).value = symbol
    ws.cell(row=next_row, column=3).value = buy_price
    ws.cell(row=next_row, column=4).value = sell_price
    ws.cell(row=next_row, column=5).value = quantity
    ws.cell(row=next_row, column=6).value = round(profit, 2)
    wb.save(file_name)


# ----------------- 3-min candle O1→C3 difference logger -----------------
def log_o1_c3_diff(symbol, first_open, last_close, direction):
    """
    Logs the difference between the first candle open (O1) and the third/last
    candle close (C3) used to compute the 3-min market direction, one row per
    trade evaluation. Appended to logs/o1_c3_diff_log.csv.
    """
    try:
        log_dir = os.path.join(PROJECT_ROOT, "logs")
        os.makedirs(log_dir, exist_ok=True)
        file_name = os.path.join(log_dir, "o1_c3_diff_log.csv")

        diff = None
        pct = None
        try:
            diff = round(float(last_close) - float(first_open), 2)
            if float(first_open) != 0:
                pct = round((diff / float(first_open)) * 100, 4)
        except (TypeError, ValueError):
            pass

        write_header = not os.path.exists(file_name)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(file_name, "a", newline="", encoding="utf-8") as f:
            import csv as _csv
            writer = _csv.writer(f)
            if write_header:
                writer.writerow(["DateTime", "Symbol", "O1_Open", "C3_Close", "Diff", "Diff_%", "Direction"])
            writer.writerow([now, symbol, first_open, last_close, diff, pct, direction])

        print(f"📝 Logged O1→C3 diff: O1={first_open}, C3={last_close}, Diff={diff} ({pct}%) → {file_name}")
    except Exception as e:
        print("⚠️ Error logging O1→C3 diff:", e)


# ----------------- CSV -> JSON loader -----------------
def csv_to_json(csv_file_path, json_file_path=None):
    import csv
    if json_file_path is None:
        json_file_path = os.path.splitext(csv_file_path)[0] + ".json"
    data = []
    with open(csv_file_path, encoding='utf-8') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            data.append(row)
    with open(json_file_path, 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file, indent=4, ensure_ascii=False)
    print(f"✅ Converted '{csv_file_path}' → '{json_file_path}'")
    return data

ltp_lock = threading.Lock()

def get_ltp_for_instrument(instrument, access_token, verbose=True,segment = "FNO", delay=0.1):
    """
    Fetches the latest traded price (LTP) for a given F&O instrument using Groww's authenticated API.
    Thread-safe with a global lock to prevent too-frequent API calls.
    """

    try:
        trading_symbol = instrument.get("trading_symbol")  # e.g. NIFTY25N0425950CE
        if not trading_symbol:
            print("⚠️ Missing trading_symbol in instrument.")
            return None

        exchange = (instrument.get("exchange") or "NSE").upper()  # BSE for SENSEX, NSE otherwise
        exchange_symbol = f"{exchange}_{trading_symbol}"
        url = f"https://api.groww.in/v1/live-data/ltp?segment={segment}&exchange_symbols={exchange_symbol}"
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0"
        }

        # 🔒 Lock ensures one API call at a time
        with ltp_lock:
            # Use session for faster connection reuse
            resp = session.get(url, headers=headers, timeout=10)
            if delay > 0:
                time.sleep(delay)  # ⏳ delay to respect Groww API rate limits

        if resp.status_code != 200:
            print(f"⚠️ HTTP {resp.status_code} error fetching LTP: {resp.text}")
            return None

        data = resp.json()
        payload = data.get("payload", {})
        ltp = payload.get(exchange_symbol)

        if ltp is None:
            print(f"⚠️ No LTP found for {exchange_symbol} in payload: {payload}")
            return None
        if verbose:
            print(f"💰 LTP for {exchange_symbol}: ₹{ltp} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            send_telegram(f"💰 LTP for {exchange_symbol}: ₹{ltp} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
        return float(ltp)

    except Exception as e:
        print(f"⚠️ Error fetching LTP for {instrument.get('trading_symbol')}: {e}")
        return None

def get_index_spot_price(index_name, access_token=None, json_path=None):
    """
    Fetches live spot price for any index (NIFTY, SENSEX, BANKNIFTY, FINNIFTY) using Groww instrument data.
    For SENSEX, falls back to the option-chain API since there's no plain spot instrument.
    """
    global instruments1
    index_name = index_name.upper()

    if json_path is None:
        json_path = os.path.splitext(csv_path)[0] + ".json"

    # 🔄 Step 1: Load or convert JSON
    if convert_csv_to_json.lower() == "yes":
        instruments1 = csv_to_json(csv_path, json_path)
    else:
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON not found: {json_path}")
        with open(json_path, "r", encoding="utf-8") as jf:
            instruments1 = json.load(jf)
        print(f"ℹ️ Loaded instruments from existing JSON: {json_path}")

    index_mappings = {
        "NIFTY": ["NIFTY", "NSE-NIFTY", "NIFTY 50"],
        "SENSEX": ["SENSEX", "BSE-SENSEX", "BSE_SENSEX"],
        "BANKNIFTY": ["BANKNIFTY", "NIFTY BANK", "NSE-BANKNIFTY"],
        "FINNIFTY": ["FINNIFTY", "NIFTY FIN SERVICE", "NSE-FINNIFTY"]
    }
    search_terms = index_mappings.get(index_name, [index_name])

    try:
        # SENSEX has no plain spot instrument — pull underlying LTP from the option chain instead
        if index_name == "SENSEX":
            print(f"📊 Fetching {index_name} spot from option chain...")
            try:
                sensex_options = [item for item in instruments1
                                  if item.get("underlying_symbol", "").upper() == "SENSEX"]
                if sensex_options:
                    expiry = sensex_options[0].get("expiry_date")
                    url = f"https://api.groww.in/v1/option-chain/exchange/BSE/underlying/SENSEX?expiry_date={expiry}"
                    headers = {
                        "Accept": "application/json",
                        "Authorization": f"Bearer {access_token}",
                        "X-API-VERSION": "1.0"
                    }
                    resp = session.get(url, headers=headers, timeout=8)
                    if resp.status_code == 200:
                        data = resp.json()
                        underlying_ltp = data.get("payload", {}).get("underlying_ltp")
                        if underlying_ltp:
                            print(f"📊 Live {index_name} Spot (from option chain): {underlying_ltp}")
                            return float(underlying_ltp)
            except Exception as e:
                print(f"⚠️ Could not fetch {index_name} from option chain: {e}")

        spot_instrument = next(
            (item for item in instruments1
             if item.get("trading_symbol", "").upper() in search_terms
             or item.get("groww_symbol", "").upper() in [f"NSE-{t}" for t in search_terms]
             or item.get("groww_symbol", "").upper() in [f"BSE-{t}" for t in search_terms]
             or item.get("name", "").upper() in search_terms),
            None
        )

        if not spot_instrument:
            print(f"⚠️ {index_name} spot instrument not found in instruments1")
            return 0

        spot = get_ltp_for_instrument(spot_instrument, access_token, verbose=False, segment="CASH")
        if spot:
            print(f"📊 Live {index_name} Spot: {spot}")
            return float(spot)
        else:
            print(f"⚠️ Failed to fetch LTP for {index_name} spot")
            return 0
    except Exception as e:
        print(f"⚠️ Error fetching {index_name} spot: {e}")
        return 0


# Backward-compat wrapper
def get_nifty_spot_price(access_token=None, json_path=None):
    return get_index_spot_price("NIFTY", access_token, json_path)


# ----------------- Interactive index / expiry / price-range selection -----------------
def get_available_expiries(index_name, json_path=None):
    """
    Returns sorted list of unique upcoming (today or later) expiry dates (YYYY-MM-DD)
    found in the instrument master for the given index.
    """
    global instruments1
    index_name = index_name.upper()

    if json_path is None:
        json_path = os.path.splitext(csv_path)[0] + ".json"

    if convert_csv_to_json.lower() == "yes":
        instruments1 = csv_to_json(csv_path, json_path)
    else:
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON not found: {json_path}")
        with open(json_path, "r", encoding="utf-8") as jf:
            instruments1 = json.load(jf)

    today = datetime.now().date()
    expiries = set()
    for item in instruments1:
        if item.get("underlying_symbol", "").upper() != index_name:
            continue
        ed = item.get("expiry_date", "").strip()
        if not ed:
            continue
        try:
            if datetime.strptime(ed, "%Y-%m-%d").date() >= today:
                expiries.add(ed)
        except ValueError:
            continue

    return sorted(expiries, key=lambda e: datetime.strptime(e, "%Y-%m-%d").date())


def prompt_index_selection():
    options = {"1": "NIFTY", "2": "BANKNIFTY", "3": "SENSEX", "4": "FINNIFTY"}
    while True:
        print("\n📌 Select Index:")
        for k, v in options.items():
            print(f"   ({k}) {v}")
        choice = input("Enter choice (1-4) or type index name: ").strip().upper()
        if choice in options:
            return options[choice]
        if choice in options.values():
            return choice
        print("❌ Invalid choice, try again.")


def prompt_expiry_selection(index_name):
    expiries = get_available_expiries(index_name)
    if not expiries:
        manual = input(
            f"⚠️ No expiries found for {index_name} in instrument master. Enter expiry date (YYYY-MM-DD): "
        ).strip()
        return manual

    current_expiry = expiries[0]
    next_expiry = expiries[1] if len(expiries) > 1 else expiries[0]
    while True:
        choice = input(
            f"\n📅 Choose expiry for {index_name} — (c)urrent [{current_expiry}] or (n)ext [{next_expiry}]: "
        ).strip().lower()
        if choice in ("c", "current", ""):
            return current_expiry
        if choice in ("n", "next"):
            return next_expiry
        print("❌ Invalid choice, enter 'c' or 'n'.")


def prompt_price_range():
    while True:
        try:
            min_p = float(input("\n💰 Enter MIN premium price: ").strip())
            max_p = float(input("💰 Enter MAX premium price: ").strip())
            if min_p < max_p:
                return min_p, max_p
            print("❌ MIN must be less than MAX.")
        except ValueError:
            print("❌ Enter valid numbers.")


def prompt_lots():
    while True:
        try:
            lots = int(input("\n📦 Enter number of lots: ").strip())
            if lots > 0:
                return lots
            print("❌ Lots must be a positive integer.")
        except ValueError:
            print("❌ Enter a valid integer.")


print("\n===== Bot Configuration =====")
SELECTED_INDEX = prompt_index_selection()
SELECTED_EXPIRY = prompt_expiry_selection(SELECTED_INDEX)
SELECTED_MIN_PREMIUM, SELECTED_MAX_PREMIUM = prompt_price_range()
SELECTED_LOTS = prompt_lots()


CONFIG = {
    "index": SELECTED_INDEX,
    "expiry": SELECTED_EXPIRY,  # chosen interactively (current/next) at startup
    "min_premium": SELECTED_MIN_PREMIUM,
    "max_premium": SELECTED_MAX_PREMIUM,
    "lots": SELECTED_LOTS,  # quantity = lots * instrument's lot_size (computed at order time)
    "book_profit": 1050,
    "target_pnl": 6000,
    "spot": get_index_spot_price(SELECTED_INDEX, access_token),
    "TRAIL_START_PROFIT": 1,  # Start trailing after this profit per unit (in points) #NEWCHANGE
    "TRAIL_STEP": .75,  # Trailing step (in points) #NEWCHANGE
    "POLL_INTERVAL": .1,  # Poll interval in seconds (fast LTP monitoring)
    "MAX_TRAIL_TIME": 3600,  # Max trailing time in seconds (1 hour)
    "HARD_SL_POINTS": 5.0,  # Hard stop loss points below entry
    "user_confirmation_needed": False,   # or False
    "ENABLE_MOMENTUM_CONSISTENCY_GATE": True,  # set False at runtime to fully restore old (ungated) behavior
    "MIN_MOMENTUM_CONSISTENCY": 70,  # % of samples that must agree on direction to allow a trade
    # === O1→C3 DIFF GATE (3-min candle trend) ===
    # False -> NO change, direction used exactly as computed now (no diff filter).
    # True  -> the |C3 - O1| move must be strictly > O1C3_MIN_DIFF and < O1C3_MAX_DIFF
    #          points for the trade to be eligible; otherwise direction is treated
    #          as uncertain (None) and the trade is skipped.
    "ENABLE_O1C3_DIFF_GATE": True,
    "O1C3_MIN_DIFF": 1.0,  # diff must be strictly greater than this (points)
    "O1C3_MAX_DIFF": 4.0,  # diff must be strictly less than this (points)
    # === INDEX MONITOR (auto mode only) ===
    # False -> NO change; the trade is managed only by target / hard-SL / level as
    #          before. Index is not watched after entry.
    # True  -> after entry, continuously poll the INDEX spot and confirm it is moving
    #          in the trade's favour (CE -> index should go UP, PE -> index should go
    #          DOWN). If the index instead moves ADVERSELY by more than
    #          INDEX_ADVERSE_POINTS from the entry spot for INDEX_ADVERSE_TICKS
    #          consecutive polls (after a short grace period), the resting target
    #          order is cancelled and the position is market-exited immediately.
    #          This is a directional invalidation stop — it exits the moment the
    #          underlying stops agreeing with the option we bought, instead of
    #          waiting for the premium to bleed all the way to the hard SL.
    "ENABLE_INDEX_MONITOR": True,
    "INDEX_ADVERSE_POINTS": 5.0,    # adverse index move (pts, from entry spot) that counts as "against us"
    "INDEX_ADVERSE_TICKS": 3,       # consecutive adverse polls required before exiting (noise filter)
    "INDEX_MONITOR_GRACE_SEC": 5.0, # ignore the first N seconds after entry (let the trade breathe)
    "INDEX_POLL_INTERVAL": 1.0,     # how often (sec) to poll the index spot inside the manage loop
    # HOLD-ON-FAVOUR: needs ENABLE_INDEX_MONITOR. When the index confirms OUR
    # direction (CE -> up, PE -> down) by INDEX_ADVERSE_POINTS for INDEX_ADVERSE_TICKS
    # consecutive polls, cancel the fixed scalp target and hand the trade to the
    # existing trailing stop so a confirmed trend can run instead of booking +1.5.
    # The adverse index-exit and hard SL stay active while it rides.
    "INDEX_HOLD_ON_FAVOUR": True,
    # === LEVEL FILTER (auto mode only) ===
    # False  -> NO change whatsoever, bot runs exactly as before.
    # True   -> after an option is selected, find the next key level in the
    #           trade's favour (nearest RESISTANCE above spot for CE, nearest
    #           SUPPORT below spot for PE). If the index touches that level,
    #           book profit instantly instead of waiting for the trailing stop.
    "LEVEL_FILTER": True,
    "LEVEL_MIN_TOUCHES": 2,   # a level must be tapped this many times to count
    "LEVEL_TOL_PCT": 0.15,    # cluster tolerance as % of price (~36 pts on NIFTY)
    "LEVEL_SPIKE_MULT": 8.0,  # drop phantom candles whose range > this x median
    "LEVEL_PIVOT_LR": 10,     # pivot lookback/forward bars for swing detection
    # === QUICK MODE (auto mode only) ===
    # False -> NO change, normal trailing behaviour.
    # True  -> after BUY, target = (actual avg buy price + QUICK_POINTS), rounded
    #          to nearest 5 paise (like PROD10). Books profit the moment LTP hits
    #          that target; trailing is disabled (hard SL + max-time still apply).
    "QUICK_MODE": True,
    "QUICK_POINTS": 1.5,
    # === TARGET ORDER TYPE (quick/RR mode) ===
    # True  -> place a resting LIMIT SELL at the target during BUY time (books even
    #          on a fast spike, exact target fill).
    # False -> DON'T place a target order; just watch LTP continuously and fire a
    #          MARKET exit the moment it reaches the target price.
    "SET_TARGET_ORDER": True,
    # === RISK:REWARD MODE (auto mode only) — risk-anchored ===
    # False -> ignored.
    # True  -> from the actual avg buy price:  SL = avg - RR_RISK_POINTS,
    #          target = avg + RR_RISK_POINTS * RR_RATIO  (both rounded to 5 paise).
    #          Enforces a fixed risk:reward. Trailing disabled; both SL and target
    #          are hard market exits. Takes precedence over QUICK_MODE if both on.
    "RR_MODE": False,
    "RR_RISK_POINTS": 3.0,  # your risk per unit (SL distance below avg buy, in points)
    "RR_RATIO": 3.0,        # reward multiple -> target = risk * ratio  (1:3)
    # === ORDER VALIDATION (auto mode) === (like PROD10)
    # True  -> after BUY, wait for it to actually EXECUTE before managing the trade;
    #          abort if rejected/failed (prevents phantom target/trailing on a
    #          position that never opened, e.g. insufficient balance).
    # False -> old behaviour: proceed immediately without confirming the fill.
    "VALIDATE_ORDERS": False,
    # === PAPER TRADING === (like PROD10) — simulate all orders, NOTHING sent to Groww.
    # True  -> place_market/limit return fake EXECUTED orders at live LTP; SL/target/
    #          trailing/level/quick/RR all run on real market data but book on paper.
    #          Safe to test every mode with zero balance / no risk.
    # False -> real orders sent to the exchange.
    "PAPER_TRADING": False,
}

# Load instruments_data
def load_instruments_from_json(json_path=None):
    """
    Loads instruments from JSON (or CSV → JSON if convert_csv_to_json = 'yes'),
    but only keeps instruments:
      - matching expiry from CONFIG
      - within ±10 strikes of current index spot price
    """
    global instruments
    config = CONFIG
    INDEX = config["index"].upper()
    EXPIRY = config["expiry"].strip()

    if json_path is None:
        json_path = os.path.splitext(csv_path)[0] + ".json"

    # 🔄 Step 1: Load or convert JSON
    if convert_csv_to_json.lower() == "yes":
        instruments = csv_to_json(csv_path, json_path)
    else:
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON not found: {json_path}")
        with open(json_path, "r", encoding="utf-8") as jf:
            instruments = json.load(jf)
        print(f"ℹ️ Loaded instruments from existing JSON: {json_path}")

    # 🧩 Step 2: Get live index spot
    spot = config["spot"]

    # Determine strike step (e.g., NIFTY = 50, BANKNIFTY/SENSEX = 100)
    step = 100 if ("BANK" in INDEX or "SENSEX" in INDEX) else 50

    # Define strike range (±10 strikes)
    nearest_strike = round(spot / step) * step
    lower_bound = nearest_strike - (10 * step)
    upper_bound = nearest_strike + (10 * step)

    print(f"🎯 Filtering {INDEX} {EXPIRY} instruments between {lower_bound}–{upper_bound} (Spot={spot})")

    # 🧹 Step 3: Filter instruments by expiry and strike range
    filtered = []
    for item in instruments:
        try:
            if item.get("underlying_symbol", "").upper() != INDEX:
                continue
            # if item.get("expiry_date", "").strip() != EXPIRY:
            #     continue
            strike = float(item.get("strike_price") or 0)
            if lower_bound <= strike <= upper_bound:
                filtered.append(item)
        except Exception:
            continue

    print(f"✅ Loaded {len(filtered)} filtered instruments (out of {len(instruments)})")
    instruments = filtered
    return instruments


# initialize
instruments_data = load_instruments_from_json()


# Full, UNFILTERED instrument master. `instruments_data` above is trimmed to
# ±10 strikes of spot (for auto-mode scanning), so manual commands — which may
# trade any strike (e.g. a far-OTM 25950) — must look up against this instead.
def load_all_instruments():
    try:
        json_path = os.path.splitext(csv_path)[0] + ".json"
        with open(json_path, "r", encoding="utf-8") as jf:
            return json.load(jf)
    except Exception as e:
        print(f"⚠️ Could not load full instrument master, falling back to filtered set: {e}")
        return instruments_data


all_instruments = load_all_instruments()

# O(1) manual lookup: (index, expiry, strike, CE/PE) -> instrument. Built once so
# a manual command resolves instantly instead of scanning ~145k rows each time.
manual_index = {}
for _it in all_instruments:
    manual_index[(
        (_it.get("underlying_symbol") or "").upper(),
        _it.get("expiry_date"),
        str(_it.get("strike_price")),
        (_it.get("instrument_type") or "").upper(),
    )] = _it


# Pre-index instruments for quick lookup by internal_trading_symbol or groww_symbol or custom compact symbol
symbol_index = {}
for it in instruments_data:
    # keys: internal_trading_symbol, groww_symbol, compact like NIFTY04NOV2525950CE approximate
    try:
        k1 = it.get("internal_trading_symbol", "") or it.get("trading_symbol", "")
        k2 = it.get("groww_symbol", "")
        if k1:
            symbol_index[k1.upper()] = it
        if k2:
            symbol_index[k2.upper()] = it
    except Exception:
        pass

# ----------------- Helpers: date/expiry normalization -----------------
MONTHS = {
    'JAN': '01','FEB':'02','MAR':'03','APR':'04','MAY':'05','JUN':'06',
    'JUL':'07','AUG':'08','SEP':'09','OCT':'10','NOV':'11','DEC':'12'
}

def cmd_expiry_to_date(expiry_token):
    """
    expiry_token example: 04NOV25 or 04NOV2025 or 28AUG25 or 28AUG2025
    Return string 'DD/MM/YYYY'
    """
    m = re.match(r'(\d{1,2})([A-Z]{3})(\d{2,4})', expiry_token.upper())
    if not m:
        return None
    dd = m.group(1).zfill(2)
    mon_abbr = m.group(2)
    yy = m.group(3)
    if len(yy) == 2:
        yyyy = "20" + yy
    else:
        yyyy = yy
    mm = MONTHS.get(mon_abbr[:3], None)
    if not mm:
        return None
    return f"{dd}/{mm}/{yyyy}"

# ----------------- Command parser -----------------
def parse_cp_command(command):
    """
    Parse strings like:
      Buy 14 NIFTY04NOV2525950CE at CP and Book at 1050
    Returns dict or None
    """
    pattern = r'(?i)^\s*(Buy|Sell)\s+(\d+)\s+([A-Z]+)(\d{1,2}[A-Z]{3}\d{2})(\d+)(CE|PE)\s+at\s+CP\s+and\s+Book\s+at\s+(\d+(\.\d+)?)\s*$'
    m = re.match(pattern, command.strip())
    if not m:
        return None
    action = m.group(1).upper()
    lots = int(m.group(2))
    underlying = m.group(3).upper()
    expiry_token = m.group(4).upper()
    strike = m.group(5)
    opt_type = m.group(6).upper()
    target_profit = float(m.group(7))
    expiry_date = cmd_expiry_to_date(expiry_token)
    return {
        "action": action,
        "lots": lots,
        "underlying": underlying,
        "expiry_token": expiry_token,
        "expiry_date": expiry_date,
        "strike": strike,
        "opt_type": opt_type,
        "target_profit": target_profit
    }

# ----------------- Find instrument in instruments_data -----------------
def find_instrument_from_command(command: str, instruments: list):
    import re
    # Example command: Buy 14 NIFTY04NOV2525950CE at CP and Book at 1050
    # Year is a fixed 2-digit token (Groww uses YY, e.g. 26). Keeping it as
    # \d{2,4} lets the greedy match steal leading strike digits, e.g.
    # NIFTY21JUL2625950CE -> yr=2625, strike=950 -> instrument never found.
    pattern = r'([A-Z]+)(\d{1,2})([A-Z]{3})(\d{2})(\d+)(CE|PE)'
    match = re.search(pattern, command.upper())
    if not match:
        print("❌ Could not parse symbol from command.")
        return None

    underlying, day, mon, yr, strike, opt_type = match.groups()
    expiry_date = f"20{yr}-{mon_to_number(mon)}-{day.zfill(2)}"

    # Find match in JSON
    for inst in instruments:
        if (
            inst["underlying_symbol"].upper() == underlying
            and inst["expiry_date"] == expiry_date
            and str(inst["strike_price"]) == strike
            and inst["instrument_type"].upper() == opt_type
        ):
            return inst

    print("❌ Instrument not found in instrument master.")
    return None


def mon_to_number(mon: str):
    mapping = {
        "JAN": "01", "FEB": "02", "MAR": "03", "APR": "04",
        "MAY": "05", "JUN": "06", "JUL": "07", "AUG": "08",
        "SEP": "09", "OCT": "10", "NOV": "11", "DEC": "12"
    }
    return mapping.get(mon.upper(), "00")


# ----------------- Compact "Book <profit>" command (uses startup index/expiry) -----------------
def parse_book_command(command: str):
    """
    Parse the compact manual command that reuses the index/expiry chosen at
    startup, e.g.:   14 25950CE Book 1050
      14    -> lots for THIS trade (overrides the startup lots)
      25950 -> strike
      CE/PE -> option type
      1050  -> TOTAL rupee profit to book. target = avg_buy + 1050/qty,
               rounded to the nearest 5 paise.
    Returns dict or None (None => not this format, fall back to the old parser).
    """
    m = re.match(r'(?i)^\s*(\d+)\s+(\d+)\s*(CE|PE)\s+Book\s+(\d+(?:\.\d+)?)\s*$', command.strip())
    if not m:
        return None
    return {
        "lots": int(m.group(1)),
        "strike": m.group(2),
        "opt_type": m.group(3).upper(),
        "book_profit": float(m.group(4)),
    }


def find_instrument_by_strike(strike: str, opt_type: str):
    """Resolve an instrument from the index & expiry selected at startup
    (CONFIG['index'] / CONFIG['expiry']) plus the strike + CE/PE from the command."""
    index = (CONFIG.get("index") or "").upper()
    expiry = CONFIG.get("expiry")   # 'YYYY-MM-DD'
    if not index or not expiry:
        print("❌ Index/expiry not selected — restart the bot and choose them first.")
        return None
    inst = manual_index.get((index, expiry, str(strike), opt_type))
    if inst:
        return inst
    print(f"❌ {index} {expiry} {strike}{opt_type} not found in instrument master.")
    return None


def place_book_target_order(parsed):
    """
    Execute the compact 'Book <profit>' manual command:
      BUY <lots> of the CONFIG index/expiry strike at market, then rest a LIMIT
      SELL at the price that yields <profit> TOTAL rupees across the whole
      position (target = avg_buy + profit/qty, rounded to 5 paise), guarded by
      the hard SL (avg_buy - HARD_SL_POINTS). No trailing.
    """
    global buy_status
    t0 = time.time()
    instrument = find_instrument_by_strike(parsed["strike"], parsed["opt_type"])
    if not instrument:
        return

    lot_size = int(instrument.get("lot_size") or instrument.get("lotsize") or 1)
    quantity = parsed["lots"] * lot_size
    symbol = instrument.get("internal_trading_symbol") or instrument.get("trading_symbol")

    # 🚀 FIRE THE BUY IMMEDIATELY. Nothing slow (LTP fetch, Telegram, market-status
    # check) runs before this line, so the order reaches Groww in milliseconds.
    # Everything informational happens AFTER the order is on the wire.
    try:
        order_resp = place_market_order_groww(instrument, quantity, transaction_type="BUY", product="MIS")
        order_id = order_resp.get("payload", {}).get("groww_order_id") or order_resp.get("groww_order_id")
        print(f"✅ BUY sent in {(time.time()-t0)*1000:.0f} ms => {symbol} x{quantity} | {order_resp}")
        print(f"⏱️ BUY latency (terminal → Groww ack): {(time.time()-t0)*1000:.0f} ms")
    except Exception as e:
        print(f"❌ Buy order failed ({(time.time()-t0)*1000:.0f} ms): {e}")
        send_telegram(f"❌ Buy order failed: {e}")
        return

    # notifications AFTER the order is placed (send_telegram is now non-blocking)
    send_telegram(f"🔹 Manual BUY {quantity} {symbol} | book ₹{parsed['book_profit']:.0f} total")
    alert_market_status()

    # --- confirm the BUY actually filled BEFORE managing (VALIDATE_ORDERS) ---
    # Mirrors auto mode: with the flag on we wait for the fill and abort if the
    # order is rejected / never executes (e.g. market closed, insufficient
    # margin), so we never rest a target / manage a position that doesn't exist.
    if CONFIG.get("VALIDATE_ORDERS", True):
        if not order_id:
            print("❌ No BUY order ID received — aborting.")
            send_telegram("❌ No BUY order ID received — aborting.")
            return
        buy_status = wait_for_order_status(order_id, access_token, "BUY")
        if buy_status not in ("EXECUTED", "COMPLETED", "DELIVERY_AWAITED"):
            print(f"❌ BUY not filled (status={buy_status}) — aborting, no position to manage.")
            send_telegram(f"❌ BUY not filled ({buy_status}) — aborted, nothing to manage.")
            return

    # --- actual avg buy price (from the fill; fall back to a fresh LTP) ---
    avg_buy = None
    if order_id:
        # With VALIDATE_ORDERS on we've just confirmed the BUY EXECUTED, but the
        # trades endpoint can lag the status endpoint by a moment ("No trades
        # found"). Retry briefly so target/SL are based on the TRUE average fill
        # rather than a fallback LTP that would skew the ₹-profit target.
        _attempts = 4 if CONFIG.get("VALIDATE_ORDERS", True) else 1
        for _i in range(_attempts):
            try:
                _avg, _q = get_order_executed_price(order_id, access_token)
                if _avg:
                    avg_buy = float(_avg)
                    break
            except Exception as e:
                print(f"⚠️ Could not fetch avg fill: {e}")
            if _i < _attempts - 1:
                time.sleep(0.4)
    if avg_buy is None:
        # e.g. PAPER mode or VALIDATE_ORDERS off + not yet filled — use a live LTP
        ltp_fb = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
        avg_buy = round(float(ltp_fb), 2) if ltp_fb else None
    if avg_buy is None:
        print("❌ Could not determine avg buy price — target not set. Manage this position manually.")
        send_telegram("❌ Could not determine avg buy price — manage position manually.")
        return

    # --- target for the requested TOTAL rupee profit ---
    per_unit = parsed["book_profit"] / quantity
    target_price = round_to_nearest_5_paise(avg_buy + per_unit)
    hard_sl = round_to_nearest_5_paise(avg_buy - CONFIG.get("HARD_SL_POINTS", 5.0))
    print(f"🎯 avg buy ₹{avg_buy:.2f} | target ₹{target_price} (+{per_unit:.2f} pts "
          f"= ₹{parsed['book_profit']:.0f} total) | hard SL ₹{hard_sl}")
    send_telegram(f"🎯 avg ₹{avg_buy:.2f} | target ₹{target_price} (=₹{parsed['book_profit']:.0f}) | SL ₹{hard_sl}")

    # --- rest a LIMIT SELL at the target (books on touch / spike) ---
    sell_order_id = None
    try:
        _sresp = place_limit_order_groww(instrument, quantity, target_price,
                                         transaction_type="SELL", product="MIS")
        sell_order_id = _sresp.get("payload", {}).get("groww_order_id") or _sresp.get("groww_order_id")
        print(f"🎯 TARGET LIMIT SELL placed @ ₹{target_price} | id={sell_order_id}")
        send_telegram(f"🎯 TARGET LIMIT SELL @ ₹{target_price} | id={sell_order_id}")
    except Exception as e:
        sell_order_id = None
        print(f"⚠️ Could not place target limit sell (will market-exit on touch): {e}")
        send_telegram(f"⚠️ Target limit failed, using market-on-touch: {e}")

    # --- manage: exit on target fill or hard SL ---
    # LTP-FIRST loop: every tick fetches ONLY the price (one API call, no
    # artificial delay), so we can genuinely watch the market ~every POLL_INTERVAL
    # seconds. Order-status calls are made only when VALIDATE_ORDERS is on.
    validate = bool(CONFIG.get("VALIDATE_ORDERS", True))
    poll = CONFIG.get("POLL_INTERVAL", 0.1)
    max_time = CONFIG.get("MAX_TRAIL_TIME", 3600)
    start_time = time.time()
    print(f"🛡️ Managing — watching price every {poll}s (target ₹{target_price} / SL ₹{hard_sl})"
          f"{' | validate=off' if not validate else ''}...")
    send_telegram("🛡️ Managing — target / hard SL active")

    def _confirm_ltp():
        """One extra immediate LTP read — guards against a single stale/garbage
        tick tripping the SL the instant we start managing."""
        v = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
        return float(v) if v is not None else None

    def _actual_exit_price(order_resp, fallback):
        """Return the REAL average fill price of an exit market order for accurate
        P&L (real fills slip vs the trigger LTP). Only bothers when validating;
        retries briefly because the trades endpoint lags the order ack. Falls back
        to the trigger price if the fill can't be read."""
        if not validate:
            return fallback
        oid = None
        try:
            oid = (order_resp or {}).get("payload", {}).get("groww_order_id") or (order_resp or {}).get("groww_order_id")
        except Exception:
            oid = None
        if not oid:
            return fallback
        for _ in range(4):
            try:
                a, _ = get_order_executed_price(oid, access_token)
                if a:
                    return float(a)
            except Exception:
                pass
            time.sleep(0.4)
        return fallback

    last_status_check = 0.0
    while True:
        # When validating, the resting limit's fill is the authoritative target
        # signal (catches a fast spike our LTP sampling missed). THROTTLED to
        # ~once/second so the hot loop still ticks at ~POLL_INTERVAL on the LTP
        # fetch alone — the price check below is what fires most exits.
        if validate and sell_order_id and (time.time() - last_status_check) >= 1.0:
            last_status_check = time.time()
            st = get_order_status(sell_order_id, access_token, verbose=False)
            if st in ("EXECUTED", "COMPLETED", "DELIVERY_AWAITED"):
                profit = (target_price - avg_buy) * quantity
                print(f"💰💰💰 TARGET BOOKED @ ₹{target_price} — profit ≈ ₹{profit:.2f}")
                print(f"⏱️ TARGET latency (terminal → Groww fill): {time.time()-t0:.2f} s")
                send_telegram(f"💰 TARGET BOOKED @ ₹{target_price} — profit ≈ ₹{profit:.2f}")
                play_sound_async(SOUND_PROFIT)
                log_trade_to_excel(symbol, avg_buy, target_price, quantity, profit)
                break

        # --- the only per-tick network call: fetch LTP, no artificial delay ---
        ltp = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
        if ltp is None:
            time.sleep(poll)
            continue
        ltp = float(ltp)

        # --- target: price touched the resting limit → it fills at target ---
        if ltp >= target_price:
            if sell_order_id:
                # In validate mode, CONFIRM the resting limit actually filled our
                # full quantity before claiming profit — a lone LTP print at
                # target (thin wick) does NOT guarantee a fill. If not filled yet,
                # keep managing; the top-of-loop status check will book it.
                if validate:
                    st = get_order_status(sell_order_id, access_token, verbose=False)
                    if st not in ("EXECUTED", "COMPLETED", "DELIVERY_AWAITED"):
                        time.sleep(poll)
                        continue
                # Resting LIMIT SELL fills at the target price (or better).
                profit = (target_price - avg_buy) * quantity
                print(f"💰💰💰 TARGET BOOKED @ ₹{target_price} — profit ≈ ₹{profit:.2f}")
                send_telegram(f"💰 TARGET BOOKED @ ₹{target_price} — profit ≈ ₹{profit:.2f}")
                log_trade_to_excel(symbol, avg_buy, target_price, quantity, profit)
            else:
                # No resting limit — exit at market on the touch. Report P&L from
                # the ACTUAL fill (slippage-aware) when validating.
                print(f"🎯 TARGET reached @ ₹{ltp} — MARKET exit")
                send_telegram(f"🎯 TARGET reached @ ₹{ltp} — MARKET exit")
                _resp = place_market_order_groww(instrument, quantity, "SELL", "MIS")
                exit_px = _actual_exit_price(_resp, ltp)
                profit = (exit_px - avg_buy) * quantity
                print(f"💰 Profit: ₹{profit:.2f} (exit ₹{exit_px:.2f})")
                send_telegram(f"💰 Profit: ₹{profit:.2f} (exit ₹{exit_px:.2f})")
                log_trade_to_excel(symbol, avg_buy, exit_px, quantity, profit)
            print(f"⏱️ TARGET latency (terminal → Groww exit): {time.time()-t0:.2f} s")
            play_sound_async(SOUND_PROFIT)
            break

        # --- hard SL: confirm once so a lone bad tick can't fire it instantly ---
        if ltp <= hard_sl:
            confirm = _confirm_ltp()
            if confirm is None or confirm > hard_sl:
                # bad/transient tick — price recovered on the re-read; keep managing
                time.sleep(poll)
                continue
            ltp = confirm
            # If validating, make sure the target limit didn't already fill in a
            # fast target→SL whipsaw before we market-exit — otherwise we'd sell
            # a flat position (naked short) and mis-report P&L. Book the target.
            if validate and sell_order_id:
                st = get_order_status(sell_order_id, access_token, verbose=False)
                if st in ("EXECUTED", "COMPLETED", "DELIVERY_AWAITED"):
                    profit = (target_price - avg_buy) * quantity
                    print(f"💰💰💰 TARGET already BOOKED @ ₹{target_price} — profit ≈ ₹{profit:.2f}")
                    print(f"⏱️ TARGET latency (terminal → Groww fill): {time.time()-t0:.2f} s")
                    send_telegram(f"💰 TARGET already BOOKED @ ₹{target_price} — profit ≈ ₹{profit:.2f}")
                    play_sound_async(SOUND_PROFIT)
                    log_trade_to_excel(symbol, avg_buy, target_price, quantity, profit)
                    break
            print(f"🛑 HARD SL HIT @ ₹{ltp}")
            send_telegram(f"🛑 HARD SL HIT @ ₹{ltp}")
            # 1) Exit the long at market FIRST — protect capital immediately.
            _sl_resp = place_market_order_groww(instrument, quantity, "SELL", "MIS")
            play_sound_async(SOUND_SL)
            # 2) Kill the resting target limit so it can't later fill as a naked
            #    short. A single cancel may 400 while the order is still NEW/not
            #    live at the exchange, so retry until it's actually cancelled.
            if sell_order_id:
                cancelled = False
                for _ in range(10):
                    try:
                        if cancel_order_groww(sell_order_id, access_token):
                            cancelled = True
                            break
                    except Exception:
                        pass
                    time.sleep(0.25)   # NEW -> OPEN window; then it's cancellable
                if cancelled:
                    print("🧹 Resting target limit cancelled.")
                else:
                    print("⚠️ COULD NOT cancel resting target limit — CANCEL IT MANUALLY to avoid a naked short!")
                    send_telegram("⚠️ Target limit still resting after SL — CANCEL MANUALLY (naked-short risk)!")
            # P&L from the ACTUAL market-sell fill (slippage-aware) when validating.
            exit_px = _actual_exit_price(_sl_resp, ltp)
            profit = (exit_px - avg_buy) * quantity
            print(f"💰 P&L: ₹{profit:.2f} (exit ₹{exit_px:.2f})")
            print(f"⏱️ SL latency (terminal → Groww exit): {time.time()-t0:.2f} s")
            send_telegram(f"💰 P&L: ₹{profit:.2f} (exit ₹{exit_px:.2f})")
            log_trade_to_excel(symbol, avg_buy, exit_px, quantity, profit)
            break

        # safety time exit
        if time.time() - start_time >= max_time:
            print("⏰ Max time reached — exiting")
            send_telegram("⏰ Max time — exiting")
            if sell_order_id:
                try:
                    cancel_order_groww(sell_order_id, access_token)
                except Exception:
                    pass
            place_market_order_groww(instrument, quantity, "SELL", "MIS")
            ltp_now = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0) or avg_buy
            profit = (float(ltp_now) - avg_buy) * quantity
            print(f"💰 P&L: ₹{profit:.2f}")
            send_telegram(f"💰 P&L: ₹{profit:.2f}")
            log_trade_to_excel(symbol, avg_buy, ltp_now, quantity, profit)
            break

        time.sleep(poll)

        time.sleep(poll)


import requests, time

import requests
import json

import requests

def get_order_status(order_id, access_token, verbose=True):
    """
    Fetch the status of a Groww order (CASH, F&O, etc.)
    Works with official Groww REST API response format.
    Pass verbose=False to silence the per-call print (used in tight poll loops).
    """
    if str(order_id).startswith("PAPER"):
        return "EXECUTED"   # paper orders fill instantly
    url = f"https://api.groww.in/v1/order/status/{order_id}?segment=FNO"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0"
    }

    try:
        resp = requests.get(url, headers=headers, timeout=8)
        resp.raise_for_status()  # raises for non-200 responses

        data = resp.json()  # Proper JSON response from Groww
        if verbose:
            print("🔍 Order status response:", data)

        # ✅ Extract status cleanly
        payload = data.get("payload", {})
        status = payload.get("order_status")

        return status

    except requests.exceptions.JSONDecodeError:
        print("⚠️ Error: Non-JSON response received.")
        print("Response text:", resp.text)
        return None

    except Exception as e:
        print(f"⚠️ Error fetching order status: {e}")
        return None


import time
from datetime import datetime, timedelta, timezone

def exchange_for_index(index_name=None):
    """Returns ('NSE'|'BSE', groww exchange constant) for the given (or currently configured) index."""
    idx = (index_name or CONFIG.get("index", "NIFTY")).upper()
    if "SENSEX" in idx:
        return "BSE", groww.EXCHANGE_BSE
    return "NSE", groww.EXCHANGE_NSE


def groww_exchange_for_instrument(instrument):
    """Groww exchange constant taken from the instrument's OWN exchange field
    (BSE for SENSEX, NSE for NIFTY/BANKNIFTY/FINNIFTY). Orders then follow the
    instrument being traded rather than CONFIG['index'], so any index works even
    if they ever disagree. Falls back to the configured index if unset."""
    exch = (instrument.get("exchange") or "").upper()
    if exch == "BSE":
        return groww.EXCHANGE_BSE
    if exch == "NSE":
        return groww.EXCHANGE_NSE
    return exchange_for_index()[1]


def get_recent_market_direction(symbol, groww):
    """
    Returns 'CE' if recent 5-min direction is upward (bullish),
    'PE' if downward (bearish), or None if uncertain.
    Also prints the equivalent cURL command.
    """
    try:
        # Current time and 5 minutes earlier
        end_time = datetime.now()
        start_time = end_time - timedelta(minutes=3)

        # Convert to string format accepted by Groww API
        end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
        start_time_str = start_time.strftime("%Y-%m-%d %H:%M:%S")

        exchange_str, exchange_const = exchange_for_index()

        # Construct the Groww candle API URL
        url = (
            f"https://api.groww.in/v1/historical/candles?"
            f"exchange={exchange_str}&segment=FNO&groww_symbol={symbol}"
            f"&start_time={start_time_str}"
            f"&end_time={end_time_str}"
            f"&candle_interval=1minute"
        )

        # Print cURL command for debugging
        print("\n🌀 Generated cURL for Groww Candle API:")
        print(f"curl --location '{url}' \\")
        print("  --header 'Accept: application/json' \\")
        print(f"  --header 'Authorization: Bearer {access_token}' \\")
        print("  --header 'X-API-VERSION: 1.0'\n")

        # Fetch last 5-min candle via Groww SDK
        historical = groww.get_historical_candles(
            groww_symbol=symbol,
            exchange=exchange_const,
            segment=groww.SEGMENT_FNO,
            start_time=start_time_str,
            end_time=end_time_str,
            candle_interval="1minute" # 1-min candles for better precision
        )

        candles = historical.get("candles", [])
        if not candles:
            print("⚠️ No recent candle data found.")
            return None

        first_open = candles[0][1]
        last_close = historical.get("closing_price")

        if "PE" in symbol:
            direction = "PE" if last_close > first_open else "CE"
        else:  # CE symbol
            direction = "CE" if last_close > first_open else "PE"

        print(f"📊 3-min candle trend → {direction} (O1={first_open}, C3={last_close})")
        log_o1_c3_diff(symbol, first_open, last_close, direction)

        # === O1→C3 diff gate ===
        if CONFIG.get("ENABLE_O1C3_DIFF_GATE", False):
            min_diff = float(CONFIG.get("O1C3_MIN_DIFF", 1.0))
            max_diff = float(CONFIG.get("O1C3_MAX_DIFF", 5.0))
            abs_diff = abs(float(last_close) - float(first_open))
            if not (min_diff < abs_diff < max_diff):
                print(
                    f"🚫 O1→C3 diff gate: |{last_close} - {first_open}| = {round(abs_diff, 2)} pts "
                    f"not in ({min_diff}, {max_diff}) → skipping trade."
                )
                return None
            print(
                f"✅ O1→C3 diff gate: {round(abs_diff, 2)} pts within ({min_diff}, {max_diff}) → eligible."
            )

        return direction

    except Exception as e:
        print("⚠️ Error fetching recent market direction:", e)
        return None


# ----------------- Place orders with Groww -----------------
def place_market_order_groww(instrument, quantity, transaction_type="BUY", product="MIS"):
    """
    place market order via growwapi wrapper. Returns order response or raises.
    In PAPER_TRADING mode, simulates the order without hitting the exchange.
    """
    trading_symbol = instrument.get("internal_trading_symbol") or instrument.get("trading_symbol")
    if CONFIG.get("PAPER_TRADING", False):
        fake_id = f"PAPER-{transaction_type}-{int(time.time()*1000)}"
        print(f"📋 [PAPER] {transaction_type} MARKET {trading_symbol} x{quantity} -> {fake_id}")
        return {"groww_order_id": fake_id, "order_status": "EXECUTED"}
    try:
        exchange_const = groww_exchange_for_instrument(instrument)
        order = groww.place_order(
            trading_symbol=trading_symbol,
            quantity=quantity,
            validity=groww.VALIDITY_DAY,
            exchange=exchange_const,
            segment=groww.SEGMENT_FNO,
            product=getattr(groww, f"PRODUCT_{product}") if hasattr(groww, f"PRODUCT_{product}") else getattr(groww, "PRODUCT_MIS", product),
            order_type=groww.ORDER_TYPE_MARKET,
            transaction_type=getattr(groww, f"TRANSACTION_TYPE_{transaction_type}"),
            price=0
        )
        return order
    except Exception as e:
        raise

def place_limit_order_groww(instrument, quantity, price, transaction_type="SELL", product="MIS"):
    trading_symbol = instrument.get("internal_trading_symbol") or instrument.get("trading_symbol")
    if CONFIG.get("PAPER_TRADING", False):
        fake_id = f"PAPER-LIMIT-{transaction_type}-{int(time.time()*1000)}"
        print(f"📋 [PAPER] {transaction_type} LIMIT {trading_symbol} x{quantity} @ ₹{price} -> {fake_id}")
        return {"groww_order_id": fake_id, "order_status": "NEW"}
    try:
        exchange_const = groww_exchange_for_instrument(instrument)
        order = groww.place_order(
            trading_symbol=trading_symbol,
            quantity=quantity,
            validity=groww.VALIDITY_DAY,
            exchange=exchange_const,
            segment=groww.SEGMENT_FNO,
            product=getattr(groww, f"PRODUCT_{product}") if hasattr(groww, f"PRODUCT_{product}") else getattr(groww, "PRODUCT_MIS", product),
            order_type=groww.ORDER_TYPE_LIMIT,
            transaction_type=getattr(groww, f"TRANSACTION_TYPE_{transaction_type}"),
            price=price
        )
        return order
    except Exception as e:
        raise

# ----------------- Rounding for limits (5 paise) -----------------
def round_to_nearest_5_paise(price):
    # Round to nearest 0.05
    return round(round(price * 20) / 20, 2)


def cancel_order_groww(order_id, access_token):
    """Cancel a pending order. Returns True if cancelled, False otherwise."""
    if str(order_id).startswith("PAPER"):
        print(f"📋 [PAPER] Order {order_id} cancelled")
        return True
    url = "https://api.groww.in/v1/order/cancel"
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0",
    }
    payload = {"segment": "FNO", "groww_order_id": order_id}
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=8)
        # On a non-200 (e.g. 400), surface Groww's ACTUAL reason instead of a
        # bare HTTPError — this tells us WHY (order still NEW, already filled,
        # already cancelled, rejected, etc.) rather than guessing.
        if resp.status_code != 200:
            try:
                body = resp.json()
            except Exception:
                body = resp.text
            print(f"⚠️ Cancel rejected (HTTP {resp.status_code}) for {order_id}: {body}")
            return False
        data = resp.json()
        print(f"🔄 Cancel order response: {data}")
        if data.get("success") or data.get("payload", {}).get("order_status") == "CANCELLED":
            return True
        return False
    except Exception as e:
        print(f"⚠️ Error cancelling order {order_id}: {e}")
        return False


def _exit_cancel_then_market(instrument, qty, sell_order_id, access_token):
    """Non-target exits (SL / level / time) in quick/RR mode: cancel the resting
    target limit FIRST, then market-sell — unless the limit already filled (avoids
    double-selling). If sell_order_id is None (no resting order), just market-sells."""
    if sell_order_id:
        if not cancel_order_groww(sell_order_id, access_token):
            st = (get_order_status(sell_order_id, access_token) or "").upper()
            if st in ("EXECUTED", "COMPLETED", "DELIVERY_AWAITED"):
                print("ℹ️ Target limit already filled — skipping market sell")
                return
    place_market_order_groww(instrument, qty, "SELL", "MIS")

# ----------------- Momentum sampling -----------------
import numpy as np
import time

def momentum_check_for_symbol(instrument, MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY, threshold=0.30):
    """
    Improved short-term momentum check for Nifty options.
    - Uses multiple intermediate samples
    - Smooths noise
    - Checks direction consistency
    - Returns a cleaner momentum signal
    """
    trading_symbol = instrument.get("trading_symbol")
    prices = []

    print(f"\n🧭 Checking momentum for {trading_symbol} ({MOMENTUM_SAMPLES} samples, every {MOMENTUM_DELAY}s):")

    for i in range(MOMENTUM_SAMPLES):
        p = get_ltp_for_instrument(instrument, access_token, verbose=False)
        if p:
            price = float(p)
            prices.append(price)
            print(f"[{trading_symbol}] ⏱ Sample {i+1}/{MOMENTUM_SAMPLES}: LTP = ₹{price:.2f}")
        else:
            print(f"[{trading_symbol}] ⚠️ Sample {i+1}/{MOMENTUM_SAMPLES}: Failed to fetch LTP")
        time.sleep(MOMENTUM_DELAY)

    if len(prices) < 3:
        print(f"[{trading_symbol}] ❌ Not enough data ({len(prices)} samples)")
        return None, len(prices)

    prices = np.array(prices)

    # 1️⃣ Smooth noise with small moving average
    smooth = np.convolve(prices, np.ones(3)/3, mode="valid")

    # 2️⃣ Compute rate of change (%)
    roc = np.diff(smooth) / smooth[:-1] * 100

    # 3️⃣ Remove outliers (big spikes)
    median = np.median(roc)
    std = np.std(roc)
    filtered = roc[(roc > median - 1.5*std) & (roc < median + 1.5*std)]

    if len(filtered) < 2:
        print(f"[{trading_symbol}] ⚠️ Too noisy for reliable momentum reading")
        return None, len(prices)

    # 4️⃣ Average change and direction consistency
    avg_change = np.mean(filtered)
    direction_signs = np.sign(filtered)
    consistency = np.mean(direction_signs == np.sign(avg_change)) * 100

    # Add volatility check after computing avg_change
    price_range = (prices.max() - prices.min()) / prices.mean() * 100  # % volatility
    if price_range < 0.5:  # Too flat/choppy
        direction = "FLAT"
        print(f"[{trading_symbol}] 📊 Low volatility ({price_range:.2f}%) → {direction}")
        print(f"[{trading_symbol}] 📈 Range ₹{prices[0]:.2f} → ₹{prices[-1]:.2f}\n")
        return {
            "symbol": trading_symbol,
            "avg_change": round(avg_change, 3),
            "consistency": round(consistency, 1),
            "direction": direction
        }, len(prices)

    # 5️⃣ Decision
    if avg_change > threshold and consistency > 75:
        direction = "UP"
    elif avg_change < -threshold and consistency > 75:
        direction = "DOWN"
    else:
        direction = "FLAT"

    print(f"[{trading_symbol}] 📊 Avg Δ = {avg_change:.3f}%, Consistency = {consistency:.1f}% → {direction}")
    print(f"[{trading_symbol}] 📈 Range ₹{prices[0]:.2f} → ₹{prices[-1]:.2f}\n")

    return {"symbol": trading_symbol,
            "avg_change": round(avg_change, 3),
            "consistency": round(consistency, 1),
            "direction": direction}, len(prices)

# ----------------- Find option by premium (parallel) -----------------

def find_option_by_premium_parallel(option_type, min_premium, max_premium,
                                    lots=1, funds_buffer=0.9, momentum_threshold_pct=0.25,
                                    MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY):
    """
    Filters instruments using INDEX and EXPIRY from config,
    matches by option_type, filters by premium range,
    and runs momentum checks in parallel.
    Returns: (instrument, ltp, lot_size) or (None, None, None)
    """
    config = CONFIG
    INDEX = config["index"].upper()
    EXPIRY = config["expiry"].strip()
    candidates = []

    # 🔍 Filter based on index + expiry + type
    for item in instruments_data:
        try:
            if item.get("underlying_symbol", "").upper() != INDEX:
                continue
            if item.get("instrument_type", "").upper() != option_type.upper():
                continue
            if item.get("expiry_date", "").strip() != EXPIRY:
                continue

            lot_size = int(item.get("lot_size") or item.get("lotsize") or 1)
            ltp = get_ltp_for_instrument(item, access_token, verbose=False)
            if ltp is None:
                continue
            if not (min_premium <= ltp <= max_premium):
                continue

            candidates.append({
                "instrument": item,
                "ltp": float(ltp),
                "lot_size": lot_size
            })

        except Exception as e:
            print(f"⚠️ Error while scanning: {e}")
            continue

    if not candidates:
        print(f"⚠️ No instruments for {INDEX} {EXPIRY} {option_type}")
        return None, None, None

    # ✅ Funds check (fallback = 1.2L if not available)
    try:
        margins = getattr(groww, "get_margins", lambda: {"availablecash": 270000})()
        available_cash = float(margins.get("availablecash", 270000))
    except Exception:
        available_cash = 130000

    affordable = []
    for c in candidates:
        qty = lots * c["lot_size"]
        est_cost = c["ltp"] * qty
        if available_cash <= 0 or est_cost <= available_cash * funds_buffer:
            affordable.append(c)

    if not affordable:
        print(f"⚠️ No affordable instruments for {INDEX} {EXPIRY} {option_type}")
        return None, None, None

    if option_type.upper() == "PE":
        momentum_threshold_pct = 0.30  # PEs move sharper
    else:
        momentum_threshold_pct = 0.25

    # ✅ Sort candidates closest to mid-premium
    mid = (min_premium + max_premium) / 2.0
    affordable.sort(key=lambda x: abs(x["ltp"] - mid))
    probe_list = affordable[:12]

    momentum_results = []

    def check_momentum(cand):
        mom_result, ticks = momentum_check_for_symbol(
            cand["instrument"],
            MOMENTUM_SAMPLES=MOMENTUM_SAMPLES,
            MOMENTUM_DELAY=MOMENTUM_DELAY
        )
        if mom_result and isinstance(mom_result, dict):
            slope_pct = mom_result.get("avg_change", 0)
            direction = mom_result.get("direction", "FLAT")
            consistency = mom_result.get("consistency", 0)

            # ✅ Apply momentum filter right here
            if abs(slope_pct) >= momentum_threshold_pct and consistency >= 70 and direction != "FLAT":
                return {
                    "instrument": cand["instrument"],
                    "ltp": cand["ltp"],
                    "lot_size": cand["lot_size"],
                    "slope_pct": slope_pct,
                    "direction": direction,
                    "consistency": consistency,
                    "ticks": ticks
                }
        return None

    print(f"⚙️ Checking momentum for top {len(probe_list)} {option_type} candidates...")

    with ThreadPoolExecutor(max_workers=min(len(probe_list), 8)) as executor:
        futures = {executor.submit(check_momentum, c): c for c in probe_list}
        for future in as_completed(futures):
            res = future.result()
            if res:
                momentum_results.append(res)

    if not momentum_results:
        print(f"⚠️ No strong momentum found for {option_type} (>{momentum_threshold_pct}%, consistency >70%)")
        # fallback: pick the one closest to mid-premium
        pick = probe_list[0]
        return pick["instrument"], pick["ltp"], pick["lot_size"]

    # ✅ Rank: strongest slope first, then consistency
    momentum_results.sort(key=lambda x: (x["slope_pct"], x["consistency"]), reverse=True)
    pick = momentum_results[0]

    print(f"🏆 Selected {option_type}: {pick['instrument']['trading_symbol']} "
          f"({pick['direction']} | {pick['slope_pct']:.2f}% | Consistency {pick['consistency']}%)")

    return pick["instrument"], pick["ltp"], pick["lot_size"]


# ----------------- Detect CE/PE (parallel) -----------------
def detect_option_type_parallel(index, expiry, min_p, max_p, lots, funds_buffer=0.9):
    print(f"🔍 Detecting best option between CE and PE for {index} {expiry}…")

    def worker(opt_type):
        print(f"➡️  Searching {opt_type} between {min_p}-{max_p}")
        inst, ltp, lot_size = find_option_by_premium_parallel(opt_type, min_p, max_p, lots, funds_buffer)
        mom = None
        if inst:
            print(f"📊 Running momentum check for {opt_type} ({inst.get('trading_symbol')})")
            mom, _ = momentum_check_for_symbol(inst, MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY)
            print(f"✅ Momentum for {opt_type}: {mom}")
        else:
            print(f"⚠️ No instrument found for {opt_type}")
        return opt_type, inst, ltp, lot_size, mom

    results = {}
    with ThreadPoolExecutor(max_workers=2) as ex:
        futures = {ex.submit(worker, t): t for t in ["CE", "PE"]}
        for future in as_completed(futures):
            opt_type, inst, ltp, lot_size, mom = future.result()
            results[opt_type] = {"instrument": inst, "ltp": ltp, "lot_size": lot_size, "momentum": mom}
            print(f"🧩 Finished {opt_type}: {inst.get('trading_symbol') if inst else 'None'}, momentum={mom}")

    print("🧮 Comparing CE vs PE momentum...")
    ce_mom = results.get("CE", {}).get("momentum")
    pe_mom = results.get("PE", {}).get("momentum")

    gate_enabled = CONFIG.get("ENABLE_MOMENTUM_CONSISTENCY_GATE", True)
    min_consistency = CONFIG.get("MIN_MOMENTUM_CONSISTENCY", 70)

    def _passes_gate(mom):
        if not mom:
            return False
        if not gate_enabled:
            return True
        return mom.get("consistency", 0) >= min_consistency

    # Handle missing momentum
    if not ce_mom and not pe_mom:
        print("❌ No momentum data found for CE or PE.")
        return None
    if not ce_mom:
        selected_type, mom = "PE", pe_mom
    elif not pe_mom:
        selected_type, mom = "CE", ce_mom
    else:
        ce_val = ce_mom["avg_change"]
        pe_val = pe_mom["avg_change"]

        print(f"📈 CE momentum: {ce_val:.3f}% ({ce_mom['direction']}, {ce_mom['consistency']}%)")
        print(f"📉 PE momentum: {pe_val:.3f}% ({pe_mom['direction']}, {pe_mom['consistency']}%)")

        # selection logic
        if abs(ce_val - pe_val) >= 0.25 and ce_val > pe_val and ce_val >= 0.10:
            print("✅ Selected CE (stronger momentum)")
            selected_type, mom = "CE", ce_mom
        elif abs(pe_val - ce_val) >= 0.25 and pe_val > ce_val and pe_val >= 0.10:
            print("✅ Selected PE (stronger momentum)")
            selected_type, mom = "PE", pe_mom
        elif ce_val >= pe_val:
            print("⚖️  Momentum similar — choosing CE fallback")
            selected_type, mom = "CE", ce_mom
        else:
            print("⚖️  Momentum similar — choosing PE fallback")
            selected_type, mom = "PE", pe_mom

    if not _passes_gate(mom):
        print(f"🚫 Momentum consistency gate: {selected_type} consistency {mom.get('consistency', 0)}% "
              f"< {min_consistency}% — skipping this cycle")
        return None

    r = results[selected_type]
    return selected_type, r["instrument"], r["ltp"], r["lot_size"], mom.get("consistency", 0)



def wait_for_order_status(order_id, access_token, order_type="BUY"):
    """
    Wait indefinitely until a Groww order reaches EXECUTED / COMPLETED / DELIVERY_AWAITED.
    Returns final status (string).
    """
    print(f"🔎 Waiting for {order_type} order ({order_id}) to finish...")

    while True:
        status = get_order_status(order_id, access_token)
        print(f"🕒 {order_type} status: {status}")

        if status in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
            print(f"✅ {order_type} order executed successfully.")
            send_telegram(f"✅ {order_type} order executed successfully.")
            return status

        elif status in ["FAILED", "REJECTED", "CANCELLED"]:
            print(f"❌ {order_type} order failed with status {status}.")
            send_telegram(f"❌ {order_type} order failed ({status}).")
            return status

        # wait before next check (adjust if needed)
        time.sleep(2)


import requests

def get_order_executed_price(order_id, access_token, segment="FNO"):
    """
    Fetch executed trades for a given Groww order_id and return average price & total quantity.
    """
    if str(order_id).startswith("PAPER"):
        return None, None   # paper: caller falls back to pre-order LTP as entry
    try:
        url = f"https://api.groww.in/v1/order/trades/{order_id}?segment={segment}&page=0&page_size=50"
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0"
        }

        print(f"\n📦 Fetching trade details for order: {order_id}")
        response = requests.get(url, headers=headers)
        data = response.json()

        if data.get("status") != "SUCCESS":
            print("⚠️ Failed to fetch trade info:", data)
            return None, None

        trades = data.get("payload", {}).get("trade_list", [])
        if not trades:
            print("⚠️ No trades found for order ID.")
            return None, None

        # Compute average price & total quantity
        total_qty = sum(t["quantity"] for t in trades)
        total_value = sum(t["price"] * t["quantity"] for t in trades)
        avg_price = round(total_value / total_qty, 2)

        symbol = trades[0]["trading_symbol"]
        side = trades[0]["transaction_type"]

        print(f"✅ {side} {symbol} | Total Qty={total_qty} | Avg Price=₹{avg_price}")
        return avg_price, total_qty

    except Exception as e:
        print("❌ Error fetching order trades:", e)
        return None, None



# ----------------- LEVEL FILTER (key-level profit guard, auto mode) -----------------
# Reuse the multi-touch S/R engine from KEY_LEVELS_TERMINAL. Guarded so a missing
# or broken module can never take down the trading bot — the filter just disables.
try:
    import KEY_LEVELS_TERMINAL as _KL
except Exception as _kl_err:  # pragma: no cover
    _KL = None
    print(f"⚠️ LEVEL_FILTER unavailable (KEY_LEVELS_TERMINAL import failed: {_kl_err})")


def _kl_index_spot(access_token):
    """Fast index spot via the direct live-data/ltp CASH endpoint.
    Deliberately avoids get_index_spot_price(), which reloads/re-converts the
    23MB instrument master on every call (fine once, terrible for polling)."""
    idx = CONFIG.get("index", "NIFTY").upper()
    exch = "BSE" if "SENSEX" in idx else "NSE"
    try:
        url = (f"https://api.groww.in/v1/live-data/ltp"
               f"?segment=CASH&exchange_symbols={exch}_{idx}")
        resp = requests.get(url, headers={
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0",
        }, timeout=6)
        if resp.status_code == 200:
            payload = resp.json().get("payload", {})
            if payload:
                return float(next(iter(payload.values())))
    except Exception:
        pass
    return None


def _kl_index_candles(index_name, interval="5minute", days_back=7):
    """Fetch index CASH candles using the bot's own groww client (not KL's auth)."""
    idx = index_name.upper()
    exch_str, exch_const = exchange_for_index(index_name)
    if idx == "NIFTY":
        syms = ["NSE-NIFTY 50", "NSE-NIFTY"]
    elif "SENSEX" in idx:
        syms = ["BSE-SENSEX", "BSE-S&P BSE SENSEX"]
    elif idx == "BANKNIFTY":
        syms = ["NSE-NIFTY BANK", "NSE-BANKNIFTY"]
    elif idx == "FINNIFTY":
        syms = ["NSE-NIFTY FIN SERVICE"]
    else:
        syms = [f"{exch_str}-{idx}"]
    end_dt = datetime.now()
    start_dt = end_dt - timedelta(days=days_back)
    for sym in syms:
        try:
            r = groww.get_historical_candles(
                groww_symbol=sym, exchange=exch_const, segment="CASH",
                start_time=start_dt.strftime("%Y-%m-%d %H:%M:%S"),
                end_time=end_dt.strftime("%Y-%m-%d %H:%M:%S"),
                candle_interval=interval,
            )
            if r and r.get("candles") and len(r["candles"]) >= 3:
                return [
                    {"ts": c[0], "open": float(c[1]), "high": float(c[2]),
                     "low": float(c[3]), "close": float(c[4])}
                    for c in r["candles"]
                ]
        except Exception:
            pass
    return []


def compute_level_barrier(instrument, access_token):
    """For the selected option, return the nearest key level in the trade's
    favour that should trigger an instant profit-book:
      CE -> nearest RESISTANCE above current spot
      PE -> nearest SUPPORT below current spot
    Candidates = multi-touch S/R clusters + previous-day high/low.
    Returns {opt, spot, barrier, kind, distance} or None if unavailable."""
    if _KL is None:
        return None
    opt = (instrument.get("instrument_type") or "").upper()
    if opt not in ("CE", "PE"):
        return None
    idx = CONFIG.get("index", "NIFTY")
    candles = _kl_index_candles(idx)
    if not candles:
        return None
    candles, _ = _KL.filter_spikes(candles, CONFIG.get("LEVEL_SPIKE_MULT", 8.0))
    # Lightweight spot fetch (direct LTP endpoint) — avoids reloading the 23MB
    # instrument master that get_index_spot_price() re-converts on every call.
    spot = _kl_index_spot(access_token) or candles[-1]["close"]
    lr = int(CONFIG.get("LEVEL_PIVOT_LR", 10))
    tol = spot * float(CONFIG.get("LEVEL_TOL_PCT", 0.15)) / 100.0
    min_touches = int(CONFIG.get("LEVEL_MIN_TOUCHES", 2))

    pivots = _KL.find_pivots(candles, lr, lr)
    levels = _KL.cluster_levels(pivots, tol)
    strong = [lv["price"] for lv in levels if lv["touches"] >= min_touches]

    pdo = _KL.prev_day_from_intraday(candles)

    if opt == "CE":
        cands = [p for p in strong if p > spot]
        if pdo and pdo.get("high", 0) > spot:
            cands.append(pdo["high"])
        barrier = min(cands) if cands else None
        kind = "RESISTANCE"
    else:  # PE
        cands = [p for p in strong if p < spot]
        if pdo and pdo.get("low", 0) and pdo["low"] < spot:
            cands.append(pdo["low"])
        barrier = max(cands) if cands else None
        kind = "SUPPORT"

    if barrier is None:
        return None
    return {"opt": opt, "spot": spot, "barrier": barrier, "kind": kind,
            "distance": abs(barrier - spot)}


def level_barrier_hit(barrier_info, access_token):
    """Return (hit, current_spot). hit=True once the index reaches the barrier."""
    if not barrier_info:
        return False, None
    cur = _kl_index_spot(access_token)
    if cur is None:
        return False, None
    if barrier_info["opt"] == "CE":
        return cur >= barrier_info["barrier"], cur
    return cur <= barrier_info["barrier"], cur


# ----------------- Place CP order workflow (mirrors AngelOne logic) -----------------
def place_cp_order(command, is_auto=False):
    global buy_status
    if is_auto:
        order = command  # dict form
        symbol = order["symbol"]
        qty = order["lots"] * order["lot_size"]
        book_profit = order["book_profit"]

        # get instrument info directly from master
        instrument = next((inst for inst in instruments_data if inst["internal_trading_symbol"] == symbol), None)
        if not instrument:
            print(f"❌ Instrument {symbol} not found in master.")
            return

        print(f"🔹 Auto order => {symbol}, qty={qty}, book@{book_profit} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")

        # --- Fetch LTP ---
        ltp_before = get_ltp_for_instrument(instrument, access_token)
        if ltp_before is None:
            print("❌ Could not fetch LTP before placing order.")
            return

        entry_price = round(float(ltp_before), 2)


        # === BUY @ MARKET ===
        alert_market_status()
        try:
            order_resp = place_market_order_groww(instrument, qty, transaction_type="BUY", product="MIS")
            order_id = order_resp.get("payload", {}).get("groww_order_id") or order_resp.get("groww_order_id")
            print(f"✅ Auto Buy placed: :{order_resp} ======= [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            send_telegram(f"✅ Auto Buy placed: :{order_resp} ======= [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
        except Exception as e:
            print(f"❌ Auto BUY failed: {e}")
            send_telegram(f"❌ Auto BUY failed: {e}")
            return

        # ✅ Confirm the BUY actually filled BEFORE managing the trade (VALIDATE_ORDERS).
        # A rejected buy (insufficient funds / margin / exchange reject) must NOT
        # lead to phantom target/trailing management of a position that doesn't exist.
        if CONFIG.get("VALIDATE_ORDERS", True):
            if not order_id:
                print("❌ No BUY order ID received — aborting.")
                send_telegram("❌ No BUY order ID received — aborting.")
                return
            buy_status = wait_for_order_status(order_id, access_token, "BUY")
            if buy_status not in ("EXECUTED", "COMPLETED", "DELIVERY_AWAITED"):
                print(f"❌ BUY not filled (status={buy_status}) — aborting, no position to manage.")
                send_telegram(f"❌ BUY not filled ({buy_status}) — trade aborted, nothing to manage.")
                return


        highest_price = entry_price

        start_time = time.time()

        trail_start = CONFIG["TRAIL_START_PROFIT"]
        trail_step = CONFIG["TRAIL_STEP"]
        poll = CONFIG["POLL_INTERVAL"]
        max_time = CONFIG["MAX_TRAIL_TIME"]
        hard_sl = entry_price - CONFIG.get("HARD_SL_POINTS")

        # === QUICK MODE / RR MODE: fixed target (and SL) from actual avg buy price ===
        # quick_target != None also disables trailing further below.
        quick_target = None
        sell_order_id = None   # resting target LIMIT SELL order id (quick/RR mode)
        if CONFIG.get("RR_MODE") or CONFIG.get("QUICK_MODE"):
            avg_buy = entry_price
            try:
                _avg, _q = get_order_executed_price(order_id, access_token)
                if _avg:
                    avg_buy = float(_avg)
            except Exception as e:
                print(f"⚠️ fixed-target mode: could not fetch avg fill, using entry {entry_price}: {e}")

            if CONFIG.get("RR_MODE"):
                # risk-anchored: SL = avg - risk, target = avg + risk*ratio
                risk_pts = float(CONFIG.get("RR_RISK_POINTS", 2.0))
                rr = float(CONFIG.get("RR_RATIO", 3.0))
                hard_sl = round_to_nearest_5_paise(avg_buy - risk_pts)   # override default hard SL
                quick_target = round_to_nearest_5_paise(avg_buy + risk_pts * rr)
                msg = (f"⚖️ RR MODE 1:{rr:g} | avg buy ₹{avg_buy:.2f} | "
                       f"SL ₹{hard_sl} (-{risk_pts:g}pt) | target ₹{quick_target} (+{risk_pts*rr:g}pt) | trailing disabled")
            else:
                # quick mode: fixed +QUICK_POINTS target, default hard SL unchanged
                quick_pts = float(CONFIG.get("QUICK_POINTS", 1.5))
                quick_target = round_to_nearest_5_paise(avg_buy + quick_pts)
                msg = (f"⚡ QUICK MODE ON | avg buy ₹{avg_buy:.2f} | "
                       f"fixed target ₹{quick_target} (+{quick_pts:g}pt) | trailing disabled")
            print(msg)
            send_telegram(msg)

            # Target mechanism:
            #  - SET_TARGET_ORDER True (and live): place a resting LIMIT SELL now so
            #    profit books even on a fast spike / exact target fill.
            #  - SET_TARGET_ORDER False: no target order — just watch LTP and market
            #    exit on touch (loop's `ltp >= quick_target` path).
            #  - PAPER mode: always watch-and-market (a fake limit reports EXECUTED
            #    instantly), regardless of SET_TARGET_ORDER.
            if CONFIG.get("PAPER_TRADING", False) or not CONFIG.get("SET_TARGET_ORDER", True):
                why = "PAPER" if CONFIG.get("PAPER_TRADING", False) else "SET_TARGET_ORDER=False"
                print(f"👁️ target ₹{quick_target} — watching LTP, MARKET exit on touch ({why}, no resting order)")
                send_telegram(f"👁️ target ₹{quick_target} — watch LTP + market exit on touch ({why})")
                sell_order_id = None
            else:
                try:
                    _sresp = place_limit_order_groww(instrument, qty, quick_target,
                                                     transaction_type="SELL", product="MIS")
                    sell_order_id = _sresp.get("payload", {}).get("groww_order_id") or _sresp.get("groww_order_id")
                    print(f"🎯 TARGET LIMIT SELL placed @ ₹{quick_target} | id={sell_order_id}")
                    send_telegram(f"🎯 TARGET LIMIT SELL @ ₹{quick_target} | id={sell_order_id}")
                except Exception as e:
                    sell_order_id = None
                    print(f"⚠️ Could not place target limit sell (will use market on touch): {e}")
                    send_telegram(f"⚠️ Target limit sell failed, using market-on-touch: {e}")

        print("📈 Trailing started...")
        send_telegram("📈 Trailing started")

        # === LEVEL FILTER: find the key level to guard profit against ===
        level_barrier = None
        if CONFIG.get("LEVEL_FILTER"):
            try:
                level_barrier = compute_level_barrier(instrument, access_token)
            except Exception as e:
                print(f"⚠️ LEVEL_FILTER: barrier computation failed: {e}")
                level_barrier = None
            if level_barrier:
                msg = (f"🎯 LEVEL FILTER ON | {level_barrier['opt']} | spot={level_barrier['spot']:.2f} | "
                       f"next {level_barrier['kind']} @ {level_barrier['barrier']:.2f} "
                       f"({level_barrier['distance']:.1f} pts away) — will book profit on touch")
                print(msg)
                send_telegram(msg)
            else:
                print("🎯 LEVEL FILTER ON | no qualifying level found — trailing as normal")
                send_telegram("🎯 LEVEL FILTER ON | no qualifying level found — trailing as normal")

        # === INDEX MONITOR: capture the entry reference spot ===
        # Directional invalidation stop — if the underlying turns against the option
        # we bought (CE wants index up, PE wants index down) we bail out instantly
        # rather than waiting for premium to decay to the hard SL.
        idx_monitor = None
        if CONFIG.get("ENABLE_INDEX_MONITOR"):
            opt_dir = (instrument.get("instrument_type") or "").upper()
            entry_spot = _kl_index_spot(access_token)
            if opt_dir in ("CE", "PE") and entry_spot:
                idx_monitor = {
                    "opt": opt_dir,
                    "entry_spot": float(entry_spot),
                    "adverse_pts": float(CONFIG.get("INDEX_ADVERSE_POINTS", 5.0)),
                    "need_ticks": int(CONFIG.get("INDEX_ADVERSE_TICKS", 3)),
                    "grace": float(CONFIG.get("INDEX_MONITOR_GRACE_SEC", 5.0)),
                    "poll_every": float(CONFIG.get("INDEX_POLL_INTERVAL", 1.0)),
                    "adverse_count": 0,
                    "favour_count": 0,   # consecutive polls the index is moving IN our favour
                    "released": False,   # True once the fixed target has been released to trail (hold-on-favour)
                    "last_poll": 0.0,
                }
                want = "UP" if opt_dir == "CE" else "DOWN"
                bad = "DOWN" if opt_dir == "CE" else "UP"
                msg = (f"🛰️ INDEX MONITOR ON | {opt_dir} | entry spot={entry_spot:.2f} | "
                       f"expect index {want} → exit if index goes {bad} "
                       f">{idx_monitor['adverse_pts']:g} pts for {idx_monitor['need_ticks']} polls")
                print(msg)
                send_telegram(msg)
            else:
                print(f"🛰️ INDEX MONITOR ON but no entry spot/opt (opt={opt_dir}, spot={entry_spot}) — monitor disabled for this trade")
                send_telegram("🛰️ INDEX MONITOR: could not read entry spot — monitor disabled for this trade")

        while True:
            ltp = get_ltp_for_instrument(instrument, access_token, verbose=False ,delay = 0)
            if ltp is None:
                time.sleep(poll)
                continue

            ltp = float(ltp)

            # ⚡ QUICK / RR MODE: target handling
            if quick_target is not None:
                if sell_order_id:
                    # resting limit sell — book when it fills
                    st = (get_order_status(sell_order_id, access_token, verbose=False) or "").upper()
                    if st in ("EXECUTED", "COMPLETED", "DELIVERY_AWAITED"):
                        print(f"🎯 TARGET LIMIT FILLED @ ₹{quick_target}")
                        send_telegram(f"🎯 TARGET LIMIT FILLED @ ₹{quick_target}")
                        play_sound_async(SOUND_PROFIT)
                        profit = (quick_target - entry_price) * qty
                        print(f"💰 Profit: ₹{profit:.2f}")
                        send_telegram(f"💰 Profit: ₹{profit:.2f}")
                        log_trade_to_excel(
                            instrument.get("internal_trading_symbol"),
                            entry_price, quick_target, qty, profit
                        )
                        break
                    elif st in ("REJECTED", "FAILED", "CANCELLED"):
                        # resting target order died — don't loop forever; fall back
                        # to market-on-touch for the rest of the trade.
                        print(f"⚠️ Target limit {st} — switching to market-on-touch")
                        send_telegram(f"⚠️ Target limit {st} — switching to market-on-touch")
                        sell_order_id = None

                # market fallback on touch — fires whenever there is NO live resting
                # limit order: either it was never placed, OR it was just found
                # rejected/cancelled above in THIS SAME poll. This must NOT be `elif`
                # on `if sell_order_id`: otherwise, on the poll where the resting order
                # dies we'd only clear the id and skip booking, and if the LTP is
                # already at/above target on that poll (a fast spike) we'd miss it and
                # defer to the next poll — by which time the spike may be gone.
                if sell_order_id is None and ltp >= quick_target:
                    # limit wasn't placed (or just died) — market fallback on touch
                    print(f"🎯 TARGET ₹{quick_target} hit @ {ltp} (market fallback) — booking profit")
                    send_telegram(f"🎯 TARGET ₹{quick_target} hit @ {ltp} (market fallback) — booking profit")
                    place_market_order_groww(instrument, qty, "SELL", "MIS")
                    play_sound_async(SOUND_PROFIT)
                    profit = (ltp - entry_price) * qty
                    print(f"💰 Profit: ₹{profit:.2f}")
                    send_telegram(f"💰 Profit: ₹{profit:.2f}")
                    log_trade_to_excel(
                        instrument.get("internal_trading_symbol"),
                        entry_price, ltp, qty, profit
                    )
                    break

            # 🎯 LEVEL FILTER: index touched the key level → book profit instantly
            if level_barrier:
                hit, cur_spot = level_barrier_hit(level_barrier, access_token)
                if hit:
                    print(f"🎯 {level_barrier['kind']} @ {level_barrier['barrier']:.2f} touched "
                          f"(spot={cur_spot:.2f}) — booking profit instantly @ {ltp}")
                    send_telegram(f"🎯 {level_barrier['kind']} @ {level_barrier['barrier']:.2f} touched "
                                  f"(spot={cur_spot:.2f}) — booking profit instantly @ {ltp}")
                    _exit_cancel_then_market(instrument, qty, sell_order_id, access_token)
                    play_sound_async(SOUND_PROFIT)
                    profit = (ltp - entry_price) * qty
                    print(f"💰 Profit: ₹{profit:.2f}")
                    send_telegram(f"💰 Profit: ₹{profit:.2f}")
                    log_trade_to_excel(
                        instrument.get("internal_trading_symbol"),
                        entry_price, ltp, qty, profit
                    )
                    break

            # 🛰️ INDEX MONITOR: exit if the underlying turns against the trade
            if idx_monitor is not None:
                now_t = time.time()
                # respect grace period after entry + throttle index polling
                if (now_t - start_time) >= idx_monitor["grace"] and \
                   (now_t - idx_monitor["last_poll"]) >= idx_monitor["poll_every"]:
                    idx_monitor["last_poll"] = now_t
                    cur_spot = _kl_index_spot(access_token)
                    if cur_spot is not None:
                        move = cur_spot - idx_monitor["entry_spot"]   # +ve = index moved UP
                        # adverse = index moving OPPOSITE to what the option needs
                        if idx_monitor["opt"] == "CE":
                            adverse = move <= -idx_monitor["adverse_pts"]   # CE wants up; down is bad
                        else:
                            adverse = move >= idx_monitor["adverse_pts"]    # PE wants down; up is bad
                        if adverse:
                            idx_monitor["adverse_count"] += 1
                            print(f"⚠️ INDEX against {idx_monitor['opt']}: spot={cur_spot:.2f} "
                                  f"(move={move:+.2f} pts) "
                                  f"[{idx_monitor['adverse_count']}/{idx_monitor['need_ticks']}]")
                        else:
                            # back in our favour (or within noise) — reset the counter
                            idx_monitor["adverse_count"] = 0

                        # 🟢 HOLD-ON-FAVOUR: index confirming our direction → stop
                        # scalping, release the fixed target and let the existing
                        # trailing stop ride the runner. Only while a fixed target is
                        # still live and we haven't already released it.
                        if CONFIG.get("INDEX_HOLD_ON_FAVOUR") and quick_target is not None \
                           and not idx_monitor["released"]:
                            # favour = index moving the SAME way the option needs
                            if idx_monitor["opt"] == "CE":
                                favour = move >= idx_monitor["adverse_pts"]     # CE wants up
                            else:
                                favour = move <= -idx_monitor["adverse_pts"]    # PE wants down
                            if favour:
                                idx_monitor["favour_count"] += 1
                                print(f"🟢 INDEX favours {idx_monitor['opt']}: spot={cur_spot:.2f} "
                                      f"(move={move:+.2f} pts) "
                                      f"[{idx_monitor['favour_count']}/{idx_monitor['need_ticks']}]")
                            else:
                                idx_monitor["favour_count"] = 0
                            if idx_monitor["favour_count"] >= idx_monitor["need_ticks"]:
                                # cancel the resting scalp target, switch to trailing
                                if sell_order_id:
                                    cancel_order_groww(sell_order_id, access_token)
                                sell_order_id = None
                                quick_target = None            # re-enables the trailing block below
                                idx_monitor["released"] = True
                                print(f"🟢🚀 INDEX confirms trend ({move:+.2f} pts favour) — released fixed "
                                      f"target, now TRAILING the runner (trail_step={trail_step}, hard SL={hard_sl})")
                                send_telegram(f"🟢🚀 INDEX confirms {idx_monitor['opt']} trend ({move:+.2f} pts) "
                                              f"— target released, trailing the runner")

                        if idx_monitor["adverse_count"] >= idx_monitor["need_ticks"]:
                            print(f"🛰️🛑 INDEX MONITOR EXIT | {idx_monitor['opt']} | index moved {move:+.2f} pts "
                                  f"against the trade — cancelling target & market-exiting @ {ltp}")
                            send_telegram(f"🛰️🛑 INDEX MONITOR EXIT | index {move:+.2f} pts against {idx_monitor['opt']} "
                                          f"— cancelling target & exiting @ {ltp}")
                            _exit_cancel_then_market(instrument, qty, sell_order_id, access_token)
                            play_sound_async(SOUND_SL)
                            profit = (ltp - entry_price) * qty
                            print(f"💰 Profit: ₹{profit:.2f}")
                            send_telegram(f"💰 Profit: ₹{profit:.2f}")
                            log_trade_to_excel(
                                instrument.get("internal_trading_symbol"),
                                entry_price, ltp, qty, profit
                            )
                            break

            # 🔴 HARD STOP LOSS
            if ltp <= hard_sl:
                print(f"🛑 HARD SL HIT @ {ltp}")
                send_telegram(f"🛑 HARD SL HIT @ {ltp}")
                _exit_cancel_then_market(instrument, qty, sell_order_id, access_token)
                play_sound_async(SOUND_SL)
                profit = (ltp - entry_price) * qty
                print(f"💰 Profit: ₹{profit:.2f}")
                send_telegram(f"💰 Profit: ₹{profit:.2f}")
                log_trade_to_excel(
                    instrument.get("internal_trading_symbol"),
                    entry_price, ltp, qty, profit
                )
                break

            # 🔼 Update highest price
            if ltp > highest_price:
                highest_price = ltp
                print(f"🔼 New High: ₹{highest_price}")
                send_telegram(f"🔼 New High: ₹{highest_price}")
            # 🟢 Start trailing after ₹1 profit  (disabled in QUICK MODE — fixed target instead)
            if quick_target is None and highest_price >= entry_price + trail_start:
                trail_exit = round_to_nearest_5_paise(highest_price - trail_step)
                print(f"📉 Trail Active | LTP={ltp} | Exit={trail_exit}")
                send_telegram(f"📉 Trail Active | LTP={ltp} | Exit={trail_exit}")

                #NEWCHANGE
                # print("Waiting for 8 sec to have momentum")
                # send_telegram("Waiting for 8 sec to have momentum")
                # time.sleep(8)
                #NEWCHANGE
                ltp = get_ltp_for_instrument(instrument, access_token, verbose=False, delay = 0)
                if ltp <= trail_exit:
                    print(f"🔻 Trailing HIT @ {ltp}")
                    send_telegram(f"🔻 Trailing HIT @ {ltp}")
                    place_market_order_groww(instrument, qty, "SELL", "MIS")
                    play_sound_async(SOUND_PROFIT)

                    profit = (ltp - entry_price) * qty
                    print(f"💰 Profit: ₹{profit:.2f}")
                    send_telegram(f"💰 Profit: ₹{profit:.2f}")
                    log_trade_to_excel(
                        instrument.get("internal_trading_symbol"),
                        entry_price, ltp, qty, profit
                    )
                    print("Waiting for new moment now for 1 min")
                    time.sleep(60)
                    break

            # ⏰ SAFETY TIME EXIT
            if time.time() - start_time >= max_time:
                print("⏰ Max trail time reached — exiting")
                send_telegram("⏰ Max trail time reached — exiting")
                _exit_cancel_then_market(instrument, qty, sell_order_id, access_token)
                play_sound_async(SOUND_PROFIT)

                ltp_now = get_ltp_for_instrument(instrument, access_token, verbose=False) or entry_price
                profit = (ltp_now - entry_price) * qty
                print(f"💰 Profit: ₹{profit:.2f}")
                send_telegram(f"💰 Profit: ₹{profit:.2f}")
                log_trade_to_excel(
                    instrument.get("internal_trading_symbol"),
                    entry_price, ltp_now, qty, profit
                )
                break

            time.sleep(3)

        return  # ✅ end of auto mode execution

    else:
        # NEW compact format using the startup index/expiry: "14 25950CE Book 1050"
        book_parsed = parse_book_command(command)
        if book_parsed:
            place_book_target_order(book_parsed)
            print("Waiting for 10 seconds to get another data.")
            return

        parsed = parse_cp_command(command)
        if not parsed:
            print("❌ Could not parse command.")
            return

        instrument = find_instrument_from_command(command, all_instruments)
        if not instrument:
            print("❌ Instrument not found in instrument master.")
            return

        lot_size = int(instrument.get("lot_size") or instrument.get("lotsize") or 1)
        quantity = parsed["lots"] * lot_size

        ltp_before = get_ltp_for_instrument(instrument, access_token)
        if ltp_before is None:
            print("❌ Could not fetch LTP before placing order.")
            return

        entry_price = round(float(ltp_before), 2)
        send_telegram(f"entry price: {entry_price} | {instrument.get('internal_trading_symbol')} | qty={quantity}")
        print(f"entry price: {entry_price}")

        # Place BUY @ MARKET
        alert_market_status()
        try:
            order_resp = place_market_order_groww(instrument, quantity, transaction_type="BUY", product="MIS")
            order_id = order_resp.get("payload", {}).get("groww_order_id") or order_resp.get("groww_order_id")
            print("✅ Buy Order placed:", order_resp)
        except Exception as e:
            print(f"❌ Buy order failed: {e}")
            send_telegram(f"❌ Buy order failed: {e}")
            return

        # ================= TRAILING LOGIC =================
        highest_price = entry_price+1
        start_time = time.time()

        trail_start = CONFIG["TRAIL_START_PROFIT"]
        trail_step = CONFIG["TRAIL_STEP"]
        poll = CONFIG["POLL_INTERVAL"]
        max_time = CONFIG["MAX_TRAIL_TIME"]
        hard_sl = entry_price - CONFIG.get("HARD_SL_POINTS")

        print("📈 Trailing started...")
        send_telegram("📈 Trailing started")

        while True:
            ltp = get_ltp_for_instrument(instrument, access_token, verbose=False)
            if ltp is None:
                time.sleep(poll)
                continue

            ltp = float(ltp)

            # 🔴 HARD STOP LOSS
            if ltp <= hard_sl:
                print(f"🛑 HARD SL HIT @ {ltp}")
                send_telegram(f"🛑 HARD SL HIT @ {ltp}")
                place_market_order_groww(instrument, quantity, "SELL", "MIS")
                play_sound_async(SOUND_SL)
                profit = (ltp - entry_price) * quantity
                print(f"💰 Profit: ₹{profit:.2f}")
                send_telegram(f"💰 Profit: ₹{profit:.2f}")
                log_trade_to_excel(
                    instrument.get("internal_trading_symbol"),
                    entry_price, ltp, quantity, profit
                )
                break

            # 🔼 Update highest price
            if ltp > highest_price:
                highest_price = ltp
                print(f"🔼 New High: ₹{highest_price}")

            # 🟢 Start trailing after ₹1 profit
            if highest_price >= entry_price + trail_start:
                trail_exit = round_to_nearest_5_paise(highest_price - trail_step)
                print(f"📉 Trail Active | LTP={ltp} | Exit={trail_exit}")

                if ltp <= trail_exit:
                    print(f"🔻 Trailing HIT @ {ltp}")
                    send_telegram(f"🔻 Trailing HIT @ {ltp}")
                    place_market_order_groww(instrument, quantity, "SELL", "MIS")
                    print(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
                    send_telegram(f"💰💰💰💰💰💰 PROFIT BOOKED 💰💰💰💰💰")
                    play_sound_async(SOUND_PROFIT)

                    profit = (ltp - entry_price) * quantity
                    print(f"💰 Profit: ₹{profit:.2f}")
                    send_telegram(f"💰 Profit: ₹{profit:.2f}")
                    log_trade_to_excel(
                        instrument.get("internal_trading_symbol"),
                        entry_price, ltp, quantity, profit
                    )
                    break

            # ⏰ SAFETY TIME EXIT
            if time.time() - start_time >= max_time:
                print("⏰ Max trail time reached — exiting")
                send_telegram("⏰ Max trail time reached — exiting")
                place_market_order_groww(instrument, quantity, "SELL", "MIS")
                play_sound_async(SOUND_PROFIT)

                ltp_now = get_ltp_for_instrument(instrument, access_token, verbose=False) or entry_price
                profit = (ltp_now - entry_price) * quantity
                print(f"💰 Profit: ₹{profit:.2f}")
                send_telegram(f"💰 Profit: ₹{profit:.2f}")
                log_trade_to_excel(
                    instrument.get("internal_trading_symbol"),
                    entry_price, ltp_now, quantity, profit
                )
                break

            time.sleep(poll)

        print("Waiting for 10 seconds to get another data.")
        time.sleep(10)
        return




# ----------------- Auto mode runner (momentum + premium) -----------------


def auto_mode_runner():
    cfg = CONFIG
    print("\n--- AUTO MODE (momentum + premium) ---")
    send_telegram("\n--- AUTO MODE (momentum + premium) ---")
    index = cfg["index"]
    expiry = cfg["expiry"]
    min_p = cfg["min_premium"]
    max_p = cfg["max_premium"]
    lots = cfg["lots"]
    book_profit = cfg["book_profit"]

    target_pnl = cfg["target_pnl"]

    while True:
        alert_market_status()
        opt_result = detect_option_type_parallel(index, expiry, min_p, max_p, lots)
        if not opt_result:
            print("❌ Could not determine CE/PE momentum side. Retrying in 10 seconds...")
            send_telegram("❌ Could not determine CE/PE momentum side. Retrying in 10 seconds...")
            time.sleep(10)
            continue

        opt_type, instrument, ltp, lot_size, consistency = opt_result
        if not instrument:
            print("❌ No matching/affordable option found. Retrying...")
            send_telegram("❌ No matching/affordable option found. Retrying...")
            time.sleep(60)
            continue

        symbol = instrument.get("tradingsymbol") or instrument.get("symbol") or instrument.get("internal_trading_symbol")
        instrument_type = instrument.get("instrument_type", "NA")
        groww_symbol = instrument.get("groww_symbol")
        print(f"✅ Selected: {symbol} ({instrument_type}) | LTP={ltp} | lot_size={lot_size} | groww_symbol={groww_symbol} | consistency={consistency}%")
        send_telegram(f"✅ Selected: {symbol} ({instrument_type}) | LTP={ltp} | lot_size={lot_size} | groww_symbol={groww_symbol} | consistency={consistency}%")

        # 🚀 Directly place the order (no string parsing)
        order_details = {
            "symbol": symbol,
            "ltp": ltp,
            "lots": lots,
            "book_profit": float(book_profit),
            "lot_size": lot_size,
            "side": "BUY"
        }

        market_direction = get_recent_market_direction(groww_symbol, groww)
        print(f"Market Direction: {market_direction}")
        send_telegram(f"Market Direction: {market_direction}")

        if market_direction == instrument_type:
            print(f"✅ Market direction CONFIRMS momentum {market_direction} → proceeding with order.")

            # 👇 Manual confirmation

            #NEWCHANGE
            user_confirmation_needed = cfg.get("user_confirmation_needed", False)
            print(f"user_confirmation_needed : {user_confirmation_needed}")
            if user_confirmation_needed:
                play_sound_async(SOUND_user_input)
                user_input = input(
                    f"Confirm trade for {instrument_type}? Type Y/Yes to proceed, anything else to skip: "
                ).strip().lower()
                if user_input in ("y", "yes"):
                    print(f"➡️ Placing auto order: {order_details} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
                    send_telegram(f"➡️ Placing auto order: {order_details}")
                    place_cp_order(order_details, is_auto=True)
                else:
                    print("❌ Trade skipped by user confirmation.")
                    send_telegram("❌ Trade skipped by user confirmation.")
            else:
                print(f"➡️ Placing auto order: {order_details} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
                send_telegram(f"➡️ Placing auto order: {order_details} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
                place_cp_order(order_details, is_auto=True)
        else:
            print("❌ Skipping trade — market direction conflicts with momentum.")
            print("Waiting for 10 seconds to get another data.")
            time.sleep(10)

        time.sleep(2)


# ----------------- Main menu -----------------
if __name__ == "__main__":
    threading.Thread(target=market_hours_watcher, daemon=True).start()
    print(f"\n✨ Groww CP Bot Ready — Index={CONFIG['index']} Expiry={CONFIG['expiry']} (Groww backend)")
    print("You can run in MANUAL or AUTO mode.")
    print(f"Manual (uses {CONFIG['index']} {CONFIG['expiry']}): 14 25950CE Book 1050")
    print("   -> 14 lots of 25950 CE, book when TOTAL profit hits ₹1050 (target = avg + 1050/qty, 5-paise rounded)")
    print(f"Full form also works: Buy 14 {CONFIG['index']}04NOV2525950CE at CP and Book at 1050\n")
    while True:
        mode = input("Choose mode: (m)anual / (a)uto / (q)uit: ").strip().lower()
        if mode in ["q", "quit", "exit"]:
            print("Exiting.")
            break
        if mode in ["a", "auto"]:
            auto_mode_runner()
            continue
        if mode in ["m", "manual"]:
            user_input = input("\nEnter command (or press Enter for status, type 'back' to menu): ").strip()
            if user_input.lower() in ["back"]:
                continue
            if user_input == "":
                print("Status check not implemented for Groww PnL in this script.")
                continue
            place_cp_order(user_input)
            continue
        print("Invalid input. Choose 'm' or 'a' or 'q'.")