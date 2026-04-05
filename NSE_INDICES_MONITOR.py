#!/usr/bin/env python3
"""
NSE Indices Heatmap Monitor
Tracks real-time percentage changes across all NSE indices categories
Provides trading signals based on overall market direction
"""

import requests
import time
import json
from datetime import datetime
from collections import deque
import statistics
import os
from openpyxl import load_workbook, Workbook

# NSE API Configuration
BASE_URL = "https://www.nseindia.com/api/heatmap-index"
HEADERS = {
    'accept': '*/*',
    'accept-language': 'en-US,en;q=0.9',
    'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36',
    'referer': 'https://www.nseindia.com/market-data/live-market-indices/heatmap',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin'
}

# Categories to monitor
CATEGORIES = {
    'Broad Market': 'Broad%20Market%20Indices',
    'Sectoral': 'Sectoral%20Indices',
    'Thematic': 'Thematic%20Indices',
    'Strategy': 'Strategy%20Indices'
}

# Store historical data for comparison
history = {
    'Broad Market': deque(maxlen=5),  # Last 5 readings
    'Sectoral': deque(maxlen=5),
    'Thematic': deque(maxlen=5),
    'Strategy': deque(maxlen=5),
    'Overall': deque(maxlen=5)
}

# Excel logging configuration
EXCEL_FILE = "Lakshmi.xlsx"
SHEET_NAME = "NSE_Indices_Monitor"


def log_to_excel(timestamp, data_snapshot, signal, reason):
    """Log market data to Excel file"""
    try:
        # Check if file exists
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
        else:
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Create or get sheet
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            # Add headers
            headers = ['Timestamp', 'Broad Market %', 'Sectoral %', 'Thematic %', 
                      'Strategy %', 'Overall %', 'Signal', 'Reason']
            ws.append(headers)
        
        # Prepare row data
        row = [
            timestamp,
            data_snapshot.get('Broad Market', 0),
            data_snapshot.get('Sectoral', 0),
            data_snapshot.get('Thematic', 0),
            data_snapshot.get('Strategy', 0),
            data_snapshot.get('Overall', 0),
            signal,
            reason
        ]
        
        ws.append(row)
        
        # Save workbook
        wb.save(EXCEL_FILE)
        print(f"✅ Logged to Excel: {EXCEL_FILE} → {SHEET_NAME}")
        
    except Exception as e:
        print(f"⚠️  Excel logging failed: {e}")


def fetch_indices_data(category_url):
    """Fetch data from NSE API for a specific category"""
    try:
        url = f"{BASE_URL}?type={category_url}"
        response = requests.get(url, headers=HEADERS, timeout=10)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"❌ Error: Status {response.status_code}")
            return None
    except Exception as e:
        print(f"❌ Exception: {e}")
        return None


def calculate_category_average(data):
    """Calculate average percentage change for a category"""
    if not data:
        return None
    
    pchanges = [item['pChange'] for item in data if 'pChange' in item]
    if not pchanges:
        return None
    
    return round(statistics.mean(pchanges), 2)


def analyze_trend(current_avg, history_deque):
    """Analyze trend based on historical data"""
    if not history_deque or len(history_deque) < 2:
        return "NEUTRAL", "📊"
    
    prev_avg = history_deque[-1]
    
    # Calculate momentum
    change = current_avg - prev_avg
    
    # Determine trend
    if change > 0.1:
        return "BULLISH", "🟢"
    elif change < -0.1:
        return "BEARISH", "🔴"
    else:
        return "NEUTRAL", "⚪"


def get_trading_signal(overall_avg, history_deque):
    """Generate trading signal based on overall market direction"""
    if not history_deque or len(history_deque) < 3:
        return "⏳ WAIT", "Collecting data..."
    
    # Calculate momentum over last 3 readings
    recent = list(history_deque)[-3:]
    
    # Check if consistently declining
    if all(recent[i] > recent[i+1] for i in range(len(recent)-1)):
        strength = abs(recent[0] - recent[-1])
        if strength > 0.5:
            return "🔴 BUY PUT", f"Strong bearish trend (-{strength:.2f}%)"
        else:
            return "🟠 BUY PUT", f"Moderate bearish trend (-{strength:.2f}%)"
    
    # Check if consistently rising
    elif all(recent[i] < recent[i+1] for i in range(len(recent)-1)):
        strength = abs(recent[-1] - recent[0])
        if strength > 0.5:
            return "🟢 BUY CALL", f"Strong bullish trend (+{strength:.2f}%)"
        else:  
            return "🟡 BUY CALL", f"Moderate bullish trend (+{strength:.2f}%)"
    
    # Volatile / No clear trend
    else:
        return "⚪ NEUTRAL", "No clear directional bias"


def display_dashboard(data_snapshot, timestamp):
    """Display comprehensive dashboard"""
    print("\n" + "="*80)
    print(f"📊 NSE INDICES MONITOR - {timestamp}")
    print("="*80)
    
    # Display each category
    for category, avg in data_snapshot.items():
        if category == 'Overall':
            continue
            
        trend, emoji = analyze_trend(avg, history[category])
        hist_display = " → ".join([f"{x:.2f}%" for x in list(history[category])[-3:]])
        
        print(f"\n{category:20s} | Avg: {avg:+6.2f}% | {emoji} {trend:8s} | History: [{hist_display}]")
    
    # Overall market summary
    print("\n" + "-"*80)
    overall_avg = data_snapshot['Overall']
    signal, reason = get_trading_signal(overall_avg, history['Overall'])
    hist_display = " → ".join([f"{x:.2f}%" for x in list(history['Overall'])[-5:]])
    
    print(f"{'OVERALL MARKET':20s} | Avg: {overall_avg:+6.2f}% | {signal}")
    print(f"{'Reason':20s} | {reason}")
    print(f"{'History (last 5)':20s} | [{hist_display}]")
    print("="*80)
    
    # Log to Excel
    log_to_excel(timestamp, data_snapshot, signal, reason)


def monitor_indices(interval=60, duration=None):
    """
    Main monitoring loop
    
    Args:
        interval: Update interval in seconds (default: 60)
        duration: Total monitoring duration in seconds (None = infinite)
    """
    print("🚀 Starting NSE Indices Monitor...")
    print(f"⏱️  Update Interval: {interval} seconds")
    print(f"📡 Monitoring 4 categories with {sum(len(v) for v in [fetch_indices_data(u) or [] for u in CATEGORIES.values()])} indices")
    
    start_time = time.time()
    iteration = 0
    
    try:
        while True:
            iteration += 1
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            print(f"\n[{timestamp}] 🔄 Fetching data (Iteration #{iteration})...")
            
            # Fetch data for all categories
            data_snapshot = {}
            
            for category_name, category_url in CATEGORIES.items():
                data = fetch_indices_data(category_url)
                if data:
                    avg = calculate_category_average(data)
                    if avg is not None:
                        data_snapshot[category_name] = avg
                        history[category_name].append(avg)
                        print(f"✅ {category_name}: {avg:+.2f}% ({len(data)} indices)")
                    else:
                        print(f"⚠️  {category_name}: No data")
                else:
                    print(f"❌ {category_name}: Failed to fetch")
                
                time.sleep(1)  # Small delay between API calls
            
            # Calculate overall average
            if data_snapshot:
                overall_avg = round(statistics.mean(data_snapshot.values()), 2)
                data_snapshot['Overall'] = overall_avg
                history['Overall'].append(overall_avg)
            
            # Display dashboard
            if data_snapshot:
                display_dashboard(data_snapshot, timestamp)
            else:
                print("❌ No data available for this iteration")
            
            # Check duration limit
            if duration and (time.time() - start_time) >= duration:
                print(f"\n⏰ Monitoring duration completed ({duration}s)")
                break
            
            # Wait for next iteration
            print(f"\n⏳ Next update in {interval} seconds... (Press Ctrl+C to stop)")
            time.sleep(interval)
    
    except KeyboardInterrupt:
        print("\n\n⛔ Monitoring stopped by user")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n✅ NSE Indices Monitor stopped")


def single_snapshot():
    """Get a single snapshot of current market conditions"""
    print("📸 Taking market snapshot...")
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    data_snapshot = {}
    for category_name, category_url in CATEGORIES.items():
        data = fetch_indices_data(category_url)
        if data:
            avg = calculate_category_average(data)
            if avg is not None:
                data_snapshot[category_name] = avg
        time.sleep(1)
    
    if data_snapshot:
        overall_avg = round(statistics.mean(data_snapshot.values()), 2)
        data_snapshot['Overall'] = overall_avg
        
        print("\n" + "="*80)
        print(f"📊 MARKET SNAPSHOT - {timestamp}")
        print("="*80)
        for category, avg in data_snapshot.items():
            emoji = "🟢" if avg > 0 else "🔴" if avg < 0 else "⚪"
            print(f"{category:20s} | {emoji} {avg:+6.2f}%")
        print("="*80)
    else:
        print("❌ Failed to get market snapshot")


if __name__ == "__main__":
    import sys
    
    print("""
╔══════════════════════════════════════════════════════════════╗
║          NSE INDICES HEATMAP MONITOR v1.0                    ║
║  Real-time market direction tracker for options trading      ║
╚══════════════════════════════════════════════════════════════╝
    """)
    
    if len(sys.argv) > 1 and sys.argv[1] == "snapshot":
        single_snapshot()
    else:
        # Default: Monitor continuously with 60 second intervals
        # You can change interval here or pass as command line argument
        interval = 60  # seconds
        monitor_indices(interval=interval)
