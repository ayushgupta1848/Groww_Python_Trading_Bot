# Groww_CP_Bot.py


############### MUST READ #####################
#• To start Bot we just need to assure that config data is proper,
# Need to remove comments from "place_cp_order" method to validate the order status under comment starts from "#STATUS VALIDATION",
# And in last we need to validate that funds are matching with BOT and Groww Account unnder comment "Funds check"

#DOWNLOAD GROWW INSTRUMENTS:- https://growwapi-assets.groww.in/instruments/instrument.csv

import os
import re
import json
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import pyotp
from openpyxl import Workbook, load_workbook
from playsound3 import playsound
from datetime import datetime, timedelta
from threading import Lock
import requests
import sys
from datetime import datetime
import time
import os
import sys
from datetime import datetime
import numpy as np

# ENHANCEMENT: Use a session for persistent HTTP connections (faster polling)
session = requests.Session()

MOMENTUM_SAMPLES = 3  # Reduced from 5 for faster execution
MOMENTUM_DELAY = 0.5  # Reduced from 1 second

def setup_persistent_logger():
    """Creates a local 'logs' folder beside the script and logs all console output there."""
    # Create /logs folder in the same directory as the script
    base_dir = os.path.dirname(os.path.abspath(__file__))
    log_dir = os.path.join(base_dir, "logs", "groww_bot")
    os.makedirs(log_dir, exist_ok=True)

    # Create a timestamped log file
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_path = os.path.join(log_dir, f"Groww_Bot_{timestamp}.log")

    # Define a Tee class to write to both console and log file,
    # prefixing every new line with a [HH:MM:SS] timestamp.
    class Tee:
        def __init__(self, *streams):
            self.streams = streams
            self._at_line_start = True

        def _stamp(self):
            return f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] "

        def write(self, data):
            if not data:
                return
            out = []
            for ch in data:
                if self._at_line_start and ch != "\n":
                    out.append(self._stamp())
                    self._at_line_start = False
                out.append(ch)
                if ch == "\n":
                    self._at_line_start = True
            stamped = "".join(out)
            for s in self.streams:
                try:
                    s.write(stamped)
                    s.flush()
                except Exception:
                    pass  # Ignore on shutdown

        def flush(self):
            for s in self.streams:
                try:
                    s.flush()
                except Exception:
                    pass

    # Open log file (unbuffered, UTF-8)
    logfile = open(log_path, "a", buffering=1, encoding="utf-8")

    # Redirect both stdout and stderr
    sys.stdout = Tee(sys.stdout, logfile)
    sys.stderr = Tee(sys.stderr, logfile)

    print(f"📝 Logging started. Log file: {log_path}")

    return log_path


# --- Initialize persistent logging ---
LOG_FILE_PATH = setup_persistent_logger()

# Replace with your Groww API key (or leave and use TOTP to fetch access_token)
api_key = "eyJraWQiOiJaTUtjVXciLCJhbGciOiJFUzI1NiJ9.eyJleHAiOjI1NjQ2NTczODEsImlhdCI6MTc3NjI1NzM4MSwibmJmIjoxNzc2MjU3MzgxLCJzdWIiOiJ7XCJ0b2tlblJlZklkXCI6XCJjMjAzMmM5MS04ZGYzLTRkZDUtYjc5NS0yMGVlOWRhZDhhZjlcIixcInZlbmRvckludGVncmF0aW9uS2V5XCI6XCJlMzFmZjIzYjA4NmI0MDZjODg3NGIyZjZkODQ5NTMxM1wiLFwidXNlckFjY291bnRJZFwiOlwiMmVlMjYyMjItN2MwNS00Y2IwLWIwM2MtNzAzYWRmNWVmN2RkXCIsXCJkZXZpY2VJZFwiOlwiNjA2MzE5M2QtZWZkMC01OWViLTgzYzQtNWQ2NGZkNzdkNzQ3XCIsXCJzZXNzaW9uSWRcIjpcIjI0OWQ2OGRlLTNjZTgtNGQ4OS05ODJkLWM0N2NmYmI1YzdlNFwiLFwiYWRkaXRpb25hbERhdGFcIjpcIno1NC9NZzltdjE2WXdmb0gvS0EwYktvMDZXRlpjc241VUNmTWF5aERtNGxSTkczdTlLa2pWZDNoWjU1ZStNZERhWXBOVi9UOUxIRmtQejFFQisybTdRPT1cIixcInJvbGVcIjpcImF1dGgtdG90cFwiLFwic291cmNlSXBBZGRyZXNzXCI6XCIyNDA5OjQwYzQ6MTBhMzozN2UzOjE4NGI6N2IyOTpiMzBlOjIwZTUsMTcyLjcwLjIxOC4xMzUsMzUuMjQxLjIzLjEyM1wiLFwidHdvRmFFeHBpcnlUc1wiOjI1NjQ2NTczODE2ODYsXCJ2ZW5kb3JOYW1lXCI6XCJncm93d0FwaVwifSIsImlzcyI6ImFwZXgtYXV0aC1wcm9kLWFwcCJ9.3kotfZI_EC0lzszHKlXiRdqEQv-O8ubYFh0pgoAT0KsSfdQ1sHmts5UtlaAq4PB6DEwY4X2jZUCD8uBgc2nwXQ"
totp_gen = pyotp.TOTP('SC3YMFLEGLHBWUPHRBOYLPEEOVAT2PZ4')

# Get project root directory (folder where your script is running)
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
# Build CSV path dynamically
csv_path = os.path.join(PROJECT_ROOT, "instrument.csv")
print(csv_path)

# Instruments CSV/JSON path (script will convert CSV -> JSON if convert_csv_to_json = yes)
# csv_path = r"C:\Users\HITS\Downloads\instrument (6).csv"
convert_csv_to_json = "yes"

from whatsapp_gateway import send_whatsapp as send_telegram, start_webhook_server

# Sound files (ensure these exist in script folder or provide full path)
SOUND_PROFIT = "coin.mp3"
SOUND_SL = "SL_HIT.mp3"
SOUND_user_input = "User_input.WAV"

# Runtime quick-mode target override — set by dashboard bridge, consumed by place_quick_order loop
_QUICK_RUNTIME_TARGET: list  = [None]   # [0] = new target pts or None
_QUICK_RUNTIME_PARTIAL: list = [None]   # [0] = {"partial":bool,"partial_pct":int} or None

# Trade defaults for Groww
DEFAULT_PRODUCT = "MIS"   # intraday; change to "NRML" if you want positional
ORDER_PRODUCT_MAP = {
    "MIS": "MIS",
    "NRML": "NRML"
}
# NOTE: the growwapi wrapper constants are used from the growwapi instance below

# ----------------- import growwapi late (after auth) -----------------
try:
    from growwapi import GrowwAPI
except Exception:
    # If local module not available, user must install or place it in PYTHONPATH
    print("❗ growwapi module not found. Make sure it's installed and importable.")
    # continue; import errors will show when script runs further

from groww_token import get_access_token as get_cached_access_token

# ----------------- Groww auth & wrapper -----------------
def groww_init(api_key):
    """
    Return growwapi client instance (GrowwAPI(access_token))
    This function gets access_token using GrowwAPI.get_access_token if available.
    """
    try:
        access_token = get_cached_access_token(api_key, 'SC3YMFLEGLHBWUPHRBOYLPEEOVAT2PZ4')
        client = GrowwAPI(access_token)
        print(access_token)
        print("✅ Groww API Initialized Successfully")
        return client, access_token
    except Exception as e:
        print(f"❌ Groww login failed: {e}")
        raise

# Init groww client
groww ,access_token = groww_init(api_key)

# ── Speed patch: reuse TCP+TLS connection for order API calls ──────────────
# growwapi uses bare requests.post() (new TCP+TLS handshake every call = ~2-3s).
# Patching _request_post to use a persistent session drops order latency to <1s.
_order_session = requests.Session()
_order_session.headers.update({"Connection": "keep-alive"})

def _fast_request_post(url, json=None, headers=None, timeout=None, **kwargs):
    try:
        return _order_session.post(url=url, json=json, headers=headers, timeout=timeout, **kwargs)
    except requests.exceptions.Timeout:
        try:
            from growwapi.groww.exceptions import GrowwAPITimeoutException
            raise GrowwAPITimeoutException()
        except ImportError:
            raise

groww._request_post = _fast_request_post
print("⚡ Order session patched — persistent TCP connection enabled")
# ── End speed patch ────────────────────────────────────────────────────────


# ----------------- Utilities: Sound, Excel Logging -----------------

def play_sound_async(filename):
    try:
        if not os.path.exists(filename):
            print(f"🔇 Sound file not found: {filename}")
            return
        threading.Thread(target=playsound, args=(filename,), daemon=True).start()
    except Exception as e:
        print(f"🔇 Sound error: {e}")

def log_trade_to_excel(symbol, buy_price, sell_price, quantity, profit):
    file_name = "Lakshmi.xlsx"
    mode = "PAPER" if CONFIG.get("PAPER_TRADING", False) else ("MOCK" if CONFIG.get("MOCK_LTP_RUN", False) else "LIVE")

    # Emit machine-parseable record for ANALYZE_BOT
    _rec = {
        "ts":       datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "symbol":   str(symbol or ""),
        "buy_px":   round(float(buy_price),  2) if buy_price  is not None else None,
        "sell_px":  round(float(sell_price), 2) if sell_price is not None else None,
        "qty":      int(quantity) if quantity is not None else None,
        "pnl":      round(float(profit), 2)  if profit    is not None else None,
        "mode":     mode,
    }
    print(f"[TRADE_RECORD] {json.dumps(_rec)}")

    try:
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Lakshmi"
            ws.append(["DateTime", "Symbol", "Buy Price", "Sell Price", "Quantity", "Profit ₹", "Capital Used", "Result", "Mode"])
            wb.save(file_name)

        wb = load_workbook(file_name)
        ws = wb.active

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        capital_used = round(float(buy_price) * int(quantity), 2) if buy_price and quantity else None
        result_label = "PROFIT" if profit >= 0 else "LOSS"

        # Find first truly empty row (guards against ghost rows from Excel formatting)
        next_row = ws.max_row + 1
        while next_row > 2 and ws.cell(row=next_row - 1, column=1).value is None:
            next_row -= 1

        ws.cell(row=next_row, column=1).value = now
        ws.cell(row=next_row, column=2).value = symbol
        ws.cell(row=next_row, column=3).value = buy_price
        ws.cell(row=next_row, column=4).value = sell_price
        ws.cell(row=next_row, column=5).value = quantity
        ws.cell(row=next_row, column=6).value = round(profit, 2)
        ws.cell(row=next_row, column=7).value = capital_used
        ws.cell(row=next_row, column=8).value = result_label
        ws.cell(row=next_row, column=9).value = mode
        wb.save(file_name)
        print(f"📊 Excel logged: {symbol}  {result_label}  ₹{profit:.0f}  row {next_row}")
    except Exception as exc:
        print(f"⚠️  Excel log failed: {exc}")


# ----------------- CSV -> JSON loader -----------------
def csv_to_json(csv_file_path, json_file_path=None):
    """
    Converts CSV to JSON only if JSON doesn't exist or CSV is newer.
    🚀 PERFORMANCE: Skips conversion if JSON is up-to-date
    """
    import csv
    if json_file_path is None:
        json_file_path = os.path.splitext(csv_file_path)[0] + ".json"
    
    # 🚀 Skip conversion if JSON exists and is newer than CSV
    if os.path.exists(json_file_path) and os.path.exists(csv_file_path):
        json_time = os.path.getmtime(json_file_path)
        csv_time = os.path.getmtime(csv_file_path)
        if json_time >= csv_time:
            print(f"⚡ Using existing JSON (up-to-date): '{json_file_path}'")
            with open(json_file_path, 'r', encoding='utf-8') as jf:
                return json.load(jf)
    
    # Convert CSV to JSON
    data = []
    with open(csv_file_path, encoding='utf-8') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            data.append(row)
    with open(json_file_path, 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file, indent=4, ensure_ascii=False)
    print(f"✅ Converted '{csv_file_path}' → '{json_file_path}'")
    return data

ltp_lock = threading.Lock()

# ── Token-bucket rate limiter for Groww Live Data API ──────────────────────
# Groww limit: 10 req/sec burst, 300 req/min (= 5 req/sec avg).
# We cap at 4 req/sec here so the tracker + fib bots can share the budget.
class _RateLimiter:
    def __init__(self, rate: float):
        self._rate   = rate      # tokens/second
        self._tokens = rate      # start full
        self._last   = time.monotonic()
        self._lock   = threading.Lock()

    def acquire(self):
        with self._lock:
            now = time.monotonic()
            self._tokens = min(self._rate, self._tokens + (now - self._last) * self._rate)
            self._last = now
            if self._tokens >= 1.0:
                self._tokens -= 1.0
                return
            wait = (1.0 - self._tokens) / self._rate
        time.sleep(wait)
        with self._lock:
            self._tokens = max(0.0, self._tokens - 1.0)

_live_data_limiter = _RateLimiter(rate=4.0)  # 4 req/sec = 240/min from this bot

def get_ltp_for_instrument(instrument, access_token, verbose=True, segment="FNO", delay=0.05, max_retries=2):
    """
    Fetches the latest traded price (LTP) for a given F&O instrument using Groww's authenticated API.
    Automatically detects exchange (NSE/BSE) from instrument data.
    Thread-safe with a global lock to prevent too-frequent API calls.
    """

    try:
        trading_symbol = instrument.get("trading_symbol")  # e.g. NIFTY25N0425950CE or SENSEX2621283500CE
        if not trading_symbol:
            print("⚠️ Missing trading_symbol in instrument.")
            return None

        # Detect exchange from instrument data
        exchange = instrument.get("exchange", "NSE").upper()  # NSE or BSE
        exchange_symbol = f"{exchange}_{trading_symbol}"
        
        url = f"https://api.groww.in/v1/live-data/ltp?segment={segment}&exchange_symbols={exchange_symbol}"
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0"
        }

        # Throttle to 4 req/sec so this bot stays within Groww's 300 req/min Live Data budget
        _live_data_limiter.acquire()

        # 🔒 Lock ensures one API call at a time
        with ltp_lock:
            resp = session.get(url, headers=headers, timeout=5)  # Use session + reduced timeout
            if delay > 0:
                time.sleep(delay)  # ⏳ short delay to respect Groww API rate limits

        if resp.status_code == 429:
            print(f"⚠️ HTTP 429 error fetching LTP: {resp.text}")
            time.sleep(3)  # back off so the rate-limit ban can expire
            return None
        if resp.status_code != 200:
            print(f"⚠️ HTTP {resp.status_code} error fetching LTP: {resp.text}")
            return None

        data = resp.json()
        payload = data.get("payload", {})
        ltp = payload.get(exchange_symbol)

        if ltp is None:
            print(f"⚠️ No LTP found for {exchange_symbol} in payload: {payload}")
            return None
        if verbose:
            print(f"💰 LTP for {exchange_symbol}: ₹{ltp} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            try:
                send_telegram(f"💰 LTP for {exchange_symbol}: ₹{ltp} ====== [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
            except:
                pass
        return float(ltp)

    except Exception as e:
        print(f"⚠️ Error fetching LTP for {instrument.get('trading_symbol')}: {e}")
        return None

def get_user_positions(access_token):
    """
    Fetches user's current positions from Groww API.
    Returns: dict with positions data or None
    """
    try:
        url = "https://api.groww.in/v1/positions/user"
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0"
        }
        
        resp = session.get(url, headers=headers, timeout=10)
        if resp.status_code != 200:
            print(f"⚠️ HTTP {resp.status_code} error fetching positions: {resp.text}")
            return None
        
        data = resp.json()
        if data.get("status") == "SUCCESS":
            return data.get("payload", {})
        else:
            print(f"⚠️ Failed to fetch positions: {data}")
            return None
            
    except Exception as e:
        print(f"⚠️ Error fetching positions: {e}")
        return None

def get_user_margins(access_token):
    """
    Fetches user's margin details from Groww API.
    Returns: dict with margin data or None
    """
    try:
        url = "https://api.groww.in/v1/margins/detail/user"
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0"
        }
        
        resp = session.get(url, headers=headers, timeout=10)
        if resp.status_code != 200:
            print(f"⚠️ HTTP {resp.status_code} error fetching margins: {resp.text}")
            return None
        
        data = resp.json()
        if data.get("status") == "SUCCESS":
            return data.get("payload", {})
        else:
            print(f"⚠️ Failed to fetch margins: {data}")
            return None
            
    except Exception as e:
        print(f"⚠️ Error fetching margins: {e}")
        return None

def display_account_summary(access_token):
    """
    Fetches and displays total realised P&L and available balance after trade execution.
    Uses clear visual indicators: 🟢 for profit, 🔴 for loss.
    """
    print("\n" + "="*60)
    print("📊 ACCOUNT SUMMARY AFTER TRADE")
    print("="*60)
    
    # Fetch positions for realised P&L
    positions_data = get_user_positions(access_token)
    total_realised_pnl = 0.0
    
    if positions_data and "positions" in positions_data:
        for pos in positions_data["positions"]:
            realised_pnl = pos.get("realised_pnl", 0)
            if realised_pnl:
                total_realised_pnl += float(realised_pnl)
                trading_symbol = pos.get("trading_symbol", "N/A")
                
                # Use clear visual indicators
                if float(realised_pnl) < 0:
                    print(f"  📈 {trading_symbol}: 🔴 LOSS: ₹{realised_pnl} ❌")
                else:
                    print(f"  📈 {trading_symbol}: 🟢 PROFIT: ₹{realised_pnl} ✅")
        
        # Total P&L with clear indicators
        print("\n" + "-"*60)
        if total_realised_pnl < 0:
            print(f"💰 Total Realised P&L: 🔴 LOSS: ₹{total_realised_pnl:.2f} ❌")
        else:
            print(f"💰 Total Realised P&L: 🟢 PROFIT: ₹{total_realised_pnl:.2f} ✅")
        print("-"*60)
    else:
        print("⚠️ Could not fetch positions data")
    
    # Fetch margins for available balance
    margins_data = get_user_margins(access_token)
    
    if margins_data:
        fno_details = margins_data.get("fno_margin_details", {})
        option_buy_balance = fno_details.get("option_buy_balance_available", 0)
        
        print(f"💵 Option Buy Balance Available: ₹{option_buy_balance:.2f}")
        
        # Also show clear cash
        clear_cash = margins_data.get("clear_cash", 0)
        print(f"💸 Clear Cash: ₹{clear_cash:.2f}")
    else:
        print("⚠️ Could not fetch margins data")
    
    print("="*60 + "\n")
    
    # Send to Telegram
    try:
        summary_msg = f"""
📊 ACCOUNT SUMMARY
━━━━━━━━━━━━━━━━
💰 Total Realised P&L: ₹{total_realised_pnl:.2f}
💵 Option Buy Balance: ₹{option_buy_balance:.2f}
💸 Clear Cash: ₹{clear_cash:.2f}
        """
        send_telegram(summary_msg)
    except:
        pass

def get_index_spot_price(index_name, access_token=None, json_path=None):
    """
    Fetches live spot price for any index (NIFTY, SENSEX, BANKNIFTY, etc.) using Groww instrument data.
    For SENSEX, fetches from option chain if spot instrument not available.
    🚀 OPTIMIZED: Uses cached instruments if available
    
    Args:
        index_name: 'NIFTY', 'SENSEX', 'BANKNIFTY', etc.
        access_token: Groww API access token
        json_path: Path to instruments JSON
    
    Returns:
        float: Spot price or 0 if failed
    """
    global instruments1, _instruments_cache
    index_name = index_name.upper()

    # 🚀 Try to use cached instruments first (much faster)
    if _instruments_cache.get("data"):
        # Check if we have instruments for this index in cache
        cached_underlying = _instruments_cache.get("index", "").upper()
        if cached_underlying == index_name or not cached_underlying:
            instruments1 = _instruments_cache["data"]
            # print(f"⚡ Using cached instruments for {index_name} spot lookup")
        else:
            # Need to load fresh for different index
            instruments1 = None
    else:
        instruments1 = None

    # Load from file if not cached
    if not instruments1:
        if json_path is None:
            json_path = os.path.splitext(csv_path)[0] + ".json"

        # 🔄 Load or convert JSON
        if convert_csv_to_json.lower() == "yes":
            instruments1 = csv_to_json(csv_path, json_path)
        else:
            if not os.path.exists(json_path):
                raise FileNotFoundError(f"JSON not found: {json_path}")
            with open(json_path, "r", encoding="utf-8") as jf:
                instruments1 = json.load(jf)
            # print(f"ℹ️ Loaded instruments for spot lookup")

    # Map index names to their various identifiers
    index_mappings = {
        "NIFTY": ["NIFTY", "NSE-NIFTY", "NIFTY 50"],
        "SENSEX": ["SENSEX", "BSE-SENSEX", "SENSEX", "BSE_SENSEX"],
        "BANKNIFTY": ["BANKNIFTY", "NIFTY BANK", "NSE-BANKNIFTY"],
        "FINNIFTY": ["FINNIFTY", "NIFTY FIN SERVICE", "NSE-FINNIFTY"]
    }
    
    search_terms = index_mappings.get(index_name, [index_name])

    try:
        # Special handling for SENSEX - use option chain to get underlying price
        if index_name == "SENSEX":
            print(f"📊 Fetching {index_name} spot from option chain...")
            try:
                # Get current/near expiry from instruments
                sensex_options = [item for item in instruments1 
                                 if item.get("underlying_symbol", "").upper() == "SENSEX" 
                                 and item.get("segment") == "FNO"]
                
                if sensex_options:
                    # Get first available expiry
                    expiry = sensex_options[0].get("expiry_date")
                    url = f"https://api.groww.in/v1/option-chain/exchange/BSE/underlying/SENSEX?expiry_date={expiry}"
                    headers = {
                        "Accept": "application/json",
                        "Authorization": f"Bearer {access_token}",
                        "X-API-VERSION": "1.0"
                    }
                    resp = session.get(url, headers=headers, timeout=8)
                    if resp.status_code == 200:
                        data = resp.json()
                        if data.get("status") == "SUCCESS":
                            underlying_ltp = data.get("payload", {}).get("underlying_ltp")
                            if underlying_ltp:
                                print(f"📊 Live {index_name} Spot (from option chain): {underlying_ltp}")
                                return float(underlying_ltp)
            except Exception as e:
                print(f"⚠️ Could not fetch {index_name} from option chain: {e}")

        # Try standard spot instrument lookup for NSE indices
        spot_instrument = next(
            (item for item in instruments1
             if item.get("trading_symbol", "").upper() in search_terms
             or item.get("groww_symbol", "").upper() in [f"NSE-{t}" for t in search_terms]
             or item.get("groww_symbol", "").upper() in [f"BSE-{t}" for t in search_terms]
             or item.get("name", "").upper() in search_terms),
            None
        )

        if not spot_instrument:
            print(f"⚠️ {index_name} spot instrument not found in instruments")
            return 0

        # SENSEX is on BSE, others typically on NSE
        segment = "CASH"
        spot = get_ltp_for_instrument(spot_instrument, access_token, verbose=False, segment=segment)
        
        if spot:
            print(f"📊 Live {index_name} Spot: {spot}")
            return float(spot)
        else:
            print(f"⚠️ Failed to fetch LTP for {index_name} spot")
            return 0
    except Exception as e:
        print(f"⚠️ Error fetching {index_name} spot: {e}")
        return 0

# Backward compatibility wrapper
def get_nifty_spot_price(access_token=None, json_path=None):
    return get_index_spot_price("NIFTY", access_token, json_path)


CONFIG = {
    "index": "NIFTY",  # Change to "SENSEX", "BANKNIFTY", "FINNIFTY" as needed
    "expiry": "2026-04-28",  #this needs to be same as expiry_date in json file of instruments # format DD/MM/YYYY to match instruments JSON (example)
    "min_premium": 90,
    "max_premium": 230,
    "lots": 16,
    "book_profit": 1050,
    "target_pnl": 6000,
    "spot": 0,  # Will be fetched dynamically below
    "TRAIL_START_PROFIT": 1,  # Start trailing after this profit per unit (in points)
    "TRAIL_STEP": .75,  # Trailing step (in points) — used when TRAIL_SL_ATR_BASED is False
    "TRAIL_SL_ATR_BASED": False,     # If True, trail step = ATR × TRAIL_SL_ATR_MULTIPLIER (adapts to volatility)
    "TRAIL_SL_ATR_MULTIPLIER": 1.0,  # How many ATRs to use as the trail distance (0.5=tight, 1.5=loose)
    "POLL_INTERVAL": 0.50,  # Poll interval in seconds (Optimized for speed)
    "MAX_TRAIL_TIME": 3600,  # Max trailing time in seconds (1 hour)
    "HARD_SL_POINTS": 6.0,        # Hard stop loss points below entry (also acts as floor when ATR-based SL is active)
    "HARD_SL_ATR_MULTIPLIER": 1.5,  # ATR multiplier for HIST ATR SL (5-min ATR × this = raw SL pts, floored at HARD_SL_POINTS)
    "VALIDATE_ORDERS": False,   # ✅ LIVE TRADING: Set True to validate BUY/SELL execution (RECOMMENDED)
    "PAPER_TRADING": True,   # ✅ Set True to simulate trades without placing real orders (safe for testing all modes)
    "QUICK_TRAIL_BUFFER": 1.0,  # DEPRECATED (unused): quick mode is now a hard target — no trailing past target
    "QUICK_TRAIL_GAP": 1.5,     # DEPRECATED (unused): quick mode is now a hard target — no trailing past target
    "user_confirmation_needed": False,   # or False
    "ENABLE_EMA_CHECK": False,
    "ENABLE_ADX_CHECK": False,
    "ENABLE_RSI_CHECK": False,
    "ENABLE_VWAP_CHECK": False,
    "ENABLE_LOGICAL_CONDITIONS_CHECK": False,
    # Directional Mode Settings
    "DIRECTIONAL_MODE": {
        "prefer_mid_premium": True,  # Pick option closest to mid-range of min/max premium (legacy, not used in new format)
    }
}

# 🚀 PERFORMANCE: Global cache to avoid reloading instruments (must be defined before usage)
_instruments_cache = {
    "data": None,
    "index": None,
    "expiry": None,
    "spot_range": None
}

# Dynamically fetch spot price based on selected index
CONFIG["spot"] = get_index_spot_price(CONFIG["index"], access_token)
print(f"🎯 Using {CONFIG['index']} with spot price: {CONFIG['spot']}")

# Load instruments_data
def load_instruments_from_json(json_path=None, force_reload=False):
    """
    Loads instruments from JSON (or CSV → JSON if convert_csv_to_json = 'yes'),
    but only keeps instruments:
      - matching expiry from CONFIG
      - within ±10 strikes of current index spot price
    
    🚀 CACHED: Returns cached data if index/expiry/spot haven't changed
    """
    global instruments, _instruments_cache
    config = CONFIG
    INDEX = config["index"].upper()
    EXPIRY = config["expiry"].strip()
    spot = config["spot"]
    
    # Determine strike step based on index
    if "BANK" in INDEX:
        step = 100
    elif "SENSEX" in INDEX:
        step = 100
    elif "FINNIFTY" in INDEX:
        step = 50
    else:  # NIFTY default
        step = 50
    
    nearest_strike = round(spot / step) * step
    lower_bound = nearest_strike - (10 * step)
    upper_bound = nearest_strike + (10 * step)
    spot_range = (lower_bound, upper_bound)
    
    # 🚀 Check cache first
    if (not force_reload and 
        _instruments_cache["data"] is not None and
        _instruments_cache["index"] == INDEX and
        _instruments_cache["expiry"] == EXPIRY and
        _instruments_cache["spot_range"] == spot_range):
        print(f"⚡ Using cached {INDEX} instruments ({len(_instruments_cache['data'])} loaded)")
        return _instruments_cache["data"]
    
    print(f"💾 Loading instruments from file...")

    if json_path is None:
        json_path = os.path.splitext(csv_path)[0] + ".json"

    # 🔄 Step 1: Load or convert JSON
    if convert_csv_to_json.lower() == "yes":
        instruments = csv_to_json(csv_path, json_path)
    else:
        if not os.path.exists(json_path):
            raise FileNotFoundError(f"JSON not found: {json_path}")
        with open(json_path, "r", encoding="utf-8") as jf:
            instruments = json.load(jf)
        print(f"ℹ️ Loaded instruments from existing JSON: {json_path}")

    # 🧩 Step 2: Get live index spot
    spot = config["spot"]

    # Determine strike step based on index
    # NIFTY = 50, BANKNIFTY = 100, SENSEX = 100, FINNIFTY = 50
    if "BANK" in INDEX:
        step = 100
    elif "SENSEX" in INDEX:
        step = 100
    elif "FINNIFTY" in INDEX:
        step = 50
    else:  # NIFTY default
        step = 50

    # Define strike range (±20 strikes for wider OTM coverage)
    nearest_strike = round(spot / step) * step
    lower_bound = nearest_strike - (20 * step)
    upper_bound = nearest_strike + (20 * step)

    print(f"🎯 Filtering {INDEX} {EXPIRY} instruments between {lower_bound}–{upper_bound} (Spot={spot})")

    # 🧹 Step 3: Filter instruments by expiry and strike range
    filtered = []
    for item in instruments:
        try:
            if item.get("underlying_symbol", "").upper() != INDEX:
                continue
            # if item.get("expiry_date", "").strip() != EXPIRY:
            #     continue
            strike = float(item.get("strike_price") or 0)
            if lower_bound <= strike <= upper_bound:
                filtered.append(item)
        except Exception:
            continue

    print(f"✅ Loaded {len(filtered)} filtered instruments (out of {len(instruments)})")
    instruments = filtered
    
    # 🚀 Save to cache
    _instruments_cache["data"] = instruments
    _instruments_cache["index"] = INDEX
    _instruments_cache["expiry"] = EXPIRY
    _instruments_cache["spot_range"] = spot_range
    
    return instruments


# initialize
instruments_data = load_instruments_from_json()


# Pre-index instruments for quick lookup by internal_trading_symbol or groww_symbol or custom compact symbol
symbol_index = {}
for it in instruments_data:
    # keys: internal_trading_symbol, groww_symbol, compact like NIFTY04NOV2525950CE approximate
    try:
        k1 = it.get("internal_trading_symbol", "") or it.get("trading_symbol", "")
        k2 = it.get("groww_symbol", "")
        if k1:
            symbol_index[k1.upper()] = it
        if k2:
            symbol_index[k2.upper()] = it
    except Exception:
        pass

# ----------------- Helpers: date/expiry normalization -----------------
MONTHS = {
    'JAN': '01', 'JANUARY': '01',
    'FEB': '02', 'FEBRUARY': '02',
    'MAR': '03', 'MARCH': '03',
    'APR': '04', 'APRIL': '04',
    'MAY': '05',
    'JUN': '06', 'JUNE': '06',
    'JUL': '07', 'JULY': '07',
    'AUG': '08', 'AUGUST': '08',
    'SEP': '09', 'SEPTEMBER': '09',
    'OCT': '10', 'OCTOBER': '10',
    'NOV': '11', 'NOVEMBER': '11',
    'DEC': '12', 'DECEMBER': '12'
}

def cmd_expiry_to_date(expiry_token):
    """
    expiry_token example: 04NOV25 or 04NOV2025 or 02MARCH2026 or 28AUG25 or 28AUG2025
    Return string 'DD/MM/YYYY'
    """
    m = re.match(r'(\d{1,2})([A-Z]+)(\d{2,4})', expiry_token.upper())
    if not m:
        return None
    dd = m.group(1).zfill(2)
    mon_abbr = m.group(2)
    yy = m.group(3)
    if len(yy) == 2:
        yyyy = "20" + yy
    else:
        yyyy = yy
    mm = MONTHS.get(mon_abbr, None)
    if not mm:
        return None
    return f"{yyyy}-{mm}-{dd}"


# ----------------- Command parser -----------------
def parse_cp_command(command):
    """
    Parse strings like:
      14 NIFTY30DEC2525950CE
    Returns dict or None
    """
    # Pattern to match: <lots> <TRADING_SYMBOL>
    pattern = r'^\s*(\d+)\s+([A-Z0-9]+)\s*$'
    m = re.match(pattern, command.strip())
    if not m:
        return None

    lots = int(m.group(1))
    trading_symbol_str = m.group(2).upper()

    return {
        "lots": lots,
        "trading_symbol_str": trading_symbol_str,
    }

def parse_trading_symbol_string(trading_symbol_str: str):
    """
    Parses a trading symbol string like 'NIFTY30DEC2525950CE' or 'NIFTY02MARCH202625300PE'
    into its components.
    Returns dict or None.
    """
    # Pattern: UNDERLYING(NIFTY) DAY(30) MONTH(DEC or MARCH) YEAR(25) STRIKE(25950) TYPE(CE)
    pattern = r'([A-Z]+)(\d{1,2}[A-Z]+\d{2,4})(\d+)(CE|PE)'
    m = re.match(pattern, trading_symbol_str)
    if not m:
        print(f"❌ Could not parse trading symbol string: {trading_symbol_str}")
        return None

    underlying = m.group(1).upper()
    expiry_token = m.group(2).upper()
    strike = m.group(3)
    opt_type = m.group(4).upper()
    expiry_date = cmd_expiry_to_date(expiry_token)

    if not expiry_date:
        print(f"❌ Could not derive expiry date from token: {expiry_token}")
        return None

    return {
        "underlying": underlying,
        "expiry_token": expiry_token,
        "expiry_date": expiry_date,
        "strike": strike,
        "opt_type": opt_type
    }

def find_instrument_by_details(underlying, expiry_date, strike, opt_type, instruments: list):
    """
    Finds an instrument in the master list based on parsed details.
    """
    print(f"🔍 Searching for: {underlying} | Expiry: {expiry_date} | Strike: {strike} | Type: {opt_type}")
    print(f"📦 Searching in {len(instruments)} loaded instruments...")
    
    for inst in instruments:
        if (
            inst["underlying_symbol"].upper() == underlying
            and inst["expiry_date"] == expiry_date
            and str(inst["strike_price"]) == strike
            and inst["instrument_type"].upper() == opt_type
        ):
            print(f"✅ Found: {inst.get('trading_symbol')} ({inst.get('groww_symbol')})")
            return inst
    
    # Debug: Show what we have for this underlying
    matching_underlying = [i for i in instruments if i["underlying_symbol"].upper() == underlying]
    if matching_underlying:
        print(f"⚠️ Found {len(matching_underlying)} {underlying} instruments, but none match the criteria")
        # Show a sample
        sample = matching_underlying[0]
        print(f"📋 Sample instrument format: Expiry={sample.get('expiry_date')}, Strike={sample.get('strike_price')}, Type={sample.get('instrument_type')}")
    else:
        print(f"⚠️ No {underlying} instruments found in loaded data")
    
    print(f"❌ Instrument not found: {underlying} {expiry_date} {strike} {opt_type}")
    return None


def calculate_sma(prices, period):
    if len(prices) < period:
        return None
    return np.mean(prices[-period:])


def calculate_ema(prices, period):
    if len(prices) < period:
        return None
    # Initial SMA
    ema = np.mean(prices[:period])
    multiplier = 2 / (period + 1)
    for price in prices[period:]:
        ema = (price - ema) * multiplier + ema
    return ema


def calculate_rsi(prices, period=14):
    prices = np.array(prices)
    if len(prices) < period + 1:
        return 50  # Default neutral

    deltas = np.diff(prices)
    gains = np.maximum(deltas, 0)
    losses = -np.minimum(deltas, 0)

    avg_gain = np.mean(gains[:period])
    avg_loss = np.mean(losses[:period])

    if avg_loss == 0:
        return 100

    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))

    # Wilder's Smoothing
    for i in range(period, len(deltas)):
        avg_gain = (avg_gain * (period - 1) + gains[i]) / period
        avg_loss = (avg_loss * (period - 1) + losses[i]) / period

        if avg_loss == 0:
            rsi = 100
        else:
            rs = avg_gain / avg_loss
            rsi = 100 - (100 / (1 + rs))

    return rsi


def calculate_adx(highs, lows, closes, period=14):
    if len(highs) < period * 2:
        return 25  # Default

    highs = np.array(highs)
    lows = np.array(lows)
    closes = np.array(closes)

    tr = np.zeros(len(highs))
    plus_dm = np.zeros(len(highs))
    minus_dm = np.zeros(len(highs))

    for i in range(1, len(highs)):
        tr[i] = max(highs[i] - lows[i], abs(highs[i] - closes[i - 1]), abs(lows[i] - closes[i - 1]))

        up_move = highs[i] - highs[i - 1]
        down_move = lows[i - 1] - lows[i]

        if up_move > down_move and up_move > 0:
            plus_dm[i] = up_move
        else:
            plus_dm[i] = 0

        if down_move > up_move and down_move > 0:
            minus_dm[i] = down_move
        else:
            minus_dm[i] = 0

    # Smoothing
    def smooth(data, period):
        smoothed = np.zeros_like(data)
        if len(data) > period:
            smoothed[period] = np.mean(data[1:period + 1])  # Initial SMA
            for i in range(period + 1, len(data)):
                smoothed[i] = (smoothed[i - 1] * (period - 1) + data[i]) / period
        return smoothed

    atr = smooth(tr, period)
    plus_di = 100 * smooth(plus_dm, period) / (atr + 1e-9)  # Avoid div by zero
    minus_di = 100 * smooth(minus_dm, period) / (atr + 1e-9)

    dx = 100 * np.abs(plus_di - minus_di) / (plus_di + minus_di + 1e-9)
    adx = smooth(dx, period)

    return adx[-1]


def calculate_vwap(prices, volumes):
    prices = np.array(prices)
    volumes = np.array(volumes)
    if len(prices) == 0 or len(volumes) == 0 or None in volumes:
        return None

    vwap = np.cumsum(prices * volumes) / np.cumsum(volumes)
    return vwap[-1]

def calculate_atr(high, low, close, period=14):
    """Calculates the Average True Range (ATR)."""
    if len(high) < period:
        return 0
    high = np.array(high)
    low = np.array(low)
    close = np.array(close)
    tr1 = high - low
    tr2 = np.abs(high - np.roll(close, 1))
    tr3 = np.abs(low - np.roll(close, 1))
    tr = np.amax((tr1, tr2, tr3), axis=0)
    atr = calculate_ema(tr, period)  # Use EMA for smoothing ATR
    return atr if atr is not None else 0


def get_technicals(symbol, groww_client, interval="1minute", segment="FNO", timeout=5, instrument=None, lookback_minutes=60):
    """
    Fetch technical indicators with timeout protection (optimized to 5s).
    Returns None if API call fails or times out.
    lookback_minutes: how far back to fetch candles (increase for wider intervals, e.g. 150 for 5-min)
    """
    try:
        # Detect exchange dynamically (BSE for SENSEX, NSE for others)
        if instrument:
            exch_str = instrument.get("exchange", "NSE").upper()
            exchange_const = groww_client.EXCHANGE_BSE if exch_str == "BSE" else groww_client.EXCHANGE_NSE
        else:
            exchange_const = groww_client.EXCHANGE_NSE

        end_time = datetime.now()
        start_time = end_time - timedelta(minutes=lookback_minutes)

        end_str   = end_time.strftime("%Y-%m-%d %H:%M:%S")
        start_str = start_time.strftime("%Y-%m-%d %H:%M:%S")

        print(f"🔄 Fetching historical candles for {symbol}...")
        
        # Add timeout protection using threading
        import signal
        
        def timeout_handler(signum, frame):
            raise TimeoutError("Historical candles fetch timed out")
        
        # Set timeout alarm (Unix-based systems only)
        try:
            signal.signal(signal.SIGALRM, timeout_handler)
            signal.alarm(timeout)
        except:
            pass  # Windows doesn't support signal.SIGALRM
        
        try:
            historical = groww_client.get_historical_candles(
                groww_symbol=symbol,
                exchange=exchange_const,
                segment=segment,
                start_time=start_str,
                end_time=end_str,
                candle_interval=interval
            )
        finally:
            try:
                signal.alarm(0)  # Cancel alarm
            except:
                pass

        if not historical:
            print("⚠️ No historical data returned")
            return None
            
        candles = historical.get("candles", [])
        if not candles or len(candles) < 20:  # Reduced from 30 for faster processing
            print(f"⚠️ Insufficient candles: {len(candles) if candles else 0}")
            return None

        print(f"✅ Fetched {len(candles)} candles")
        
        # Groww candles: [timestamp, open, high, low, close, volume]
        opens = [c[1] for c in candles]
        highs = [c[2] for c in candles]
        lows = [c[3] for c in candles]
        close_prices = [c[4] for c in candles]

        vwap = None
        # VWAP is only applicable for instruments with volume
        if segment == "FNO":
            volumes = [c[5] for c in candles]
            vwap = calculate_vwap(close_prices, volumes)

        # If VWAP couldn't be calculated, use last price as a fallback
        if vwap is None:
            vwap = close_prices[-1]

        sma_20 = calculate_sma(close_prices, 20)
        ema_9 = calculate_ema(close_prices, 9)
        rsi_14 = calculate_rsi(close_prices, 14)
        adx_14 = calculate_adx(highs, lows, close_prices, 14)
        atr = calculate_atr(highs, lows, close_prices)


        current_price = close_prices[-1]

        return {
            "sma_20": sma_20,
            "ema_9": ema_9,
            "rsi": rsi_14,
            "adx": adx_14,
            "vwap": vwap,
            "ltp": current_price,
            "atr": atr
        }
    except Exception as e:
        print(f"⚠️ Error fetching technicals: {e}")
        return None

# --- START: Caching layer to prevent API rate limiting ---
_option_chain_cache = {}
_option_chain_cache_lock = threading.Lock()
CACHE_EXPIRY_SECONDS = 15  # Cache for 15 seconds
_last_api_call_time = 0
_api_call_lock = threading.Lock()

def _get_full_option_chain_cached(underlying, expiry_date, access_token):
    """Cached option chain fetcher with rate limit handling"""
    global _last_api_call_time
    cache_key = (underlying, expiry_date)
    now = time.time()

    # Return cached if fresh
    cached_payload, timestamp = _option_chain_cache.get(cache_key, (None, 0))
    if cached_payload and (now - timestamp) < CACHE_EXPIRY_SECONDS:
        return cached_payload

    with _option_chain_cache_lock:
        # Double-check after lock
        cached_payload, timestamp = _option_chain_cache.get(cache_key, (None, 0))
        if cached_payload and (time.time() - timestamp) < CACHE_EXPIRY_SECONDS:
            return cached_payload

        url = f"https://api.groww.in/v1/option-chain/exchange/NSE/underlying/{underlying}?expiry_date={expiry_date}"
        headers = {
            "Accept": "application/json",
            "Authorization": f"Bearer {access_token}",
            "X-API-VERSION": "1.0"
        }

        with _api_call_lock:
            time_since_last = time.time() - _last_api_call_time
            if time_since_last < 0.2:
                time.sleep(0.2 - time_since_last)
            
            resp = session.get(url, headers=headers, timeout=8)
            _last_api_call_time = time.time()

        if resp.status_code != 200:
            raise Exception(f"HTTP {resp.status_code}: {resp.text}")

        data = resp.json()
        if data.get("status") != "SUCCESS":
            raise Exception("Failed to fetch option chain")

        payload = data["payload"]
        _option_chain_cache[cache_key] = (payload, time.time())
        return payload

# --- END: Caching layer ---


def get_option_data_from_trading_symbol(
    trading_symbol: str,
    exchange: str = "NSE",
    underlying: str = "NIFTY"
):
    """
    Fetch delta, theta, OI, LTP, IV, volume etc. for a given trading_symbol
    using a cached Groww Option Chain API call
    """
    expiry_date = CONFIG["expiry"].strip()

    try:
        payload = _get_full_option_chain_cached(underlying, expiry_date, access_token)
    except Exception as e:
        raise e

    strikes = payload["strikes"]
    underlying_ltp = payload["underlying_ltp"]

    # 🔍 Find this trading_symbol in option chain
    for strike_str, opt_data in strikes.items():
        for opt_type in ("CE", "PE"):
            opt = opt_data.get(opt_type)
            if not opt:
                continue

            if opt.get("trading_symbol") == trading_symbol:
                greeks = opt.get("greeks", {})

                return {
                    "trading_symbol": trading_symbol,
                    "option_type": opt_type,
                    "strike": int(strike_str),
                    "expiry": expiry_date,
                    "ltp": opt.get("ltp"),
                    "open_interest": opt.get("open_interest"),
                    "volume": opt.get("volume"),
                    "delta": greeks.get("delta"),
                    "theta": greeks.get("theta"),
                    "iv": greeks.get("iv"),
                    "gamma": greeks.get("gamma"),
                    "vega": greeks.get("vega"),
                    "rho": greeks.get("rho"),
                    "underlying_ltp": underlying_ltp
                }

    raise ValueError(f"{trading_symbol} not found in option chain")

# ----------------- Prefetcher (background) -----------------
def _option_chain_prefetcher_loop():
    """Daemon loop to keep option-chain cache warm"""
    cfg = CONFIG
    underlying = cfg.get("index", "NIFTY")
    expiry = cfg.get("expiry")
    interval = 10  # Prefetch every 10 seconds

    while True:
        try:
            _get_full_option_chain_cached(underlying, expiry, access_token)
            time.sleep(interval)
        except Exception as e:
            print(f"⚠️ Prefetcher error: {e}")
            time.sleep(30)

def start_option_chain_prefetcher():
    t = threading.Thread(target=_option_chain_prefetcher_loop, daemon=True, name="OptionChainPrefetcher")
    t.start()

# Start background prefetcher
try:
    start_option_chain_prefetcher()
    print("🔁 Option-chain prefetcher started.")
except Exception as e:
    print(f"⚠️ Failed to start prefetcher: {e}")

# ----------------- End prefetcher -----------------


# ----------------- Paper trading state -----------------
_paper_orders = {}       # fake_order_id -> {"price": float, "qty": int, "symbol": str, "type": str}
_paper_order_counter = [0]

def _paper_order_id():
    _paper_order_counter[0] += 1
    return f"PAPER_{_paper_order_counter[0]:04d}"

# ----------------- Place orders with Groww -----------------
def place_market_order_groww(instrument, quantity, transaction_type="BUY", product="MIS"):
    """
    place market order via growwapi wrapper. Returns order response or raises.
    In PAPER_TRADING mode, simulates the order without hitting the exchange.
    """
    trading_symbol = instrument.get("internal_trading_symbol") or instrument.get("trading_symbol")

    if CONFIG.get("PAPER_TRADING", False):
        ltp = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0) or 0.0
        fake_id = _paper_order_id()
        _paper_orders[fake_id] = {"price": float(ltp), "qty": quantity, "symbol": trading_symbol, "type": transaction_type}
        print(f"📋 [PAPER] MARKET {transaction_type} {quantity} × {trading_symbol} @ ₹{ltp:.2f} | ID: {fake_id}")
        send_telegram(f"📋 [PAPER] {transaction_type} {quantity} × {trading_symbol} @ ₹{ltp:.2f}")
        return {"payload": {"groww_order_id": fake_id}}

    # Detect exchange dynamically (BSE for SENSEX, NSE for others)
    exch_str = instrument.get("exchange", "NSE").upper()
    exchange_const = groww.EXCHANGE_BSE if exch_str == "BSE" else groww.EXCHANGE_NSE
    try:
        order = groww.place_order(
            trading_symbol=trading_symbol,
            quantity=quantity,
            validity=groww.VALIDITY_DAY,
            exchange=exchange_const,
            segment=groww.SEGMENT_FNO,
            product=getattr(groww, f"PRODUCT_{product}") if hasattr(groww, f"PRODUCT_{product}") else getattr(groww, "PRODUCT_MIS", product),
            order_type=groww.ORDER_TYPE_MARKET,
            transaction_type=getattr(groww, f"TRANSACTION_TYPE_{transaction_type}"),
            price=0
        )
        return order
    except Exception as e:
        raise

def place_limit_order_groww(instrument, quantity, price, transaction_type="SELL", product="MIS"):
    trading_symbol = instrument.get("internal_trading_symbol") or instrument.get("trading_symbol")

    if CONFIG.get("PAPER_TRADING", False):
        fake_id = _paper_order_id()
        _paper_orders[fake_id] = {"price": float(price), "qty": quantity, "symbol": trading_symbol, "type": transaction_type}
        print(f"📋 [PAPER] LIMIT {transaction_type} {quantity} × {trading_symbol} @ ₹{price:.2f} | ID: {fake_id}")
        return {"payload": {"groww_order_id": fake_id}}

    # Detect exchange dynamically (BSE for SENSEX, NSE for others)
    exch_str = instrument.get("exchange", "NSE").upper()
    exchange_const = groww.EXCHANGE_BSE if exch_str == "BSE" else groww.EXCHANGE_NSE
    try:
        order = groww.place_order(
            trading_symbol=trading_symbol,
            quantity=quantity,
            validity=groww.VALIDITY_DAY,
            exchange=exchange_const,
            segment=groww.SEGMENT_FNO,
            product=getattr(groww, f"PRODUCT_{product}") if hasattr(groww, f"PRODUCT_{product}") else getattr(groww, "PRODUCT_MIS", product),
            order_type=groww.ORDER_TYPE_LIMIT,
            transaction_type=getattr(groww, f"TRANSACTION_TYPE_{transaction_type}"),
            price=price
        )
        return order
    except Exception as e:
        raise

def cancel_order_groww(order_id, access_token):
    """
    Cancel a pending order using Groww API
    Returns True if cancellation successful, False otherwise
    """
    if CONFIG.get("PAPER_TRADING", False):
        _paper_orders.pop(str(order_id), None)
        print(f"📋 [PAPER] Order {order_id} cancelled")
        return True

    url = "https://api.groww.in/v1/order/cancel"
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0"
    }
    
    payload = {
        "segment": "FNO",
        "groww_order_id": order_id
    }
    
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=8)
        resp.raise_for_status()
        data = resp.json()
        print(f"🔄 Cancel order response: {data}")
        
        # Check if cancellation was successful
        if data.get("success") or data.get("payload", {}).get("order_status") == "CANCELLED":
            return True
        return False
    except Exception as e:
        print(f"⚠️ Error cancelling order {order_id}: {e}")
        return False

# ----------------- Rounding for limits (5 paise) -----------------
def round_to_nearest_5_paise(price):
    # Round to nearest 0.05
    return round(round(price * 20) / 20, 2)

# ----------------- Momentum sampling -----------------

def momentum_check_for_symbol(instrument, MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY, threshold=0.25):
    """
    Improved short-term momentum check for Nifty options.
    - Uses multiple intermediate samples
    - Smooths noise
    - Checks direction consistency
    - Returns a cleaner momentum signal
    """
    trading_symbol = instrument.get("trading_symbol")
    prices = []

    print(f"\n🧭 Checking momentum for {trading_symbol} ({MOMENTUM_SAMPLES} samples, every {MOMENTUM_DELAY}s):")

    for i in range(MOMENTUM_SAMPLES):
        p = get_ltp_for_instrument(instrument, access_token, verbose=False)
        if p:
            price = float(p)
            prices.append(price)
            print(f"[{trading_symbol}] ⏱ Sample {i+1}/{MOMENTUM_SAMPLES}: LTP = ₹{price:.2f}")
        else:
            print(f"[{trading_symbol}] ⚠️ Sample {i+1}/{MOMENTUM_SAMPLES}: Failed to fetch LTP")
        time.sleep(MOMENTUM_DELAY)

    if len(prices) < 3:
        print(f"[{trading_symbol}] ❌ Not enough data ({len(prices)} samples)")
        return None, len(prices)

    prices = np.array(prices)

    # 1️⃣ Smooth noise with small moving average
    smooth = np.convolve(prices, np.ones(3)/3, mode="valid")

    # 2️⃣ Compute rate of change (%)
    roc = np.diff(smooth) / smooth[:-1] * 100

    # 3️⃣ Remove outliers (big spikes)
    median = np.median(roc)
    std = np.std(roc)
    filtered = roc[(roc > median - 1.5*std) & (roc < median + 1.5*std)]

    if len(filtered) < 2:
        print(f"[{trading_symbol}] ⚠️ Too noisy for reliable momentum reading")
        return None, len(prices)

    # 4️⃣ Average change and direction consistency
    avg_change = np.mean(filtered)
    direction_signs = np.sign(filtered)
    consistency = np.mean(direction_signs == np.sign(avg_change)) * 100

    # 5️⃣ Decision
    if avg_change > threshold and consistency > 70:
        direction = "UP"
    elif avg_change < -threshold and consistency > 70:
        direction = "DOWN"
    else:
        direction = "FLAT"

    print(f"[{trading_symbol}] 📊 Avg Δ = {avg_change:.3f}%, Consistency = {consistency:.1f}% → {direction}")
    print(f"[{trading_symbol}] 📈 Range ₹{prices[0]:.2f} → ₹{prices[-1]:.2f}\n")

    return {"symbol": trading_symbol,
            "avg_change": round(avg_change, 3),
            "consistency": round(consistency, 1),
            "direction": direction}, len(prices)



# ----------------- Find option by premium (parallel) -----------------

def find_option_by_premium_parallel(option_type, min_premium, max_premium,
                                    lots=1, funds_buffer=0.9, momentum_threshold_pct=0.25,
                                    MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY):
    """
    Filters instruments using INDEX and EXPIRY from config,
    matches by option_type, filters by premium range,
    and runs momentum checks in parallel.
    Returns: (instrument, ltp, lot_size) or (None, None, None)
    """
    config = CONFIG
    INDEX = config["index"].upper()
    EXPIRY = config["expiry"].strip()
    candidates = []

    # 🔍 Filter based on index + expiry + type
    for item in instruments_data:
        try:
            if item.get("underlying_symbol", "").upper() != INDEX:
                continue
            if item.get("instrument_type", "").upper() != option_type.upper():
                continue
            if item.get("expiry_date", "").strip() != EXPIRY:
                continue

            lot_size = int(item.get("lot_size") or item.get("lotsize") or 1)
            ltp = get_ltp_for_instrument(item, access_token, verbose=False)
            if ltp is None:
                continue
            if not (min_premium <= ltp <= max_premium):
                continue

            candidates.append({
                "instrument": item,
                "ltp": float(ltp),
                "lot_size": lot_size
            })

        except Exception as e:
            print(f"⚠️ Error while scanning: {e}")
            continue

    if not candidates:
        print(f"⚠️ No instruments for {INDEX} {EXPIRY} {option_type}")
        return None, None, None

    # ✅ Funds check (fallback = 1.2L if not available)
    try:
        margins = getattr(groww, "get_margins", lambda: {"availablecash": 130000})()
        available_cash = float(margins.get("availablecash", 130000))
    except Exception:
        available_cash = 130000

    affordable = []
    for c in candidates:
        qty = lots * c["lot_size"]
        est_cost = c["ltp"] * qty
        if available_cash <= 0 or est_cost <= available_cash * funds_buffer:
            affordable.append(c)

    if not affordable:
        print(f"⚠️ No affordable instruments for {INDEX} {EXPIRY} {option_type}")
        return None, None, None

    if option_type.upper() == "PE":
        momentum_threshold_pct = 0.30  # PEs move sharper
    else:
        momentum_threshold_pct = 0.25

    # ✅ Sort candidates closest to mid-premium
    mid = (min_premium + max_premium) / 2.0
    affordable.sort(key=lambda x: abs(x["ltp"] - mid))
    probe_list = affordable[:12]

    momentum_results = []

    def check_momentum(cand):
        mom_result, ticks = momentum_check_for_symbol(
            cand["instrument"],
            MOMENTUM_SAMPLES=MOMENTUM_SAMPLES,
            MOMENTUM_DELAY=MOMENTUM_DELAY
        )
        if mom_result and isinstance(mom_result, dict):
            slope_pct = mom_result.get("avg_change", 0)
            direction = mom_result.get("direction", "FLAT")
            consistency = mom_result.get("consistency", 0)

            # ✅ Apply momentum filter right here
            if abs(slope_pct) >= momentum_threshold_pct and consistency >= 70 and direction != "FLAT":
                return {
                    "instrument": cand["instrument"],
                    "ltp": cand["ltp"],
                    "lot_size": cand["lot_size"],
                    "slope_pct": slope_pct,
                    "direction": direction,
                    "consistency": consistency,
                    "ticks": ticks
                }
        return None

    print(f"⚙️ Checking momentum for top {len(probe_list)} {option_type} candidates...")

    with ThreadPoolExecutor(max_workers=min(len(probe_list), 8)) as executor:
        futures = {executor.submit(check_momentum, c): c for c in probe_list}
        for future in as_completed(futures):
            res = future.result()
            if res:
                momentum_results.append(res)

    if not momentum_results:
        print(f"⚠️ No strong momentum found for {option_type} (>{momentum_threshold_pct}%, consistency >70%)")
        # fallback: pick the one closest to mid-premium
        pick = probe_list[0]
        return pick["instrument"], pick["ltp"], pick["lot_size"]

    # ✅ Rank: strongest slope first, then consistency
    momentum_results.sort(key=lambda x: (x["slope_pct"], x["consistency"]), reverse=True)
    pick = momentum_results[0]

    print(f"🏆 Selected {option_type}: {pick['instrument']['trading_symbol']} "
          f"({pick['direction']} | {pick['slope_pct']:.2f}% | Consistency {pick['consistency']}%)")

    return pick["instrument"], pick["ltp"], pick["lot_size"]


# ----------------- Detect CE/PE (parallel) -----------------
def detect_option_type_parallel(index, expiry, min_p, max_p, lots, funds_buffer=0.9):
    print(f"🔍 Detecting best option between CE and PE for {index} {expiry}…")

    def worker(opt_type):
        print(f"➡️  Searching {opt_type} between {min_p}-{max_p}")
        inst, ltp, lot_size = find_option_by_premium_parallel(opt_type, min_p, max_p, lots, funds_buffer)
        mom = None
        if inst:
            print(f"📊 Running momentum check for {opt_type} ({inst.get('trading_symbol')})")
            mom, _ = momentum_check_for_symbol(inst, MOMENTUM_SAMPLES=MOMENTUM_SAMPLES, MOMENTUM_DELAY=MOMENTUM_DELAY)
            print(f"✅ Momentum for {opt_type}: {mom}")
        else:
            print(f"⚠️ No instrument found for {opt_type}")
        return opt_type, inst, ltp, lot_size, mom

    results = {}
    with ThreadPoolExecutor(max_workers=2) as ex:
        futures = {ex.submit(worker, t): t for t in ["CE", "PE"]}
        for future in as_completed(futures):
            opt_type, inst, ltp, lot_size, mom = future.result()
            results[opt_type] = {"instrument": inst, "ltp": ltp, "lot_size": lot_size, "momentum": mom}
            print(f"🧩 Finished {opt_type}: {inst.get('trading_symbol') if inst else 'None'}, momentum={mom}")

    print("🧮 Comparing CE vs PE momentum...")
    ce_mom = results.get("CE", {}).get("momentum")
    pe_mom = results.get("PE", {}).get("momentum")

    # Handle missing momentum
    if not ce_mom and not pe_mom:
        print("❌ No momentum data found for CE or PE.")
        return None
    if not ce_mom:
        r = results["PE"]
        return "PE", r["instrument"], r["ltp"], r["lot_size"]
    if not pe_mom:
        r = results["CE"]
        return "CE", r["instrument"], r["ltp"], r["lot_size"]

    ce_val = ce_mom["avg_change"]
    pe_val = pe_mom["avg_change"]

    print(f"📈 CE momentum: {ce_val:.3f}% ({ce_mom['direction']}, {ce_mom['consistency']}%)")
    print(f"📉 PE momentum: {pe_val:.3f}% ({pe_mom['direction']}, {pe_mom['consistency']}%)")

    # selection logic
    if abs(ce_val - pe_val) >= 0.25 and ce_val > pe_val and ce_val >= 0.10:
        print("✅ Selected CE (stronger momentum)")
        r = results["CE"]
        return "CE", r["instrument"], r["ltp"], r["lot_size"]
    if abs(pe_val - ce_val) >= 0.25 and pe_val > ce_val and pe_val >= 0.10:
        print("✅ Selected PE (stronger momentum)")
        r = results["PE"]
        return "PE", r["instrument"], r["ltp"], r["lot_size"]

    # fallback
    if ce_val >= pe_val:
        print("⚖️  Momentum similar — choosing CE fallback")
        r = results["CE"]
        return "CE", r["instrument"], r["ltp"], r["lot_size"]
    else:
        print("⚖️  Momentum similar — choosing PE fallback")
        r = results["PE"]
        return "PE", r["instrument"], r["ltp"], r["lot_size"]

def get_order_status(order_id, access_token):
    """
    Fetch the status of a Groww order (CASH, F&O, etc.)
    Works with official Groww REST API response format.
    """
    url = f"https://api.groww.in/v1/order/status/{order_id}?segment=FNO"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0"
    }

    try:
        resp = requests.get(url, headers=headers, timeout=8)
        resp.raise_for_status()  # raises for non-200 responses

        data = resp.json()  # Proper JSON response from Groww
        print("🔍 Order status response:", data)

        # ✅ Extract status cleanly
        payload = data.get("payload", {})
        status = payload.get("order_status")

        return status

    except requests.exceptions.JSONDecodeError:
        print("⚠️ Error: Non-JSON response received.")
        print("Response text:", resp.text)
        return None

    except Exception as e:
        print(f"⚠️ Error fetching order status: {e}")
        return None



def wait_for_order_status(order_id, access_token, order_type="BUY"):
    """
    Wait indefinitely until a Groww order reaches EXECUTED / COMPLETED / DELIVERY_AWAITED.
    Returns final status (string).
    """
    if CONFIG.get("PAPER_TRADING", False):
        print(f"📋 [PAPER] {order_type} order {order_id} → EXECUTED (simulated)")
        return "EXECUTED"

    print(f"🔎 Waiting for {order_type} order ({order_id}) to finish...")

    while True:
        status = get_order_status(order_id, access_token)
        print(f"🕒 {order_type} status: {status}")

        if status in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
            print(f"✅ {order_type} order executed successfully.")
            send_telegram(f"✅ {order_type} order executed successfully.")
            return status

        elif status in ["FAILED", "REJECTED", "CANCELLED"]:
            print(f"❌ {order_type} order failed with status {status}.")
            send_telegram(f"❌ {order_type} order failed ({status}).")
            return status

        # ⚡ Fast polling for buy orders (0.2s), slower for sell (1s)
        time.sleep(0.2 if order_type == "BUY" else 1.0)


import requests

def get_order_executed_price(order_id, access_token, segment="FNO"):
    """
    Fetch executed trades for a given Groww order_id and return average price & total quantity.
    """
    if CONFIG.get("PAPER_TRADING", False):
        info = _paper_orders.get(str(order_id), {})
        price = info.get("price", 0.0)
        qty = info.get("qty", 0)
        print(f"📋 [PAPER] Executed: ₹{price:.2f} × {qty} (order {order_id})")
        return float(price), int(qty)

    url = f"https://api.groww.in/v1/order/trades/{order_id}?segment={segment}&page=0&page_size=50"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {access_token}",
        "X-API-VERSION": "1.0"
    }

    print(f"\n📦 Fetching trade details for order: {order_id}")

    # The trades endpoint lags order status by a few hundred ms (eventually consistent).
    # First attempt fires immediately — retries/sleeps happen only when no trades came back.
    _backoff = (0, 0.25, 0.5, 1.0, 1.25)
    for _attempt, _delay in enumerate(_backoff, start=1):
        if _delay:
            time.sleep(_delay)
        try:
            response = requests.get(url, headers=headers, timeout=5)
            data = response.json()
        except Exception as e:
            print(f"⚠️ Trades fetch attempt {_attempt}/{len(_backoff)} error: {e}")
            continue

        if data.get("status") != "SUCCESS":
            print(f"⚠️ Trades fetch attempt {_attempt}/{len(_backoff)} failed:", data)
            continue

        trades = data.get("payload", {}).get("trade_list", [])
        if not trades:
            print(f"⚠️ No trades yet for order ID (attempt {_attempt}/{len(_backoff)}).")
            continue

        # Compute average price & total quantity
        total_qty = sum(t["quantity"] for t in trades)
        total_value = sum(t["price"] * t["quantity"] for t in trades)
        avg_price = round(total_value / total_qty, 2)

        symbol = trades[0]["trading_symbol"]
        side = trades[0]["transaction_type"]

        print(f"✅ {side} {symbol} | Total Qty={total_qty} | Avg Price=₹{avg_price}")
        return avg_price, total_qty

    # Fallback: the order-status payload carries average_price for executed orders
    try:
        status_url = f"https://api.groww.in/v1/order/status/{order_id}?segment={segment}"
        s_payload = requests.get(status_url, headers=headers, timeout=5).json().get("payload", {}) or {}
        avg_price = s_payload.get("average_price") or s_payload.get("avg_price") or 0
        qty = s_payload.get("filled_quantity") or s_payload.get("quantity") or 0
        if avg_price and qty:
            avg_price = round(float(avg_price), 2)
            qty = int(qty)
            print(f"✅ Fallback via order-status: Qty={qty} | Avg Price=₹{avg_price}")
            return avg_price, qty
        print("⚠️ Order-status fallback had no average_price:", s_payload)
    except Exception as e:
        print("❌ Order-status fallback failed:", e)

    return None, None



# ----------------- ATR-based trail step helper -----------------
def _resolve_trail_step(atr_value):
    """
    Returns the effective trail step:
    - If TRAIL_SL_ATR_BASED is True and a valid ATR is supplied, returns ATR × multiplier.
    - Otherwise falls back to the fixed CONFIG["TRAIL_STEP"].
    """
    cfg = CONFIG
    if cfg.get("TRAIL_SL_ATR_BASED", False) and atr_value and atr_value > 0:
        step = round(float(atr_value) * cfg.get("TRAIL_SL_ATR_MULTIPLIER", 1.0), 2)
        print(f"📐 ATR trail step: ₹{step:.2f}  (ATR={atr_value:.2f} × {cfg.get('TRAIL_SL_ATR_MULTIPLIER', 1.0)})")
        return step
    return cfg["TRAIL_STEP"]

def _fetch_atr_sync(instrument, timeout=3):
    """Fetches ATR for an instrument in a background thread. Returns None on timeout/failure."""
    import queue as _queue
    q = _queue.Queue()
    def _worker():
        try:
            techs = get_technicals(instrument["groww_symbol"], groww, segment="FNO", instrument=instrument)
            q.put(techs.get("atr") if techs else None)
        except Exception:
            q.put(None)
    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    t.join(timeout=timeout)
    return q.get() if not q.empty() else None

# ----------------- Place CP order workflow (mirrors AngelOne logic) -----------------
def place_quick_order(command, atr_based=True, quick_pts=1.5, atr_source="candle", partial=False, partial_pct=50, ltp_hint=0):
    """Quick mode: Buy at market, set limit sell at +quick_pts, ATR-based hard SL + full trail.
    partial=True: sell partial_pct% of qty when price reverses from 60%-of-target sub-peak."""
    global buy_status, instruments_data, CONFIG
    
    parsed_command = parse_cp_command(command)
    if not parsed_command:
        print("❌ Invalid command format. Expected: <lots> <TRADING_SYMBOL>")
        return

    lots = parsed_command["lots"]
    trading_symbol_str = parsed_command["trading_symbol_str"]

    parsed_symbol_details = parse_trading_symbol_string(trading_symbol_str)
    if not parsed_symbol_details:
        return

    requested_index = parsed_symbol_details["underlying"]
    current_index = CONFIG.get("index", "").upper()
    
    if requested_index != current_index:
        print(f"🔄 Detected index change: {current_index} → {requested_index}")
        print(f"📦 Reloading instruments for {requested_index}...")
        CONFIG["index"] = requested_index
        CONFIG["expiry"] = parsed_symbol_details["expiry_date"]
        CONFIG["spot"] = get_index_spot_price(requested_index, access_token)
        instruments_data = load_instruments_from_json()
        print(f"✅ Switched to {requested_index} | Spot: {CONFIG['spot']}")

    instrument = find_instrument_by_details(
        parsed_symbol_details["underlying"],
        parsed_symbol_details["expiry_date"],
        parsed_symbol_details["strike"],
        parsed_symbol_details["opt_type"],
        instruments_data
    )
    if not instrument:
        return

    lot_size = int(instrument.get("lot_size") or instrument.get("lotsize") or 1)
    quantity = lots * lot_size

    _atr_src_label = ('HIST ATR' if atr_source=='candle' else 'TICK RNG') if atr_based else 'OFF'

    # Use dashboard chain LTP as reference (no pre-fetch needed — market order fills at exchange price)
    _ref_ltp = round(float(ltp_hint), 2) if ltp_hint and ltp_hint > 0 else None
    if _ref_ltp:
        print(f"⚡ QUICK MODE: Ref LTP=₹{_ref_ltp} (from dashboard) | +{quick_pts}pt target  ATR-SL={_atr_src_label}")
    else:
        print(f"⚡ QUICK MODE: Placing MARKET order immediately | +{quick_pts}pt target  ATR-SL={_atr_src_label}")

    # Place BUY order immediately — no LTP pre-fetch, target set from actual fill price
    try:
        order_resp = place_market_order_groww(instrument, quantity, transaction_type="BUY", product="MIS")
        order_id = order_resp.get("payload", {}).get("groww_order_id") or order_resp.get("groww_order_id")
        print(f"✅ Buy Order placed:", order_resp, {datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
    except Exception as e:
        print(f"❌ Buy order failed: {e}")
        send_telegram(f"❌ Buy order failed: {e}")
        return

    # Get actual fill price — use for target (more accurate than pre-order LTP estimate)
    avg_price = None
    if CONFIG.get("VALIDATE_ORDERS", True):
        if not order_id:
            print("❌ No BUY order ID received.")
            return
        buy_status = wait_for_order_status(order_id, access_token, "BUY")
        if buy_status not in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
            print(f"⚠️ BUY failed: {buy_status}")
            send_telegram(f"⚠️ BUY failed: {buy_status}")
            return
        avg_price, executed_qty = get_order_executed_price(order_id, access_token)
        if not avg_price or not executed_qty:
            # BUY is confirmed EXECUTED — never abandon the position. Estimate entry from LTP.
            _est = _ref_ltp or get_ltp_for_instrument(instrument, access_token, verbose=True, delay=0)
            if not _est:
                print(f"🚨 BUY {order_id} EXECUTED but no price available — POSITION OPEN & UNMANAGED. Exit manually!")
                send_telegram(f"🚨 BUY {order_id} EXECUTED but price unavailable — POSITION OPEN & UNMANAGED. Exit manually NOW!")
                return
            avg_price = round(float(_est), 2)
            executed_qty = quantity
            print(f"⚠️ Avg price unavailable for BUY {order_id} — using LTP estimate ₹{avg_price}; managing position.")
            send_telegram(f"⚠️ BUY {order_id} executed; avg price unavailable — using LTP ₹{avg_price} as entry estimate. Position is being managed.")
        quantity = executed_qty
        target_price = round_to_nearest_5_paise(avg_price + quick_pts)
        print(f"🎯 BUY EXECUTED @ ₹{avg_price} | Target: ₹{target_price} (+{quick_pts}pt)")
    else:
        # No-validate mode: use dashboard LTP or fallback to a quick post-order fetch
        if _ref_ltp:
            avg_price = _ref_ltp
            print(f"⚡ No-validate: using dashboard ref LTP ₹{avg_price} as entry estimate")
        else:
            _fetched = get_ltp_for_instrument(instrument, access_token, verbose=True, delay=0)
            avg_price = round(float(_fetched), 2) if _fetched else 0
            if not avg_price:
                print("⚠️ Could not determine entry price — using 0 as fallback")
        target_price = round_to_nearest_5_paise(avg_price + quick_pts)
        print(f"⚠️ No-validate mode: entry estimate ₹{avg_price}, target ₹{target_price}")

    send_telegram(f"⚡ QUICK BUY PLACED: fill≈₹{avg_price} | Target: ₹{target_price} | {instrument.get('internal_trading_symbol')} | qty={quantity}")

    # ── SL calculation — supports HIST ATR (candle) and TICK RNG (scan) ──
    atr_val     = 0.0
    hard_sl_pts = CONFIG.get("HARD_SL_POINTS", 5)
    mult        = CONFIG.get("HARD_SL_ATR_MULTIPLIER", 1.5)

    if atr_based:
        if atr_source == "scan":
            # ── TICK RNG: sample LTP for ~8 s, compute high-low range ────────
            print(f"📡 TICK RNG: sampling LTP for ~8 seconds…")
            import queue as _q
            _scan_q = _q.Queue()
            def _do_tick_scan():
                _ticks = []
                for _ in range(8):
                    try:
                        _v = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
                        if _v: _ticks.append(float(_v))
                    except Exception:
                        pass
                    time.sleep(1)
                _scan_q.put(_ticks)
            threading.Thread(target=_do_tick_scan, daemon=True).start()
            try:
                _ticks = _scan_q.get(timeout=10)
                if len(_ticks) >= 2:
                    atr_val = round(max(_ticks) - min(_ticks), 2)
                    print(f"✅ Tick range sampled: {atr_val:.2f}  (hi={max(_ticks):.2f} lo={min(_ticks):.2f} n={len(_ticks)})")
            except Exception:
                pass
            if atr_val > 0:
                raw = round(atr_val * mult, 2)
                sl_pts = max(3.0, raw)   # 3-pt floor for tick range (same as MOMENTUM_AUTO_BOT)
                sl_price = round(avg_price - sl_pts, 2)
                print(f"🛡️ TICK RNG Hard SL: ₹{sl_price}  (range={atr_val:.2f} × {mult:.1f}={raw:.2f}, floor=3.0)")
            else:
                sl_price = round(avg_price - hard_sl_pts, 2)
                print(f"🛡️ Tick range unavailable → Fixed Hard SL: ₹{sl_price}  ({hard_sl_pts}pts)")
        else:
            # ── HIST ATR: 14-period EMA ATR from 5-min candles (150 min lookback = 30 candles) ──
            try:
                from threading import Thread
                import queue
                result_queue = queue.Queue()
                def fetch_technicals():
                    try:
                        techs = get_technicals(instrument['groww_symbol'], groww, segment="FNO", instrument=instrument,
                                               interval="5minute", lookback_minutes=150)
                        result_queue.put(techs)
                    except Exception:
                        result_queue.put(None)
                thread = Thread(target=fetch_technicals, daemon=True)
                thread.start()
                thread.join(timeout=5)
                if not result_queue.empty():
                    techs = result_queue.get()
                    if techs and techs.get("atr"):
                        atr_val = float(techs["atr"])
                        print(f"✅ Hist ATR fetched (5-min): {atr_val:.2f}")
            except Exception:
                pass
            if atr_val > 0:
                raw_sl = mult * atr_val
                sl_pts = max(hard_sl_pts, raw_sl)  # floor at HARD_SL_POINTS so SL is never tighter than fixed fallback
                sl_price = round(avg_price - sl_pts, 2)
                print(f"🛡️ HIST ATR Hard SL: ₹{sl_price}  ({mult:.1f} × ATR {atr_val:.2f} = {raw_sl:.2f}, floor={hard_sl_pts})")
            else:
                sl_price = round(avg_price - hard_sl_pts, 2)
                print(f"🛡️ Hist ATR unavailable → Fixed Hard SL: ₹{sl_price}  ({hard_sl_pts}pts)")
    else:
        # ATR-SL OFF — use fixed points directly
        sl_price = round(avg_price - hard_sl_pts, 2)
        print(f"🛡️ Fixed Hard SL: ₹{sl_price}  ({hard_sl_pts}pts, ATR-SL OFF)")

    # Place LIMIT SELL order instantly at target
    try:
        sell_resp = place_limit_order_groww(instrument, quantity, target_price, transaction_type="SELL", product="MIS")
        sell_order_id = sell_resp.get("payload", {}).get("groww_order_id") or sell_resp.get("groww_order_id")
        print(f"✅ LIMIT SELL placed @ ₹{target_price}:", sell_resp)
        send_telegram(f"🎯 LIMIT SELL @ ₹{target_price} | Order ID: {sell_order_id}")
        
        # Validate SELL order placement if flag is true
        if CONFIG.get("VALIDATE_ORDERS", True) and sell_order_id:
            print(f"🔎 Validating SELL order placement...")
            initial_status = get_order_status(sell_order_id, access_token)
            print(f"📋 SELL order status: {initial_status}")
            if initial_status in ["FAILED", "REJECTED", "CANCELLED"]:
                print(f"❌ SELL order placement failed: {initial_status}")
                send_telegram(f"❌ SELL order failed: {initial_status}")
                return
    except Exception as e:
        print(f"❌ Limit SELL order failed: {e}")
        send_telegram(f"❌ Limit SELL failed: {e}")
        return

    # Reversal step for PARTIAL profit exits (ATR-based or CONFIG["TRAIL_STEP"]).
    # Quick target mode is a HARD TARGET: the resting limit sell caps the exit at
    # target and is the guaranteed exit even if this bot dies. There is no trailing
    # past target — a resting limit sell fills the instant price touches target, so
    # price can never run above it while the order is open.
    trail_gap  = _resolve_trail_step(atr_val)
    _trail_src = ('ATR×'+str(CONFIG.get('TRAIL_SL_ATR_MULTIPLIER',1.0))) if CONFIG.get('TRAIL_SL_ATR_BASED') and atr_val>0 else 'fixed CONFIG[TRAIL_STEP]'
    print(f"📐 Partial reversal step: ₹{trail_gap:.2f}  ({_trail_src})")

    # Monitor price until target or SL is hit
    print(f"⏳ Monitoring price... Target: ₹{target_price} | SL: ₹{sl_price}")
    start_time = time.time()
    max_monitor_time = 3600  # 1 hour max

    limit_sell_alive = True   # False while the resting limit sell is temporarily cancelled (partial exit)

    # ── Partial profit state ─────────────────────────────────────────────
    lot_size         = int(instrument.get("lot_size", 1))
    remaining_qty    = quantity
    partial_booked   = False    # True after partial exit executed
    partial_sub_peak = 0.0      # tracks local high once partial zone is entered
    # Partial trigger = 60% of target pts from avg; on reversal drop by trail_gap → sell
    partial_trigger_pts  = round(quick_pts * 0.60, 2)
    partial_trigger_lvl  = round(avg_price + partial_trigger_pts, 2)
    partial_qty_raw      = max(lot_size, round(quantity * (partial_pct / 100) / lot_size) * lot_size)
    partial_qty          = partial_qty_raw if partial_qty_raw < quantity else 0  # 0 = disabled
    if partial:
        print(f"📊 PARTIAL mode ON | trigger≥₹{partial_trigger_lvl} (+{partial_trigger_pts}pt) "
              f"| sell {partial_pct}% = {partial_qty} qty | drop = {trail_gap}pt")

    while True:
        if time.time() - start_time >= max_monitor_time:
            print("⏰ Max monitoring time reached (1 hour)")
            break

        try:
            ltp = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
            if ltp is None:
                time.sleep(1)
                continue
            ltp = float(ltp)

            # ── RUNTIME PARTIAL UPDATE (dashboard SET button) ───────────────────
            if _QUICK_RUNTIME_PARTIAL[0] is not None and not partial_booked:
                _rp = _QUICK_RUNTIME_PARTIAL[0]
                _QUICK_RUNTIME_PARTIAL[0] = None
                partial     = bool(_rp.get("partial", False))
                partial_pct = int(_rp.get("partial_pct", 50))
                partial_qty_raw = max(lot_size, round(quantity * (partial_pct / 100) / lot_size) * lot_size)
                partial_qty = partial_qty_raw if partial_qty_raw < quantity else 0
                print(f"\n📊 [RUNTIME] Partial updated → {'ON' if partial else 'OFF'} {partial_pct}% = {partial_qty} qty")
                if partial:
                    print(f"   trigger≥₹{partial_trigger_lvl} | drop={trail_gap}pt")

            # ── PARTIAL PROFIT EXIT ──────────────────────────────────────────────
            if partial and not partial_booked and limit_sell_alive and partial_qty > 0:
                if ltp >= partial_trigger_lvl:
                    # Price in partial zone — track local peak
                    if ltp > partial_sub_peak:
                        partial_sub_peak = ltp
                    elif partial_sub_peak > 0 and ltp <= partial_sub_peak - trail_gap:
                        # Reversal confirmed — execute partial exit
                        print(f"\n📊 PARTIAL EXIT triggered | LTP ₹{ltp} dropped {trail_gap}pt from peak ₹{partial_sub_peak}")
                        send_telegram(f"📊 PARTIAL EXIT @ ₹{ltp} | peak was ₹{partial_sub_peak:.2f} | selling {partial_qty} qty")
                        try:
                            # 1. Cancel existing full-qty limit sell
                            if sell_order_id:
                                _cs = get_order_status(sell_order_id, access_token) if CONFIG.get("VALIDATE_ORDERS", True) else None
                                if _cs in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                                    print(f"ℹ️  Limit sell already filled before partial exit — skipping")
                                    partial_booked = True
                                    break
                                cancel_order_groww(sell_order_id, access_token)
                                limit_sell_alive = False

                            # 2. Market sell partial qty
                            _psell = place_market_order_groww(instrument, partial_qty, "SELL", "MIS")
                            _psell_id = _psell.get("payload", {}).get("groww_order_id") or _psell.get("groww_order_id")
                            _partial_sell_price = ltp
                            if CONFIG.get("VALIDATE_ORDERS", True) and _psell_id:
                                _pst = wait_for_order_status(_psell_id, access_token, "SELL")
                                if _pst in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                                    _pp, _pq = get_order_executed_price(_psell_id, access_token)
                                    if _pp: _partial_sell_price = _pp
                            _partial_pnl = (_partial_sell_price - avg_price) * partial_qty
                            print(f"💰 PARTIAL PROFIT: ₹{_partial_pnl:.2f} | {partial_qty} qty @ ₹{_partial_sell_price:.2f}")
                            send_telegram(f"💰 PARTIAL PROFIT ₹{_partial_pnl:.2f} | {partial_qty}×₹{_partial_sell_price:.2f}")
                            log_trade_to_excel(instrument.get('internal_trading_symbol') or instrument.get('trading_symbol'),
                                               avg_price, _partial_sell_price, partial_qty, _partial_pnl)

                            # 3. Update remaining qty, raise SL to partial trigger, replace limit sell
                            remaining_qty = quantity - partial_qty
                            sl_price = max(sl_price, partial_trigger_lvl)   # raise SL floor
                            print(f"🔼 SL raised to ₹{sl_price} (partial trigger level) | remaining qty: {remaining_qty}")
                            _new_sell = place_limit_order_groww(instrument, remaining_qty, target_price, transaction_type="SELL", product="MIS")
                            sell_order_id  = _new_sell.get("payload", {}).get("groww_order_id") or _new_sell.get("groww_order_id")
                            limit_sell_alive = True
                            quantity       = remaining_qty  # update qty for trail exits
                            partial_booked = True
                            print(f"✅ New limit sell ({remaining_qty} qty) @ ₹{target_price} placed")
                        except Exception as _pe:
                            print(f"❌ Partial exit failed: {_pe}")

            # ── RUNTIME TARGET UPDATE (dashboard "SET" button mid-trade) ────────
            if limit_sell_alive and _QUICK_RUNTIME_TARGET[0] is not None:
                _new_tgt_pts = float(_QUICK_RUNTIME_TARGET[0])
                _QUICK_RUNTIME_TARGET[0] = None   # consume immediately
                _new_target = round_to_nearest_5_paise(avg_price + _new_tgt_pts)
                if _new_target != target_price:
                    print(f"\n🔄 [RUNTIME] Target updated: ₹{target_price} → ₹{_new_target} (+{_new_tgt_pts}pt)")
                    # Cancel old limit sell, replace with new target
                    _cancelled = False
                    if sell_order_id:
                        try:
                            if CONFIG.get("VALIDATE_ORDERS", True):
                                _cur_st = get_order_status(sell_order_id, access_token)
                                if _cur_st in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                                    print(f"⚠️  Limit sell already filled at old target ₹{target_price} — ignoring update")
                                    continue
                            cancel_order_groww(sell_order_id, access_token)
                            _cancelled = True
                        except Exception as _ce:
                            print(f"⚠️  Cancel for target update failed: {_ce}")
                    if _cancelled or not sell_order_id:
                        target_price   = _new_target
                        try:
                            _new_sell_resp  = place_limit_order_groww(instrument, quantity, target_price, transaction_type="SELL", product="MIS")
                            sell_order_id   = _new_sell_resp.get("payload", {}).get("groww_order_id") or _new_sell_resp.get("groww_order_id")
                            send_telegram(f"🔄 Target updated → ₹{target_price} (+{_new_tgt_pts}pt)")
                            print(f"✅ New LIMIT SELL @ ₹{target_price} placed")
                        except Exception as _ne:
                            print(f"❌ New limit sell failed after target update: {_ne}")

            # ── SL CHECK ─────────────────────────────────────────────────────
            if ltp <= sl_price:
                print(f"🛑 SL HIT! LTP: ₹{ltp}")
                send_telegram(f"🛑 SL HIT @ ₹{ltp}")
                play_sound_async(SOUND_SL)
                try:
                    print(f"🔄 Cancelling target order and placing market SELL...")
                    if sell_order_id and limit_sell_alive:
                        cancel_success = cancel_order_groww(sell_order_id, access_token)
                        if cancel_success:
                            print(f"✅ Target order {sell_order_id} cancelled successfully")
                            send_telegram(f"✅ Target order cancelled")
                        else:
                            print(f"⚠️ Could not cancel target order {sell_order_id}, it may have already executed")
                    market_sell_resp = place_market_order_groww(instrument, quantity, "SELL", "MIS")
                    market_sell_id = market_sell_resp.get("payload", {}).get("groww_order_id") or market_sell_resp.get("groww_order_id")
                    print(f"✅ Market SELL placed: {market_sell_id}")
                    if CONFIG.get("VALIDATE_ORDERS", True) and market_sell_id:
                        final_status = wait_for_order_status(market_sell_id, access_token, "SELL")
                        if final_status in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                            sell_price, sold_qty = get_order_executed_price(market_sell_id, access_token)
                            if sell_price and sold_qty:
                                loss = (sell_price - avg_price) * sold_qty
                                print(f"💸 LOSS: ₹{loss:.2f} (Buy @ ₹{avg_price}, Sell @ ₹{sell_price})")
                                send_telegram(f"💸 LOSS: ₹{loss:.2f}")
                                log_trade_to_excel(instrument.get('internal_trading_symbol') or instrument.get('trading_symbol'), avg_price, sell_price, sold_qty, loss)
                                break
                    loss = (ltp - avg_price) * quantity
                    print(f"💸 Estimated LOSS: ₹{loss:.2f}")
                    log_trade_to_excel(instrument.get('internal_trading_symbol') or instrument.get('trading_symbol'), avg_price, ltp, quantity, loss)
                except Exception as e:
                    print(f"❌ SL execution failed: {e}")
                break

            # ── TARGET HIT (resting limit sell fills at target — hard cap) ───
            if ltp >= target_price:
                print(f"🎯 TARGET HIT! LTP: ₹{ltp}")
                send_telegram(f"🎯 TARGET HIT @ ₹{ltp}")
                play_sound_async(SOUND_PROFIT)
                if CONFIG.get("VALIDATE_ORDERS", True) and sell_order_id:
                    final_status = wait_for_order_status(sell_order_id, access_token, "SELL")
                    if final_status in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                        sell_price, sold_qty = get_order_executed_price(sell_order_id, access_token)
                        if sell_price and sold_qty:
                            profit = (sell_price - avg_price) * sold_qty
                            print(f"💰 PROFIT: ₹{profit:.2f} (Buy @ ₹{avg_price}, Sell @ ₹{sell_price})")
                            send_telegram(f"💰 PROFIT: ₹{profit:.2f}")
                            log_trade_to_excel(instrument.get('internal_trading_symbol') or instrument.get('trading_symbol'), avg_price, sell_price, sold_qty, profit)
                        else:
                            profit = (target_price - avg_price) * quantity
                            print(f"💰 Estimated PROFIT: ₹{profit:.2f}")
                            log_trade_to_excel(instrument.get('internal_trading_symbol') or instrument.get('trading_symbol'), avg_price, target_price, quantity, profit)
                else:
                    profit = (target_price - avg_price) * quantity
                    print(f"💰 Estimated PROFIT: ₹{profit:.2f}")
                    log_trade_to_excel(instrument.get('internal_trading_symbol') or instrument.get('trading_symbol'), avg_price, target_price, quantity, profit)
                break

            time.sleep(1)  # Poll every second

        except Exception as e:
            print(f"⚠️ Monitoring error: {e}")
            time.sleep(2)
    
    print("✅ Quick order complete. Ready for next command.")


def place_cp_order(command, is_auto=False):
    global buy_status, instruments_data, CONFIG
    
    # Start timing for manual mode
    command_start_time = datetime.now()
    print(f"[{command_start_time.strftime('%H:%M:%S.%f')[:-3]}] ⏱️  Command entered: {command}")
    
    if is_auto:
        print("Auto mode not supported in this bot, only manual mode")
    else:
        parsed_command = parse_cp_command(command)
        if not parsed_command:
            print("❌ Invalid command format. Expected: <lots> <TRADING_SYMBOL>")
            return

        lots = parsed_command["lots"]
        trading_symbol_str = parsed_command["trading_symbol_str"]
        
        print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 🔍 Parsing symbol: {trading_symbol_str}")

        parsed_symbol_details = parse_trading_symbol_string(trading_symbol_str)
        if not parsed_symbol_details:
            return # Error message already printed by parse_trading_symbol_string

        # 🔄 Auto-detect if index changed and reload instruments if needed
        requested_index = parsed_symbol_details["underlying"]
        current_index = CONFIG.get("index", "").upper()
        
        if requested_index != current_index:
            print(f"🔄 Detected index change: {current_index} → {requested_index}")
            print(f"📦 Reloading instruments for {requested_index}...")
            
            # Update CONFIG
            CONFIG["index"] = requested_index
            CONFIG["expiry"] = parsed_symbol_details["expiry_date"]
            CONFIG["spot"] = get_index_spot_price(requested_index, access_token)
            
            # Reload instruments with new index
            instruments_data = load_instruments_from_json()
            print(f"✅ Switched to {requested_index} | Spot: {CONFIG['spot']}")

        instrument = find_instrument_by_details(
            parsed_symbol_details["underlying"],
            parsed_symbol_details["expiry_date"],
            parsed_symbol_details["strike"],
            parsed_symbol_details["opt_type"],
            instruments_data
        )
        if not instrument:
            return # Error message already printed by find_instrument_by_details

        lot_size = int(instrument.get("lot_size") or instrument.get("lotsize") or 1)
        quantity = lots * lot_size
        
        # ⚡ Try option-chain cache first (instant) — prefetcher keeps it warm
        _ltp_cache_hit = None
        try:
            _oc_payload, _oc_ts = _option_chain_cache.get(
                (parsed_symbol_details["underlying"], parsed_symbol_details["expiry_date"]), (None, 0))
            if _oc_payload and (time.time() - _oc_ts) < 10:
                _sd = _oc_payload.get("strikes", {}).get(str(int(float(parsed_symbol_details["strike"]))), {})
                _lv = float(_sd.get(parsed_symbol_details["opt_type"], {}).get("ltp", 0) or 0)
                if _lv > 0:
                    _ltp_cache_hit = _lv
        except Exception:
            pass

        if _ltp_cache_hit:
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 💰 LTP from cache: ₹{_ltp_cache_hit}")
            ltp_before = _ltp_cache_hit
        else:
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 💰 Fetching LTP (cache miss)...")
            ltp_before = get_ltp_for_instrument(instrument, access_token, verbose=True, delay=0)
            if ltp_before is None:
                print("❌ Could not fetch LTP before placing order.")
                return

        entry_price = round(float(ltp_before), 2)
        print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 💵 Entry price: ₹{entry_price}")

        # ⚡ PLACE BUY ORDER IMMEDIATELY (no delays before this)
        try:
            order_start = datetime.now()
            print(f"[{order_start.strftime('%H:%M:%S.%f')[:-3]}] 🔄 Placing BUY order for {quantity} units...")
            order_resp = place_market_order_groww(instrument, quantity, transaction_type="BUY", product="MIS")
            print(f"🔍 DEBUG order_resp keys: {list(order_resp.keys()) if isinstance(order_resp, dict) else order_resp}")
            order_id = order_resp.get("payload", {}).get("groww_order_id") or order_resp.get("groww_order_id")
            order_placed = datetime.now()
            order_duration = (order_placed - order_start).total_seconds()
            print(f"[{order_placed.strftime('%H:%M:%S.%f')[:-3]}] ✅ BUY Order placed: {order_id} (took {order_duration:.2f}s)")
            send_telegram(f"entry price: {entry_price} | {instrument.get('internal_trading_symbol')} | qty={quantity}")
        except Exception as e:
            print(f"❌ Buy order failed: {e}")
            send_telegram(f"❌ Buy order failed: {e}")
            return

        # Fetch technicals for ATR in background (non-blocking)
        atr = CONFIG["HARD_SL_POINTS"]  # Default fallback
        try:
            from threading import Thread
            import queue
            
            result_queue = queue.Queue()
            
            def fetch_technicals():
                try:
                    techs = get_technicals(instrument['groww_symbol'], groww, segment="FNO", instrument=instrument)
                    result_queue.put(techs)
                except Exception as e:
                    result_queue.put(None)
            
            # Start background fetch (don't wait)
            thread = Thread(target=fetch_technicals, daemon=True)
            thread.start()
        except Exception as e:
            pass

        # STATUS VALIDATION
        # ✅ LIVE TRADING: Wait until BUY order is EXECUTED (controlled by VALIDATE_ORDERS)
        if CONFIG.get("VALIDATE_ORDERS", True):
            if order_id:
                validation_start = datetime.now()
                buy_status = wait_for_order_status(order_id, access_token, "BUY")
                if buy_status not in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                    print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] ⚠️ Skipping trade monitoring due to BUY status: {buy_status}")
                    send_telegram(f"⚠️ BUY failed: {buy_status}")
                    return
            else:
                print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] ❌ No BUY order ID received. Aborting trade.")
                send_telegram("❌ No BUY order ID received")
                return
            
            # Fetch actual executed price and quantity
            avg_price, executed_qty = get_order_executed_price(order_id, access_token)
            if not avg_price or not executed_qty:
                # BUY is confirmed EXECUTED — never abandon the position. Use the
                # pre-order LTP as the entry estimate so target/SL management still runs.
                avg_price = entry_price
                executed_qty = quantity
                print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] ⚠️ Avg price unavailable for BUY {order_id} — using pre-order LTP ₹{entry_price}; managing position.")
                send_telegram(f"⚠️ BUY {order_id} executed; avg price unavailable — using entry estimate ₹{entry_price}. Position is being managed.")
            quantity = executed_qty # Use the actual executed quantity
            validation_end = datetime.now()
            validation_duration = (validation_end - validation_start).total_seconds()
            total_duration = (validation_end - command_start_time).total_seconds()
            
            print(f"[{validation_end.strftime('%H:%M:%S.%f')[:-3]}] 🎯 Executed avg price: ₹{avg_price}, Qty: {quantity}")
            print(f"[{validation_end.strftime('%H:%M:%S.%f')[:-3]}] ⏱️  Total time: {total_duration:.2f}s (Order: {order_duration:.2f}s, Validation: {validation_duration:.2f}s)")
            send_telegram(f"🎯 BUY EXECUTED @ ₹{avg_price} | Qty={quantity}")
        else:
            # Testing mode: Use entry price estimate
            avg_price = entry_price
            executed_qty = quantity
            print(f"⚠️ Testing mode: Using entry price estimate ₹{avg_price}")

        # Check if ATR fetch completed in background
        try:
            if not result_queue.empty():
                techs = result_queue.get()
                if techs and techs.get("atr"):
                    atr = techs["atr"]
                    print(f"✅ ATR fetched: {atr:.2f}")
        except:
            pass

        highest_price = avg_price
        start_time = time.time()
        last_trail_exit = None  # Track last printed trail exit to avoid spam
        last_heartbeat = time.time()  # Track heartbeat for alive signal

        trail_start = CONFIG["TRAIL_START_PROFIT"]
        trail_step = _resolve_trail_step(atr)  # ATR-based if flag set, else fixed TRAIL_STEP
        poll = CONFIG["POLL_INTERVAL"]
        max_time = CONFIG["MAX_TRAIL_TIME"]
        hard_sl = entry_price - (1.5 * atr)  # Dynamic SL based on 1.5 * ATR

        print(f"📈 Trailing started... Dynamic SL: {hard_sl:.2f} (based on ATR: {atr:.2f})")
        try:
            send_telegram(f"📈 Trailing started... Dynamic SL: {hard_sl:.2f}")
        except:
            pass

        # ⚡ CRITICAL: This loop is optimized for ZERO-DELAY sell execution
        # - LTP fetch: delay=0 (no artificial delays)
        # - Poll interval: 0.15s (fast response)
        # - Telegram: wrapped in try-except (won't block)
        # - Sell order: placed IMMEDIATELY when condition met (no pre-checks)
        # ── Mock LTP simulation (used when CONFIG["MOCK_LTP_RUN"] is True) ──────
        _mock_tick   = 0
        _mock_entry  = avg_price
        def _next_mock_ltp():
            nonlocal _mock_tick
            t = _mock_tick; _mock_tick += 1
            # t 0-2 : stable at entry (trail not yet active)
            # t 3-6 : rise +5 pts → trail activates, trail_exit = entry+5-trail_step
            # t 7-8 : rise +8 pts → trail_exit = entry+8-trail_step
            # t 9+  : drop to entry+3 → below trail_exit → TRAIL HIT
            if   t <  3: offset = 0.0
            elif t <  7: offset = 5.0
            elif t <  9: offset = 8.0
            else:         offset = 3.0
            v = round(_mock_entry + offset, 2)
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 🎭 MOCK LTP tick={t}: ₹{v}")
            time.sleep(1.0)   # 1s between ticks — makes timing visible
            return v
        # ─────────────────────────────────────────────────────────────────────

        _ltp_fail_streak = 0
        while True:
            # Heartbeat every 30 seconds to show script is alive
            if time.time() - last_heartbeat >= 30:
                print(f"💓 Monitoring... LTP last seen: ₹{ltp if 'ltp' in locals() else 'fetching...'}")
                last_heartbeat = time.time()

            if CONFIG.get("MOCK_LTP_RUN"):
                ltp = _next_mock_ltp()
            else:
                try:
                    # ⚡ Zero delay for fastest trailing response
                    ltp = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
                except Exception as e:
                    _ltp_fail_streak += 1
                    _backoff = min(30, poll * (2 ** min(_ltp_fail_streak, 6)))
                    if _ltp_fail_streak <= 3 or _ltp_fail_streak % 10 == 0:
                        print(f"⚠️ LTP fetch error (streak={_ltp_fail_streak}, retry in {_backoff:.1f}s): {e}")
                    time.sleep(_backoff)
                    continue

                if ltp is None:
                    _ltp_fail_streak += 1
                    _backoff = min(30, poll * (2 ** min(_ltp_fail_streak, 6)))
                    if _ltp_fail_streak <= 3 or _ltp_fail_streak % 10 == 0:
                        print(f"⚠️ LTP is None (streak={_ltp_fail_streak}, retry in {_backoff:.1f}s)")
                    time.sleep(_backoff)
                    continue
                _ltp_fail_streak = 0

            ltp = float(ltp)
            sell_reason = None

            # 1. Check exit conditions
            if ltp <= hard_sl:
                sell_reason = f"🛑 DYNAMIC SL HIT @ {ltp}"
            elif time.time() - start_time >= max_time:
                sell_reason = "⏰ Max trail time reached — exiting"
            else:
                if ltp > highest_price:
                    highest_price = ltp
                    print(f"🔼 New High: ₹{highest_price}")
                    # Non-blocking notification (won't delay trading)
                    try:
                        send_telegram(f"🔼 New High: ₹{highest_price}")
                    except:
                        pass

                if highest_price >= avg_price + trail_start:
                    trail_exit = round_to_nearest_5_paise(highest_price - trail_step)
                    
                    # Only print if trail exit changed (avoid spam)
                    if trail_exit != last_trail_exit:
                        print(f"📉 Trail Active | LTP={ltp} | High={highest_price} | Exit={trail_exit}")
                        # Non-blocking notification
                        try:
                            send_telegram(f"📉 Trail Active | LTP={ltp} | High={highest_price} | Exit={trail_exit}")
                        except:
                            pass
                        last_trail_exit = trail_exit
                    
                    if ltp <= trail_exit:
                        sell_reason = f"🔻 Trailing HIT @ ₹{ltp}  (trail_exit=₹{trail_exit}  high=₹{highest_price})"
                        sell_price = ltp
                        sold_qty = quantity

            # 2. If an exit condition is met, place sell order IMMEDIATELY
            if sell_reason:
                _exit_ts = datetime.now().strftime('%H:%M:%S.%f')[:-3]
                print(f"[{_exit_ts}] {sell_reason}")
                
                # 🔊 Play sound immediately on exit trigger
                if "SL HIT" in sell_reason or "DYNAMIC SL" in sell_reason:
                    play_sound_async(SOUND_SL)
                elif "Trailing HIT" in sell_reason or "PROFIT" in sell_reason:
                    play_sound_async(SOUND_PROFIT)
                else:
                    play_sound_async(SOUND_SL)  # Default to SL sound

                try:
                    # ⚡ IMMEDIATE SELL - No delays before this
                    _sell_start = datetime.now()
                    print(f"[{_sell_start.strftime('%H:%M:%S.%f')[:-3]}] 🔄 Placing SELL order for {quantity} units @ market price...")
                    order_resp = place_market_order_groww(instrument, quantity, "SELL", "MIS")
                    sell_order_id = order_resp.get("payload", {}).get("groww_order_id") or order_resp.get("groww_order_id")
                    _sell_placed = datetime.now()
                    _sell_ms = (_sell_placed - _sell_start).total_seconds()
                    print(f"[{_sell_placed.strftime('%H:%M:%S.%f')[:-3]}] ✅ SELL Order placed: {sell_order_id} (took {_sell_ms:.3f}s)")
                    
                    # Send notification AFTER order is placed (non-blocking)
                    try:
                        send_telegram(f"{sell_reason}\n✅ SELL Order: {sell_order_id}")
                    except:
                        pass
                    
                    # ✅ LIVE TRADING: Wait for SELL order to execute (controlled by VALIDATE_ORDERS)
                    if CONFIG.get("VALIDATE_ORDERS", True) and sell_order_id:
                        sell_status = wait_for_order_status(sell_order_id, access_token, "SELL")
                        if sell_status in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                            # Fetch actual executed price
                            sell_price, sold_qty = get_order_executed_price(sell_order_id, access_token)
                            _exec_ts = datetime.now().strftime('%H:%M:%S.%f')[:-3]
                            if sell_price and sold_qty:
                                profit = (sell_price - avg_price) * sold_qty
                                print(f"[{_exec_ts}] 💰 SELL EXECUTED @ ₹{sell_price} | P&L: ₹{profit:.2f} (Buy ₹{avg_price} → Sell ₹{sell_price})")
                                send_telegram(f"💰 PROFIT: ₹{profit:.2f}")
                                play_sound_async(SOUND_PROFIT if profit > 0 else SOUND_SL)
                                log_trade_to_excel(
                                    instrument.get("internal_trading_symbol"),
                                    avg_price, sell_price, sold_qty, profit
                                )
                                # Display account summary
                                display_account_summary(access_token)
                            else:
                                print("⚠️ Could not get executed SELL price. Logging with LTP.")
                                profit = (ltp - avg_price) * quantity
                                play_sound_async(SOUND_PROFIT if profit > 0 else SOUND_SL)
                                log_trade_to_excel(
                                    instrument.get("internal_trading_symbol"),
                                    avg_price, ltp, quantity, profit
                                )
                                # Display account summary
                                display_account_summary(access_token)
                        else:
                            print(f"⚠️ SELL order failed with status: {sell_status}")
                            send_telegram(f"⚠️ SELL failed: {sell_status}")
                            play_sound_async(SOUND_SL)  # Play sound on failure
                    elif not CONFIG.get("VALIDATE_ORDERS", True):
                        # Testing mode: Use LTP estimate
                        sell_price = ltp
                        sold_qty = quantity
                        profit = (sell_price - avg_price) * sold_qty
                        print(f"⚠️ Testing mode: Estimated profit ₹{profit:.2f}")
                        play_sound_async(SOUND_PROFIT if profit > 0 else SOUND_SL)
                        log_trade_to_excel(
                            instrument.get("internal_trading_symbol"),
                            avg_price, sell_price, sold_qty, profit
                        )
                        # Display account summary
                        display_account_summary(access_token)
                    else:
                        print("⚠️ No SELL order ID received.")
                        
                except Exception as e:
                    print(f"❌ SELL order placement failed: {e}")
                    send_telegram(f"❌ SELL failed: {e}")

                print("✅ Trade cycle completed. Ready for next trade.")
                break  # Exit the trailing loop

            time.sleep(poll)


# ----------------- Directional Mode -----------------

def directional_mode():
    """
    Directional mode: User types premium value and direction
    Example: '150 c' = Find Call option with premium closest to ₹150
    Example: '200 p' = Find Put option with premium closest to ₹200
    """
    cfg = CONFIG
    dir_cfg = cfg["DIRECTIONAL_MODE"]
    
    print("\n" + "="*60)
    print("📍 DIRECTIONAL MODE")
    print("="*60)
    print(f"Index: {cfg['index']} | Expiry: {cfg['expiry']}")
    print(f"Lots: {cfg['lots']}")
    print("\n💡 Usage: <premium> <direction>")
    print("   Example: 150 c  → Find Call option near ₹150")
    print("   Example: 200 p  → Find Put option near ₹200")
    print("="*60)
    
    while True:
        user_input = input("\nEnter command (e.g., '150 c') or 'back' to menu: ").strip().lower()
        
        if user_input == "back":
            return
        
        # Parse input: expect "<premium> <c/p>"
        parts = user_input.split()
        if len(parts) != 2:
            print("⚠️ Invalid format. Use: <premium> <c/p>  (e.g., '150 c')")
            continue
        
        try:
            target_premium = float(parts[0])
        except ValueError:
            print("⚠️ Invalid premium value. Use a number (e.g., '150 c')")
            continue
        
        if parts[1] not in ["c", "p"]:
            print("⚠️ Invalid direction. Use 'c' for Call or 'p' for Put")
            continue
        
        option_type = "CE" if parts[1] == "c" else "PE"
        direction_name = "CALL (Bullish)" if parts[1] == "c" else "PUT (Bearish)"
        
        # Start timing
        start_time = datetime.now()
        print(f"\n[{start_time.strftime('%H:%M:%S.%f')[:-3]}] 🎯 Direction: {direction_name} | Target Premium: ₹{target_premium:.2f}")
        print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 🔍 Searching for matching option...")
        
        try:
            # Use spot price from CONFIG (already fetched at startup)
            spot_price = cfg["spot"]
            if not spot_price or spot_price <= 0:
                # Fallback: try to fetch fresh spot price
                spot_price = get_index_spot_price(cfg["index"], access_token)
                if not spot_price:
                    print("❌ Could not fetch spot price. Try again.")
                    continue
            
            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 📊 {cfg['index']} Spot: ₹{spot_price:.2f}")
            
            # Find all matching options for this type and expiry
            matching_options = []
            for instrument in instruments_data:
                if (instrument.get("underlying_symbol", "").upper() == cfg["index"].upper() and
                    instrument.get("expiry_date", "") == cfg["expiry"] and
                    instrument.get("instrument_type", "").upper() == option_type):
                    matching_options.append(instrument)
            
            if not matching_options:
                print(f"❌ No {option_type} options found for {cfg['index']} expiry {cfg['expiry']}")
                continue
            
            # Sort options by distance from ATM to scan smartly (ATM first, then expanding)
            step = 100 if "SENSEX" in cfg["index"].upper() else 50
            atm_strike = round(spot_price / step) * step
            matching_options.sort(key=lambda x: abs(float(x.get("strike_price", 0)) - atm_strike))
            
            scan_start = datetime.now()
            print(f"[{scan_start.strftime('%H:%M:%S.%f')[:-3]}] 📋 Scanning {len(matching_options)} options starting from ATM...")
            
            # Parallel LTP fetching for speed (fetch 5 options at once)
            max_checks = 30  # Maximum options to check
            batch_size = 5  # Fetch 5 options in parallel
            
            best_option = None
            best_diff = float('inf')
            best_ltp = None
            checked_count = 0
            
            # Helper function to fetch LTP with option data
            def fetch_ltp_with_option(opt):
                try:
                    ltp = get_ltp_for_instrument(opt, access_token, verbose=False, delay=0)
                    if ltp and ltp > 0:
                        return (opt, float(ltp), abs(float(ltp) - target_premium))
                except:
                    pass
                return None
            
            # Process options in batches
            for batch_start in range(0, min(len(matching_options), max_checks), batch_size):
                batch_end = min(batch_start + batch_size, len(matching_options), max_checks)
                batch = matching_options[batch_start:batch_end]
                
                # Fetch LTPs in parallel
                with ThreadPoolExecutor(max_workers=batch_size) as executor:
                    futures = [executor.submit(fetch_ltp_with_option, opt) for opt in batch]
                    
                    for future in as_completed(futures):
                        result = future.result()
                        if result:
                            opt, ltp, diff = result
                            checked_count += 1
                            
                            # Update best if closer
                            if diff < best_diff:
                                best_diff = diff
                                best_option = opt
                                best_ltp = ltp
                            
                            # Early exit if found very close match (within ₹3)
                            if diff <= 3.0:
                                print(f"✅ Found close match after checking {checked_count} options (diff: ₹{diff:.2f})")
                                break
                
                # Exit if found good match
                if best_option and best_diff <= 3.0:
                    break
            
            if best_option:
                print(f"✅ Best match found after checking {checked_count} options")
            
            if not best_option or not best_ltp:
                print("❌ Could not find any options with valid prices. Try again.")
                continue
            
            selected_option = best_option
            ltp = best_ltp
            
            selection_time = datetime.now()
            scan_duration = (selection_time - scan_start).total_seconds()
            
            # Display selected option
            symbol = selected_option.get("internal_trading_symbol") or selected_option.get("trading_symbol")
            strike = float(selected_option.get("strike_price", 0))
            lot_size = int(selected_option.get("lot_size", 25))
            quantity = cfg["lots"] * lot_size
            total_value = ltp * quantity
            
            # Determine ITM/OTM/ATM
            step = 100 if "SENSEX" in cfg["index"].upper() else 50
            atm_strike = round(spot_price / step) * step
            if option_type == "CE":
                position = "ITM" if strike < atm_strike else "OTM" if strike > atm_strike else "ATM"
            else:
                position = "ITM" if strike > atm_strike else "OTM" if strike < atm_strike else "ATM"
            
            print(f"\n[{selection_time.strftime('%H:%M:%S.%f')[:-3]}] ⏱️  Scan completed in {scan_duration:.2f}s")
            print("\n" + "="*60)
            print("✅ OPTION SELECTED")
            print("="*60)
            print(f"Symbol: {symbol}")
            print(f"Strike: {strike} {option_type} ({position})")
            print(f"Premium: ₹{ltp:.2f} (Target: ₹{target_premium:.2f}, Diff: ₹{abs(ltp - target_premium):.2f})")
            print(f"Lots: {cfg['lots']} × {lot_size} = {quantity} units")
            print(f"Total Value: ₹{total_value:,.2f}")
            print(f"SL: ₹{ltp - cfg['HARD_SL_POINTS']:.2f} ({cfg['HARD_SL_POINTS']} points)")
            print(f"Trail: Activates at ₹{ltp + cfg['TRAIL_START_PROFIT']:.2f}")
            print("="*60)
            
            # Auto-execute trade directly without command parsing
            print(f"\n[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 🚀 Executing trade automatically...")
            
            try:
                # Place BUY order
                order_start = datetime.now()
                print(f"[{order_start.strftime('%H:%M:%S.%f')[:-3]}] 🔄 Placing BUY order for {quantity} units @ market price...")
                buy_order_resp = place_market_order_groww(selected_option, quantity, "BUY", "MIS")
                buy_order_id = buy_order_resp.get("payload", {}).get("groww_order_id") or buy_order_resp.get("groww_order_id")
                
                if not buy_order_id:
                    print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] ❌ Failed to get BUY order ID")
                    continue
                
                order_placed = datetime.now()
                order_duration = (order_placed - order_start).total_seconds()
                print(f"[{order_placed.strftime('%H:%M:%S.%f')[:-3]}] ✅ BUY Order placed: {buy_order_id} (took {order_duration:.2f}s)")
                send_telegram(f"✅ BUY {symbol} @ ₹{ltp:.2f} | Qty: {quantity}")
                
                # Wait for BUY order execution
                if cfg.get("VALIDATE_ORDERS", True):
                    validation_start = datetime.now()
                    buy_status = wait_for_order_status(buy_order_id, access_token, "BUY")
                    if buy_status not in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                        print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] ❌ BUY order failed with status: {buy_status}")
                        continue
                    
                    # Get executed price
                    avg_price, bought_qty = get_order_executed_price(buy_order_id, access_token)
                    if not avg_price:
                        avg_price = ltp
                        bought_qty = quantity
                    
                    validation_end = datetime.now()
                    validation_duration = (validation_end - validation_start).total_seconds()
                else:
                    avg_price = ltp
                    bought_qty = quantity
                    validation_duration = 0
                
                execution_time = datetime.now()
                total_duration = (execution_time - start_time).total_seconds()
                print(f"[{execution_time.strftime('%H:%M:%S.%f')[:-3]}] ✅ BUY executed @ ₹{avg_price:.2f} | Qty: {bought_qty}")
                print(f"[{execution_time.strftime('%H:%M:%S.%f')[:-3]}] ⏱️  Total time: {total_duration:.2f}s (Scan: {scan_duration:.2f}s, Order: {order_duration:.2f}s, Validation: {validation_duration:.2f}s)")
                
                # Start trailing stop monitoring (inline logic from place_cp_order)
                trail_start_time = datetime.now()
                print(f"\n[{trail_start_time.strftime('%H:%M:%S.%f')[:-3]}] 🔄 Starting trailing stop monitoring...")
                
                # Fetch ATR for dynamic trail step (3s timeout, non-blocking feel)
                _dir_atr = _fetch_atr_sync(selected_option, timeout=3)

                # Calculate dynamic SL using ATR
                hard_sl_points = cfg["HARD_SL_POINTS"]
                hard_sl = round_to_nearest_5_paise(avg_price - hard_sl_points)

                print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 📈 Trailing started... Hard SL: ₹{hard_sl:.2f}")
                send_telegram(f"📈 Trailing started | Entry: ₹{avg_price:.2f} | SL: ₹{hard_sl:.2f}")

                # Trailing parameters
                trail_start = cfg["TRAIL_START_PROFIT"]
                trail_step = _resolve_trail_step(_dir_atr)  # ATR-based if flag set, else fixed TRAIL_STEP
                poll = cfg["POLL_INTERVAL"]
                max_time = cfg["MAX_TRAIL_TIME"]
                
                highest_price = avg_price
                start_time = time.time()
                last_heartbeat = time.time()
                last_trail_exit = None
                
                while True:
                    # Heartbeat every 30 seconds
                    if time.time() - last_heartbeat > 30:
                        print(f"💓 Monitoring... LTP last seen: ₹{ltp if 'ltp' in locals() else 'fetching...'}")
                        last_heartbeat = time.time()
                    
                    try:
                        ltp = get_ltp_for_instrument(selected_option, access_token, verbose=False, delay=0)
                    except Exception as e:
                        print(f"⚠️ LTP fetch error (retrying): {e}")
                        time.sleep(poll)
                        continue
                    
                    if ltp is None:
                        print("⚠️ LTP is None (retrying...)")
                        time.sleep(poll)
                        continue
                    
                    ltp = float(ltp)
                    sell_reason = None
                    
                    # Check exit conditions
                    if ltp <= hard_sl:
                        sell_reason = f"🛑 DYNAMIC SL HIT @ {ltp}"
                    elif time.time() - start_time >= max_time:
                        sell_reason = "⏰ Max trail time reached"
                    else:
                        if ltp > highest_price:
                            highest_price = ltp
                            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 🔼 New High: ₹{highest_price}")
                        
                        if highest_price >= avg_price + trail_start:
                            trail_exit = round_to_nearest_5_paise(highest_price - trail_step)
                            
                            if trail_exit != last_trail_exit:
                                print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] 📉 Trail Active | LTP={ltp} | High={highest_price} | Exit={trail_exit}")
                                last_trail_exit = trail_exit
                            
                            if ltp <= trail_exit:
                                sell_reason = f"🔻 Trailing HIT @ {ltp}"
                    
                    # If exit condition met, place SELL order
                    if sell_reason:
                        exit_time = datetime.now()
                        monitoring_duration = (exit_time - trail_start_time).total_seconds()
                        print(f"[{exit_time.strftime('%H:%M:%S.%f')[:-3]}] {sell_reason}")
                        print(f"[{exit_time.strftime('%H:%M:%S.%f')[:-3]}] ⏱️  Monitored for {monitoring_duration:.2f}s")
                        if "SL HIT" in sell_reason or "DYNAMIC SL" in sell_reason:
                            play_sound_async(SOUND_SL)
                        else:
                            play_sound_async(SOUND_PROFIT)
                        
                        try:
                            sell_order_start = datetime.now()
                            print(f"[{sell_order_start.strftime('%H:%M:%S.%f')[:-3]}] 🔄 Placing SELL order for {bought_qty} units...")
                            sell_order_resp = place_market_order_groww(selected_option, bought_qty, "SELL", "MIS")
                            sell_order_id = sell_order_resp.get("payload", {}).get("groww_order_id") or sell_order_resp.get("groww_order_id")
                            sell_order_placed = datetime.now()
                            sell_order_duration = (sell_order_placed - sell_order_start).total_seconds()
                            print(f"[{sell_order_placed.strftime('%H:%M:%S.%f')[:-3]}] ✅ SELL Order placed: {sell_order_id} (took {sell_order_duration:.2f}s)")
                            send_telegram(f"{sell_reason}\n✅ SELL Order: {sell_order_id}")
                            
                            if cfg.get("VALIDATE_ORDERS", True) and sell_order_id:
                                sell_validation_start = datetime.now()
                                sell_status = wait_for_order_status(sell_order_id, access_token, "SELL")
                                if sell_status in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                                    sell_price, sold_qty = get_order_executed_price(sell_order_id, access_token)
                                    if sell_price and sold_qty:
                                        sell_executed = datetime.now()
                                        sell_validation_duration = (sell_executed - sell_validation_start).total_seconds()
                                        profit = (sell_price - avg_price) * sold_qty
                                        
                                        # Calculate complete trade duration
                                        complete_trade_duration = (sell_executed - start_time).total_seconds()
                                        
                                        print(f"[{sell_executed.strftime('%H:%M:%S.%f')[:-3]}] 💰 PROFIT: ₹{profit:.2f} (Buy @ ₹{avg_price}, Sell @ ₹{sell_price})")
                                        print(f"[{sell_executed.strftime('%H:%M:%S.%f')[:-3]}] ⏱️  SELL validation: {sell_validation_duration:.2f}s | Complete trade: {complete_trade_duration:.2f}s")
                                        send_telegram(f"💰 PROFIT: ₹{profit:.2f}")
                                        play_sound_async(SOUND_PROFIT if profit > 0 else SOUND_SL)
                                        log_trade_to_excel(symbol, avg_price, sell_price, sold_qty, profit)
                                        display_account_summary(access_token)
                            
                            print(f"[{datetime.now().strftime('%H:%M:%S.%f')[:-3]}] ✅ Trade cycle completed.")
                            break
                        
                        except Exception as sell_error:
                            print(f"❌ SELL order failed: {sell_error}")
                            send_telegram(f"❌ SELL failed: {sell_error}")
                            break
                    
                    time.sleep(poll)
                
            except Exception as trade_error:
                print(f"❌ Trade execution failed: {trade_error}")
                import traceback
                traceback.print_exc()
            
        except Exception as e:
            print(f"❌ Error in directional mode: {e}")
            import traceback
            traceback.print_exc()


# ─────────────────────────────────────────────────────────────────────────────
#  AUTO MODE v2 — Multi-signal consensus engine
#  Sources: MASTER_SIGNAL_BOT log  +  Fibonacci (1h + 15m)  +  premium momentum
#  Entry: minimum 4 agreeing votes before placing any order
#  Exit:  %-based SL/target, signal-reversal detection, market-close gate
# ─────────────────────────────────────────────────────────────────────────────

AUTO_V2_CONFIG = {
    # Signal thresholds
    "MIN_MASTER_CONFIDENCE":   65,   # % — MASTER_SIGNAL_BOT minimum to count as a vote
    "MIN_VOTES":                4,   # total votes needed (max attainable ~8)
    "MASTER_SIGNAL_MAX_AGE_S": 90,   # seconds — ignore stale MASTER_SIGNAL entries
    # Scan behaviour
    "SCAN_WAIT_SEC":           15,   # wait between signal-check cycles when no trade
    "SCAN_TIMEOUT_MIN":        25,   # report and pause after X min of no signal
    # Risk — SL and target are % of entry premium
    "SL_PCT_HIGH":            0.12,  # 12% SL for HIGH-confidence signal
    "SL_PCT_MEDIUM":          0.09,  # 9% SL for MEDIUM-confidence signal
    "TARGET_MULTIPLIER":       2.0,  # target = SL * multiplier  (enforces R:R ≥ 2)
    # Trade duration
    "MAX_TRADE_MIN":           75,   # hard max time in a single trade (minutes)
    "SIGNAL_RECHECK_SEC":      30,   # re-check MASTER_SIGNAL during hold (seconds)
    # Market-hours gate
    "NO_TRADE_BEFORE": (9, 30),      # (hour, minute) — no entry before this time
    "NO_TRADE_AFTER":  (15, 0),      # (hour, minute) — no entry after this time
}


def _read_master_signal_latest(index_name: str):
    """
    Return the most recent valid signal record from logs/master_signal/*.log.
    Returns None if no file exists, file is empty, or signal is stale.
    """
    base    = os.path.dirname(os.path.abspath(__file__))
    log_dir = os.path.join(base, "logs", "master_signal")
    if not os.path.isdir(log_dir):
        return None
    files = sorted(
        [f for f in os.listdir(log_dir)
         if f.endswith(".log") and "Master_Signal" in f],
        key=lambda f: os.path.getmtime(os.path.join(log_dir, f)),
        reverse=True,
    )
    if not files:
        return None
    log_path = os.path.join(log_dir, files[0])
    try:
        with open(log_path, "r", encoding="utf-8") as fh:
            lines = [l.strip() for l in fh if l.strip()]
        for line in reversed(lines):
            try:
                rec = json.loads(line)
            except json.JSONDecodeError:
                continue
            if rec.get("index", "").upper() != index_name.upper():
                continue
            try:
                age = (datetime.now() - datetime.strptime(rec["ts"], "%Y-%m-%dT%H:%M:%S")).total_seconds()
                if age > AUTO_V2_CONFIG["MASTER_SIGNAL_MAX_AGE_S"]:
                    return None
            except Exception:
                pass
            return rec
    except Exception:
        pass
    return None


def _fetch_index_candles_auto(index_name: str, interval: str, hours_back: int) -> list:
    """Fetch CASH candles for the index using the global groww client."""
    idx = index_name.upper()
    sym_map = {
        "NIFTY":     ["NSE-NIFTY 50", "NSE-NIFTY"],
        "SENSEX":    ["BSE-SENSEX"],
        "BANKNIFTY": ["NSE-BANKNIFTY", "NSE-NIFTY BANK"],
        "FINNIFTY":  ["NSE-NIFTY FIN SERVICE", "NSE-FINNIFTY"],
    }
    candidates = sym_map.get(idx, [f"NSE-{idx}"])
    exc = groww.EXCHANGE_BSE if "SENSEX" in idx else groww.EXCHANGE_NSE
    end_dt   = datetime.now()
    start_dt = end_dt - timedelta(hours=hours_back)
    for sym in candidates:
        try:
            r = groww.get_historical_candles(
                groww_symbol=sym, exchange=exc, segment="CASH",
                start_time=start_dt.strftime("%Y-%m-%d %H:%M:%S"),
                end_time=end_dt.strftime("%Y-%m-%d %H:%M:%S"),
                candle_interval=interval,
            )
            if r and r.get("candles") and len(r["candles"]) >= 5:
                return [{"ts": c[0], "open": float(c[1]), "high": float(c[2]),
                         "low": float(c[3]), "close": float(c[4])} for c in r["candles"]]
        except Exception:
            continue
    return []


def _fib_score_auto(spot: float, candles: list) -> int:
    """
    Fibonacci position score: -3 (price at/below swing low) to +3 (at/above swing high).
    Uses most recent 20-candle swing high/low as the reference range.
    """
    if len(candles) < 10:
        return 0
    n  = min(20, len(candles))
    sh = max(c["high"]  for c in candles[-n:])
    sl = min(c["low"]   for c in candles[-n:])
    rng = sh - sl
    if rng < 1:
        return 0
    pos = (spot - sl) / rng   # 0.0 = at swing low, 1.0 = at swing high
    if pos >= 0.786: return  3
    if pos >= 0.618: return  2
    if pos >= 0.500: return  1
    if pos >= 0.382: return -1
    if pos >= 0.236: return -2
    return -3


def _rsi_auto(candles: list, period: int = 14):
    """Simple Wilder RSI from candle closes. Returns float or None."""
    closes = [c["close"] for c in candles]
    if len(closes) < period + 1:
        return None
    gains  = [max(closes[i] - closes[i-1], 0.0) for i in range(1, len(closes))]
    losses = [max(closes[i-1] - closes[i], 0.0) for i in range(1, len(closes))]
    avg_g  = sum(gains[:period])  / period
    avg_l  = sum(losses[:period]) / period
    for i in range(period, len(gains)):
        avg_g = (avg_g * (period - 1) + gains[i])  / period
        avg_l = (avg_l * (period - 1) + losses[i]) / period
    return round(100.0 - 100.0 / (1.0 + avg_g / avg_l), 1) if avg_l else 100.0


def _premium_momentum_dir(instrument, samples: int = 4, delay: float = 0.35) -> str:
    """
    Sample option LTP `samples` times and return 'UP', 'DOWN', or 'FLAT'.
    """
    prices = []
    for _ in range(samples):
        p = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
        if p:
            prices.append(float(p))
        time.sleep(delay)
    if len(prices) < 3:
        return "FLAT"
    changes  = [prices[i] - prices[i-1] for i in range(1, len(prices))]
    avg_ch   = sum(changes) / len(changes)
    pos_frac = sum(1 for c in changes if c > 0) / len(changes)
    if avg_ch > 0.1 and pos_frac >= 0.6:
        return "UP"
    if avg_ch < -0.1 and pos_frac <= 0.4:
        return "DOWN"
    return "FLAT"


def _within_trading_hours_auto() -> bool:
    """True when current time is inside the allowed entry window."""
    now  = datetime.now()
    h, m = now.hour, now.minute
    cfg  = AUTO_V2_CONFIG
    return cfg["NO_TRADE_BEFORE"] <= (h, m) < cfg["NO_TRADE_AFTER"]


def _get_nearest_expiry_auto(index_name: str) -> str:
    """
    Scan instruments_data for the nearest upcoming expiry for the given index.
    Falls back to CONFIG["expiry"] if nothing is found.
    Guarantees auto mode always trades on a valid live expiry.
    """
    today  = datetime.now().date()
    idx    = index_name.upper()
    expiries = set()
    for item in instruments_data:
        if item.get("underlying_symbol", "").upper() != idx:
            continue
        ed = item.get("expiry_date", "").strip()
        if not ed:
            continue
        try:
            if datetime.strptime(ed, "%Y-%m-%d").date() >= today:
                expiries.add(ed)
        except ValueError:
            continue
    if expiries:
        nearest = min(expiries, key=lambda e: datetime.strptime(e, "%Y-%m-%d").date())
        return nearest
    return CONFIG["expiry"]   # fallback to whatever is in CONFIG


def _get_spot_via_chain(index_name: str, expiry: str) -> float:
    """
    Get live spot price from the option-chain API (reliable for all indices).
    Falls back to 0 on error.
    """
    exc = "BSE" if "SENSEX" in index_name.upper() else "NSE"
    url = (f"https://api.groww.in/v1/option-chain/exchange/{exc}"
           f"/underlying/{index_name}?expiry_date={expiry}")
    hdrs = {"Accept": "application/json",
            "Authorization": f"Bearer {access_token}", "X-API-VERSION": "1.0"}
    try:
        resp = session.get(url, headers=hdrs, timeout=8)
        if resp.status_code == 200:
            ltp = resp.json().get("payload", {}).get("underlying_ltp")
            if ltp:
                return float(ltp)
    except Exception:
        pass
    return 0.0


def _find_option_quiet(direction: str, index_name: str, expiry: str,
                        min_p: float, max_p: float, lots: int):
    """
    Silently find the best option in premium range — NO momentum sampling,
    NO print output. Used by auto mode to avoid verbose momentum spam.
    Returns (instrument, ltp, lot_size) or (None, 0.0, 1).
    """
    idx = index_name.upper()
    mid = (min_p + max_p) / 2
    candidates = []
    for item in instruments_data:
        if item.get("underlying_symbol", "").upper() != idx:
            continue
        if item.get("instrument_type", "").upper() != direction.upper():
            continue
        if item.get("expiry_date", "").strip() != expiry:
            continue
        ltp = get_ltp_for_instrument(item, access_token, verbose=False)
        if ltp is None:
            continue
        if not (min_p <= ltp <= max_p):
            continue
        lot_size = int(item.get("lot_size") or item.get("lotsize") or 1)
        if lots * lot_size * ltp > get_available_margin(access_token) * 0.9:
            continue
        candidates.append((abs(ltp - mid), item, ltp, lot_size))
    if not candidates:
        return None, 0.0, 1
    candidates.sort(key=lambda x: x[0])
    _, best_inst, best_ltp, best_lot = candidates[0]
    return best_inst, best_ltp, best_lot


def _quick_ltp_direction(instrument) -> str:
    """
    Two LTP samples 1.5 s apart — returns 'UP', 'DOWN', or 'FLAT'.
    Threshold: 0.20% move. Silent (no prints).
    """
    if not instrument:
        return "FLAT"
    try:
        ltp1 = get_ltp_for_instrument(instrument, access_token, verbose=False)
        time.sleep(1.5)
        ltp2 = get_ltp_for_instrument(instrument, access_token, verbose=False)
        if ltp1 and ltp2 and ltp1 > 0:
            chg = (ltp2 - ltp1) / ltp1
            if chg >= 0.002:
                return "UP"
            if chg <= -0.002:
                return "DOWN"
    except Exception:
        pass
    return "FLAT"


def _collect_signals_auto(index_name: str, expiry: str,
                           lots: int, min_p: float, max_p: float) -> dict:
    """
    Gather all available signals and compute a vote-weighted consensus.

    Vote sources (max attainable ≈ 8):
      MASTER_SIGNAL_BOT  →  +3 if conf ≥ 75%, +2 if conf ≥ 65%
                            −1 each side if WAIT AND conf ≥ 50%
                            (ignored if conf < 50% — no strong opinion)
      Fibonacci 1-hour   →  +1 or +2 per side
      Fibonacci 15-min   →  +1 or +2 per side
      Premium momentum   →  +1 (silent 2-sample LTP check)

    RSI dampening: −1 from over-extended side (RSI ≥ 72 or ≤ 28).

    Returns dict with keys: direction, confidence, votes_ce, votes_pe,
                             instrument, ltp, lot_size, detail, sl_pct, master
    """
    lines    = []
    votes_ce = 0
    votes_pe = 0

    # ── 1. MASTER SIGNAL BOT ────────────────────────────────────────────────
    master = _read_master_signal_latest(index_name)
    if master:
        m_dir  = master.get("direction", "WAIT")
        m_conf = float(master.get("confidence", 0))
        thr    = AUTO_V2_CONFIG["MIN_MASTER_CONFIDENCE"]
        lines.append(
            f"  📡 MASTER: {m_dir}  conf={m_conf:.0f}%  "
            f"1h={master.get('s1h',0):+}  15m={float(master.get('s15m',0)):+.1f}  "
            f"5m={master.get('s5m',0):+}  R:R={master.get('rr',0):.1f}"
        )
        if m_dir == "CE" and m_conf >= thr:
            w = 3 if m_conf >= 75 else 2
            votes_ce += w
        elif m_dir == "PE" and m_conf >= thr:
            w = 3 if m_conf >= 75 else 2
            votes_pe += w
        elif m_dir == "WAIT" and m_conf >= 50:
            # Only penalise when MASTER is reasonably confident about WAIT.
            # At conf < 50% it has no strong opinion — ignoring is safer than
            # penalising, otherwise FIBO signals get unfairly blocked.
            votes_ce -= 1; votes_pe -= 1
            lines.append(f"  ⚠️  MASTER WAIT (conf={m_conf:.0f}% ≥ 50) → both sides −1")
        else:
            lines.append(f"  ℹ️  MASTER WAIT at low conf ({m_conf:.0f}%) — ignored")
    else:
        lines.append("  📡 MASTER: no recent signal (start MASTER_SIGNAL_BOT)")

    # ── 2. Fibonacci scores — use option-chain spot (works for all indices) ──
    spot = _get_spot_via_chain(index_name, expiry)
    if spot:
        for tf_label, interval, hours in [("1h", "1hour", 48), ("15m", "15minute", 26)]:
            candles = _fetch_index_candles_auto(index_name, interval, hours)
            if candles:
                sc  = _fib_score_auto(spot, candles)
                rsi = _rsi_auto(candles)
                rsi_str = f"{rsi:.0f}" if rsi else "N/A"
                lines.append(
                    f"  📐 FIBO {tf_label}: score={sc:+d}  "
                    f"RSI={rsi_str}  spot={spot:.0f}")
                if sc >= 2:    votes_ce += 2
                elif sc == 1:  votes_ce += 1
                elif sc <= -2: votes_pe += 2
                elif sc == -1: votes_pe += 1
                if rsi:
                    if rsi >= 72:
                        votes_ce -= 1
                        lines.append(f"    ⚠️  {tf_label} RSI overbought ({rsi:.0f}) → CE −1")
                    elif rsi <= 28:
                        votes_pe -= 1
                        lines.append(f"    ⚠️  {tf_label} RSI oversold ({rsi:.0f}) → PE −1")
            else:
                lines.append(f"  📐 FIBO {tf_label}: candle data unavailable")
    else:
        lines.append(f"  📐 FIBO: spot unavailable for {index_name} {expiry}")

    # ── 3. Premium momentum — silent 2-sample check, no verbose output ───────
    # Find best CE and PE instruments quietly (no momentum spam).
    ce_inst, ce_ltp, ce_lot = _find_option_quiet("CE", index_name, expiry, min_p, max_p, lots)
    pe_inst, pe_ltp, pe_lot = _find_option_quiet("PE", index_name, expiry, min_p, max_p, lots)

    # 2-sample direction check: compare which side is rising
    ce_dir = _quick_ltp_direction(ce_inst)
    pe_dir = _quick_ltp_direction(pe_inst)

    prem_note = f"CE={ce_dir}(₹{ce_ltp:.0f})  PE={pe_dir}(₹{pe_ltp:.0f})"
    if ce_dir == "UP" and pe_dir != "UP":
        votes_ce += 1
        lines.append(f"  📊 Premium momentum: CE rising → CE +1  [{prem_note}]")
        opt_instrument, opt_ltp, lot_size = ce_inst, ce_ltp, ce_lot
    elif pe_dir == "UP" and ce_dir != "UP":
        votes_pe += 1
        lines.append(f"  📊 Premium momentum: PE rising → PE +1  [{prem_note}]")
        opt_instrument, opt_ltp, lot_size = pe_inst, pe_ltp, pe_lot
    else:
        lines.append(f"  📊 Premium momentum: flat / no clear winner  [{prem_note}]")
        # Use whichever side votes favour so far
        if votes_ce >= votes_pe:
            opt_instrument, opt_ltp, lot_size = ce_inst, ce_ltp, ce_lot
        else:
            opt_instrument, opt_ltp, lot_size = pe_inst, pe_ltp, pe_lot

    # ── 4. Resolve direction ─────────────────────────────────────────────────
    min_v = AUTO_V2_CONFIG["MIN_VOTES"]
    if votes_ce >= min_v and votes_ce >= votes_pe:
        direction = "CE"
        if opt_instrument and \
                opt_instrument.get("instrument_type", "").upper() != "CE":
            opt_instrument, opt_ltp, lot_size = ce_inst, ce_ltp, ce_lot
    elif votes_pe >= min_v and votes_pe > votes_ce:
        direction = "PE"
        if opt_instrument and \
                opt_instrument.get("instrument_type", "").upper() != "PE":
            opt_instrument, opt_ltp, lot_size = pe_inst, pe_ltp, pe_lot
    else:
        direction      = "WAIT"
        opt_instrument = None
        opt_ltp        = 0.0
        lot_size       = 1

    winning_votes = votes_ce if direction == "CE" else (votes_pe if direction == "PE" else 0)
    losing_votes  = votes_pe if direction == "CE" else (votes_ce if direction == "PE" else 0)
    margin        = winning_votes - max(losing_votes, 0)

    if direction != "WAIT":
        confidence = "HIGH"   if (winning_votes >= 7 or margin >= 4) else \
                     "MEDIUM" if (winning_votes >= 5 or margin >= 2) else "LOW"
    else:
        confidence = "NONE"

    sl_pct = (AUTO_V2_CONFIG["SL_PCT_HIGH"] if confidence == "HIGH"
              else AUTO_V2_CONFIG["SL_PCT_MEDIUM"])

    lines.append(
        f"  🗳  VOTES  CE={votes_ce}  PE={votes_pe}  →  {direction}  [{confidence}]"
    )

    return {
        "direction":   direction,
        "confidence":  confidence,
        "votes_ce":    votes_ce,
        "votes_pe":    votes_pe,
        "instrument":  opt_instrument,
        "ltp":         opt_ltp,
        "lot_size":    lot_size,
        "detail":      "\n".join(lines),
        "sl_pct":      sl_pct,
        "master":      master,
    }


_AUTO_STATUS_FILE = os.path.join(PROJECT_ROOT, ".auto_mode_status.json")


def _write_auto_status(state: str, **kwargs):
    """
    Write a lightweight JSON status file that the Live Dashboard can poll
    to display auto-mode activity without scanning the raw log stream.

    Fields always written: state, ts, mode_label (PAPER / LIVE)
    Optional kwargs are merged in (direction, confidence, votes_ce, votes_pe,
    instrument_symbol, ltp, total_pnl, trade_count, detail_line, etc.)
    """
    payload = {"state": state, "ts": time.time()}
    payload.update(kwargs)
    try:
        with open(_AUTO_STATUS_FILE, "w") as _f:
            import json as _json
            _json.dump(payload, _f)
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────────

def auto_mode_runner():
    """
    AUTO MODE v2 — multi-signal consensus before every entry.

    Entry gate: MASTER_SIGNAL + Fibonacci (1h + 15m) + premium momentum
                must reach MIN_VOTES consensus threshold.
    Exit gates: %-based SL / target, trailing stop, MASTER_SIGNAL direction
                reversal, max-trade-time, market-close guard.
    """
    cfg       = CONFIG
    paper     = cfg.get("PAPER_TRADING", False)
    mode_lbl  = "PAPER" if paper else "LIVE"
    index     = cfg["index"]
    min_p     = cfg["min_premium"]
    max_p     = cfg["max_premium"]
    lots      = cfg["lots"]
    target_pnl = cfg["target_pnl"]
    poll       = cfg["POLL_INTERVAL"]

    # Always detect the nearest live expiry — CONFIG["expiry"] may be stale
    expiry = _get_nearest_expiry_auto(index)
    if expiry != cfg["expiry"]:
        print(f"⚠️  CONFIG expiry {cfg['expiry']} is stale — using nearest: {expiry}")
        cfg["expiry"] = expiry   # update in-memory so find_option_by_premium_parallel sees it

    _write_auto_status("STARTING", mode_label=mode_lbl, index=index)
    print("\n" + "="*68)
    print(f"🤖 AUTO MODE v2 [{mode_lbl}] — Multi-signal consensus")
    print(f"   Index: {index}  |  Expiry: {expiry}  (auto-detected)")
    print(f"   Premium range: ₹{min_p}–₹{max_p}  |  Lots: {lots}")
    print(f"   Target P&L: ₹{target_pnl}  |  Min votes: {AUTO_V2_CONFIG['MIN_VOTES']}")
    print(f"   MASTER conf threshold: {AUTO_V2_CONFIG['MIN_MASTER_CONFIDENCE']}%")
    print(f"   SL: HIGH={AUTO_V2_CONFIG['SL_PCT_HIGH']*100:.0f}%  "
          f"MEDIUM={AUTO_V2_CONFIG['SL_PCT_MEDIUM']*100:.0f}%  "
          f"Target R:R ≥ {AUTO_V2_CONFIG['TARGET_MULTIPLIER']:.1f}:1")
    print("="*68)
    start_webhook_server()
    send_telegram(f"🤖 AUTO v2 [{mode_lbl}] started  |  {index}")

    total_pnl   = 0.0
    trade_count = 0
    scan_start  = time.time()

    while True:

        # ── P&L target gate ──────────────────────────────────────────────────
        if total_pnl >= target_pnl:
            print(f"🎯 Target P&L ₹{target_pnl} reached! Stopping.")
            send_telegram(f"🎯 AUTO v2 done — Total P&L: ₹{total_pnl:.2f}")
            _write_auto_status("STOPPED", mode_label=mode_lbl, index=index,
                               total_pnl=total_pnl, trade_count=trade_count,
                               stop_reason="target_reached")
            break

        # ── Market-hours gate ────────────────────────────────────────────────
        if not _within_trading_hours_auto():
            nh, nm = AUTO_V2_CONFIG["NO_TRADE_BEFORE"]
            print(f"⏳ Outside trading window. Waiting for {nh:02d}:{nm:02d}...")
            time.sleep(60)
            continue

        # ── Signal collection ────────────────────────────────────────────────
        print(f"\n{'─'*60}")
        print(f"🔍 Trade #{trade_count+1} | Scanning signals"
              f"  (P&L: ₹{total_pnl:.2f} / ₹{target_pnl:.2f})")
        signals = _collect_signals_auto(index, expiry, lots, min_p, max_p)
        print(signals["detail"])

        direction  = signals["direction"]
        confidence = signals["confidence"]

        _write_auto_status(
            "SCANNING",
            mode_label=mode_lbl,
            index=index,
            expiry=expiry,
            direction=direction,
            confidence=confidence,
            votes_ce=signals["votes_ce"],
            votes_pe=signals["votes_pe"],
            total_pnl=total_pnl,
            trade_count=trade_count,
            last_scan_detail=signals["detail"],
        )

        if direction == "WAIT" or confidence in ("LOW", "NONE"):
            elapsed = (time.time() - scan_start) / 60
            print(f"  ⏳ No actionable signal ({direction} | {confidence}). "
                  f"Waiting {AUTO_V2_CONFIG['SCAN_WAIT_SEC']}s…  "
                  f"({elapsed:.0f}/{AUTO_V2_CONFIG['SCAN_TIMEOUT_MIN']} min)")
            time.sleep(AUTO_V2_CONFIG["SCAN_WAIT_SEC"])
            if elapsed >= AUTO_V2_CONFIG["SCAN_TIMEOUT_MIN"]:
                print(f"⚠️  No signal for {AUTO_V2_CONFIG['SCAN_TIMEOUT_MIN']} min — pausing 5 min")
                time.sleep(300)
                scan_start = time.time()
            continue

        # ── Good signal — validate instrument ────────────────────────────────
        scan_start  = time.time()
        instrument  = signals["instrument"]
        ltp         = signals["ltp"]
        lot_size    = signals["lot_size"]
        sl_pct      = signals["sl_pct"]
        master      = signals["master"]

        if not instrument or not ltp:
            print(f"⚠️  No valid {direction} instrument in range ₹{min_p}–₹{max_p}. Retrying…")
            time.sleep(AUTO_V2_CONFIG["SCAN_WAIT_SEC"])
            continue

        symbol   = instrument.get("internal_trading_symbol") or instrument.get("trading_symbol")
        quantity = lots * lot_size

        print(f"\n✅ SIGNAL CONFIRMED: {direction}  [{confidence}]")
        print(f"   Option : {symbol}")
        print(f"   LTP    : ₹{ltp:.2f}  |  Qty: {quantity}")
        if master:
            print(f"   MASTER : conf={master.get('confidence',0):.0f}%  "
                  f"stop={master.get('stop',0):.1f}  "
                  f"target={master.get('target',0):.1f}  "
                  f"R:R={master.get('rr',0):.1f}")

        # ── Optional user confirmation ────────────────────────────────────────
        if cfg.get("user_confirmation_needed", False):
            ans = input(f"\nProceed BUY {direction} {symbol}? (y/n): ").strip().lower()
            if ans != "y":
                print("Skipped.")
                time.sleep(5)
                continue

        # ── BUY order ─────────────────────────────────────────────────────────
        try:
            buy_resp     = place_market_order_groww(instrument, quantity, "BUY", "MIS")
            buy_order_id = (buy_resp.get("payload", {}).get("groww_order_id")
                            or buy_resp.get("groww_order_id"))
            print(f"✅ BUY placed: {buy_order_id}")
            send_telegram(
                f"🤖 AUTO v2 BUY [{mode_lbl}]: {symbol} @ ₹{ltp:.2f}  "
                f"Qty:{quantity}  [{direction}|{confidence}]"
            )
        except Exception as e:
            print(f"❌ BUY failed: {e}. Retrying in 10s…")
            time.sleep(10)
            continue

        # ── Wait for BUY execution ────────────────────────────────────────────
        if cfg.get("VALIDATE_ORDERS", True) and buy_order_id:
            b_status = wait_for_order_status(buy_order_id, access_token, "BUY")
            if b_status not in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                print(f"⚠️  BUY not executed ({b_status}). Skipping trade.")
                continue
            avg_price, bought_qty = get_order_executed_price(buy_order_id, access_token)
            if not avg_price:
                avg_price, bought_qty = ltp, quantity
        else:
            avg_price, bought_qty = ltp, quantity

        print(f"🎯 BUY executed @ ₹{avg_price:.2f}  Qty: {bought_qty}")
        _write_auto_status(
            "IN_TRADE",
            mode_label=mode_lbl,
            index=index,
            expiry=expiry,
            direction=direction,
            confidence=confidence,
            symbol=symbol,
            entry_price=avg_price,
            quantity=bought_qty,
            total_pnl=total_pnl,
            trade_count=trade_count + 1,
        )

        # ── Compute SL and target ─────────────────────────────────────────────
        # %-based SL from confidence level; target enforces R:R ≥ TARGET_MULTIPLIER
        sl_abs     = avg_price * sl_pct
        hard_sl    = round_to_nearest_5_paise(avg_price - sl_abs)
        target_p   = round_to_nearest_5_paise(avg_price + sl_abs * AUTO_V2_CONFIG["TARGET_MULTIPLIER"])
        # Also apply CONFIG's absolute hard-SL as a tighter floor
        cfg_sl     = round_to_nearest_5_paise(avg_price - cfg["HARD_SL_POINTS"])
        hard_sl    = max(hard_sl, cfg_sl)

        trail_start_p = cfg["TRAIL_START_PROFIT"]
        trail_step    = cfg["TRAIL_STEP"]
        max_trade_sec = AUTO_V2_CONFIG["MAX_TRADE_MIN"] * 60

        print(f"📈 SL ₹{hard_sl:.2f} ({sl_pct*100:.0f}%)  "
              f"|  Target ₹{target_p:.2f}  "
              f"|  Trail after +{trail_start_p}pt")
        send_telegram(
            f"📈 AUTO v2 — Entry ₹{avg_price:.2f}  "
            f"SL ₹{hard_sl:.2f}  Target ₹{target_p:.2f}  [{direction}]"
        )

        # ── Monitor loop ──────────────────────────────────────────────────────
        highest_price   = avg_price
        mon_start       = time.time()
        last_recheck    = time.time()
        last_heartbeat  = time.time()
        last_trail_exit = None
        sell_price      = avg_price
        entry_dir       = direction

        while True:
            # Heartbeat every 30 s
            if time.time() - last_heartbeat > 30:
                print(f"  💓 Monitoring {entry_dir} @ ₹{sell_price:.2f}"
                      f"  SL:{hard_sl:.2f}  Tgt:{target_p:.2f}"
                      f"  High:{highest_price:.2f}")
                last_heartbeat = time.time()

            try:
                cur_ltp = get_ltp_for_instrument(instrument, access_token, verbose=False, delay=0)
            except Exception:
                time.sleep(poll)
                continue
            if cur_ltp is None:
                time.sleep(poll)
                continue

            cur_ltp    = float(cur_ltp)
            elapsed    = time.time() - mon_start
            sell_price = cur_ltp
            sell_reason = None

            # Hard SL
            if cur_ltp <= hard_sl:
                sell_reason = f"🛑 Hard SL hit @ ₹{cur_ltp:.2f}"

            # Target hit
            elif cur_ltp >= target_p:
                sell_reason = f"🎯 Target hit @ ₹{cur_ltp:.2f}"

            # Max trade duration
            elif elapsed >= max_trade_sec:
                sell_reason = (f"⏰ Max trade time ({AUTO_V2_CONFIG['MAX_TRADE_MIN']}min)"
                               f" @ ₹{cur_ltp:.2f}")

            # Market-close guard
            elif not _within_trading_hours_auto():
                sell_reason = f"🔔 Market closing — squaring off @ ₹{cur_ltp:.2f}"

            # Signal-reversal check (every SIGNAL_RECHECK_SEC)
            elif time.time() - last_recheck >= AUTO_V2_CONFIG["SIGNAL_RECHECK_SEC"]:
                last_recheck = time.time()
                fresh = _read_master_signal_latest(index)
                if fresh:
                    new_dir  = fresh.get("direction", "WAIT")
                    new_conf = float(fresh.get("confidence", 0))
                    if (new_dir not in (entry_dir, "WAIT")
                            and new_conf >= AUTO_V2_CONFIG["MIN_MASTER_CONFIDENCE"]):
                        sell_reason = (
                            f"🔄 MASTER flipped to {new_dir} ({new_conf:.0f}%)"
                            f" — exiting {entry_dir}"
                        )

            # Trailing stop (kicks in after trail_start_p profit)
            if not sell_reason:
                if cur_ltp > highest_price:
                    highest_price = cur_ltp
                if highest_price >= avg_price + trail_start_p:
                    trail_exit = round_to_nearest_5_paise(highest_price - trail_step)
                    if trail_exit != last_trail_exit:
                        print(f"  📉 Trail | LTP={cur_ltp:.2f}"
                              f"  High={highest_price:.2f}  Exit={trail_exit:.2f}")
                        last_trail_exit = trail_exit
                    if cur_ltp <= trail_exit:
                        sell_reason = f"🔻 Trail stop hit @ ₹{cur_ltp:.2f}"

            if sell_reason:
                print(sell_reason)
                is_profit = sell_price > avg_price
                play_sound_async(SOUND_PROFIT if is_profit else SOUND_SL)

                # ── SELL order ────────────────────────────────────────────────
                try:
                    sell_resp     = place_market_order_groww(instrument, bought_qty, "SELL", "MIS")
                    sell_order_id = (sell_resp.get("payload", {}).get("groww_order_id")
                                     or sell_resp.get("groww_order_id"))
                    if cfg.get("VALIDATE_ORDERS", True) and sell_order_id:
                        s_stat = wait_for_order_status(sell_order_id, access_token, "SELL")
                        if s_stat in ["EXECUTED", "COMPLETED", "DELIVERY_AWAITED"]:
                            sp, sq = get_order_executed_price(sell_order_id, access_token)
                            if sp and sq:
                                sell_price, bought_qty = sp, sq
                except Exception as ex:
                    print(f"❌ SELL failed: {ex}")

                profit       = (sell_price - avg_price) * bought_qty
                total_pnl   += profit
                trade_count += 1
                icon = "🟢" if profit >= 0 else "🔴"
                print(f"{icon} Trade #{trade_count} [{mode_lbl}]"
                      f"  Entry ₹{avg_price:.2f}  Exit ₹{sell_price:.2f}"
                      f"  P&L ₹{profit:.2f}  |  Total ₹{total_pnl:.2f}")
                send_telegram(
                    f"🤖 AUTO v2 SELL [{mode_lbl}]: {symbol}"
                    f" @ ₹{sell_price:.2f}  P&L: ₹{profit:.2f}"
                    f"  Total: ₹{total_pnl:.2f}"
                )
                log_trade_to_excel(symbol, avg_price, sell_price, bought_qty, profit)
                _write_auto_status(
                    "TRADE_CLOSED",
                    mode_label=mode_lbl,
                    index=index,
                    expiry=expiry,
                    direction=direction,
                    symbol=symbol,
                    entry_price=avg_price,
                    exit_price=sell_price,
                    trade_pnl=profit,
                    total_pnl=total_pnl,
                    trade_count=trade_count,
                    exit_reason=sell_reason,
                )
                break

            time.sleep(poll)

        print("⏳ Cooldown 15s before next signal scan…")
        time.sleep(15)


# ----------------- Main menu -----------------
if __name__ == "__main__":
    print("\n" + "="*60)
    print("✨ Groww Multi-Index Options Trading Bot Ready")
    print("="*60)
    print(f"📊 Index: {CONFIG['index']} | Expiry: {CONFIG['expiry']}")
    print(f"💰 Lots: {CONFIG['lots']} | Poll: {CONFIG['POLL_INTERVAL']}s")
    
    if CONFIG.get("PAPER_TRADING", False):
        print("📋 PAPER TRADING MODE: All orders are SIMULATED — no real trades")
        print("   → LTP is fetched live; P&L and Excel log work normally")
        print("   → Set PAPER_TRADING = False in CONFIG for live trading")
    elif CONFIG.get("VALIDATE_ORDERS", True):
        print("✅ LIVE TRADING MODE: Order validation ENABLED")
        print("   → BUY/SELL orders will be verified before proceeding")
    else:
        print("⚠️  TESTING MODE: Order validation DISABLED")
        print("   → Using estimated prices (NOT recommended for live)")

    print("="*60)
    
    # Display initial account summary
    print("\n📊 Fetching initial account summary...")
    display_account_summary(access_token)
    
    print("="*60)
    print("Supported: NIFTY (NSE) | SENSEX (BSE) | BANKNIFTY | FINNIFTY")
    print("Manual example: 20 NIFTY17MAR202623150CE")
    print("                50 SENSEX12MAR202674600CE")
    print("Directional: c (Call) / p (Put) - Auto-selects option\n")

    # ── Live Dashboard Command Bridge ────────────────────────────────────────
    _BRIDGE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".prod10_bridge_cmd.json")
    _bridge_lock = threading.Lock()

    # Single-instance guard — only ONE PROD10 process may own the bridge.
    # (2026-08-04: two live instances both consumed the same dashboard click and
    # both placed real orders; the duplicates were only rejected by luck.)
    _BRIDGE_OWNER_LOCK = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".prod10_bridge.lock")

    def _claim_bridge_ownership():
        """Exclusive flock held for process lifetime; auto-released on exit/crash."""
        try:
            import fcntl
            fd = open(_BRIDGE_OWNER_LOCK, "w")
            fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
            fd.write(str(os.getpid()))
            fd.flush()
            return fd
        except Exception:
            return None

    _bridge_owner_fd = _claim_bridge_ownership()

    def _dashboard_bridge_watcher():
        import json as _json
        _hb_last = 0.0
        while True:
            _now = time.time()
            if _now - _hb_last >= 60:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] 🌐 [DASHBOARD] Bridge alive — idle, ready for commands")
                _hb_last = _now
            try:
                if os.path.exists(_BRIDGE_FILE):
                    # Atomic claim: rename succeeds for exactly one process, so a
                    # command can never be consumed twice even if two bots run.
                    _claim = f"{_BRIDGE_FILE}.claimed.{os.getpid()}"
                    try:
                        os.rename(_BRIDGE_FILE, _claim)
                    except OSError:
                        time.sleep(0.01)
                        continue  # another instance claimed it first
                    with open(_claim) as _f:
                        _data = _json.load(_f)
                    os.remove(_claim)
                    _cmd        = (_data.get("command") or "").strip()
                    _mode       = _data.get("mode", "manual")
                    _paper      = _data.get("paper", None)      # None = keep CONFIG default
                    _mock       = bool(_data.get("mock", False))
                    _validate   = _data.get("validate_orders", None)  # None = keep CONFIG default
                    _atr_flag    = bool(_data.get("atr", True))   # ATR-based SL on/off
                    _atr_src     = str(_data.get("atr_source", "candle"))  # "candle" or "scan"
                    _quick_pts   = float(_data.get("quick_pts", 1.5))
                    _partial     = bool(_data.get("partial", False))
                    _partial_pct = int(_data.get("partial_pct", 50))
                    _ltp_hint    = float(_data.get("ltp", 0) or 0)  # chain LTP from dashboard
                    # ── Runtime target update (no new trade, just changes target mid-run) ──
                    if _cmd == "set_quick_pts":
                        _new_tgt = float(_data.get("quick_pts", 0))
                        if _new_tgt > 0:
                            _QUICK_RUNTIME_TARGET[0] = _new_tgt
                            print(f"\n🎯 [DASHBOARD] Runtime target update → +{_new_tgt}pt (will apply on next LTP tick)")
                        continue
                    if _cmd == "set_partial":
                        _QUICK_RUNTIME_PARTIAL[0] = {
                            "partial":     bool(_data.get("partial", False)),
                            "partial_pct": int(_data.get("partial_pct", 50)),
                        }
                        print(f"\n📊 [DASHBOARD] Runtime partial update → {'ON' if _data.get('partial') else 'OFF'} {_data.get('partial_pct',50)}%")
                        continue
                    # auto mode needs no command string; all others require one
                    _actionable = _cmd or _mode == "auto"
                    if _actionable:
                        if _bridge_lock.acquire(blocking=False):
                            _vflag = '' if _validate is None else (' VALIDATE' if _validate else ' NO-VALIDATE')
                            print(f"\n\n🌐 [DASHBOARD] Command received: {_cmd or '(auto)'}  (mode={_mode}{' PAPER' if _paper else ''}{' MOCK-RUN' if _mock else ''}{_vflag})")
                            def _run(_c=_cmd, _m=_mode, _p=_paper, _mk=_mock, _v=_validate, _lk=_bridge_lock,
                                     _af=_atr_flag, _as=_atr_src, _qp=_quick_pts,
                                     _pt=_partial, _pp=_partial_pct, _lh=_ltp_hint):
                                _orig       = CONFIG.get("PAPER_TRADING")
                                _orig_mock  = CONFIG.get("MOCK_LTP_RUN", False)
                                _orig_val   = CONFIG.get("VALIDATE_ORDERS", False)
                                try:
                                    if _p is not None:
                                        CONFIG["PAPER_TRADING"] = bool(_p)
                                    CONFIG["MOCK_LTP_RUN"] = bool(_mk)
                                    if _v is not None:
                                        CONFIG["VALIDATE_ORDERS"] = bool(_v)
                                    if _m == "auto":
                                        auto_mode_runner()
                                    elif _m == "quick":
                                        place_quick_order(_c, atr_based=_af, quick_pts=_qp, atr_source=_as, partial=_pt, partial_pct=_pp, ltp_hint=_lh)
                                    else:
                                        place_cp_order(_c)
                                except Exception as _exc:
                                    import traceback as _tb
                                    print(f"\n❌ [DASHBOARD] {_m} mode crashed: {_exc}")
                                    _tb.print_exc()
                                finally:
                                    if _p is not None:
                                        CONFIG["PAPER_TRADING"] = _orig
                                    CONFIG["MOCK_LTP_RUN"] = _orig_mock
                                    if _v is not None:
                                        CONFIG["VALIDATE_ORDERS"] = _orig_val
                                    _lk.release()
                            threading.Thread(target=_run, daemon=True, name="DashBridge").start()
                        else:
                            print("\n⚠️  [DASHBOARD] Ignored — bot is already executing an order.")
            except Exception:
                pass
            time.sleep(0.01)

    if _bridge_owner_fd:
        threading.Thread(target=_dashboard_bridge_watcher, daemon=True, name="DashboardBridgeWatcher").start()
        print("🌐 Live Dashboard bridge active — select a strike in the Dashboard and click → PROD10")
    else:
        try:
            _owner_pid = open(_BRIDGE_OWNER_LOCK).read().strip() or "?"
        except Exception:
            _owner_pid = "?"
        print(f"🚫 Dashboard bridge DISABLED in this instance — another PROD10 (PID {_owner_pid}) already owns it.")
        print("   Dashboard clicks go to that instance only. Close it and restart this one to take over.")
        send_telegram(f"⚠️ PROD10 started with bridge DISABLED — another instance (PID {_owner_pid}) owns the dashboard bridge.")
    # ── End Bridge ───────────────────────────────────────────────────────────

    while True:
        mode = input("Choose mode: (m)anual / (q)uick / (d)irectional / (a)uto / (e)xit: ").strip().lower()
        if mode in ["e", "exit", "quit"]:
            print("Exiting.")
            break
        if mode in ["a", "auto"]:
            auto_mode_runner()
            continue
        if mode in ["d", "directional", "dir"]:
            directional_mode()
            continue
        if mode in ["q", "quick"]:
            user_input = input("\n⚡ QUICK MODE - Enter command (buy + instant 1.5pt target): ").strip()
            command_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
            print(f"⏱️  Command entered at: {command_time}")
            if user_input.lower() in ["back"]:
                continue
            if user_input == "":
                continue
            place_quick_order(user_input)
            continue
        if mode in ["m", "manual"]:
            user_input = input("\nEnter command (or press Enter for status, type 'back' to menu): ").strip()
            command_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]  # Include milliseconds
            print(f"⏱️  Command entered at: {command_time}")
            if user_input.lower() in ["back"]:
                continue
            if user_input == "":
                print("Status check not implemented for Groww PnL in this script.")
                continue
            place_cp_order(user_input)
            continue
        print("Invalid input. Choose 'm' (manual), 'q' (quick), 'a' (auto), or 'e' (exit).")