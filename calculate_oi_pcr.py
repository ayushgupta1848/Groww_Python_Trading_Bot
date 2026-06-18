#!/usr/bin/env python3
import requests
import time
import json
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import PatternFill
import os

# Set your Anthropic API key as an environment variable before running:
#   export ANTHROPIC_API_KEY="your-key-here"


def next_expiry(weekday=3):
    """Return the nearest upcoming weekday (0=Mon…6=Sun). NIFTY=Thursday(3)."""
    today = datetime.now().date()
    days_ahead = (weekday - today.weekday()) % 7
    if days_ahead == 0:
        days_ahead = 7
    return (today + timedelta(days=days_ahead)).strftime("%-d-%b-%Y")


# ================= CONFIG =================
URL = "https://www.nseindia.com/api/NextApi/apiClient/GetQuoteApi"
PARAMS = {
    "functionName": "getOptionChainData",
    "symbol": "NIFTY",
    "params": f"expiryDate={next_expiry(weekday=1)}"  # NIFTY expires on Tuesday (weekday=1)
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",   # no 'br' — requests cannot decode Brotli
    "Referer": "https://www.nseindia.com/",
    "Connection": "keep-alive",
}

REFRESH_SECONDS = 60
EXCEL_FILE = "oi_pcr_dashboard.xlsx"
LOT_SIZE = 75  # NIFTY lot size (change to 20 for SENSEX, 35 for BANKNIFTY, etc.)

V3_URL    = "https://www.nseindia.com/api/option-chain-v3"
V3_SYMBOL = PARAMS.get("symbol", "NIFTY")

session = requests.Session()
session.headers.update(HEADERS)

prev_price = None
prev_ce_oi = None
prev_pe_oi = None
prev_strike_data = {}  # {strike: {"ce_oi": x, "pe_oi": x}} — persists across ticks for writer tracking
prev_ltp_data    = {}  # {strike: {"ce_ltp": x, "pe_ltp": x}} — for writing detection (OI ↑ + LTP ↓)
prev_iv_data     = {}  # {strike: {"ce_iv": x, "pe_iv": x}} — for IV change detection
prev_pcr_all     = None  # previous tick's PCR — for PCR change signal

# ================= HELPERS =================
def fetch_data():
    r = session.get(URL, params=PARAMS, timeout=15)
    if r.status_code == 403 or len(r.content) == 0:
        # NSE cookie expired — visit homepage to refresh, then retry
        try: session.get("https://www.nseindia.com/", timeout=10)
        except Exception: pass
        time.sleep(1)
        r = session.get(URL, params=PARAMS, timeout=15)
    if len(r.content) == 0:
        raise ValueError(f"NSE returned empty body (status {r.status_code})")
    r.raise_for_status()
    return r.json()

def fetch_v3_extras(atm):
    """Fetch IV, Volume, Max Pain from NSE option-chain-v3 API (runs alongside main fetch)."""
    expiry = next_expiry(weekday=1)
    try:
        r = session.get(V3_URL,
                        params={"type": "Indices", "symbol": V3_SYMBOL, "expiry": expiry},
                        headers={"Referer": "https://www.nseindia.com/option-chain"},
                        timeout=15)
        if r.status_code == 403 or len(r.content) == 0:
            return {}
        r.raise_for_status()
        data = r.json().get("records", {}).get("data", [])

        # Max Pain — strike where total options-writer monetary loss is minimised
        strikes = sorted({d.get("strikePrice", 0) for d in data if d.get("strikePrice")})
        pain = {}
        for exp_s in strikes:
            pain[exp_s] = sum(
                d.get("CE", {}).get("openInterest", 0) * max(0, exp_s - d.get("strikePrice", 0)) +
                d.get("PE", {}).get("openInterest", 0) * max(0, d.get("strikePrice", 0) - exp_s)
                for d in data
            )
        max_pain = min(pain, key=pain.get) if pain else 0

        # Volume PCR — intraday sentiment (faster than OI PCR)
        ce_vol = sum(d.get("CE", {}).get("totalTradedVolume", 0) for d in data)
        pe_vol = sum(d.get("PE", {}).get("totalTradedVolume", 0) for d in data)
        vol_pcr = round(pe_vol / ce_vol, 3) if ce_vol else 0

        # ATM Implied Volatility + skew
        atm_row = next((d for d in data if d.get("strikePrice") == atm), {})
        atm_ce_iv = round(atm_row.get("CE", {}).get("impliedVolatility", 0), 1)
        atm_pe_iv = round(atm_row.get("PE", {}).get("impliedVolatility", 0), 1)
        iv_skew   = round(atm_ce_iv - atm_pe_iv, 1)   # +ve = CE priced higher = upside fear

        # Per-strike IV, LTP, and volume for ATM±3 + full IV map for change detection
        atm_extras = {}
        iv_map = {}   # {strike: {"ce_iv": x, "pe_iv": x}} — all strikes, for prev IV comparison
        for d in data:
            sp = d.get("strikePrice", 0)
            if not sp:
                continue
            ce_iv_v = round(d.get("CE", {}).get("impliedVolatility", 0), 1)
            pe_iv_v = round(d.get("PE", {}).get("impliedVolatility", 0), 1)
            iv_map[sp] = {"ce_iv": ce_iv_v, "pe_iv": pe_iv_v}
            if abs(sp - atm) <= 150:
                atm_extras[sp] = {
                    "ce_iv":  ce_iv_v,
                    "pe_iv":  pe_iv_v,
                    "ce_ltp": d.get("CE", {}).get("lastPrice", 0),
                    "pe_ltp": d.get("PE", {}).get("lastPrice", 0),
                    "ce_vol": d.get("CE", {}).get("totalTradedVolume", 0),
                    "pe_vol": d.get("PE", {}).get("totalTradedVolume", 0),
                }

        # Smart Money Flow — top 5 CE and PE strikes by today's session OI addition
        ce_oi_adds, pe_oi_adds = [], []
        for d in data:
            sp = d.get("strikePrice", 0)
            if not sp:
                continue
            ce_chg_v3 = d.get("CE", {}).get("changeinOpenInterest", 0)
            pe_chg_v3 = d.get("PE", {}).get("changeinOpenInterest", 0)
            if ce_chg_v3 > 0:
                ce_oi_adds.append({"strike": sp, "oi_change": ce_chg_v3,
                                   "ltp": d.get("CE", {}).get("lastPrice", 0),
                                   "vol": d.get("CE", {}).get("totalTradedVolume", 0)})
            if pe_chg_v3 > 0:
                pe_oi_adds.append({"strike": sp, "oi_change": pe_chg_v3,
                                   "ltp": d.get("PE", {}).get("lastPrice", 0),
                                   "vol": d.get("PE", {}).get("totalTradedVolume", 0)})
        ce_oi_adds.sort(key=lambda x: x["oi_change"], reverse=True)
        pe_oi_adds.sort(key=lambda x: x["oi_change"], reverse=True)

        return {
            "max_pain":      max_pain,
            "vol_pcr":       vol_pcr,
            "total_ce_vol":  ce_vol,
            "total_pe_vol":  pe_vol,
            "atm_ce_iv":     atm_ce_iv,
            "atm_pe_iv":     atm_pe_iv,
            "iv_skew":       iv_skew,
            "atm_extras":    atm_extras,
            "smart_money_ce": ce_oi_adds[:5],
            "smart_money_pe": pe_oi_adds[:5],
            "iv_map":         iv_map,
        }
    except Exception as e:
        print(f"⚠️  v3 extras fetch failed: {e}")
        return {}


def coi_pct(ce, pe):
    total = abs(ce) + abs(pe)
    if total == 0:
        return 0, 0
    return abs(ce) * 100 / total, abs(pe) * 100 / total

def activity(price, prev_price, oi, prev_oi):
    if prev_price is None or prev_oi is None:
        return "N/A"
    if price > prev_price and oi > prev_oi:
        return "LONG BUILDUP"
    if price < prev_price and oi > prev_oi:
        return "SHORT BUILDUP"
    if price > prev_price and oi < prev_oi:
        return "SHORT COVERING"
    if price < prev_price and oi < prev_oi:
        return "LONG UNWINDING"
    return "SIDEWAYS"

def power(price, prev_price, chg):
    if prev_price is None:
        return "N/A"
    if chg > 0 and price < prev_price:
        return "WRITERS"
    if chg > 0 and price > prev_price:
        return "BUYERS"
    return "NEUTRAL"

# ================= RULE-BASED TRADING SUMMARY =================
def generate_trading_summary(m, price_change):
    ce = m["total_oi_ce"]
    pe = m["total_oi_pe"]
    ce_chg = m["total_chg_ce"]
    pe_chg = m["total_chg_pe"]

    # 1. Market bias
    if pe > ce * 1.1:
        bias_line = "Bullish bias — Put OI dominates Call OI."
        bias = "BULLISH"
    elif ce > pe * 1.1:
        bias_line = "Bearish bias — Call OI dominates Put OI."
        bias = "BEARISH"
    else:
        bias_line = "Neutral — Call and Put OI are balanced."
        bias = "NEUTRAL"

    # 2. Writing activity (change OI)
    if pe_chg > 0 and ce_chg <= 0:
        writing_line = "Strong put writing with call unwinding — support forming."
        bias = "BULLISH"
    elif ce_chg > 0 and pe_chg <= 0:
        writing_line = "Strong call writing with put unwinding — resistance forming."
        bias = "BEARISH"
    elif pe_chg > ce_chg and pe_chg > 0:
        writing_line = "Put writing observed — support building up."
    elif ce_chg > pe_chg and ce_chg > 0:
        writing_line = "Call writing observed — resistance building up."
    else:
        writing_line = "No significant writing activity detected."

    # 3. Price-OI confirmation
    if price_change > 0 and pe_chg > 0:
        price_line = "Price up with OI build-up — uptrend confirmed."
    elif price_change < 0 and ce_chg > 0:
        price_line = "Price down with OI build-up — downtrend continuation."
    elif price_change > 0 and ce_chg < 0:
        price_line = "Price up with call OI falling — short covering rally."
    elif price_change < 0 and pe_chg < 0:
        price_line = "Price down with put OI falling — long unwinding."
    else:
        price_line = "Price movement not showing strong OI confirmation."

    # 4. ATM strike diff dominance
    atm_d = m["atm_strikes_oi"].get(m["atm"], {})
    atm_diff = atm_d.get("pe_oi", 0) - atm_d.get("ce_oi", 0)
    if atm_diff > 0:
        final_bias = "Bias: Prefer CALL side — PE dominant at ATM."
    elif atm_diff < 0:
        final_bias = "Bias: Prefer PUT side — CE dominant at ATM."
    else:
        final_bias = f"Bias: Prefer {'CALL' if bias == 'BULLISH' else 'PUT' if bias == 'BEARISH' else 'WAIT'} side."

    return f"{bias_line}\n{writing_line}\n{price_line}\n{final_bias}"


# ================= WRITER ACTIVITY ANALYSIS =================
def analyze_writer_activity(ce_map, pe_map, prev_data):
    """
    Compare current tick OI to previous tick at every strike.
    Identifies where writers are adding/removing positions and
    derives a net directional bias from that activity.
    Returns None on first tick (no baseline yet).
    """
    if not prev_data:
        return None

    results = {}
    ce_writing, pe_writing = [], []
    ce_unwinding, pe_unwinding = [], []

    for strike in sorted(set(ce_map) | set(pe_map)):
        curr_ce = ce_map.get(strike, 0)
        curr_pe = pe_map.get(strike, 0)
        prev = prev_data.get(strike, {})
        prev_ce = prev.get("ce_oi", curr_ce)
        prev_pe = prev.get("pe_oi", curr_pe)

        ce_delta = curr_ce - prev_ce
        pe_delta = curr_pe - prev_pe

        ce_action = "WRITING" if ce_delta > 0 else "UNWINDING" if ce_delta < 0 else "—"
        pe_action = "WRITING" if pe_delta > 0 else "UNWINDING" if pe_delta < 0 else "—"

        results[strike] = {
            "ce_oi": curr_ce, "pe_oi": curr_pe,
            "ce_delta": ce_delta, "pe_delta": pe_delta,
            "ce_action": ce_action, "pe_action": pe_action,
        }

        if ce_delta > 0:
            ce_writing.append((strike, ce_delta))
        elif ce_delta < 0:
            ce_unwinding.append((strike, abs(ce_delta)))

        if pe_delta > 0:
            pe_writing.append((strike, pe_delta))
        elif pe_delta < 0:
            pe_unwinding.append((strike, abs(pe_delta)))

    ce_writing.sort(key=lambda x: x[1], reverse=True)
    pe_writing.sort(key=lambda x: x[1], reverse=True)

    # Bullish: puts being written (support) + calls unwinding (bears exiting)
    # Bearish: calls being written (resistance) + puts unwinding (bulls exiting)
    bullish_score = sum(d for _, d in pe_writing) + sum(d for _, d in ce_unwinding)
    bearish_score = sum(d for _, d in ce_writing) + sum(d for _, d in pe_unwinding)

    if bullish_score > bearish_score * 1.15:
        writer_bias = "BULLISH"
    elif bearish_score > bullish_score * 1.15:
        writer_bias = "BEARISH"
    else:
        writer_bias = "NEUTRAL"

    return {
        "strike_activity": results,
        "ce_writing": ce_writing[:5],
        "pe_writing": pe_writing[:5],
        "ce_unwinding": ce_unwinding[:5],
        "pe_unwinding": pe_unwinding[:5],
        "bullish_score": round(bullish_score / 1e6, 2),
        "bearish_score": round(bearish_score / 1e6, 2),
        "writer_bias": writer_bias,
    }


def print_writer_activity(wa, price, atm):
    """Pretty-print writer activity analysis to console."""
    if wa is None:
        print("\n--- WRITER ACTIVITY --- (first tick — no baseline yet)")
        return

    print("\n--- WRITER ACTIVITY (tick-over-tick OI change) ---")
    print(f"  Writer Bias  : {wa['writer_bias']}")
    print(f"  Bullish Score: {wa['bullish_score']:.2f}M  (PE writing + CE unwinding)")
    print(f"  Bearish Score: {wa['bearish_score']:.2f}M  (CE writing + PE unwinding)")

    # CALL writers ADDING = resistance building = market may not cross = PE favours (bearish)
    if wa["ce_writing"]:
        strikes_str = "  |  ".join(f"{s} (+{d/1e6:.2f}M)" for s, d in wa["ce_writing"][:3])
        print(f"\n  CALL writers ADDING   → {strikes_str}  [PE favours — Resistance building, market capped]")
    # CALL writers EXITING = resistance weakening = upside opening = CE favours (bullish)
    if wa["ce_unwinding"]:
        strikes_str = "  |  ".join(f"{s} (-{d/1e6:.2f}M)" for s, d in wa["ce_unwinding"][:3])
        print(f"  CALL writers EXITING  → {strikes_str}  [CE favours — Resistance weakening, upside open]")
    # PUT writers ADDING = support building = market held up = CE favours (bullish)
    if wa["pe_writing"]:
        strikes_str = "  |  ".join(f"{s} (+{d/1e6:.2f}M)" for s, d in wa["pe_writing"][:3])
        print(f"  PUT  writers ADDING   → {strikes_str}  [CE favours — Support building, floor forming]")
    # PUT writers EXITING = support weakening = downside opening = PE favours (bearish)
    if wa["pe_unwinding"]:
        strikes_str = "  |  ".join(f"{s} (-{d/1e6:.2f}M)" for s, d in wa["pe_unwinding"][:3])
        print(f"  PUT  writers EXITING  → {strikes_str}  [PE favours — Support weakening, downside risk]")

    # ATM strike breakdown for above/below current price
    above_ce_writes = [(s, d) for s, d in wa["ce_writing"] if s > price][:2]
    below_pe_writes = [(s, d) for s, d in wa["pe_writing"] if s < price][:2]
    if above_ce_writes:
        print(f"\n  KEY RESISTANCE (call writing above {price:.0f}): "
              + "  |  ".join(f"{s}" for s, _ in above_ce_writes)
              + "  → PE favours if price stays below")
    if below_pe_writes:
        print(f"  KEY SUPPORT   (put writing below {price:.0f}): "
              + "  |  ".join(f"{s}" for s, _ in below_pe_writes)
              + "  → CE favours if price holds above")


def compute_oi_signals(m, wa, extras, pcr_all, prev_price_val, prev_ltp):
    """
    10-factor bull/bear score (0-100 each) → market direction signal.
    Returns signal dict: market_signal, bull_score_v2, bear_score_v2,
    momentum_score, signal_list, call_writing, put_writing.
    """
    bull = 0
    bear = 0
    signal_list = []

    def add(key, label, direction, pts):
        nonlocal bull, bear
        if direction == "bull":
            bull += pts
        elif direction == "bear":
            bear += pts
        signal_list.append({"key": key, "label": label, "dir": direction, "pts": pts})

    price   = m["price"]
    ce_oi   = m["total_oi_ce"]
    pe_oi   = m["total_oi_pe"]
    ce_sess = m["total_chg_ce"]
    pe_sess = m["total_chg_pe"]

    # ── 1. PCR overall (0–20) ──
    if pcr_all >= 1.5:
        add("pcr", f"PCR {pcr_all:.2f} ≥1.5 — STRONG BULLISH", "bull", 20)
    elif pcr_all >= 1.2:
        add("pcr", f"PCR {pcr_all:.2f} ≥1.2 — BULLISH", "bull", 14)
    elif pcr_all >= 1.0:
        add("pcr", f"PCR {pcr_all:.2f} — mild BULLISH", "bull", 7)
    elif pcr_all <= 0.6:
        add("pcr", f"PCR {pcr_all:.2f} ≤0.6 — STRONG BEARISH", "bear", 20)
    elif pcr_all <= 0.8:
        add("pcr", f"PCR {pcr_all:.2f} ≤0.8 — BEARISH", "bear", 14)
    else:
        add("pcr", f"PCR {pcr_all:.2f} — mild BEARISH", "bear", 7)

    # ── 2. OI Imbalance (0–12) ──
    if ce_oi + pe_oi > 0:
        imb = (pe_oi - ce_oi) / (ce_oi + pe_oi)
        if imb > 0.2:
            add("oi_imb", f"PE OI {imb*100:.0f}% dominant — BULLISH", "bull", 12)
        elif imb > 0.1:
            add("oi_imb", f"PE OI {imb*100:.0f}% dominant — mild BULLISH", "bull", 7)
        elif imb < -0.2:
            add("oi_imb", f"CE OI {abs(imb)*100:.0f}% dominant — BEARISH", "bear", 12)
        elif imb < -0.1:
            add("oi_imb", f"CE OI {abs(imb)*100:.0f}% dominant — mild BEARISH", "bear", 7)
        else:
            add("oi_imb", "OI balanced — NEUTRAL", "neutral", 0)

    # ── 3. Session OI change direction (0–12) ──
    if pe_sess > 0 and ce_sess <= 0:
        add("sess_chg", f"Put addition +{pe_sess/1e6:.1f}M + Call unwind — BULLISH", "bull", 12)
    elif pe_sess > 0 and pe_sess > ce_sess * 1.3:
        add("sess_chg", f"PE session add {pe_sess/1e6:.1f}M > CE — BULLISH", "bull", 7)
    elif ce_sess > 0 and pe_sess <= 0:
        add("sess_chg", f"Call addition +{ce_sess/1e6:.1f}M + Put unwind — BEARISH", "bear", 12)
    elif ce_sess > 0 and ce_sess > pe_sess * 1.3:
        add("sess_chg", f"CE session add {ce_sess/1e6:.1f}M > PE — BEARISH", "bear", 7)
    else:
        add("sess_chg", f"Session OI neutral (CE:{ce_sess/1e6:.1f}M PE:{pe_sess/1e6:.1f}M)", "neutral", 0)

    # ── 4. Tick writer activity (0–15) ──
    if wa:
        bw = wa.get("bullish_score", 0)
        brw = wa.get("bearish_score", 0)
        if bw > brw * 1.5:
            add("writer", f"PE writing {bw:.1f}M > CE {brw:.1f}M — WRITERS BULLISH", "bull", 15)
        elif bw > brw:
            add("writer", f"Mild PE writer bias ({bw:.1f}M vs {brw:.1f}M)", "bull", 7)
        elif brw > bw * 1.5:
            add("writer", f"CE writing {brw:.1f}M > PE {bw:.1f}M — WRITERS BEARISH", "bear", 15)
        elif brw > bw:
            add("writer", f"Mild CE writer bias ({brw:.1f}M vs {bw:.1f}M)", "bear", 7)
        else:
            add("writer", f"Writer neutral (bull:{bw:.1f}M bear:{brw:.1f}M)", "neutral", 0)
    else:
        add("writer", "Writer activity — awaiting 2nd tick", "neutral", 0)

    # ── 5. Smart Money Flow — session OI additions (0–15) ──
    sm_ce = extras.get("smart_money_ce", []) if extras else []
    sm_pe = extras.get("smart_money_pe", []) if extras else []
    sm_ce_tot = sum(x.get("oi_change", 0) for x in sm_ce)
    sm_pe_tot = sum(x.get("oi_change", 0) for x in sm_pe)
    if sm_pe_tot + sm_ce_tot > 0:
        if sm_pe_tot > sm_ce_tot * 1.3:
            add("smart_money", f"Smart money: PE {sm_pe_tot/1e3:.0f}K > CE {sm_ce_tot/1e3:.0f}K lots — BULLISH", "bull", 15)
        elif sm_pe_tot > sm_ce_tot:
            add("smart_money", f"Smart money: mild PE bias ({sm_pe_tot/1e3:.0f}K lots)", "bull", 7)
        elif sm_ce_tot > sm_pe_tot * 1.3:
            add("smart_money", f"Smart money: CE {sm_ce_tot/1e3:.0f}K > PE {sm_pe_tot/1e3:.0f}K lots — BEARISH", "bear", 15)
        elif sm_ce_tot > sm_pe_tot:
            add("smart_money", f"Smart money: mild CE bias ({sm_ce_tot/1e3:.0f}K lots)", "bear", 7)
        else:
            add("smart_money", "Smart money: balanced", "neutral", 0)
    else:
        add("smart_money", "Smart money — v3 data pending", "neutral", 0)

    # ── 6. Volume PCR (0–10) ──
    vol_pcr = extras.get("vol_pcr", 0) if extras else 0
    if vol_pcr >= 1.2:
        add("vol_pcr", f"Vol PCR {vol_pcr:.2f} — strong PE volume BULLISH", "bull", 10)
    elif vol_pcr >= 1.0:
        add("vol_pcr", f"Vol PCR {vol_pcr:.2f} — mild PE volume", "bull", 4)
    elif 0 < vol_pcr <= 0.8:
        add("vol_pcr", f"Vol PCR {vol_pcr:.2f} — strong CE volume BEARISH", "bear", 10)
    elif 0 < vol_pcr < 1.0:
        add("vol_pcr", f"Vol PCR {vol_pcr:.2f} — mild CE volume", "bear", 4)
    else:
        add("vol_pcr", "Vol PCR — no data", "neutral", 0)

    # ── 7. Max Pain position (0–5) ──
    max_pain = extras.get("max_pain", 0) if extras else 0
    if max_pain:
        mp_dist = price - max_pain
        if mp_dist > 150:
            add("max_pain", f"Spot {mp_dist:+.0f}pts above Max Pain {max_pain} — BULLISH", "bull", 5)
        elif mp_dist < -150:
            add("max_pain", f"Spot {mp_dist:+.0f}pts below Max Pain {max_pain} — BEARISH", "bear", 5)
        else:
            add("max_pain", f"Spot near Max Pain {max_pain} ({mp_dist:+.0f}pts) — expiry magnet", "neutral", 0)
    else:
        add("max_pain", "Max Pain — v3 data pending", "neutral", 0)

    # ── 8. IV Skew (0–8) ──
    iv_skew = extras.get("iv_skew", 0) if extras else 0
    if iv_skew > 3:
        add("iv_skew", f"IV Skew +{iv_skew}% CE costly — upside expectation (BULLISH)", "bull", 8)
    elif iv_skew > 1:
        add("iv_skew", f"IV Skew +{iv_skew}% mild CE premium", "bull", 3)
    elif iv_skew < -3:
        add("iv_skew", f"IV Skew {iv_skew}% PE costly — downside fear (BEARISH)", "bear", 8)
    elif iv_skew < -1:
        add("iv_skew", f"IV Skew {iv_skew}% mild PE premium", "bear", 3)
    elif iv_skew != 0:
        add("iv_skew", f"IV Skew {iv_skew:+.1f}% — neutral", "neutral", 0)
    else:
        add("iv_skew", "IV Skew — no data", "neutral", 0)

    # ── 9. Price action / Buildup type (0–10) ──
    writer_bias = wa.get("writer_bias", "NEUTRAL") if wa else "NEUTRAL"
    if prev_price_val is not None:
        pc = price - prev_price_val
        if pc > 5 and writer_bias == "BULLISH":
            add("buildup", f"LONG BUILDUP — Price ↑{pc:.0f} + PE writers adding", "bull", 10)
        elif pc < -5 and writer_bias == "BEARISH":
            add("buildup", f"SHORT BUILDUP — Price ↓{abs(pc):.0f} + CE writers adding", "bear", 10)
        elif pc > 5 and writer_bias == "BEARISH":
            add("buildup", f"SHORT COVERING — Price ↑{pc:.0f} + CE writers exiting", "bull", 6)
        elif pc < -5 and writer_bias == "BULLISH":
            add("buildup", f"LONG UNWINDING — Price ↓{abs(pc):.0f} + PE writers exiting", "bear", 6)
        else:
            add("buildup", f"SIDEWAYS — ΔPrice {pc:+.0f}pts, Writers: {writer_bias}", "neutral", 0)
    else:
        add("buildup", "Buildup detection — awaiting 2nd tick", "neutral", 0)

    # ── 10. Volume Spike at strike level (0–5) ──
    ce_vol_map = m.get("ce_vol_map", {})
    pe_vol_map = m.get("pe_vol_map", {})
    if ce_vol_map and pe_vol_map:
        top_ce_sp = max(ce_vol_map, key=ce_vol_map.get)
        top_pe_sp = max(pe_vol_map, key=pe_vol_map.get)
        top_ce_v  = ce_vol_map[top_ce_sp]
        top_pe_v  = pe_vol_map[top_pe_sp]
        if top_pe_v > top_ce_v * 1.5:
            add("vol_spike", f"PE vol spike at {top_pe_sp} ({top_pe_v:,} lots) — BULLISH", "bull", 5)
        elif top_ce_v > top_pe_v * 1.5:
            add("vol_spike", f"CE vol spike at {top_ce_sp} ({top_ce_v:,} lots) — BEARISH", "bear", 5)
        else:
            add("vol_spike", f"Volume balanced. Top CE:{top_ce_sp}({top_ce_v:,}) PE:{top_pe_sp}({top_pe_v:,})", "neutral", 0)
    else:
        add("vol_spike", "Volume data unavailable", "neutral", 0)

    # ── Writing detection — tick LTP confirmation (CE OI↑ + LTP↓) ──
    call_writing = []
    put_writing  = []
    ce_chg_map = m.get("ce_chg_map", {})
    pe_chg_map = m.get("pe_chg_map", {})
    ce_ltp_map = m.get("ce_ltp_map", {})
    pe_ltp_map = m.get("pe_ltp_map", {})
    atm = m["atm"]

    for sp, chg in sorted(ce_chg_map.items()):
        if chg <= 0:
            continue
        ltp = ce_ltp_map.get(sp, 0)
        prev_ce_ltp = prev_ltp.get(sp, {}).get("ce_ltp", 0)
        ltp_chg = round(ltp - prev_ce_ltp, 2) if prev_ce_ltp else 0
        confirmed = bool(prev_ce_ltp and ltp < prev_ce_ltp)
        tag = "CONFIRMED" if confirmed else ("OTM" if sp > atm else "ITM")
        call_writing.append({"strike": sp, "oi_change": chg, "ltp": ltp, "ltp_chg": ltp_chg, "tag": tag})

    for sp, chg in sorted(pe_chg_map.items()):
        if chg <= 0:
            continue
        ltp = pe_ltp_map.get(sp, 0)
        prev_pe_ltp = prev_ltp.get(sp, {}).get("pe_ltp", 0)
        ltp_chg = round(ltp - prev_pe_ltp, 2) if prev_pe_ltp else 0
        confirmed = bool(prev_pe_ltp and ltp < prev_pe_ltp)
        tag = "CONFIRMED" if confirmed else ("OTM" if sp < atm else "ITM")
        put_writing.append({"strike": sp, "oi_change": chg, "ltp": ltp, "ltp_chg": ltp_chg, "tag": tag})

    call_writing.sort(key=lambda x: (x["tag"] == "CONFIRMED", x["oi_change"]), reverse=True)
    put_writing.sort(key=lambda x: (x["tag"] == "CONFIRMED", x["oi_change"]), reverse=True)

    conf_calls = [x for x in call_writing if x["tag"] == "CONFIRMED"]
    conf_puts  = [x for x in put_writing  if x["tag"] == "CONFIRMED"]
    if conf_calls:
        bear += 8
        signal_list.append({"key": "writing", "label": f"CALL WRITING CONFIRMED at {[x['strike'] for x in conf_calls[:2]]} — BEARISH", "dir": "bear", "pts": 8})
    elif conf_puts:
        bull += 8
        signal_list.append({"key": "writing", "label": f"PUT WRITING CONFIRMED at {[x['strike'] for x in conf_puts[:2]]} — BULLISH", "dir": "bull", "pts": 8})
    else:
        otm_c = [x["strike"] for x in call_writing if x["tag"] == "OTM"][:3]
        otm_p = [x["strike"] for x in put_writing  if x["tag"] == "OTM"][:3]
        if otm_c and not otm_p:
            bear += 3
            signal_list.append({"key": "writing", "label": f"OTM Call OI adding at {otm_c} (mild BEARISH)", "dir": "bear", "pts": 3})
        elif otm_p and not otm_c:
            bull += 3
            signal_list.append({"key": "writing", "label": f"OTM Put OI adding at {otm_p} (mild BULLISH)", "dir": "bull", "pts": 3})
        else:
            signal_list.append({"key": "writing", "label": "Writing detection — awaiting LTP confirmation", "dir": "neutral", "pts": 0})

    bull = min(bull, 100)
    bear = min(bear, 100)

    if bull >= 70:
        market_signal = "STRONG BULLISH"
    elif bull >= 45 and bull > bear + 10:
        market_signal = "BULLISH"
    elif bear >= 70:
        market_signal = "STRONG BEARISH"
    elif bear >= 45 and bear > bull + 10:
        market_signal = "BEARISH"
    else:
        market_signal = "NEUTRAL"

    total_oi  = m["total_oi_ce"] + m["total_oi_pe"]
    total_chg = abs(m["total_chg_ce"]) + abs(m["total_chg_pe"])
    momentum_score = min(100, int(total_chg / max(total_oi, 1) * 300))

    return {
        "bull_score_v2":  bull,
        "bear_score_v2":  bear,
        "market_signal":  market_signal,
        "momentum_score": momentum_score,
        "signal_list":    signal_list,
        "call_writing":   call_writing[:8],
        "put_writing":    put_writing[:8],
    }


def compute_strike_signals(m, extras, prev_strike, prev_ltp, prev_iv, prev_pcr, pcr_all):
    """
    Per-strike buildup analysis (ATM ±3), IV change detection,
    PCR change, and ATM momentum signal.
    Returns dict with: strike_buildups, iv_changes, pcr_change, atm_momentum.
    """
    atm = m["atm"]
    ce_map = m["ce_map"]
    pe_map = m["pe_map"]
    ce_ltp_map = m["ce_ltp_map"]
    pe_ltp_map = m["pe_ltp_map"]

    # ── Per-strike buildup (ATM ±3) ──
    atm_strikes = sorted(m["atm_strikes_oi"].keys())
    strike_buildups = []
    for sp in atm_strikes:
        prev = prev_strike.get(sp, {})
        prev_ce_oi = prev.get("ce_oi", ce_map.get(sp, 0))
        prev_pe_oi = prev.get("pe_oi", pe_map.get(sp, 0))
        curr_ce_oi = ce_map.get(sp, 0)
        curr_pe_oi = pe_map.get(sp, 0)
        prev_ce_ltp = prev_ltp.get(sp, {}).get("ce_ltp", 0)
        prev_pe_ltp = prev_ltp.get(sp, {}).get("pe_ltp", 0)
        curr_ce_ltp = ce_ltp_map.get(sp, 0)
        curr_pe_ltp = pe_ltp_map.get(sp, 0)

        ce_oi_chg = curr_ce_oi - prev_ce_oi
        pe_oi_chg = curr_pe_oi - prev_pe_oi

        ce_buildup = None
        if prev_ce_ltp and ce_oi_chg != 0:
            if curr_ce_ltp > prev_ce_ltp and ce_oi_chg > 0:
                ce_buildup = "LONG BUILDUP"
            elif curr_ce_ltp < prev_ce_ltp and ce_oi_chg > 0:
                ce_buildup = "SHORT BUILDUP"
            elif curr_ce_ltp > prev_ce_ltp and ce_oi_chg < 0:
                ce_buildup = "SHORT COVERING"
            elif curr_ce_ltp < prev_ce_ltp and ce_oi_chg < 0:
                ce_buildup = "LONG UNWINDING"

        pe_buildup = None
        if prev_pe_ltp and pe_oi_chg != 0:
            if curr_pe_ltp > prev_pe_ltp and pe_oi_chg > 0:
                pe_buildup = "LONG BUILDUP"
            elif curr_pe_ltp < prev_pe_ltp and pe_oi_chg > 0:
                pe_buildup = "SHORT BUILDUP"
            elif curr_pe_ltp > prev_pe_ltp and pe_oi_chg < 0:
                pe_buildup = "SHORT COVERING"
            elif curr_pe_ltp < prev_pe_ltp and pe_oi_chg < 0:
                pe_buildup = "LONG UNWINDING"

        if ce_buildup or pe_buildup:
            strike_buildups.append({
                "strike":     sp,
                "is_atm":     sp == atm,
                "ce_buildup": ce_buildup,
                "pe_buildup": pe_buildup,
                "ce_oi_chg":  ce_oi_chg,
                "pe_oi_chg":  pe_oi_chg,
                "ce_ltp":     curr_ce_ltp,
                "pe_ltp":     curr_pe_ltp,
            })

    # ── IV Change detection (sudden IV spike = panic / covering) ──
    iv_changes = []
    iv_map = extras.get("iv_map", {}) if extras else {}
    SPIKE_THRESHOLD = 1.5   # IV change > 1.5% in one tick = spike
    for sp, curr_ivs in iv_map.items():
        prev_ivs = prev_iv.get(sp, {})
        if not prev_ivs:
            continue
        ce_iv_chg = round(curr_ivs.get("ce_iv", 0) - prev_ivs.get("ce_iv", 0), 1)
        pe_iv_chg = round(curr_ivs.get("pe_iv", 0) - prev_ivs.get("pe_iv", 0), 1)
        if abs(ce_iv_chg) >= SPIKE_THRESHOLD or abs(pe_iv_chg) >= SPIKE_THRESHOLD:
            signal = ""
            if pe_iv_chg >= SPIKE_THRESHOLD:
                signal = "PANIC BUYING PUTS — fear spike (BEARISH)"
            elif pe_iv_chg <= -SPIKE_THRESHOLD:
                signal = "PUT IV COOLING — fear fading (BULLISH)"
            elif ce_iv_chg >= SPIKE_THRESHOLD:
                signal = "CALL IV SPIKE — short covering rally (BULLISH)"
            elif ce_iv_chg <= -SPIKE_THRESHOLD:
                signal = "CALL IV COOLING — rally fading (BEARISH)"
            iv_changes.append({
                "strike":    sp,
                "ce_iv_chg": ce_iv_chg,
                "pe_iv_chg": pe_iv_chg,
                "signal":    signal,
                "is_atm":    sp == atm,
            })
    iv_changes.sort(key=lambda x: max(abs(x["ce_iv_chg"]), abs(x["pe_iv_chg"])), reverse=True)

    # ── PCR Change ──
    pcr_change = None
    pcr_direction = "—"
    if prev_pcr is not None:
        delta = round(pcr_all - prev_pcr, 3)
        pcr_direction = f"PCR {pcr_all:.2f} ({'+' if delta>=0 else ''}{delta:.3f} from last tick)"
        if delta > 0.05:
            pcr_direction += " ↑ BULLISH momentum (more put writing added)"
        elif delta < -0.05:
            pcr_direction += " ↓ BEARISH momentum (puts being shed)"
        pcr_change = {"delta": delta, "current": pcr_all, "label": pcr_direction}

    # ── ATM Momentum Signal ──
    atm_momentum = None
    atm_ce_ltp = ce_ltp_map.get(atm, 0)
    atm_pe_ltp = pe_ltp_map.get(atm, 0)
    prev_atm_ce = prev_ltp.get(atm, {}).get("ce_ltp", 0)
    prev_atm_pe = prev_ltp.get(atm, {}).get("pe_ltp", 0)
    atm_ce_oi = ce_map.get(atm, 0)
    atm_pe_oi = pe_map.get(atm, 0)
    prev_atm_ce_oi = prev_strike.get(atm, {}).get("ce_oi", atm_ce_oi)
    prev_atm_pe_oi = prev_strike.get(atm, {}).get("pe_oi", atm_pe_oi)

    if prev_atm_ce and prev_atm_pe:
        ce_ltp_chg = atm_ce_ltp - prev_atm_ce
        pe_ltp_chg = atm_pe_ltp - prev_atm_pe
        ce_oi_chg  = atm_ce_oi  - prev_atm_ce_oi
        pe_oi_chg  = atm_pe_oi  - prev_atm_pe_oi

        # CE momentum: price rising + OI rising = real buyers (BUY CE)
        # PE momentum: price rising + OI rising = real buyers (BUY PE)
        ce_momentum = 0
        pe_momentum = 0
        if ce_ltp_chg > 0 and ce_oi_chg > 0:   ce_momentum += 40   # CE long buildup
        if ce_ltp_chg > 0 and ce_oi_chg < 0:   ce_momentum += 20   # short covering
        if pe_ltp_chg < 0 and pe_oi_chg > 0:   ce_momentum += 25   # put writing = bullish
        if pe_ltp_chg < 0 and pe_oi_chg < 0:   ce_momentum += 10   # put long unwinding

        if pe_ltp_chg > 0 and pe_oi_chg > 0:   pe_momentum += 40   # PE long buildup
        if pe_ltp_chg > 0 and pe_oi_chg < 0:   pe_momentum += 20   # short covering
        if ce_ltp_chg < 0 and ce_oi_chg > 0:   pe_momentum += 25   # call writing = bearish
        if ce_ltp_chg < 0 and ce_oi_chg < 0:   pe_momentum += 10   # call unwinding

        ce_momentum = min(ce_momentum, 100)
        pe_momentum = min(pe_momentum, 100)

        if ce_momentum >= 60:
            action = "🚀 BUY CE NOW"
            reason = f"ATM CE: LTP {'+' if ce_ltp_chg>=0 else ''}{ce_ltp_chg:.2f}, OI {'↑' if ce_oi_chg>0 else '↓'} — momentum confirmed"
            target = round(atm_ce_ltp * 1.06, 1)   # +6%
            stop   = round(atm_ce_ltp * 0.97, 1)   # -3%
        elif pe_momentum >= 60:
            action = "🔻 BUY PE NOW"
            reason = f"ATM PE: LTP {'+' if pe_ltp_chg>=0 else ''}{pe_ltp_chg:.2f}, OI {'↑' if pe_oi_chg>0 else '↓'} — momentum confirmed"
            target = round(atm_pe_ltp * 1.06, 1)
            stop   = round(atm_pe_ltp * 0.97, 1)
        else:
            action = "⏳ WAIT"
            reason = f"ATM CE score {ce_momentum} PE score {pe_momentum} — no clear momentum at ATM yet"
            target = stop = None

        atm_momentum = {
            "action": action, "reason": reason,
            "ce_momentum": ce_momentum, "pe_momentum": pe_momentum,
            "atm": atm, "ce_ltp": atm_ce_ltp, "pe_ltp": atm_pe_ltp,
            "target": target, "stop": stop,
            "ce_ltp_chg": round(ce_ltp_chg, 2), "pe_ltp_chg": round(pe_ltp_chg, 2),
        }

    return {
        "strike_buildups": strike_buildups,
        "iv_changes":      iv_changes[:5],
        "pcr_change":      pcr_change,
        "atm_momentum":    atm_momentum,
    }


def print_oi_signals(sig):
    print("\n--- MARKET DIRECTION SIGNAL ---")
    print(f"  Signal       : {sig['market_signal']}")
    print(f"  Bull Score   : {sig['bull_score_v2']}/100")
    print(f"  Bear Score   : {sig['bear_score_v2']}/100")
    print(f"  Momentum     : {sig['momentum_score']}/100")
    print("\n  Factor Breakdown:")
    for s in sig["signal_list"]:
        icon = "🟢" if s["dir"] == "bull" else "🔴" if s["dir"] == "bear" else "⬜"
        pts  = f"+{s['pts']}" if s["pts"] > 0 else "  "
        print(f"  {icon} [{s['key']:12}] {pts:5} | {s['label']}")
    cw = [x for x in sig.get("call_writing", []) if x.get("tag") == "CONFIRMED"]
    pw = [x for x in sig.get("put_writing",  []) if x.get("tag") == "CONFIRMED"]
    if cw:
        print("\n  ✓ CALL WRITING CONFIRMED:")
        for x in cw[:3]:
            print(f"    Strike {x['strike']} | OI+{x['oi_change']/1e3:.0f}K | ₹{x['ltp']} LTP{x['ltp_chg']:+.2f}")
    if pw:
        print("\n  ✓ PUT WRITING CONFIRMED:")
        for x in pw[:3]:
            print(f"    Strike {x['strike']} | OI+{x['oi_change']/1e3:.0f}K | ₹{x['ltp']} LTP{x['ltp_chg']:+.2f}")


def compute_breakout_signal(m):
    """
    Detect if spot has crossed resistance or support with significantly high volume.
    Returns dict with 'resistance_breakout' and 'support_breakdown' keys (or None if not active).
    """
    price         = m.get("price", 0)
    ce_vol_map    = m.get("ce_vol_map", {})
    pe_vol_map    = m.get("pe_vol_map", {})
    res_levels    = m.get("resistance", [])   # sorted by CE OI desc
    sup_levels    = m.get("support",    [])   # sorted by PE OI desc

    result = {"resistance_breakout": None, "support_breakdown": None}
    if not price or not ce_vol_map or not pe_vol_map:
        return result

    avg_ce = sum(ce_vol_map.values()) / len(ce_vol_map) if ce_vol_map else 0
    avg_pe = sum(pe_vol_map.values()) / len(pe_vol_map) if pe_vol_map else 0

    # Resistance breakout: spot above the primary resistance wall
    if res_levels and avg_ce > 0:
        res_strike = res_levels[0]
        if price > res_strike:
            vol_at   = ce_vol_map.get(res_strike, 0)
            vol_ratio = round(vol_at / avg_ce, 2)
            confirmed = vol_ratio >= 1.5
            result["resistance_breakout"] = {
                "strike":    res_strike,
                "vol_ratio": vol_ratio,
                "confirmed": confirmed,
                "label":     "CONFIRMED" if confirmed else "TENTATIVE",
            }

    # Support breakdown: spot below the primary support floor
    if sup_levels and avg_pe > 0:
        sup_strike = sup_levels[0]
        if price < sup_strike:
            vol_at   = pe_vol_map.get(sup_strike, 0)
            vol_ratio = round(vol_at / avg_pe, 2)
            confirmed = vol_ratio >= 1.5
            result["support_breakdown"] = {
                "strike":    sup_strike,
                "vol_ratio": vol_ratio,
                "confirmed": confirmed,
                "label":     "CONFIRMED" if confirmed else "TENTATIVE",
            }

    return result


def write_oi_snapshot(m, wa, sentiment, pcr_all, pcr_atm, extras=None, signals=None, strike_sig=None, breakout_sig=None):
    """Write a compact OI snapshot JSON for MOMENTUM_AUTO_BOT to consume."""
    snapshot = {
        "time": m["time"],
        "timestamp": time.time(),
        "price": m["price"],
        "atm": m["atm"],
        "sentiment": sentiment,
        "pcr_all": round(pcr_all, 3),
        "pcr_atm": round(pcr_atm, 3),
        "total_oi_ce": m["total_oi_ce"],
        "total_oi_pe": m["total_oi_pe"],
        "total_chg_ce": m["total_chg_ce"],
        "total_chg_pe": m["total_chg_pe"],
        "resistance": m["resistance"],
        "support": m["support"],
        # Strength: [(strike, oi_at_strike, total_oi_at_strike)]
        "resistance_strength": [{"strike": s, "ce_oi": c, "total_oi": t}
                                 for s, c, t in m.get("resistance_strength", [])],
        "support_strength":    [{"strike": s, "pe_oi": p, "total_oi": t}
                                 for s, p, t in m.get("support_strength", [])],
        "atm_strikes_oi": {str(k): v for k, v in m["atm_strikes_oi"].items()},
        "writer_bias": wa["writer_bias"] if wa else "NEUTRAL",
        "bullish_score": wa["bullish_score"] if wa else 0,
        "bearish_score": wa["bearish_score"] if wa else 0,
        "ce_writing_strikes": [s for s, _ in wa["ce_writing"][:3]] if wa else [],
        "pe_writing_strikes": [s for s, _ in wa["pe_writing"][:3]] if wa else [],
        # v3 extras: IV, Volume, Max Pain, Smart Money
        "max_pain":      extras.get("max_pain", 0)       if extras else 0,
        "vol_pcr":       extras.get("vol_pcr", 0)        if extras else 0,
        "total_ce_vol":  extras.get("total_ce_vol", 0)   if extras else 0,
        "total_pe_vol":  extras.get("total_pe_vol", 0)   if extras else 0,
        "atm_ce_iv":     extras.get("atm_ce_iv", 0)      if extras else 0,
        "atm_pe_iv":     extras.get("atm_pe_iv", 0)      if extras else 0,
        "iv_skew":       extras.get("iv_skew", 0)        if extras else 0,
        "atm_extras":    {str(k): v for k, v in extras.get("atm_extras", {}).items()} if extras else {},
        "smart_money_ce": extras.get("smart_money_ce", []) if extras else [],
        "smart_money_pe": extras.get("smart_money_pe", []) if extras else [],
        # 10-factor market direction signal
        "market_signal":  signals.get("market_signal",  "NEUTRAL") if signals else "NEUTRAL",
        "bull_score_v2":  signals.get("bull_score_v2",  0)         if signals else 0,
        "bear_score_v2":  signals.get("bear_score_v2",  0)         if signals else 0,
        "momentum_score": signals.get("momentum_score", 0)         if signals else 0,
        "signal_list":    signals.get("signal_list",    [])        if signals else [],
        "call_writing":   signals.get("call_writing",   [])        if signals else [],
        "put_writing":    signals.get("put_writing",    [])        if signals else [],
        # Per-strike buildup, IV change, PCR change, ATM momentum
        "strike_buildups": strike_sig.get("strike_buildups", []) if strike_sig else [],
        "iv_changes":      strike_sig.get("iv_changes",      []) if strike_sig else [],
        "pcr_change":      strike_sig.get("pcr_change",      None) if strike_sig else None,
        "atm_momentum":    strike_sig.get("atm_momentum",    None) if strike_sig else None,
        # Volume-backed breakout / breakdown detection
        "resistance_breakout": breakout_sig.get("resistance_breakout") if breakout_sig else None,
        "support_breakdown":   breakout_sig.get("support_breakdown")   if breakout_sig else None,
    }
    try:
        with open("oi_snapshot.json", "w") as f:
            json.dump(snapshot, f, indent=2)
    except Exception as e:
        print(f"⚠️  Could not write oi_snapshot.json: {e}")


# ================= CONDITIONAL FORMATTING =================
def apply_conditional_formatting(ws):
    ws.conditional_formatting._cf_rules.clear()

    ws.conditional_formatting.add(
        "M2:M10000",
        ColorScaleRule(
            start_type="num", start_value=0.7, start_color="FF6666",
            mid_type="num", mid_value=1.0, mid_color="FFFF99",
            end_type="num", end_value=1.3, end_color="66FF66"
        )
    )

    ws.conditional_formatting.add(
        "N2:N10000",
        ColorScaleRule(
            start_type="num", start_value=0.7, start_color="FF6666",
            mid_type="num", mid_value=1.0, mid_color="FFFF99",
            end_type="num", end_value=1.3, end_color="66FF66"
        )
    )

    ws.conditional_formatting.add(
        "Q2:Q10000",
        ColorScaleRule(
            start_type="num", start_value=40, start_color="66FF66",
            mid_type="num", mid_value=50, mid_color="FFFF99",
            end_type="num", end_value=65, end_color="FF6666"
        )
    )

    ws.conditional_formatting.add(
        "R2:R10000",
        ColorScaleRule(
            start_type="num", start_value=40, start_color="FF6666",
            mid_type="num", mid_value=50, mid_color="FFFF99",
            end_type="num", end_value=65, end_color="66FF66"
        )
    )

    # === SENTIMENT ===
    ws.conditional_formatting.add("Y2:Y10000",
                                  FormulaRule(formula=['$Y2="BULLISH"'],
                                              fill=PatternFill("solid", fgColor="00C853")))

    ws.conditional_formatting.add("Y2:Y10000",
                                  FormulaRule(formula=['$Y2="BEARISH"'],
                                              fill=PatternFill("solid", fgColor="D50000")))

    ws.conditional_formatting.add("Y2:Y10000",
                                  FormulaRule(formula=['$Y2="NEUTRAL"'],
                                              fill=PatternFill("solid", fgColor="FFD54F")))

    ws.conditional_formatting.add(
        "L2:L10000",
        ColorScaleRule(
            start_type="num", start_value=0.5, start_color="FFEB84",
            mid_type="num", mid_value=1.0, mid_color="9BC2E6",
            end_type="num", end_value=2.0, end_color="63BE7B"
        )
    )

    # === TOTAL OI CE vs PE (D & E) ===
    ws.conditional_formatting.add("D2:D10000",
                                  FormulaRule(formula=["$E2>$D2"],
                                              fill=PatternFill("solid", fgColor="F8696B")))

    ws.conditional_formatting.add("E2:E10000",
                                  FormulaRule(formula=["$E2>$D2"],
                                              fill=PatternFill("solid", fgColor="63BE7B")))

    ws.conditional_formatting.add("D2:D10000",
                                  FormulaRule(formula=["$D2>$E2"],
                                              fill=PatternFill("solid", fgColor="63BE7B")))

    ws.conditional_formatting.add("E2:E10000",
                                  FormulaRule(formula=["$D2>$E2"],
                                              fill=PatternFill("solid", fgColor="F8696B")))

    ws.conditional_formatting.add("D2:E10000",
                                  FormulaRule(formula=["ABS($D2-$E2)/MAX($D2,$E2)<0.05"],
                                              fill=PatternFill("solid", fgColor="FFF59D")))

    ws.conditional_formatting.add(
        "M2:O10000",
        ColorScaleRule(
            start_type="num", start_value=30, start_color="63BE7B",
            mid_type="num", mid_value=50, mid_color="FFEB84",
            end_type="num", end_value=70, end_color="F8696B"
        )
    )

    ws.conditional_formatting.add(
        "N2:P10000",
        ColorScaleRule(
            start_type="num", start_value=30, start_color="F8696B",
            mid_type="num", mid_value=50, mid_color="FFEB84",
            end_type="num", end_value=70, end_color="63BE7B"
        )
    )

    ws.conditional_formatting.add(
        "S2:T10000",
        FormulaRule(
            formula=['S2="BUYERS"'],
            fill=PatternFill("solid", fgColor="63BE7B")
        )
    )

    ws.conditional_formatting.add(
        "S2:T10000",
        FormulaRule(
            formula=['S2="WRITERS"'],
            fill=PatternFill("solid", fgColor="F8696B")
        )
    )


# ================= EXCEL INIT =================
def init_excel():
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "DATA"

        ws.append([
            "Time","Price","ATM",
            "Total OI CE","Total OI PE",
            "Total Chg CE","Total Chg PE",
            "ATM OI CE","ATM OI PE",
            "ATM Chg CE","ATM Chg PE",
            "PCR All","PCR ATM ±3","PCR Chg ATM",
            "COI CE % All","COI PE % All",
            "COI CE % ATM","COI PE % ATM",
            "CE Activity","PE Activity",
            "CE Power","PE Power",
            "Resistance","Support","Sentiment"
        ])

        apply_conditional_formatting(ws)
        wb.save(EXCEL_FILE)

    return wb

# ================= CALCULATION =================
def calculate(j):
    price = j["underlyingValue"]
    data = j["data"]

    strikes = sorted(i["strikePrice"] for i in data if "strikePrice" in i)
    if not strikes:
        raise ValueError("No strike data returned — expiry date may be wrong or market is closed.")
    atm = min(strikes, key=lambda x: abs(x - price))
    idx = strikes.index(atm)
    atm_range = strikes[max(0, idx-3): idx+4]

    t_oi_ce = t_oi_pe = t_chg_ce = t_chg_pe = 0
    a_oi_ce = a_oi_pe = a_chg_ce = a_chg_pe = 0

    ce_map, pe_map = {}, {}
    ce_chg_map, pe_chg_map = {}, {}
    ce_ltp_map, pe_ltp_map = {}, {}
    ce_vol_map, pe_vol_map = {}, {}

    for i in data:
        s = i["strikePrice"]
        ce = i.get("CE", {})
        pe = i.get("PE", {})

        ce_oi = ce.get("openInterest", 0) * LOT_SIZE
        pe_oi = pe.get("openInterest", 0) * LOT_SIZE
        ce_chg = ce.get("changeinOpenInterest", 0) * LOT_SIZE
        pe_chg = pe.get("changeinOpenInterest", 0) * LOT_SIZE

        ce_map[s] = ce_oi
        pe_map[s] = pe_oi
        ce_chg_map[s] = ce_chg
        pe_chg_map[s] = pe_chg
        ce_ltp_map[s] = ce.get("lastPrice", 0)
        pe_ltp_map[s] = pe.get("lastPrice", 0)
        ce_vol_map[s] = ce.get("totalTradedVolume", 0)
        pe_vol_map[s] = pe.get("totalTradedVolume", 0)

        t_oi_ce += ce_oi
        t_oi_pe += pe_oi
        t_chg_ce += ce_chg
        t_chg_pe += pe_chg

        if s in atm_range:
            a_oi_ce += ce_oi
            a_oi_pe += pe_oi
            a_chg_ce += ce_chg
            a_chg_pe += pe_chg

    atm_strikes_oi = {s: {"ce_oi": ce_map.get(s, 0), "pe_oi": pe_map.get(s, 0)} for s in atm_range}

    # Calculate level strength for resistance/support
    resistance_levels = sorted(ce_map, key=ce_map.get, reverse=True)[:3]
    support_levels = sorted(pe_map, key=pe_map.get, reverse=True)[:3]
    resistance_strength = [(level, ce_map[level], ce_map[level] + pe_map.get(level, 0)) for level in resistance_levels]
    support_strength = [(level, pe_map[level], pe_map[level] + ce_map.get(level, 0)) for level in support_levels]

    return {
        "time": datetime.now().strftime("%H:%M:%S"),
        "price": price,
        "atm": atm,
        "total_oi_ce": t_oi_ce,
        "total_oi_pe": t_oi_pe,
        "total_chg_ce": t_chg_ce,
        "total_chg_pe": t_chg_pe,
        "atm_oi_ce": a_oi_ce,
        "atm_oi_pe": a_oi_pe,
        "atm_chg_ce": a_chg_ce,
        "atm_chg_pe": a_chg_pe,
        "resistance": resistance_levels,
        "support": support_levels,
        "resistance_strength": resistance_strength,
        "support_strength": support_strength,
        "atm_strikes_oi": atm_strikes_oi,
        "ce_map": ce_map,
        "pe_map": pe_map,
        "ce_chg_map": ce_chg_map,
        "pe_chg_map": pe_chg_map,
        "ce_ltp_map": ce_ltp_map,
        "pe_ltp_map": pe_ltp_map,
        "ce_vol_map": ce_vol_map,
        "pe_vol_map": pe_vol_map,
    }

# ================= MAIN LOOP =================

wb = init_excel()
ws = wb["DATA"]

# Store 15-min candle closes for breakout/breakdown detection
candle_history = {}
candle_interval = 900  # 15 min in seconds
last_candle_close_time = None

atm_strike_history = {}  # {strike: [(time, ce_oi, pe_oi), ...]}

while True:
    try:
        j = fetch_data()
        m = calculate(j)
        extras = fetch_v3_extras(m["atm"])   # IV, Vol PCR, Max Pain from v3 API

        pcr_all = m["total_oi_pe"] / m["total_oi_ce"] if m["total_oi_ce"] else 0
        pcr_atm = m["atm_oi_pe"] / m["atm_oi_ce"] if m["atm_oi_ce"] else 0
        pcr_chg = abs(m["atm_chg_pe"]) / abs(m["atm_chg_ce"]) if m["atm_chg_ce"] else 0

        coi_ce_all, coi_pe_all = coi_pct(m["total_chg_ce"], m["total_chg_pe"])
        coi_ce_atm, coi_pe_atm = coi_pct(m["atm_chg_ce"], m["atm_chg_pe"])

        ce_act = activity(m["price"], prev_price, m["atm_oi_ce"], prev_ce_oi)
        pe_act = activity(m["price"], prev_price, m["atm_oi_pe"], prev_pe_oi)

        ce_pow = power(m["price"], prev_price, m["atm_chg_ce"])
        pe_pow = power(m["price"], prev_price, m["atm_chg_pe"])

        sentiment = "BULLISH" if pcr_atm > 1.1 else "BEARISH" if pcr_atm < 0.9 else "NEUTRAL"

        # Store ATM strike history (last 5)
        atm = m["atm"]
        atm_strike_history.setdefault(atm, []).append((m["time"], m["atm_strikes_oi"][atm]["ce_oi"], m["atm_strikes_oi"][atm]["pe_oi"]))
        if len(atm_strike_history[atm]) > 5:
            atm_strike_history[atm].pop(0)

        # ================= PRINT =================
        print("\n" + "="*75)
        print(f"Time                         : {m['time']}")
        print(f"Current Market Price         : {m['price']}")
        print(f"ATM Strike                   : {m['atm']}")

        print("\n--- ALL STRIKES ---")
        print(f"1. Total OI CE                : {m['total_oi_ce']:,}")
        print(f"2. Total OI PE                : {m['total_oi_pe']:,}")
        print(f"3. Total Change OI CE         : {m['total_chg_ce']:,}")
        print(f"4. Total Change OI PE         : {m['total_chg_pe']:,}")

        print("\n--- ATM ±3 STRIKES ---")
        print(f"5. Total OI CE (ATM ±3)       : {m['atm_oi_ce']:,}")
        print(f"6. Total OI PE (ATM ±3)       : {m['atm_oi_pe']:,}")
        print(f"7. Change OI CE (ATM ±3)      : {m['atm_chg_ce']:,}")
        print(f"8. Change OI PE (ATM ±3)      : {m['atm_chg_pe']:,}")

        print("\n  Strike-wise OI (ATM ±3):")
        print(f"  {'Strike':>8} | {'CE OI':>14} | {'PE OI':>14} | {'Diff (PE-CE)':>16}")
        print(f"  {'-'*8}-+-{'-'*14}-+-{'-'*14}-+-{'-'*16}")
        for strike in sorted(m["atm_strikes_oi"]):
            marker = " <-- ATM" if strike == m["atm"] else ""
            d = m["atm_strikes_oi"][strike]
            diff = d["pe_oi"] - d["ce_oi"]
            print(f"  {strike:>8} | {d['ce_oi']:>14,} | {d['pe_oi']:>14,} | {diff:>+16,}{marker}")

        hist = atm_strike_history.get(m["atm"], [])
        if len(hist) > 1:
            print(f"\n  ATM Strike {m['atm']} — Last {len(hist)} readings:")
            print(f"  {'Time':>10} | {'CE OI':>14} | {'PE OI':>14} | {'Diff (PE-CE)':>16}")
            print(f"  {'-'*10}-+-{'-'*14}-+-{'-'*14}-+-{'-'*16}")
            for t, c, p in hist:
                print(f"  {t:>10} | {c:>14,} | {p:>14,} | {(p - c):>+16,}")

        print("\n--- PCR ---")
        print(f"9. PCR (All Strikes)          : {pcr_all:.2f}")
        print(f"10. PCR (ATM ±3)              : {pcr_atm:.2f}")
        print(f"11. PCR Change OI (ATM ±3)    : {pcr_chg:.2f}")

        print("\n--- COI IMBALANCE (%) ---")
        print(f"12. CALL COI % (Overall)      : {coi_ce_all:.2f}%")
        print(f"13. PUT  COI % (Overall)      : {coi_pe_all:.2f}%")
        print(f"14. CALL COI % (ATM ±3)       : {coi_ce_atm:.2f}%")
        print(f"15. PUT  COI % (ATM ±3)       : {coi_pe_atm:.2f}%")

        print("\n--- MARKET SENTIMENT ---")
        print(f"SENTIMENT                    : {sentiment}")

        if extras:
            print("\n--- IV + VOLUME + MAX PAIN ---")
            mp   = extras.get("max_pain", 0)
            spot = m["price"]
            mp_dist = f"{spot - mp:+.0f} pts {'above' if spot > mp else 'below'} max pain"
            print(f"Max Pain                     : {mp}  ({mp_dist})")
            print(f"Volume PCR                   : {extras.get('vol_pcr', 0):.2f}  "
                  f"(CE vol: {extras.get('total_ce_vol',0):,}  PE vol: {extras.get('total_pe_vol',0):,})")
            print(f"ATM IV  CE: {extras.get('atm_ce_iv',0):.1f}%   "
                  f"PE: {extras.get('atm_pe_iv',0):.1f}%   "
                  f"Skew (CE-PE): {extras.get('iv_skew',0):+.1f}%  "
                  f"{'↑ upside fear' if extras.get('iv_skew',0) > 1 else '↓ downside fear' if extras.get('iv_skew',0) < -1 else '≈ neutral'}")
            atm_ex = extras.get("atm_extras", {})
            if atm_ex:
                print(f"\n  {'Strike':>8} | {'CE IV':>7} | {'PE IV':>7} | {'CE LTP':>9} | {'PE LTP':>9}")
                print(f"  {'-'*8}-+-{'-'*7}-+-{'-'*7}-+-{'-'*9}-+-{'-'*9}")
                for sp in sorted(atm_ex):
                    ex = atm_ex[sp]
                    mk = " ← ATM" if sp == m["atm"] else ""
                    print(f"  {sp:>8} | {ex.get('ce_iv',0):>6.1f}% | {ex.get('pe_iv',0):>6.1f}% | ₹{ex.get('ce_ltp',0):>8.2f} | ₹{ex.get('pe_ltp',0):>8.2f}{mk}")

        # === LEVELS & STRENGTH ===
        print("\n--- RESISTANCE LEVELS & STRENGTH ---")
        for level, ce_oi, total_oi in m["resistance_strength"]:
            print(f"Resistance: {level} | CE OI: {ce_oi:,} | Total OI: {total_oi:,}")
        print("\n--- SUPPORT LEVELS & STRENGTH ---")
        for level, pe_oi, total_oi in m["support_strength"]:
            print(f"Support: {level} | PE OI: {pe_oi:,} | Total OI: {total_oi:,}")


        # === Live signal update on each run ===
        breakout_live = None
        breakdown_live = None
        breakout_prob = 0
        breakdown_prob = 0
        signal_strength = {}
        signal_consistency = {}

        # Calculate distance to resistance/support and probability
        for level, ce_oi, total_oi in m["resistance_strength"]:
            dist = m["price"] - level
            strength = ce_oi
            consistency = min(1.0, ce_oi / (total_oi + 1e-6))
            prob = max(0, min(1, 1 - abs(dist) / (level * 0.01) + consistency * 0.5))  # crude estimate
            signal_strength[level] = strength
            signal_consistency[level] = consistency
            if dist > 0:
                breakout_live = (level, dist, strength, consistency, prob)
                breakout_prob = prob
        for level, pe_oi, total_oi in m["support_strength"]:
            dist = m["price"] - level
            strength = pe_oi
            consistency = min(1.0, pe_oi / (total_oi + 1e-6))
            prob = max(0, min(1, 1 - abs(dist) / (level * 0.01) + consistency * 0.5))
            signal_strength[level] = strength
            signal_consistency[level] = consistency
            if dist < 0:
                breakdown_live = (level, dist, strength, consistency, prob)
                breakdown_prob = prob

        # Print live signal info
        print("\n--- LIVE SIGNAL UPDATE ---")
        if breakout_live:
            print(f"Breakout above {breakout_live[0]} | Distance: {breakout_live[1]:.2f} | Strength: {breakout_live[2]:,} | Consistency: {breakout_live[3]*100:.1f}% | Probability: {breakout_live[4]*100:.1f}%")
        else:
            print("No breakout imminent.")
        if breakdown_live:
            print(f"Breakdown below {breakdown_live[0]} | Distance: {breakdown_live[1]:.2f} | Strength: {breakdown_live[2]:,} | Consistency: {breakdown_live[3]*100:.1f}% | Probability: {breakdown_live[4]*100:.1f}%")
        else:
            print("No breakdown imminent.")

        # === 15-min candle close breakout/breakdown detection ===
        now = time.time()
        breakout_signal = None
        breakdown_signal = None
        if last_candle_close_time is None or now - last_candle_close_time >= candle_interval:
            for level in m["resistance"] + m["support"]:
                candle_history.setdefault(level, []).append(m["price"])
            last_candle_close_time = now
            for level in m["resistance"]:
                if candle_history[level][-1] > level:
                    breakout_signal = (level, m["price"])
            for level in m["support"]:
                if candle_history[level][-1] < level:
                    breakdown_signal = (level, m["price"])
            if breakout_signal:
                print(f"\n🚀 BREAKOUT SIGNAL: Price closed above resistance {breakout_signal[0]} | Market Direction: UP | Target: Next resistance")
            if breakdown_signal:
                print(f"\n🔻 BREAKDOWN SIGNAL: Price closed below support {breakdown_signal[0]} | Market Direction: DOWN | Target: Next support")

        # === AI Trading Summary ===
        price_change = (m["price"] - prev_price) if prev_price is not None else 0.0
        print("\n--- AI OI SUMMARY ---")
        print(generate_trading_summary(m, price_change))

        # === Writer Activity Analysis ===
        wa = analyze_writer_activity(m["ce_map"], m["pe_map"], prev_strike_data)
        print_writer_activity(wa, m["price"], m["atm"])

        # === 10-Factor Market Direction Signal ===
        sig = compute_oi_signals(m, wa, extras, pcr_all, prev_price, prev_ltp_data)
        print_oi_signals(sig)

        # === Per-strike buildup / IV change / PCR change / ATM momentum ===
        strike_sig = compute_strike_signals(
            m, extras, prev_strike_data, prev_ltp_data, prev_iv_data, prev_pcr_all, pcr_all
        )

        # === Update per-strike baselines for next tick ===
        prev_strike_data.clear()
        for s in set(m["ce_map"]) | set(m["pe_map"]):
            prev_strike_data[s] = {
                "ce_oi": m["ce_map"].get(s, 0),
                "pe_oi": m["pe_map"].get(s, 0),
            }
        prev_ltp_data.clear()
        for s in set(m["ce_ltp_map"]) | set(m["pe_ltp_map"]):
            prev_ltp_data[s] = {
                "ce_ltp": m["ce_ltp_map"].get(s, 0),
                "pe_ltp": m["pe_ltp_map"].get(s, 0),
            }
        if extras and "iv_map" in extras:
            prev_iv_data.clear()
            prev_iv_data.update(extras["iv_map"])
        prev_pcr_all = pcr_all

        # === Write OI snapshot for MOMENTUM_AUTO_BOT ===
        breakout_sig = compute_breakout_signal(m)
        write_oi_snapshot(m, wa, sentiment, pcr_all, pcr_atm, extras=extras, signals=sig, strike_sig=strike_sig, breakout_sig=breakout_sig)

        # === Save to Excel ===
        ws.append([
            m["time"], m["price"], m["atm"],
            m["total_oi_ce"], m["total_oi_pe"],
            m["total_chg_ce"], m["total_chg_pe"],
            m["atm_oi_ce"], m["atm_oi_pe"],
            m["atm_chg_ce"], m["atm_chg_pe"],
            pcr_all, pcr_atm, pcr_chg,
            coi_ce_all, coi_pe_all,
            coi_ce_atm, coi_pe_atm,
            ce_act, pe_act, ce_pow, pe_pow,
            str(m["resistance"]), str(m["support"]), sentiment,
            str(m["resistance_strength"]), str(m["support_strength"]),
            "BREAKOUT" if breakout_signal else "", "BREAKDOWN" if breakdown_signal else "",
            breakout_live[0] if breakout_live else "", breakout_live[1] if breakout_live else "", breakout_live[2] if breakout_live else "", breakout_live[3] if breakout_live else "", breakout_live[4] if breakout_live else "",
            breakdown_live[0] if breakdown_live else "", breakdown_live[1] if breakdown_live else "", breakdown_live[2] if breakdown_live else "", breakdown_live[3] if breakdown_live else "", breakdown_live[4] if breakdown_live else ""
        ])

        wb.save(EXCEL_FILE)

        prev_price = m["price"]
        prev_ce_oi = m["atm_oi_ce"]
        prev_pe_oi = m["atm_oi_pe"]

        time.sleep(REFRESH_SECONDS)

    except Exception as e:
        print("Error:", e)
        time.sleep(10)
